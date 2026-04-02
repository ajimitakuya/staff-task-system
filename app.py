from common import now_jst, mask_secret_text, safe_text, heart_label, parse_time_range, _to_minutes, _normalize_weekday_label, is_time_overlap, get_saturday_dates_for_month, get_sheet_name_candidates, get_sheet_name, get_next_numeric_id, normalize_company_scoped_df, filter_by_company_id
from data_access import load_db, save_db, get_companies_df, get_users_df, get_user_company_permissions_df, get_task_required_cols, get_tasks_df, get_urgent_tasks_df, get_resident_master_df, get_resident_schedule_df, get_resident_notes_df, get_attendance_logs_df, get_attendance_logs_df, get_attendance_display_settings_df
import streamlit as st
import pandas as pd
import base64
import time
import random
import streamlit.components.v1 as components
import json
from io import BytesIO
from datetime import datetime, timedelta, timezone, date
import calendar as py_calendar
from openpyxl import load_workbook
from streamlit_gsheets import GSheetsConnection
from streamlit_calendar import calendar as st_calendar
import google.generativeai as genai
import tempfile
from contextlib import contextmanager
from openpyxl import Workbook
from run_assistance import (
    build_chrome_driver,
    get_knowbe_login_credentials,
    manual_login_wait,
    fetch_support_record_text_for_month,
    run_support_record_kind_export,
)

JST = timezone(timedelta(hours=9))

def get_genai_client():
    api_key = ""

    try:
        api_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        api_key = ""

    if not api_key:
        import os
        api_key = os.environ.get("GEMINI_API_KEY", "")

    if not api_key:
        raise ValueError("GEMINI_API_KEY が設定されていません")

    import google.generativeai as genai
    genai.configure(api_key=api_key)

    return genai

# --- ページ基本設定 ---
st.set_page_config(page_title="作業管理システム", layout="wide")
# st.caption("APP_VERSION = 2026-03-21-knowbe-debug-01")

# --- 🔌 スプレッドシート接続設定 ---
conn = st.connection("gsheets", type=GSheetsConnection)

COMMON_SHEETS = {
    "active_users",
    "companies",
    "users",
    "user_company_permissions",
    "admin_logs",
    "attendance_logs",
    "attendance_display_settings",    
}

COMPANY_SCOPED_SHEETS = {
    "task",
    "chat",
    "manual",
    "record_status",
    "calendar",
    "chat_rooms",
    "chat_messages",
    "warehouse_files",
    "archive_files",
    "resident_master",
    "resident_schedule",
    "resident_notes",
    "document_master",
    "external_contacts",
    "resident_links",
    "saved_documents",
    "diary_input_rules",
    "staff_examples",
    "personal_rules",
    "assistant_plans",
    "piecework_master",
    "piecework_entries",
    "piecework_production",
    "piecework_clients",
}

def render_attendance_page():
    st.header("勤怠管理")

    if "attendance_mode" not in st.session_state:
        st.session_state["attendance_mode"] = False
    if "attendance_users_df" not in st.session_state:
        st.session_state["attendance_users_df"] = None
    if "attendance_permissions_df" not in st.session_state:
        st.session_state["attendance_permissions_df"] = None
    if "attendance_logs_df" not in st.session_state:
        st.session_state["attendance_logs_df"] = None
    if "attendance_settings_df" not in st.session_state:
        st.session_state["attendance_settings_df"] = None
    if "attendance_companies_df" not in st.session_state:
        st.session_state["attendance_companies_df"] = None

    top_reload_cols = st.columns([1, 1, 4])

    with top_reload_cols[0]:
        if st.button("勤怠管理表示", key="attendance_mode_button", use_container_width=True):
            st.session_state["attendance_mode"] = True

            st.session_state["attendance_users_df"] = get_users_df()
            st.session_state["attendance_permissions_df"] = get_user_company_permissions_df()
            st.session_state["attendance_logs_df"] = get_attendance_logs_df()
            st.session_state["attendance_settings_df"] = get_attendance_display_settings_df()
            st.session_state["attendance_companies_df"] = get_companies_df()

            st.rerun()

    with top_reload_cols[1]:
        if st.button("再読込", key="attendance_reload_button", use_container_width=True):
            st.session_state["attendance_users_df"] = get_users_df()
            st.session_state["attendance_permissions_df"] = get_user_company_permissions_df()
            st.session_state["attendance_logs_df"] = get_attendance_logs_df()
            st.session_state["attendance_settings_df"] = get_attendance_display_settings_df()
            st.session_state["attendance_companies_df"] = get_companies_df()
            st.rerun()

    if not st.session_state.get("attendance_mode", False):
        st.info("「勤怠管理表示」を押すと読み込みます。")
        return

    users_df = st.session_state.get("attendance_users_df")
    permissions_df = st.session_state.get("attendance_permissions_df")
    attendance_logs_df = st.session_state.get("attendance_logs_df")
    settings_df = st.session_state.get("attendance_settings_df")
    companies_df = st.session_state.get("attendance_companies_df")

    if users_df is None or permissions_df is None or attendance_logs_df is None or settings_df is None or companies_df is None:
        st.warning("まだ勤怠データが読み込まれていません。『勤怠管理表示』を押してください。")
        return

    # ===== 必須列補完 =====
    if settings_df is None or settings_df.empty:
        settings_df = pd.DataFrame(columns=[
            "setting_id", "group_id", "slot_no", "company_id",
            "status", "created_at", "registered_by"
        ])
    else:
        for col in ["setting_id", "group_id", "slot_no", "company_id", "status", "created_at", "registered_by"]:
            if col not in settings_df.columns:
                settings_df[col] = ""

    if attendance_logs_df is None or attendance_logs_df.empty:
        attendance_logs_df = pd.DataFrame(columns=[
            "attendance_id", "date", "user_id", "company_id",
            "action", "timestamp", "device_name", "recorded_by"
        ])
    else:
        for col in ["attendance_id", "date", "user_id", "company_id", "action", "timestamp", "device_name", "recorded_by"]:
            if col not in attendance_logs_df.columns:
                attendance_logs_df[col] = ""

    # ===== ログイン中情報 =====
    current_company_id = str(st.session_state.get("company_id", "")).strip()
    current_user_id = str(st.session_state.get("user_id", "")).strip()

    # ===== group_id取得 =====
    group_id = None
    if current_company_id and companies_df is not None and not companies_df.empty:
        row = companies_df[companies_df["company_id"].astype(str).str.strip() == current_company_id]
        if not row.empty and "group_id" in row.columns:
            group_id = str(row.iloc[0]["group_id"]).strip()

    if not group_id:
        st.error("group_id が取得できません")
        return

    st.subheader("事業所登録（最大5件）")

    existing_settings = settings_df[
        (settings_df["group_id"].astype(str).str.strip() == group_id) &
        (settings_df["status"].astype(str).str.strip() == "active")
    ].copy()

    for i in range(1, 6):
        slot_row = existing_settings[
            pd.to_numeric(existing_settings["slot_no"], errors="coerce").fillna(0).astype(int) == i
        ]

        registered_company_name = ""
        registered_company_id = ""

        if not slot_row.empty:
            registered_company_id = str(slot_row.iloc[0]["company_id"]).strip()
            comp_row = companies_df[
                companies_df["company_id"].astype(str).str.strip() == registered_company_id
            ]
            if not comp_row.empty:
                registered_company_name = str(comp_row.iloc[0].get("company_name", "")).strip()

        st.markdown(f"### 事業所{i}")

        if registered_company_id:
            col_a, col_b = st.columns([4, 1])
            with col_a:
                st.success(f"登録済み: {registered_company_name}（{registered_company_id}）")
            with col_b:
                if st.button("削除", key=f"delete_slot_{i}", use_container_width=True):
                    idx = slot_row.index
                    settings_df.loc[idx, "status"] = "inactive"
                    save_db(settings_df, "attendance_display_settings")
                    st.rerun()
        else:
            col1, col2, col3 = st.columns([2, 2, 1])

            with col1:
                login_id = st.text_input(f"事業所ID_{i}", key=f"attendance_id_{i}")

            with col2:
                password = st.text_input(f"事業所PASS_{i}", type="password", key=f"attendance_pass_{i}")

            with col3:
                st.write("")
                if st.button(f"登録_{i}", key=f"attendance_register_{i}", use_container_width=True):
                    comp = companies_df[
                        (companies_df["company_login_id"].astype(str).str.strip() == str(login_id).strip()) &
                        (companies_df["company_login_password"].astype(str).str.strip() == str(password).strip()) &
                        (companies_df["status"].astype(str).str.strip() == "active")
                    ]

                    if comp.empty:
                        st.error("認証失敗")
                    else:
                        cid = str(comp.iloc[0]["company_id"]).strip()

                        same_active = settings_df[
                            (settings_df["group_id"].astype(str).str.strip() == group_id) &
                            (settings_df["company_id"].astype(str).str.strip() == cid) &
                            (settings_df["status"].astype(str).str.strip() == "active")
                        ]
                        if not same_active.empty:
                            st.warning("その事業所はもう登録済みです")
                        else:
                            new_row = {
                                "setting_id": f"ADS{len(settings_df) + 1:04}",
                                "group_id": group_id,
                                "slot_no": i,
                                "company_id": cid,
                                "status": "active",
                                "created_at": now_jst().strftime("%Y-%m-%d %H:%M:%S"),
                                "registered_by": current_user_id,
                            }
                            settings_df = pd.concat([settings_df, pd.DataFrame([new_row])], ignore_index=True)
                            save_db(settings_df, "attendance_display_settings")
                            st.success(f"{cid} を登録しました。")
                            st.rerun()

    target_settings = settings_df[
        (settings_df["group_id"].astype(str).str.strip() == group_id) &
        (settings_df["status"].astype(str).str.strip() == "active")
    ].copy()

    if target_settings.empty:
        st.info("登録済み事業所がまだありません。")
        return

    target_settings["slot_no_num"] = pd.to_numeric(target_settings["slot_no"], errors="coerce").fillna(9999)
    target_settings = target_settings.sort_values("slot_no_num")

    company_ids = target_settings["company_id"].astype(str).str.strip().tolist()

    company_name_map = {}
    for cid in company_ids:
        comp_row = companies_df[companies_df["company_id"].astype(str).str.strip() == cid]
        if not comp_row.empty:
            company_name_map[cid] = str(comp_row.iloc[0].get("company_name", cid)).strip()
        else:
            company_name_map[cid] = cid

    selected_company = st.radio(
        "事業所選択",
        options=company_ids,
        format_func=lambda x: company_name_map.get(x, x),
        horizontal=True,
        key="attendance_selected_company"
    )

    valid_users = users_df
    if "attendance_enabled" in valid_users.columns:
        valid_users = valid_users[
            pd.to_numeric(valid_users["attendance_enabled"], errors="coerce").fillna(0).astype(int) == 1
        ]
    if "status" in valid_users.columns:
        valid_users = valid_users[
            valid_users["status"].astype(str).str.strip() == "active"
        ]

    merged = pd.merge(valid_users, permissions_df, on="user_id", how="inner", suffixes=("", "_perm"))
    merged = merged[
        merged["company_id_perm"].astype(str).str.strip() == selected_company
    ].copy()

    if "can_use_perm" in merged.columns:
        merged = merged[
            pd.to_numeric(merged["can_use_perm"], errors="coerce").fillna(0) == 1
        ]
    elif "can_use" in merged.columns:
        merged = merged[
            merged["can_use"].astype(str).str.strip().isin(["1", "TRUE", "True", "true"])
        ]

    if "display_order" in merged.columns:
        merged["display_order_num"] = pd.to_numeric(merged["display_order"], errors="coerce").fillna(9999)
        merged = merged.sort_values("display_order_num")
    else:
        merged = merged.sort_values("display_name")

    st.markdown("---")

    if "attendance_action_mode" not in st.session_state:
        st.session_state["attendance_action_mode"] = "in"

    if "attendance_local_status" not in st.session_state:
        st.session_state["attendance_local_status"] = {}

    def get_status(user_id):
        user_id = str(user_id).strip()

        local_map = st.session_state.get("attendance_local_status", {})
        if user_id in local_map:
            return local_map[user_id]

        logs = attendance_logs_df[
            (attendance_logs_df["user_id"].astype(str).str.strip() == user_id) &
            (attendance_logs_df["company_id"].astype(str).str.strip() == selected_company)
        ].copy()

        if logs.empty:
            return "退勤"

        if "timestamp" in logs.columns:
            logs = logs.sort_values("timestamp")

        last = logs.iloc[-1]
        last_action = str(last.get("action", "")).strip()

        action_map = {
            "in": "出勤",
            "out": "退勤",
            "break_start": "休憩中",
            "break_end": "出勤",
        }
        return action_map.get(last_action, "退勤")

    st.markdown("""
    <style>
    div[data-testid="stButton"] > button {
        height: 56px !important;
        width: 100% !important;
        font-size: 18px !important;
        font-weight: 600 !important;
        border-radius: 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("### 勤怠モード選択")

    mode_cols = st.columns(4)

    with mode_cols[0]:
        if st.button(
            "出勤",
            key="attendance_mode_in",
            use_container_width=True,
            type="primary" if st.session_state["attendance_action_mode"] == "in" else "secondary",
        ):
            st.session_state["attendance_action_mode"] = "in"
            st.rerun()

    with mode_cols[1]:
        if st.button(
            "退勤",
            key="attendance_mode_out",
            use_container_width=True,
            type="primary" if st.session_state["attendance_action_mode"] == "out" else "secondary",
        ):
            st.session_state["attendance_action_mode"] = "out"
            st.rerun()

    with mode_cols[2]:
        if st.button(
            "休憩開始",
            key="attendance_mode_break_start",
            use_container_width=True,
            type="primary" if st.session_state["attendance_action_mode"] == "break_start" else "secondary",
        ):
            st.session_state["attendance_action_mode"] = "break_start"
            st.rerun()

    with mode_cols[3]:
        if st.button(
            "休憩終了",
            key="attendance_mode_break_end",
            use_container_width=True,
            type="primary" if st.session_state["attendance_action_mode"] == "break_end" else "secondary",
        ):
            st.session_state["attendance_action_mode"] = "break_end"
            st.rerun()

    mode_label_map = {
        "in": "出勤モード",
        "out": "退勤モード",
        "break_start": "休憩開始モード",
        "break_end": "休憩終了モード",
    }

    st.info(f"現在のモード: {mode_label_map.get(st.session_state['attendance_action_mode'], '出勤モード')}")

    users_list = merged.to_dict(orient="records")

    for start_idx in range(0, len(users_list), 5):
        cols = st.columns(5)

        for col_idx in range(5):
            data_idx = start_idx + col_idx

            with cols[col_idx]:
                if data_idx >= len(users_list):
                    st.write("")
                    continue

                row = users_list[data_idx]

                uid = str(row["user_id"]).strip()
                name = str(row.get("display_name", uid)).strip()
                status = get_status(uid)

                button_type = "primary" if status in ["出勤", "休憩中"] else "secondary"
                button_label = f"{name}\n{status}"

                if st.button(
                    button_label,
                    key=f"attendance_user_{selected_company}_{uid}",
                    use_container_width=True,
                    type=button_type,
                ):
                    now = now_jst()
                    action = st.session_state.get("attendance_action_mode", "in")

                    action_text_map = {
                        "in": "出勤",
                        "out": "退勤",
                        "break_start": "休憩中",
                        "break_end": "出勤",
                    }

                    # 先に画面上の状態を即更新
                    st.session_state["attendance_local_status"][uid] = action_text_map.get(action, "退勤")

                    new_log = {
                        "attendance_id": f"A{len(attendance_logs_df) + 1:04}",
                        "date": now.strftime("%Y-%m-%d"),
                        "user_id": uid,
                        "company_id": selected_company,
                        "action": action,
                        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
                        "device_name": "tablet",
                        "recorded_by": current_user_id,
                    }

                    attendance_logs_df = pd.concat(
                        [attendance_logs_df, pd.DataFrame([new_log])],
                        ignore_index=True
                    )
                    try:
                        save_db(attendance_logs_df, "attendance_logs")
                        st.session_state["attendance_logs_df"] = attendance_logs_df
                        st.success(f"{name} → {result_text_map.get(action, action)}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"勤怠保存でエラーです: {e}")

                    result_text_map = {
                        "in": "出勤",
                        "out": "退勤",
                        "break_start": "休憩開始",
                        "break_end": "休憩終了",
                    }

                    st.success(f"{name} → {result_text_map.get(action, action)}")
                    st.rerun()

def render_support_record_audit_page():
    st.header("過去日誌照合")
    st.caption("Knowbeの支援記録を期間指定で読み込み、登録区分と日誌内容の判定結果を一覧化します。")

    if not st.session_state.get("is_admin", False):
        st.error("このページは管理者専用です。")
        return

    company_id = get_current_company_id() 
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者マスタが見つかりません。")
        return

    work = master_df.copy()
    if "resident_name" not in work.columns:
        st.error("resident_master に resident_name 列がありません。")
        return

    if "status" in work.columns:
        active_df = work[
            work["status"].astype(str).str.strip().isin(["active", "利用中", ""])
        ].copy()
        if not active_df.empty:
            work = active_df

    work["resident_name"] = work["resident_name"].fillna("").astype(str).str.strip()
    work = work[work["resident_name"] != ""].copy()

    if work.empty:
        st.warning("選択できる利用者がいません。")
        return

    resident_options = sorted(work["resident_name"].unique().tolist())

    st.markdown("### 条件設定")

    col1, col2 = st.columns(2)

    with col1:
        resident_name = st.selectbox(
            "利用者",
            resident_options,
            key="support_audit_resident_name"
        )

        start_year = st.number_input(
            "開始年",
            min_value=2024,
            max_value=2035,
            value=2025,
            step=1,
            key="support_audit_start_year"
        )

        start_month = st.number_input(
            "開始月",
            min_value=1,
            max_value=12,
            value=8,
            step=1,
            key="support_audit_start_month"
        )

    with col2:
        end_year = st.number_input(
            "終了年",
            min_value=2024,
            max_value=2035,
            value=2026,
            step=1,
            key="support_audit_end_year"
        )

        end_month = st.number_input(
            "終了月",
            min_value=1,
            max_value=12,
            value=3,
            step=1,
            key="support_audit_end_month"
        )

    file_name = f"過去日誌照合_{resident_name}_{start_year}_{start_month:02d}_to_{end_year}_{end_month:02d}.xlsx"

    if st.button("過去日誌照合を実行", key="run_support_record_audit", use_container_width=True):
        if (int(start_year), int(start_month)) > (int(end_year), int(end_month)):
            st.error("開始年月が終了年月より後になっています。")
            return

        api_key = get_gemini_api_key_from_app()
        if not api_key:
            st.error("GEMINI_API_KEY が見つかりません。")
            return

        try:
            from google import genai
        except Exception as e:
            st.error(f"google.genai の読み込みに失敗しました: {e}")
            return

        try:
            client = genai.Client(api_key=api_key)
        except Exception as e:
            st.error(f"Geminiクライアント作成に失敗しました: {e}")
            return

        try:
            with st.spinner("Knowbeから支援記録を読み込み中です。少し待ってください…"):
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                tmp_path = tmp.name
                tmp.close()

                rows = run_support_record_kind_export(
                    resident_name=resident_name,
                    start_year=int(start_year),
                    start_month=int(start_month),
                    end_year=int(end_year),
                    end_month=int(end_month),
                    output_path=tmp_path,
                    gemini_client=client,
                )

                with open(tmp_path, "rb") as f:
                    result_bytes = f.read()

            st.session_state["support_record_audit_result_bytes"] = result_bytes
            st.session_state["support_record_audit_result_name"] = file_name
            st.session_state["support_record_audit_rows"] = rows

            st.success("照合が完了したある。")

        except Exception as e:
            st.error(f"照合中にエラーが出ました: {e}")

    rows = st.session_state.get("support_record_audit_rows", [])
    if rows:
        st.markdown("### 結果プレビュー")

        preview_df = pd.DataFrame(rows).copy()

        show_cols = []
        rename_map = {}

        if "resident_name" in preview_df.columns:
            show_cols.append("resident_name")
            rename_map["resident_name"] = "利用者"

        if "year" in preview_df.columns and "month" in preview_df.columns and "day" in preview_df.columns:
            preview_df["date_str"] = (
                preview_df["year"].astype(str) + "/" +
                preview_df["month"].astype(str) + "/" +
                preview_df["day"].astype(str)
            )
            show_cols.append("date_str")
            rename_map["date_str"] = "日付"

        if "registered_kind" in preview_df.columns:
            show_cols.append("registered_kind")
            rename_map["registered_kind"] = "登録"

        if "diary_kind" in preview_df.columns:
            show_cols.append("diary_kind")
            rename_map["diary_kind"] = "日誌"

        if "weekday" in preview_df.columns:
            show_cols.append("weekday")
            rename_map["weekday"] = "曜日"

        if show_cols:
            show_df = preview_df[show_cols].copy().rename(columns=rename_map)
            st.dataframe(show_df, use_container_width=True)

        total_count = len(preview_df)

        cannot_count = 0
        if "diary_kind" in preview_df.columns:
            cannot_count = int((preview_df["diary_kind"].astype(str) == "判定できず").sum())

        mismatch_count = 0
        if "registered_kind" in preview_df.columns and "diary_kind" in preview_df.columns:
            judge_df = preview_df[
                preview_df["diary_kind"].astype(str) != "判定できず"
            ].copy()
            if not judge_df.empty:
                mismatch_count = int(
                    (judge_df["registered_kind"].astype(str) != judge_df["diary_kind"].astype(str)).sum()
                )

        c1, c2, c3 = st.columns(3)
        with c1:
            st.info(f"総件数: {total_count}")
        with c2:
            st.warning(f"判定できず: {cannot_count}")
        with c3:
            st.error(f"不一致件数: {mismatch_count}")

    if st.session_state.get("support_record_audit_result_bytes"):
        st.download_button(
            label="Excelダウンロード",
            data=st.session_state["support_record_audit_result_bytes"],
            file_name=st.session_state.get("support_record_audit_result_name", "過去日誌照合.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_support_record_audit_excel",
            use_container_width=True,
        )




def generate_with_gemini(prompt: str):
    api_key = get_gemini_api_key_from_app()
    if not api_key:
        raise RuntimeError("APIキーありません")

    genai.configure(api_key=api_key)

    model_candidates = ["gemini-2.5-flash"]
    errors = []

    for model_name in model_candidates:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            text = str(getattr(response, "text", "")).strip()

            if text:
                return text

            errors.append(f"{model_name}: empty response")

        except Exception as e:
            errors.append(f"{model_name}: {e}")

    raise RuntimeError("Gemini生成失敗: " + " | ".join(errors))

def edit_text_with_gemini_for_start_memo(resident_name: str, original_text: str, remark_text: str = ""):
    original_text = str(original_text or "").strip()
    remark_text = str(remark_text or "").strip()

    if not original_text:
        return ""

    prompt = f"""
利用者名: {resident_name}

以下の開始メモを、就労継続支援B型の支援記録として自然な日本語に整えてください。
事実を変えすぎず、伝聞調ばかりになりすぎないように、簡潔で読みやすくしてください。

備考:
{remark_text}

開始メモ:
{original_text}

出力は本文のみ。
"""
    return generate_with_gemini(prompt)


def edit_text_with_gemini_for_end_memo(resident_name: str, original_text: str, remark_text: str = ""):
    original_text = str(original_text or "").strip()
    remark_text = str(remark_text or "").strip()

    if not original_text:
        return ""

    prompt = f"""
利用者名: {resident_name}

以下の終了メモを、就労継続支援B型の職員考察として自然な日本語に整えてください。
事実を変えすぎず、簡潔で読みやすくしてください。

備考:
{remark_text}

終了メモ:
{original_text}

出力は本文のみ。
"""
    return generate_with_gemini(prompt)

def generate_json_with_gemini(prompt: str):
    api_key = get_gemini_api_key_from_app()
    if not api_key:
        raise RuntimeError("APIキーありません")

    genai.configure(api_key=api_key)

    model_candidates = [
        "gemini-2.5-flash",
    ]

    last_error = None

    for model_name in model_candidates:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            text = str(getattr(response, "text", "")).strip()

            if not text:
                continue

            text = text.replace("```json", "").replace("```", "").strip()
            return json.loads(text)

        except Exception as e:
            last_error = e
            continue

    raise RuntimeError(f"Gemini JSON生成失敗: {last_error}")

def build_home_eval_from_support_record_prompt(resident_name: str, year_val: str, month_val: str, support_record_text: str):
    return f"""
利用者名: {resident_name}
対象年月: {year_val}年{month_val}月

以下はKnowbeの支援記録ページ本文です。
この内容だけを根拠に、在宅評価シート用の内容をJSONのみで作成してください。

【支援記録本文】
{support_record_text}

【作成ルール】
- 在宅評価シートの形式に合わせる
- 目標は2〜3個
- 月間評価は2〜3個
- 週報は第1週〜第5週をすべて作る
- 記録が少ない週は、無理に長くせず自然で簡潔にする
- 「通所」は在宅利用の意味である可能性があるため、来所と断定しない
- 事実にないことを勝手に足しすぎない
- 出力はJSONのみ

【出力形式】
{{
  "goals": ["目標1", "目標2", "目標3"],
  "monthly_evaluations": ["月間評価1", "月間評価2", "月間評価3"],
  "weekly_reports": {{
    "1": "第1週の評価",
    "2": "第2週の評価",
    "3": "第3週の評価",
    "4": "第4週の評価",
    "5": "第5週の評価"
  }}
}}
"""

def build_bulk_plan_prompt(resident_name: str):
    return f"""
利用者名: {resident_name}

就労継続支援B型の個別支援計画をJSONのみで作成してください。

出力形式:
{{
  "policy": "サービス等利用計画の総合的な方針",
  "long_goal": "長期目標",
  "short_goal": "短期目標",
  "rows": [
    {{
      "target": "具体的到達目標1",
      "role": "本人の役割1",
      "support": "支援内容1",
      "period": "支援期間1",
      "person": "担当者1",
      "priority": "1"
    }},
    {{
      "target": "具体的到達目標2",
      "role": "本人の役割2",
      "support": "支援内容2",
      "period": "支援期間2",
      "person": "担当者2",
      "priority": "2"
    }},
    {{
      "target": "具体的到達目標3",
      "role": "本人の役割3",
      "support": "支援内容3",
      "period": "支援期間3",
      "person": "担当者3",
      "priority": "3"
    }}
  ]
}}
"""
def build_bulk_meeting_prompt(resident_name: str, plan_json: dict, meeting_info: str, attendees_text: str):
    return f"""
利用者名: {resident_name}

以下の個別支援計画案をもとに、サービス担当者会議記録をJSONのみで作成してください。

計画案:
{json.dumps(plan_json, ensure_ascii=False)}

開催情報:
{meeting_info}

出席者:
{attendees_text}

出力形式:
{{
  "agenda": "議題",
  "discussion": "検討内容",
  "issues_left": "残された課題",
  "conclusion": "結論"
}}
"""
def build_bulk_plan_from_monitoring_prompt(resident_name: str, monitoring_json: dict):
    return f"""
利用者名: {resident_name}

以下の直近モニタリング内容をもとに、就労継続支援B型の個別支援計画案をJSONのみで作成してください。

参照元（直近モニタリング）:
{json.dumps(monitoring_json, ensure_ascii=False)}

出力ルール:
- モニタリング内容から自然につながる個別支援計画案を作成する
- 支援期間は空でもよい
- 担当者は原則「全職員」でよい
- 優先順位は 1,2,3 を入れる
- 出力はJSONのみ

出力形式:
{{
  "policy": "サービス等利用計画の総合的な方針",
  "long_goal": "長期目標",
  "short_goal": "短期目標",
  "rows": [
    {{
      "target": "具体的到達目標1",
      "role": "本人の役割1",
      "support": "支援内容1",
      "period": "支援期間1",
      "person": "担当者1",
      "priority": "1"
    }},
    {{
      "target": "具体的到達目標2",
      "role": "本人の役割2",
      "support": "支援内容2",
      "period": "支援期間2",
      "person": "担当者2",
      "priority": "2"
    }},
    {{
      "target": "具体的到達目標3",
      "role": "本人の役割3",
      "support": "支援内容3",
      "period": "支援期間3",
      "person": "担当者3",
      "priority": "3"
    }}
  ]
}}
"""


def build_bulk_final_plan_prompt(resident_name: str, draft_plan_json: dict, meeting_json: dict):
    return f"""
利用者名: {resident_name}

以下の個別支援計画案とサービス担当者会議記録をもとに、
就労継続支援B型の個別支援計画（本計画）をJSONのみで作成してください。

個別支援計画案:
{json.dumps(draft_plan_json, ensure_ascii=False)}

サービス担当者会議:
{json.dumps(meeting_json, ensure_ascii=False)}

出力ルール:
- 会議の検討内容・結論を踏まえて、計画案を必要に応じて調整する
- 計画案をベースにしつつ、本計画として自然な内容に整える
- 支援期間は空でもよい
- 担当者は原則「全職員」でよい
- 優先順位は 1,2,3 を入れる
- 出力はJSONのみ

出力形式:
{{
  "policy": "サービス等利用計画の総合的な方針",
  "long_goal": "長期目標",
  "short_goal": "短期目標",
  "rows": [
    {{
      "target": "具体的到達目標1",
      "role": "本人の役割1",
      "support": "支援内容1",
      "period": "支援期間1",
      "person": "担当者1",
      "priority": "1"
    }},
    {{
      "target": "具体的到達目標2",
      "role": "本人の役割2",
      "support": "支援内容2",
      "period": "支援期間2",
      "person": "担当者2",
      "priority": "2"
    }},
    {{
      "target": "具体的到達目標3",
      "role": "本人の役割3",
      "support": "支援内容3",
      "period": "支援期間3",
      "person": "担当者3",
      "priority": "3"
    }}
  ]
}}
"""


def apply_bulk_plan_overrides(plan_json, periods, persons):
    result = json.loads(json.dumps(plan_json, ensure_ascii=False))

    rows = result.get("rows", [])
    while len(rows) < 3:
        rows.append({})

    for i in range(3):
        if i < len(periods):
            rows[i]["period"] = str(periods[i]).strip()

        if i < len(persons):
            person_val = str(persons[i]).strip()
            rows[i]["person"] = person_val if person_val else "全職員"

        if not str(rows[i].get("person", "")).strip():
            rows[i]["person"] = "全職員"

        if not str(rows[i].get("priority", "")).strip():
            rows[i]["priority"] = str(i + 1)

    result["rows"] = rows
    return result

def build_plan_cell_data_from_json(plan_json, resident_name, year_val, month_val, day_val, manager_val):
    rows = plan_json.get("rows", [])
    while len(rows) < 3:
        rows.append({})

    return {
        "E5": resident_name,
        "M5": year_val,
        "O5": month_val,
        "Q5": day_val,

        "B8": str(plan_json.get("policy", "")),
        "B10": str(plan_json.get("long_goal", "")),
        "B12": str(plan_json.get("short_goal", "")),

        "C17": str(rows[0].get("target", "")),
        "G17": str(rows[0].get("role", "")),
        "J17": str(rows[0].get("support", "")),
        "M17": str(rows[0].get("period", "")),
        "O17": str(rows[0].get("person", "")),
        "Q17": str(rows[0].get("priority", "")),

        "C18": str(rows[1].get("target", "")),
        "G18": str(rows[1].get("role", "")),
        "J18": str(rows[1].get("support", "")),
        "M18": str(rows[1].get("period", "")),
        "O18": str(rows[1].get("person", "")),
        "Q18": str(rows[1].get("priority", "")),

        "C19": str(rows[2].get("target", "")),
        "G19": str(rows[2].get("role", "")),
        "J19": str(rows[2].get("support", "")),
        "M19": str(rows[2].get("period", "")),
        "O19": str(rows[2].get("person", "")),
        "Q19": str(rows[2].get("priority", "")),

        "N21": manager_val,
    }

def build_home_eval_cell_data(
    resident_name,
    create_year,
    create_month,
    manager_name,
    goals,
    monthly_evaluations,
    weekly_dates,
    weekly_reports,
    weekly_visits,
):
    goals = list(goals or [])
    monthly_evaluations = list(monthly_evaluations or [])

    while len(goals) < 3:
        goals.append("")

    while len(monthly_evaluations) < 3:
        monthly_evaluations.append("")

    return {
        "B3": resident_name,
        "J3": create_year,
        "L3": create_month,

        "B7": goals[0],
        "B8": goals[1],
        "B9": goals[2],

        "H11": manager_name,

        "B12": monthly_evaluations[0],
        "B13": monthly_evaluations[1],
        "B14": monthly_evaluations[2],

        "A19": weekly_dates.get("1", ""),
        "C19": weekly_reports.get("1", ""),
        "J20": weekly_visits.get("1", ""),

        "A21": weekly_dates.get("2", ""),
        "C21": weekly_reports.get("2", ""),
        "J22": weekly_visits.get("2", ""),

        "A23": weekly_dates.get("3", ""),
        "C23": weekly_reports.get("3", ""),
        "J24": weekly_visits.get("3", ""),

        "A25": weekly_dates.get("4", ""),
        "C25": weekly_reports.get("4", ""),
        "J26": weekly_visits.get("4", ""),

        "A27": weekly_dates.get("5", ""),
        "C27": weekly_reports.get("5", ""),
        "J28": weekly_visits.get("5", ""),
    }

def build_home_eval_week_ranges(year_val, month_val):
    try:
        y = int(str(year_val).strip())
        m = int(str(month_val).strip())
    except Exception:
        return {
            "1": "",
            "2": "",
            "3": "",
            "4": "",
            "5": "",
        }

    last_day = py_calendar.monthrange(y, m)[1]

    return {
        "1": f"{m}/1〜{m}/7",
        "2": f"{m}/8〜{m}/14",
        "3": f"{m}/15〜{m}/21",
        "4": f"{m}/22〜{m}/28",
        "5": f"{m}/29〜{m}/{last_day}",
    }

def apply_bulk_plan_overrides(plan_json, periods, persons):
    result = json.loads(json.dumps(plan_json, ensure_ascii=False))

    rows = result.get("rows", [])
    while len(rows) < 3:
        rows.append({})

    for i in range(3):
        if i < len(periods):
            rows[i]["period"] = str(periods[i]).strip()
        if i < len(persons):
            person_val = str(persons[i]).strip()
            rows[i]["person"] = person_val if person_val else "全職員"

    result["rows"] = rows
    return result

def build_meeting_cell_data_from_json(
    meeting_json,
    resident_name,
    create_year,
    create_month,
    create_day,
    meeting_year,
    meeting_month,
    meeting_day,
    meeting_info,
    attendees,
    meeting_creator,
):
    return {
        # 作成年月日
        "M3": create_year,
        "O3": create_month,
        "Q3": create_day,

        # 利用者名・作成者
        "C4": resident_name,
        "M4": meeting_creator,

        # 開催日時・開催場所
        "C5": meeting_year,
        "E5": meeting_month,
        "G5": meeting_day,
        "M5": meeting_info,

        # 会議出席者
        "E8": attendees.get("admin", ""),
        "J8": attendees.get("staff", ""),
        "O8": attendees.get("user", ""),

        "E9": attendees.get("caremanager", ""),
        "J9": attendees.get("nurse", ""),
        "O9": attendees.get("family", ""),

        "E10": attendees.get("manager", ""),
        "J10": attendees.get("consultant", ""),
        "O10": attendees.get("keyperson", ""),

        # 会議内容
        "C11": meeting_json.get("agenda", ""),
        "C12": meeting_json.get("discussion", ""),
        "C13": meeting_json.get("issues_left", ""),
        "C14": meeting_json.get("conclusion", ""),
    }



def sync_resident_master_from_assessment(resident_id: str, welfare_status: str):
    master_df = load_db("resident_master")
    if master_df is None or master_df.empty:
        return

    master_df = master_df.fillna("").copy()

    if "public_assistance" not in master_df.columns:
        master_df["public_assistance"] = ""

    mask = master_df["resident_id"].astype(str).str.strip() == str(resident_id).strip()
    if not mask.any():
        return

    master_df.loc[mask, "public_assistance"] = str(welfare_status).strip()
    master_df.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(master_df, "resident_master")

@st.cache_data(ttl=60)
def get_chat_rooms_df_cached():
    df = load_db("chat_rooms")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "room_id",
            "room_name",
            "room_type",
            "room_password",
            "created_by_user_id",
            "created_by_company_id",
            "description",
            "status",
            "created_at",
            "updated_at",
        ])
    else:
        for col in [
            "room_id",
            "room_name",
            "room_type",
            "room_password",
            "created_by_user_id",
            "created_by_company_id",
            "description",
            "status",
            "created_at",
            "updated_at",
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def get_chat_rooms_df():
    return get_chat_rooms_df_cached().copy()

def render_excel_download_block(doc_title, file_name, template_name, cell_data):
    import json

    file_key = f"{doc_title}_excel_file"
    sig_key = f"{doc_title}_excel_signature"

    # 入力内容の変化を判定するための署名
    current_signature = json.dumps(
        cell_data,
        ensure_ascii=False,
        sort_keys=True,
        default=str
    )

    # 入力内容が1文字でも変わったら、古いExcelを消す
    if st.session_state.get(sig_key) != current_signature:
        st.session_state[sig_key] = current_signature
        st.session_state.pop(file_key, None)

    # Excel作成
    if st.button("Excelを作成", key=f"{doc_title}_make_excel"):
        st.session_state[file_key] = create_excel_file(template_name, cell_data)
        st.session_state[sig_key] = current_signature

    # 作成済みのときだけダウンロードボタン表示
    if st.session_state.get(file_key):
        st.download_button(
            label="📥 ダウンロード",
            data=st.session_state[file_key],
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{doc_title}_download_excel"
        )

@st.cache_data(ttl=60)
def get_warehouse_files_df_cached():
    df = load_db("warehouse_files")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "file_id",
            "title",
            "description",
            "category_main",
            "category_sub",
            "tags",
            "file_name",
            "file_data",
            "file_type",
            "uploaded_by_user_id",
            "uploaded_by_company_id",
            "source_room_id",
            "visibility_type",
            "download_password",
            "is_searchable",
            "is_deleted",
            "created_at",
            "updated_at",
            "deleted_by_user_id",
            "deleted_at",
        ])
    else:
        for col in [
            "file_id",
            "title",
            "description",
            "category_main",
            "category_sub",
            "tags",
            "file_name",
            "file_data",
            "file_type",
            "uploaded_by_user_id",
            "uploaded_by_company_id",
            "source_room_id",
            "visibility_type",
            "download_password",
            "is_searchable",
            "is_deleted",
            "created_at",
            "updated_at",
            "deleted_by_user_id",
            "deleted_at",
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")

def get_warehouse_files_df():
    return get_warehouse_files_df_cached().copy()

def get_next_warehouse_file_id():
    df = get_warehouse_files_df()
    if df is None or df.empty:
        return "W0001"

    nums = []
    for x in df["file_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("W"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"W{next_num:04d}"


def save_warehouse_file(
    title,
    description,
    category_main,
    category_sub,
    tags,
    uploaded_file,
    visibility_type="public",
    download_password="",
    is_searchable="1",
    source_room_id="",
):
    df = get_warehouse_files_df()

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")
    file_id = get_next_warehouse_file_id()

    file_bytes = uploaded_file.read()
    file_data_base64 = base64.b64encode(file_bytes).decode("utf-8")
    file_name = str(uploaded_file.name).strip()
    lower_name = file_name.lower()

    if lower_name.endswith(".xlsx"):
        file_type = "xlsx"
    elif lower_name.endswith(".xls"):
        file_type = "xls"
    elif lower_name.endswith(".pdf"):
        file_type = "pdf"
    elif lower_name.endswith(".docx"):
        file_type = "docx"
    elif lower_name.endswith(".doc"):
        file_type = "doc"
    elif lower_name.endswith(".jpg") or lower_name.endswith(".jpeg"):
        file_type = "jpg"
    elif lower_name.endswith(".png"):
        file_type = "png"
    else:
        file_type = "other"

    new_row = pd.DataFrame([{
        "file_id": file_id,
        "title": str(title).strip(),
        "description": str(description).strip(),
        "category_main": str(category_main).strip(),
        "category_sub": str(category_sub).strip(),
        "tags": str(tags).strip(),
        "file_name": file_name,
        "file_data": file_data_base64,
        "file_type": file_type,
        "uploaded_by_user_id": str(st.session_state.get("user_id", "")).strip(),
        "uploaded_by_company_id": str(st.session_state.get("company_id", "")).strip(),
        "source_room_id": str(source_room_id).strip(),
        "visibility_type": str(visibility_type).strip(),
        "download_password": str(download_password).strip(),
        "is_searchable": str(is_searchable).strip(),
        "is_deleted": "0",
        "created_at": now_str,
        "updated_at": now_str,
        "deleted_by_user_id": "",
        "deleted_at": "",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "warehouse_files")
    return file_id


def soft_delete_warehouse_file(file_id):
    df = get_warehouse_files_df()
    if df is None or df.empty:
        return False

    mask = df["file_id"].astype(str) == str(file_id).strip()
    if not mask.any():
        return False

    df.loc[mask, "is_deleted"] = "1"
    df.loc[mask, "deleted_by_user_id"] = str(st.session_state.get("user_id", "")).strip()
    df.loc[mask, "deleted_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")
    df.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(df, "warehouse_files")
    return True


def get_warehouse_download_data(row):
    file_data_base64 = str(row.get("file_data", "")).strip()
    file_name = str(row.get("file_name", "")).strip()

    if not file_data_base64 or not file_name:
        return None, None, None

    file_bytes = base64.b64decode(file_data_base64)

    lower_name = file_name.lower()
    if lower_name.endswith(".xlsx"):
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif lower_name.endswith(".xls"):
        mime = "application/vnd.ms-excel"
    elif lower_name.endswith(".pdf"):
        mime = "application/pdf"
    elif lower_name.endswith(".docx"):
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif lower_name.endswith(".doc"):
        mime = "application/msword"
    elif lower_name.endswith(".jpg") or lower_name.endswith(".jpeg"):
        mime = "image/jpeg"
    elif lower_name.endswith(".png"):
        mime = "image/png"
    else:
        mime = "application/octet-stream"

    return file_bytes, file_name, mime

def render_warehouse_page():
    st.title("🏭 倉庫")
    st.caption("全事業所共通の資料置き場です。検索して、共有して、必要なら限定公開もできます。")

    current_company_id = str(st.session_state.get("company_id", "")).strip()
    current_user_id = str(st.session_state.get("user_id", "")).strip()
    is_admin = bool(st.session_state.get("is_admin", False))

    if "warehouse_unlocked_files" not in st.session_state:
        st.session_state.warehouse_unlocked_files = []

    top_cols = st.columns([1, 1])

    with top_cols[0]:
        if st.button("← 休憩室へ戻る", key="back_from_warehouse", use_container_width=True):
            st.session_state.current_page = "休憩室"
            st.rerun()

    with top_cols[1]:
        st.info(f"ログイン中: {st.session_state.get('company_name', '')} / {st.session_state.get('user', '')}")

    st.divider()

    with st.expander("＋ 新しい資料を登録する"):
        title = st.text_input("タイトル", key="warehouse_title")
        description = st.text_area("説明", key="warehouse_description", height=80)
        category_main = st.text_input("カテゴリ大", key="warehouse_category_main")
        category_sub = st.text_input("カテゴリ小", key="warehouse_category_sub")
        tags = st.text_input("タグ（カンマ区切りでもOK）", key="warehouse_tags")
        visibility_type = st.selectbox(
            "公開設定",
            ["public", "limited", "private"],
            key="warehouse_visibility_type"
        )
        download_password = st.text_input(
            "ダウンロードパスワード（limited/privateなら設定）",
            key="warehouse_download_password"
        )
        is_searchable = st.selectbox(
            "検索に表示するか",
            ["1", "0"],
            format_func=lambda x: "表示する" if x == "1" else "表示しない",
            key="warehouse_is_searchable"
        )
        uploaded_file = st.file_uploader("ファイルを選択", key="warehouse_uploaded_file")

        if st.button("倉庫へ保存", key="save_warehouse_button", use_container_width=True):
            if not title.strip():
                st.error("タイトルを入れてください。")
            elif uploaded_file is None:
                st.error("ファイルを選んでください。")
            elif visibility_type in ["limited", "private"] and not download_password.strip():
                st.error("その公開設定ならダウンロードパスワードが必要です。")
            else:
                file_id = save_warehouse_file(
                    title=title,
                    description=description,
                    category_main=category_main,
                    category_sub=category_sub,
                    tags=tags,
                    uploaded_file=uploaded_file,
                    visibility_type=visibility_type,
                    download_password=download_password,
                    is_searchable=is_searchable,
                    source_room_id="",
                )
                st.success(f"保存しました！ {file_id}")
                st.rerun()

    st.divider()

    df = get_warehouse_files_df()
    if df is None or df.empty:
        st.info("まだ倉庫に資料がありません。")
        return

    work = df.copy()
    work = work[work["is_deleted"].astype(str) != "1"].copy()

    keyword_cols = st.columns([2, 1, 1])

    with keyword_cols[0]:
        keyword = st.text_input("検索", key="warehouse_search_keyword")

    with keyword_cols[1]:
        filter_main = st.text_input("カテゴリ大で絞る", key="warehouse_filter_main")

    with keyword_cols[2]:
        filter_sub = st.text_input("カテゴリ小で絞る", key="warehouse_filter_sub")

    # 検索可能なものだけ対象
    work = work[work["is_searchable"].astype(str) == "1"].copy()

    if keyword.strip():
        kw = keyword.strip()
        work = work[
            work["title"].astype(str).str.contains(kw, case=False, na=False) |
            work["description"].astype(str).str.contains(kw, case=False, na=False) |
            work["tags"].astype(str).str.contains(kw, case=False, na=False) |
            work["file_name"].astype(str).str.contains(kw, case=False, na=False)
        ].copy()

    if filter_main.strip():
        work = work[
            work["category_main"].astype(str).str.contains(filter_main.strip(), case=False, na=False)
        ].copy()

    if filter_sub.strip():
        work = work[
            work["category_sub"].astype(str).str.contains(filter_sub.strip(), case=False, na=False)
        ].copy()

    try:
        work = work.sort_values(["updated_at", "created_at"], ascending=[False, False])
    except Exception:
        pass

    st.markdown(f"### 一覧（{len(work)}件）")

    if work.empty:
        st.info("条件に合う資料がありません。")
        return

    users_df = get_users_df()

    for _, row in work.iterrows():
        file_id = str(row.get("file_id", "")).strip()
        title = str(row.get("title", "")).strip()
        description = str(row.get("description", "")).strip()
        category_main = str(row.get("category_main", "")).strip()
        category_sub = str(row.get("category_sub", "")).strip()
        tags = str(row.get("tags", "")).strip()
        file_name = str(row.get("file_name", "")).strip()
        file_type = str(row.get("file_type", "")).strip()
        uploaded_by_user_id = str(row.get("uploaded_by_user_id", "")).strip()
        uploaded_by_company_id = str(row.get("uploaded_by_company_id", "")).strip()
        visibility_type = str(row.get("visibility_type", "")).strip()
        created_at = str(row.get("created_at", "")).strip()
        updated_at = str(row.get("updated_at", "")).strip()

        uploader_name = uploaded_by_user_id
        try:
            target_user = users_df[users_df["user_id"].astype(str) == uploaded_by_user_id]
            if not target_user.empty:
                uploader_name = str(target_user.iloc[0].get("display_name", uploaded_by_user_id)).strip()
        except Exception:
            pass

        st.markdown("---")
        st.markdown(f"## {title}")
        st.caption(f"{file_id} / {file_name} / {file_type} / 公開設定: {visibility_type}")

        meta_cols = st.columns([2, 2, 2])
        with meta_cols[0]:
            st.write(f"カテゴリ: {category_main} / {category_sub}")
        with meta_cols[1]:
            st.write(f"登録者: {uploader_name}")
        with meta_cols[2]:
            st.write(f"更新: {updated_at or created_at}")

        st.caption(f"登録事業所: {uploaded_by_company_id}")

        if description:
            st.write(description)
        if tags:
            st.caption(f"タグ: {tags}")

        file_bytes, dl_name, mime = get_warehouse_download_data(row)

        can_delete = (uploaded_by_user_id == current_user_id) or (is_admin and uploaded_by_company_id == current_company_id)

        is_unlocked = (file_id in st.session_state.warehouse_unlocked_files)

        # public は即DL可
        if visibility_type == "public":
            is_unlocked = True

        # private は検索結果に出さない設計でもいいけど、
        # 今回は is_searchable=1 のものだけここに出るので、DL時だけPW要求にしてる
        if visibility_type in ["limited", "private"] and not is_unlocked:
            pw_cols = st.columns([2, 1])
            with pw_cols[0]:
                input_pw = st.text_input(
                    f"{file_id} のダウンロードパスワード",
                    type="password",
                    key=f"warehouse_pw_{file_id}"
                )
            with pw_cols[1]:
                st.write("")
                if st.button("解除", key=f"unlock_warehouse_{file_id}", use_container_width=True):
                    real_pw = str(row.get("download_password", "")).strip()
                    if str(input_pw).strip() == real_pw:
                        st.session_state.warehouse_unlocked_files.append(file_id)
                        st.success("ダウンロード可能になりました。")
                        st.rerun()
                    else:
                        st.error("パスワードが違います。")

        action_cols = st.columns([1, 1, 1])

        with action_cols[0]:
            if is_unlocked and file_bytes is not None:
                st.download_button(
                    label="ダウンロード",
                    data=file_bytes,
                    file_name=dl_name,
                    mime=mime,
                    key=f"warehouse_download_{file_id}",
                    use_container_width=True
                )

        with action_cols[1]:
            if can_delete:
                if st.button("削除", key=f"warehouse_delete_{file_id}", use_container_width=True):
                    ok = soft_delete_warehouse_file(file_id)
                    if ok:
                        create_admin_log(
                            action_type="delete_warehouse_file",
                            target_type="warehouse_file",
                            target_id=file_id,
                            action_detail=f"title={title}"
                        )
                        st.success("削除しました。")
                        st.rerun()
                    else:
                        st.error("削除に失敗しました。")

        with action_cols[2]:
            st.write("")

@st.cache_data(ttl=60)
def get_admin_logs_df_cached():
    df = load_db("admin_logs")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "log_id",
            "company_id",
            "acted_by_user_id",
            "acted_by_display_name",
            "action_type",
            "target_type",
            "target_id",
            "action_detail",
            "created_at",
        ])
    else:
        for col in [
            "log_id",
            "company_id",
            "acted_by_user_id",
            "acted_by_display_name",
            "action_type",
            "target_type",
            "target_id",
            "action_detail",
            "created_at",
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def get_admin_logs_df():
    return get_admin_logs_df_cached().copy()


def get_next_admin_log_id():
    df = get_admin_logs_df()
    if df is None or df.empty:
        return "L0001"

    nums = []
    for x in df["log_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("L"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"L{next_num:04d}"


def create_admin_log(action_type, target_type, target_id, action_detail=""):
    df = get_admin_logs_df()

    new_row = pd.DataFrame([{
        "log_id": get_next_admin_log_id(),
        "company_id": str(st.session_state.get("company_id", "")).strip(),
        "acted_by_user_id": str(st.session_state.get("user_id", "")).strip(),
        "acted_by_display_name": str(st.session_state.get("user", "")).strip(),
        "action_type": str(action_type).strip(),
        "target_type": str(target_type).strip(),
        "target_id": str(target_id).strip(),
        "action_detail": str(action_detail).strip(),
        "created_at": now_jst().strftime("%Y-%m-%d %H:%M:%S"),
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "admin_logs")

@st.cache_data(ttl=60)
def get_archive_files_df_cached():
    df = load_db("archive_files")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "archive_file_id",
            "company_id",
            "title",
            "description",
            "category_main",
            "category_sub",
            "tags",
            "file_name",
            "file_data",
            "file_type",
            "uploaded_by_user_id",
            "visibility_type",
            "download_password",
            "is_deleted",
            "created_at",
            "updated_at",
            "deleted_by_user_id",
            "deleted_at",
        ])
    else:
        for col in [
            "archive_file_id",
            "company_id",
            "title",
            "description",
            "category_main",
            "category_sub",
            "tags",
            "file_name",
            "file_data",
            "file_type",
            "uploaded_by_user_id",
            "visibility_type",
            "download_password",
            "is_deleted",
            "created_at",
            "updated_at",
            "deleted_by_user_id",
            "deleted_at",
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")

def search_shared_documents(keyword: str):
    keyword = str(keyword).strip()
    if not keyword:
        return pd.DataFrame()

    company_id = str(st.session_state.get("company_id", "")).strip()

    archive_df = get_archive_files_df()
    warehouse_df = get_warehouse_files_df()

    result_rows = []

    # 書類アップロード（事業所内限定）
    if archive_df is not None and not archive_df.empty:
        work = archive_df.copy()
        work = work[
            (work["company_id"].astype(str) == company_id) &
            (work["is_deleted"].astype(str) != "1")
        ].copy()

        hit = work[
            work["title"].astype(str).str.contains(keyword, case=False, na=False) |
            work["description"].astype(str).str.contains(keyword, case=False, na=False) |
            work["tags"].astype(str).str.contains(keyword, case=False, na=False) |
            work["file_name"].astype(str).str.contains(keyword, case=False, na=False)
        ].copy()

        for _, row in hit.iterrows():
            result_rows.append({
                "source": "書類アップロード",
                "id": str(row.get("archive_file_id", "")).strip(),
                "title": str(row.get("title", "")).strip(),
                "description": str(row.get("description", "")).strip(),
                "category_main": str(row.get("category_main", "")).strip(),
                "category_sub": str(row.get("category_sub", "")).strip(),
                "file_name": str(row.get("file_name", "")).strip(),
                "visibility_type": "normal",
                "updated_at": str(row.get("updated_at", "")).strip(),
            })

    # 倉庫（全体共通）
    if warehouse_df is not None and not warehouse_df.empty:
        work = warehouse_df.copy()
        work = work[
            (work["is_deleted"].astype(str) != "1") &
            (work["is_searchable"].astype(str) == "1")
        ].copy()

        hit = work[
            work["title"].astype(str).str.contains(keyword, case=False, na=False) |
            work["description"].astype(str).str.contains(keyword, case=False, na=False) |
            work["tags"].astype(str).str.contains(keyword, case=False, na=False) |
            work["file_name"].astype(str).str.contains(keyword, case=False, na=False)
        ].copy()

        for _, row in hit.iterrows():
            result_rows.append({
                "source": "倉庫",
                "id": str(row.get("file_id", "")).strip(),
                "title": str(row.get("title", "")).strip(),
                "description": str(row.get("description", "")).strip(),
                "category_main": str(row.get("category_main", "")).strip(),
                "category_sub": str(row.get("category_sub", "")).strip(),
                "file_name": str(row.get("file_name", "")).strip(),
                "visibility_type": str(row.get("visibility_type", "")).strip(),
                "updated_at": str(row.get("updated_at", "")).strip(),
            })

    if not result_rows:
        return pd.DataFrame()

    df = pd.DataFrame(result_rows)

    try:
        df = df.sort_values(["updated_at"], ascending=[False])
    except Exception:
        pass

    return df

def get_archive_files_df():
    return get_archive_files_df_cached().copy()

def get_next_archive_file_id():
    df = get_archive_files_df()
    if df is None or df.empty:
        return "A0001"

    nums = []
    for x in df["archive_file_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("A"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"A{next_num:04d}"


def save_archive_file(
    title,
    description,
    category_main,
    category_sub,
    tags,
    uploaded_file,
    visibility_type="normal",
    download_password="",
):
    df = get_archive_files_df()

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")
    archive_file_id = get_next_archive_file_id()

    file_bytes = uploaded_file.read()
    file_data_base64 = base64.b64encode(file_bytes).decode("utf-8")
    file_name = str(uploaded_file.name).strip()
    lower_name = file_name.lower()

    if lower_name.endswith(".xlsx"):
        file_type = "xlsx"
    elif lower_name.endswith(".xls"):
        file_type = "xls"
    elif lower_name.endswith(".pdf"):
        file_type = "pdf"
    elif lower_name.endswith(".docx"):
        file_type = "docx"
    elif lower_name.endswith(".doc"):
        file_type = "doc"
    elif lower_name.endswith(".jpg") or lower_name.endswith(".jpeg"):
        file_type = "jpg"
    elif lower_name.endswith(".png"):
        file_type = "png"
    else:
        file_type = "other"

    new_row = pd.DataFrame([{
        "archive_file_id": archive_file_id,
        "company_id": str(st.session_state.get("company_id", "")).strip(),
        "title": str(title).strip(),
        "description": str(description).strip(),
        "category_main": str(category_main).strip(),
        "category_sub": str(category_sub).strip(),
        "tags": str(tags).strip(),
        "file_name": file_name,
        "file_data": file_data_base64,
        "file_type": file_type,
        "uploaded_by_user_id": str(st.session_state.get("user_id", "")).strip(),
        "visibility_type": str(visibility_type).strip(),
        "download_password": str(download_password).strip(),
        "is_deleted": "0",
        "created_at": now_str,
        "updated_at": now_str,
        "deleted_by_user_id": "",
        "deleted_at": "",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "archive_files")
    return archive_file_id


def soft_delete_archive_file(archive_file_id):
    df = get_archive_files_df()
    if df is None or df.empty:
        return False

    mask = df["archive_file_id"].astype(str) == str(archive_file_id).strip()
    if not mask.any():
        return False

    df.loc[mask, "is_deleted"] = "1"
    df.loc[mask, "deleted_by_user_id"] = str(st.session_state.get("user_id", "")).strip()
    df.loc[mask, "deleted_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")
    df.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(df, "archive_files")
    return True


def get_archive_download_data(row):
    file_data_base64 = str(row.get("file_data", "")).strip()
    file_name = str(row.get("file_name", "")).strip()

    if not file_data_base64 or not file_name:
        return None, None, None

    file_bytes = base64.b64decode(file_data_base64)

    lower_name = file_name.lower()
    if lower_name.endswith(".xlsx"):
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif lower_name.endswith(".xls"):
        mime = "application/vnd.ms-excel"
    elif lower_name.endswith(".pdf"):
        mime = "application/pdf"
    elif lower_name.endswith(".docx"):
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif lower_name.endswith(".doc"):
        mime = "application/msword"
    elif lower_name.endswith(".jpg") or lower_name.endswith(".jpeg"):
        mime = "image/jpeg"
    elif lower_name.endswith(".png"):
        mime = "image/png"
    else:
        mime = "application/octet-stream"

    return file_bytes, file_name, mime

def render_archive_page():
    st.title("📤 書類アップロード")
    st.caption("この事業所だけで共有する資料置き場です。")

    current_company_id = str(st.session_state.get("company_id", "")).strip()
    current_user_id = str(st.session_state.get("user_id", "")).strip()
    is_admin = bool(st.session_state.get("is_admin", False))

    top_cols = st.columns([1, 1])

    with top_cols[0]:
        if st.button("← 休憩室へ戻る", key="back_from_archive", use_container_width=True):
            st.session_state.current_page = "休憩室"
            st.rerun()

    with top_cols[1]:
        st.info(f"事業所: {st.session_state.get('company_name', '')}")

    st.divider()

    with st.expander("＋ 新しい資料を登録する"):
        title = st.text_input("タイトル", key="archive_title")
        description = st.text_area("説明", key="archive_description", height=80)
        category_main = st.text_input("カテゴリ大", key="archive_category_main")
        category_sub = st.text_input("カテゴリ小", key="archive_category_sub")
        tags = st.text_input("タグ（カンマ区切りでもOK）", key="archive_tags")
        uploaded_file = st.file_uploader(
            "ファイルを選択",
            key="archive_uploaded_file"
        )

        if st.button("書庫へ保存", key="save_archive_button", use_container_width=True):
            if not title.strip():
                st.error("タイトルを入れてください。")
            elif uploaded_file is None:
                st.error("ファイルを選んでください。")
            else:
                archive_file_id = save_archive_file(
                    title=title,
                    description=description,
                    category_main=category_main,
                    category_sub=category_sub,
                    tags=tags,
                    uploaded_file=uploaded_file,
                    visibility_type="normal",
                    download_password="",
                )
                st.success(f"保存しました！ {archive_file_id}")
                st.rerun()

    st.divider()

    df = get_archive_files_df()

    if df is None or df.empty:
        st.info("まだ書庫に資料がありません。")
        return

    work = df.copy()
    work = work[
        (work["company_id"].astype(str) == current_company_id) &
        (work["is_deleted"].astype(str) != "1")
    ].copy()

    if work.empty:
        st.info("この事業所の書庫にはまだ資料がありません。")
        return

    search_cols = st.columns([2, 1, 1])

    with search_cols[0]:
        keyword = st.text_input("検索", key="archive_search_keyword")

    with search_cols[1]:
        filter_main = st.text_input("カテゴリ大で絞る", key="archive_filter_main")

    with search_cols[2]:
        filter_sub = st.text_input("カテゴリ小で絞る", key="archive_filter_sub")

    if keyword.strip():
        kw = keyword.strip()
        work = work[
            work["title"].astype(str).str.contains(kw, case=False, na=False) |
            work["description"].astype(str).str.contains(kw, case=False, na=False) |
            work["tags"].astype(str).str.contains(kw, case=False, na=False) |
            work["file_name"].astype(str).str.contains(kw, case=False, na=False)
        ].copy()

    if filter_main.strip():
        work = work[
            work["category_main"].astype(str).str.contains(filter_main.strip(), case=False, na=False)
        ].copy()

    if filter_sub.strip():
        work = work[
            work["category_sub"].astype(str).str.contains(filter_sub.strip(), case=False, na=False)
        ].copy()

    try:
        work = work.sort_values(["updated_at", "created_at"], ascending=[False, False])
    except Exception:
        pass

    st.markdown(f"### 一覧（{len(work)}件）")

    if work.empty:
        st.info("条件に合う資料がありません。")
        return

    users_df = get_users_df()

    for _, row in work.iterrows():
        archive_file_id = str(row.get("archive_file_id", "")).strip()
        title = str(row.get("title", "")).strip()
        description = str(row.get("description", "")).strip()
        category_main = str(row.get("category_main", "")).strip()
        category_sub = str(row.get("category_sub", "")).strip()
        tags = str(row.get("tags", "")).strip()
        file_name = str(row.get("file_name", "")).strip()
        file_type = str(row.get("file_type", "")).strip()
        uploaded_by_user_id = str(row.get("uploaded_by_user_id", "")).strip()
        created_at = str(row.get("created_at", "")).strip()
        updated_at = str(row.get("updated_at", "")).strip()


        uploader_name = uploaded_by_user_id
        try:
            target_user = users_df[users_df["user_id"].astype(str) == uploaded_by_user_id]
            if not target_user.empty:
                uploader_name = str(target_user.iloc[0].get("display_name", uploaded_by_user_id)).strip()
        except Exception:
            pass

        st.markdown("---")
        st.markdown(f"## {title}")
        st.caption(f"{archive_file_id} / {file_name} / {file_type}")

        meta_cols = st.columns([2, 2, 2])
        with meta_cols[0]:
            st.write(f"カテゴリ: {category_main} / {category_sub}")
        with meta_cols[1]:
            st.write(f"登録者: {uploader_name}")
        with meta_cols[2]:
            st.write(f"更新: {updated_at or created_at}")

        if description:
            st.write(description)
        if tags:
            st.caption(f"タグ: {tags}")

        file_bytes, dl_name, mime = get_archive_download_data(row)

        action_cols = st.columns([1, 1, 1])

        with action_cols[0]:
            if file_bytes is not None:
                st.download_button(
                    label="ダウンロード",
                    data=file_bytes,
                    file_name=dl_name,
                    mime=mime,
                    key=f"archive_download_{archive_file_id}",
                    use_container_width=True
                )

        can_delete = (uploaded_by_user_id == current_user_id) or (
            is_admin and str(row.get("company_id", "")).strip() == current_company_id
        )

        with action_cols[1]:
            if can_delete:
                if st.button("削除", key=f"archive_delete_{archive_file_id}", use_container_width=True):
                    ok = soft_delete_archive_file(archive_file_id)
                    if ok:
                        create_admin_log(
                            action_type="delete_archive_file",
                            target_type="archive_file",
                            target_id=archive_file_id,
                            action_detail=f"title={title}"
                        )
                        st.success("削除しました。")
                        st.rerun()
                    else:
                        st.error("削除に失敗しました。")

        with action_cols[2]:
            st.write("")

@st.cache_data(ttl=60)
def get_chat_messages_df_cached():
    df = load_db("chat_messages")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "message_id",
            "room_id",
            "user_id",
            "display_name",
            "company_id",
            "message_text",
            "has_attachment",
            "attachment_type",
            "linked_file_id",
            "is_deleted",
            "created_at",
            "updated_at",
            "deleted_by_user_id",
            "deleted_at",
        ])
    else:
        for col in [
            "message_id",
            "room_id",
            "user_id",
            "display_name",
            "company_id",
            "message_text",
            "has_attachment",
            "attachment_type",
            "linked_file_id",
            "is_deleted",
            "created_at",
            "updated_at",
            "deleted_by_user_id",
            "deleted_at",
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def get_chat_messages_df():
    return get_chat_messages_df_cached().copy()

def get_next_room_id():
    df = get_chat_rooms_df()
    if df is None or df.empty:
        return "R0001"

    nums = []
    for x in df["room_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("R"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"R{next_num:04d}"


def get_next_message_id():
    df = get_chat_messages_df()
    if df is None or df.empty:
        return "M0001"

    nums = []
    for x in df["message_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("M"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"M{next_num:04d}"


def create_chat_room(
    room_name,
    room_type,
    room_password="",
    description="",
):
    df = get_chat_rooms_df()

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")
    room_id = get_next_room_id()

    new_row = pd.DataFrame([{
        "room_id": room_id,
        "room_name": str(room_name).strip(),
        "room_type": str(room_type).strip(),
        "room_password": str(room_password).strip(),
        "created_by_user_id": str(st.session_state.get("user_id", "")).strip(),
        "created_by_company_id": str(st.session_state.get("company_id", "")).strip(),
        "description": str(description).strip(),
        "status": "active",
        "created_at": now_str,
        "updated_at": now_str,
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "chat_rooms")
    return room_id


def create_chat_message(room_id, message_text, attached_file=None):
    df = get_chat_messages_df()

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")
    message_id = get_next_message_id()

    has_attachment = "0"
    attachment_type = ""
    linked_file_id = ""

    if attached_file is not None:
        # チャット添付は倉庫へ自動保存
        visibility_type = "public"
        room_df = get_chat_rooms_df()
        room_row = room_df[room_df["room_id"].astype(str) == str(room_id).strip()].copy()

        if not room_row.empty:
            room_type = str(room_row.iloc[0].get("room_type", "")).strip()
            room_pw = str(room_row.iloc[0].get("room_password", "")).strip()

            if room_type == "limited":
                visibility_type = "limited"
            else:
                visibility_type = "public"

            linked_file_id = save_warehouse_file(
                title=f"[チャット添付] {attached_file.name}",
                description=f"チャットルーム {room_id} から自動保存",
                category_main="チャット添付",
                category_sub=str(room_id).strip(),
                tags="チャット添付,自動保存",
                uploaded_file=attached_file,
                visibility_type=visibility_type,
                download_password=room_pw if room_type == "limited" else "",
                is_searchable="1",
                source_room_id=str(room_id).strip(),
            )

            has_attachment = "1"
            lower_name = str(attached_file.name).lower()
            if "." in lower_name:
                attachment_type = lower_name.rsplit(".", 1)[-1]
            else:
                attachment_type = "other"

    new_row = pd.DataFrame([{
        "message_id": message_id,
        "room_id": str(room_id).strip(),
        "user_id": str(st.session_state.get("user_id", "")).strip(),
        "display_name": str(st.session_state.get("user", "")).strip(),
        "company_id": str(st.session_state.get("company_id", "")).strip(),
        "message_text": str(message_text).strip(),
        "has_attachment": has_attachment,
        "attachment_type": attachment_type,
        "linked_file_id": linked_file_id,
        "is_deleted": "0",
        "created_at": now_str,
        "updated_at": now_str,
        "deleted_by_user_id": "",
        "deleted_at": "",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "chat_messages")
    return message_id

def render_break_room_page():
    st.title("☕ 休憩室")
    st.caption("ここから チャットルーム・書類アップロード・倉庫 に入る。")

    st.markdown(
        f"""
        **現在の事業所**: {st.session_state.get("company_name", "")}  
        **ログイン中**: {st.session_state.get("user", "")}
        """
    )

    st.divider()

    cols = st.columns(3)

    with cols[0]:
        st.markdown("## 🚪 チャットルーム")
        st.caption("全事業所共通の交流・共有スペースです。")
        if st.button("チャットルームへ", key="go_chat_rooms", use_container_width=True):
            st.session_state.current_page = "休憩室_チャットルーム"
            st.rerun()

    with cols[1]:
        st.markdown("## 🚪 書類アップロード")
        st.caption("この事業所だけで共有する資料を登録・閲覧できます。")
        if st.button("書類アップロードへ", key="go_archive_page", use_container_width=True):
            st.session_state.current_page = "休憩室_書類アップロード"
            st.rerun()

    with cols[2]:
        st.markdown("## 🚪 倉庫")
        st.caption("全事業所共通の資料置き場です。")
        if st.button("倉庫へ", key="go_warehouse_page", use_container_width=True):
            st.session_state.current_page = "休憩室_倉庫"
            st.rerun()


def render_contact_page():
    st.title("📩 お問い合わせ")
    st.caption("不具合報告・ご要望・導入相談などを送れます。")

    company_id = str(st.session_state.get("company_id", "")).strip()
    company_name = str(st.session_state.get("company_name", "")).strip()
    user_id = str(st.session_state.get("user_id", "")).strip()
    user_name = str(st.session_state.get("user", "")).strip()

    top_cols = st.columns([1, 1])

    with top_cols[0]:
        if st.button("← 戻る", key="back_from_contact_page", use_container_width=True):
            st.session_state.current_page = "① 未着手の任務（掲示板）"
            st.rerun()

    with top_cols[1]:
        st.info(f"事業所: {company_name} / ログイン中: {user_name}")

    st.divider()

    contact_type = st.selectbox(
        "お問い合わせ種別",
        ["不具合", "使い方", "要望", "導入相談", "その他"],
        key="contact_type"
    )

    message = st.text_area(
        "内容",
        height=220,
        key="contact_message",
        placeholder="ここに要望や不具合内容を書いてください"
    )

    if st.button("送信する", key="send_contact_message", use_container_width=True):
        if not str(message).strip():
            st.error("内容を入力してください。")
        else:
            df = load_db("contact_messages")
            if df is None or df.empty:
                df = pd.DataFrame(columns=[
                    "id",
                    "company_id",
                    "company_name",
                    "user_id",
                    "user_name",
                    "contact_type",
                    "message",
                    "status",
                    "created_at",
                ])

            new_id = f"C{int(time.time())}"

            new_row = pd.DataFrame([{
                "id": new_id,
                "company_id": company_id,
                "company_name": company_name,
                "user_id": user_id,
                "user_name": user_name,
                "contact_type": str(contact_type).strip(),
                "message": str(message).strip(),
                "status": "未対応",
                "created_at": now_jst().strftime("%Y-%m-%d %H:%M:%S"),
            }])

            df = pd.concat([df, new_row], ignore_index=True)
            save_db(df, "contact_messages")

            st.success("お問い合わせを送信しました！")
            st.session_state.contact_message = ""
            st.rerun()

def render_chat_room_page():
    import html
    import re
    import pandas as pd
    import streamlit as st
    import streamlit.components.v1 as components

    st.title("💬 チャットルーム")
    st.caption("ルーム一覧・新規作成・投稿ができます。")

    rooms_df = get_chat_rooms_df()
    msgs_df = get_chat_messages_df()

    if rooms_df is None:
        rooms_df = pd.DataFrame()
    if msgs_df is None:
        msgs_df = pd.DataFrame()

    if "selected_room_id" not in st.session_state:
        st.session_state.selected_room_id = ""

    if "pending_room_id" not in st.session_state:
        st.session_state.pending_room_id = ""

    if "pending_room_type" not in st.session_state:
        st.session_state.pending_room_type = ""

    top_cols = st.columns([1, 1])

    with top_cols[0]:
        if st.button(
            "← 休憩室へ戻る",
            key="back_break_room",
            width="stretch",
            type="secondary",
        ):
            st.session_state.current_page = "休憩室"
            st.rerun()

    with top_cols[1]:
        if st.button(
            "選択中ルームを解除",
            key="clear_selected_room",
            width="stretch",
            type="secondary",
        ):
            st.session_state.selected_room_id = ""
            st.session_state.pending_room_id = ""
            st.session_state.pending_room_type = ""
            st.rerun()

    st.divider()

    with st.expander("＋ 新しいルームを作る"):
        room_name = st.text_input("ルーム名", key="new_room_name")
        room_type = st.selectbox(
            "公開設定",
            ["public", "limited"],
            format_func=lambda x: "公開ルーム" if x == "public" else "制限ルーム",
            key="new_room_type",
        )
        room_password = st.text_input(
            "ルームパスワード（制限ルーム用）",
            key="new_room_password",
        )
        room_description = st.text_area("説明", key="new_room_description", height=80)

        if st.button(
            "ルームを作成",
            key="create_new_room_button",
            width="stretch",
            type="secondary",
        ):
            if not str(room_name).strip():
                st.error("ルーム名を入れてください。")
            elif room_type == "limited" and not str(room_password).strip():
                st.error("制限ルームにはパスワードが必要です。")
            else:
                new_room_id = create_chat_room(
                    room_name=room_name,
                    room_type=room_type,
                    room_password=room_password,
                    description=room_description,
                )
                st.session_state.selected_room_id = new_room_id
                st.session_state.pending_room_id = ""
                st.session_state.pending_room_type = ""
                st.success(f"ルーム作成完了: {new_room_id}")
                st.rerun()

    st.divider()

    def clean_plain_text(value):
        text = "" if value is None else str(value)
        text = re.sub(r"<[^>]*>", "", text)
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        return text.strip()

    def esc_text(value):
        return html.escape(clean_plain_text(value))

    rooms_col, chat_col = st.columns([1, 2])

    with rooms_col:
        st.markdown("### ルーム一覧")

        if rooms_df.empty:
            st.info("まだルームがありません。")
        else:
            work = rooms_df.copy()

            if "status" in work.columns:
                work = work[
                    work["status"].astype(str).str.strip().str.lower() == "active"
                ].copy()

            if "room_type" in work.columns:
                work["room_type"] = (
                    work["room_type"].fillna("").astype(str).str.strip().str.lower()
                )
                work = work[work["room_type"].isin(["public", "limited"])].copy()

            try:
                if "created_at" in work.columns:
                    work = work.sort_values(["created_at"], ascending=[False])
            except Exception:
                pass

            for _, row in work.iterrows():
                room_id = str(row.get("room_id", "")).strip()
                room_name = clean_plain_text(row.get("room_name", ""))
                room_type = str(row.get("room_type", "")).strip().lower()
                desc = clean_plain_text(row.get("description", ""))

                is_selected = (
                    str(st.session_state.get("selected_room_id", "")).strip() == room_id
                )

                label = "🟢 公開ルーム" if room_type == "public" else "🩷 制限ルーム"

                with st.container(border=True):
                    st.markdown(f"### {(room_name or '名称なし')}{' ✓' if is_selected else ''}")
                    st.caption(label)
                    st.write(desc if desc else "説明なし")
                    st.caption(f"ID: {room_id}")

                    if st.button(
                        "詳細を見る",
                        key=f"select_room_{room_id}",
                        width="stretch",
                        type="primary" if is_selected else "secondary",
                    ):
                        if room_type == "limited":
                            st.session_state.pending_room_id = room_id
                            st.session_state.pending_room_type = room_type
                        else:
                            st.session_state.selected_room_id = room_id
                            st.session_state.pending_room_id = ""
                            st.session_state.pending_room_type = ""
                        st.rerun()

        if st.session_state.get("pending_room_id"):
            st.divider()
            st.markdown("### パスワード入力")
            pw = st.text_input(
                "ルームパスワード",
                type="password",
                key="room_access_password",
            )

            if st.button(
                "入室する",
                key="enter_limited_room",
                width="stretch",
                type="secondary",
            ):
                pending_room_id = str(st.session_state.get("pending_room_id", "")).strip()

                if rooms_df.empty:
                    st.error("ルームが見つかりません。")
                else:
                    target = rooms_df[
                        rooms_df["room_id"].astype(str).str.strip() == pending_room_id
                    ].copy()

                    if target.empty:
                        st.error("ルームが見つかりません。")
                    else:
                        real_pw = str(target.iloc[0].get("room_password", "")).strip()
                        if str(pw).strip() == real_pw:
                            st.session_state.selected_room_id = pending_room_id
                            st.session_state.pending_room_id = ""
                            st.session_state.pending_room_type = ""
                            st.success("入室できました。")
                            st.rerun()
                        else:
                            st.error("パスワードが違います。")

    with chat_col:
        selected_room_id = str(st.session_state.get("selected_room_id", "")).strip()

        if not selected_room_id:
            st.info("左からルームを選んでください。")
            return

        if rooms_df.empty:
            st.warning("選択中ルームが見つかりません。")
            return

        room_row = rooms_df[
            rooms_df["room_id"].astype(str).str.strip() == selected_room_id
        ].copy()

        if room_row.empty:
            st.warning("選択中ルームが見つかりません。")
            return

        room_name = clean_plain_text(room_row.iloc[0].get("room_name", ""))
        room_type = clean_plain_text(room_row.iloc[0].get("room_type", ""))
        room_desc = clean_plain_text(room_row.iloc[0].get("description", ""))

        st.markdown(f"## {room_name if room_name else '名称なし'}")
        st.caption(f"公開設定: {room_type if room_type else '-'}")
        if room_desc:
            st.write(room_desc)

        st.divider()

        text_nonce_key = f"chat_text_nonce_{selected_room_id}"
        uploader_nonce_key = f"chat_attach_nonce_{selected_room_id}"

        if text_nonce_key not in st.session_state:
            st.session_state[text_nonce_key] = 0

        if uploader_nonce_key not in st.session_state:
            st.session_state[uploader_nonce_key] = 0

        text_key = f"chat_post_text_{selected_room_id}_{st.session_state[text_nonce_key]}"
        uploader_key = f"chat_attach_{selected_room_id}_{st.session_state[uploader_nonce_key]}"

        room_msgs = msgs_df.copy()
        if not room_msgs.empty:
            room_msgs = room_msgs[
                room_msgs["room_id"].astype(str).str.strip() == selected_room_id
            ].copy()

            if "is_deleted" in room_msgs.columns:
                room_msgs = room_msgs[
                    room_msgs["is_deleted"].astype(str).str.strip() != "1"
                ].copy()

            try:
                room_msgs = room_msgs.copy()

                room_msgs["created_at_dt"] = pd.to_datetime(
                    room_msgs["created_at"].astype(str).str.strip(),
                    errors="coerce"
                )

                if "message_id" in room_msgs.columns:
                    room_msgs["message_id_num"] = (
                        room_msgs["message_id"]
                        .astype(str)
                        .str.replace("M", "", regex=False)
                        .str.strip()
                    )
                    room_msgs["message_id_num"] = pd.to_numeric(
                        room_msgs["message_id_num"],
                        errors="coerce"
                    )

                    room_msgs = room_msgs.sort_values(
                        ["created_at_dt", "message_id_num"],
                        ascending=[True, True]
                    )
                else:
                    room_msgs = room_msgs.sort_values(
                        ["created_at_dt"],
                        ascending=[True]
                    )
            except Exception:
                pass

        # -----------------------------
        # 上：投稿一覧
        # -----------------------------
        st.markdown("### 投稿一覧")

        if room_msgs.empty:
            st.info("まだ投稿がありません。")
        else:
            with st.container(height=800, border=True):
                current_user_name = str(st.session_state.get("user", "")).strip()
                current_user_id = str(st.session_state.get("user_id", "")).strip()

                warehouse_df = get_warehouse_files_df()
                if warehouse_df is None:
                    warehouse_df = pd.DataFrame()
                else:
                    warehouse_df = warehouse_df.copy()
                    if "is_deleted" in warehouse_df.columns:
                        warehouse_df = warehouse_df[
                            warehouse_df["is_deleted"].astype(str).str.strip() != "1"
                        ].copy()

                for _, msg in room_msgs.iterrows():
                    display_name = clean_plain_text(msg.get("display_name", ""))
                    message_user_id = str(msg.get("user_id", "")).strip()
                    message_text = clean_plain_text(msg.get("message_text", ""))
                    created_at = str(msg.get("created_at", "")).strip()
                    has_attachment = str(msg.get("has_attachment", "")).strip()
                    linked_file_id = clean_plain_text(msg.get("linked_file_id", ""))

                    attached_row = None
                    attached_file_bytes = None
                    attached_file_name = ""
                    attached_file_mime = ""

                    if linked_file_id and not warehouse_df.empty:
                        hit_file = warehouse_df[
                            warehouse_df["file_id"].astype(str).str.strip() == linked_file_id
                        ].copy()

                        if not hit_file.empty:
                            attached_row = hit_file.iloc[0]
                            attached_file_bytes, attached_file_name, attached_file_mime = get_warehouse_download_data(attached_row)

                    is_me = False
                    if current_user_id and message_user_id:
                        is_me = (current_user_id == message_user_id)
                    elif current_user_name and display_name:
                        is_me = (current_user_name == display_name)

                    time_text = created_at[-8:] if created_at and len(created_at) >= 8 else created_at

                    left_msg_col, spacer_col, right_msg_col = st.columns([5, 1, 5])

                    normalized_message_text = str(message_text).strip()

                    looks_like_broken_html = normalized_message_text in [
                        "</div>",
                        "<div>",
                        "<br>",
                        "</span>",
                        "<span>",
                    ]

                    is_file_only = False
                    if has_attachment == "1" and (not normalized_message_text or looks_like_broken_html):
                        is_file_only = True

                    if is_file_only:
                        bubble_text = "＜添付ファイルを送信しました＞"
                    elif normalized_message_text:
                        bubble_text = normalized_message_text
                    else:
                        bubble_text = "　"

                    bubble_html = esc_text(bubble_text).replace("\n", "<br>")

                    if is_me:
                        with right_msg_col:
                            should_show_bubble = not is_file_only and bool(str(bubble_text).strip())

                            if should_show_bubble:
                                my_bg_color = "#95EC69"
                                my_text_color = "#111827"
                                my_font_weight = "normal"

                                st.markdown(
                                    f"""<div style="
                                        background:{my_bg_color};
                                        color:{my_text_color};
                                        padding:10px 14px;
                                        border-radius:18px;
                                        border-bottom-right-radius:6px;
                                        margin:6px 0 2px auto;
                                        display:inline-block;
                                        max-width:70%;
                                        word-break:break-word;
                                        overflow-wrap:break-word;
                                        font-weight:{my_font_weight};
                                    ">{bubble_html}</div>
                                    """,
                                    unsafe_allow_html=True,
                                )

                            if attached_file_bytes is not None and attached_file_name:
                                st.download_button(
                                    label=f"📎 {attached_file_name}",
                                    data=attached_file_bytes,
                                    file_name=attached_file_name,
                                    mime=attached_file_mime,
                                    key=f"chat_download_me_{msg.get('message_id', '')}",
                                    use_container_width=False,
                                )

                            if time_text:
                                st.markdown(
                                    f"""<div style="
                                        font-size:11px;
                                        color:#777;
                                        text-align:right;
                                        margin-top:2px;
                                    ">
                                        {time_text}
                                    </div>""",
                                    unsafe_allow_html=True
                                )

                    else:
                        with left_msg_col:
                            if display_name:
                                st.caption(display_name)

                            should_show_bubble = not is_file_only and bool(str(bubble_text).strip())

                            if should_show_bubble:
                                other_bg_color = "#FFFFFF"
                                other_text_color = "#111827"
                                other_font_weight = "normal"
                                other_border = "1px solid #DADADA"

                                st.markdown(
                                    f"""<div style="
                                        background:{other_bg_color};
                                        color:{other_text_color};
                                        border:{other_border};
                                        padding:10px 14px;
                                        border-radius:18px;
                                        border-bottom-left-radius:6px;
                                        margin:6px auto 2px 0;
                                        display:inline-block;
                                        max-width:70%;
                                        word-break:break-word;
                                        overflow-wrap:break-word;
                                        font-weight:{other_font_weight};
                                    ">{bubble_html}</div>
                                    """,
                                    unsafe_allow_html=True,
                                )

                            if attached_file_bytes is not None and attached_file_name:
                                st.download_button(
                                    label=f"📎 {attached_file_name}",
                                    data=attached_file_bytes,
                                    file_name=attached_file_name,
                                    mime=attached_file_mime,
                                    key=f"chat_download_other_{msg.get('message_id', '')}",
                                    use_container_width=False,
                                )

                            if time_text:
                                st.markdown(
                                    f"""<div style="
                                        font-size:11px;
                                        color:#777;
                                        text-align:left;
                                        margin-top:2px;
                                    ">
                                        {time_text}
                                    </div>""",
                                    unsafe_allow_html=True
                                )
                            else:
                                st.markdown(
                                    f"""
                                    <div style="
                                        background:{other_bg_color};
                                        color:{other_text_color};
                                        border:{other_border};
                                        padding:10px 14px;
                                        border-radius:18px;
                                        border-bottom-left-radius:6px;
                                        margin:6px auto 2px 0;
                                        display:inline-block;
                                        max-width:70%;
                                        word-break:break-word;
                                        overflow-wrap:break-word;
                                        font-weight:{other_font_weight};
                                    ">{bubble_html}</div>
                                    """,
                                    unsafe_allow_html=True,
                                )

                            if attached_file_bytes is not None and attached_file_name:
                                st.download_button(
                                    label=f"📎 {attached_file_name}",
                                    data=attached_file_bytes,
                                    file_name=attached_file_name,
                                    mime=attached_file_mime,
                                    key=f"chat_download_other_{msg.get('message_id', '')}",
                                    use_container_width=False,
                                )

                            if time_text:
                                st.markdown(
                                    f"""<div style="
                                        font-size:11px;
                                        color:#777;
                                        text-align:left;
                                        margin-top:2px;
                                    ">
                                        {time_text}
                                    </div>""",
                                    unsafe_allow_html=True
                                )

                st.markdown('<div id="chat-bottom-anchor"></div>', unsafe_allow_html=True)

                components.html(
                    """
                    <script>
                    const anchor = window.parent.document.getElementById("chat-bottom-anchor");
                    if (anchor) {
                        anchor.scrollIntoView({ behavior: "auto", block: "end" });
                    }
                    </script>
                    """,
                    height=0,
                )

        st.divider()

        # -----------------------------
        # 下：メッセージ入力部分
        # -----------------------------
        post_text = st.text_area(
            "メッセージ",
            key=text_key,
            height=100,
        )
        attached_file = st.file_uploader(
            "添付ファイル（あれば倉庫へ自動保存）",
            key=uploader_key,
        )

        if st.button(
            "投稿する",
            key="chat_post_button",
            width="stretch",
            type="secondary",
        ):
            post_text_clean = clean_plain_text(post_text)

            if not post_text_clean and attached_file is None:
                st.error("メッセージか添付のどちらかを入れてください。")
            else:
                create_chat_message(
                    selected_room_id,
                    post_text_clean,
                    attached_file=attached_file,
                )

                st.session_state[text_nonce_key] += 1
                st.session_state[uploader_nonce_key] += 1

                st.success("投稿しました！")
                st.rerun()

        st.markdown(
            """
            <div style="
                background:#F3F4F6;
                border:1px solid #D1D5DB;
                border-radius:14px;
                padding:10px 12px;
                margin-bottom:10px;
                color:#6B7280;
                font-size:12px;
            ">
                チャット履歴
            </div>
            """,
            unsafe_allow_html=True,
        )
                        
def render_other_office_register_page():
    st.title("🪪 他事業所へ登録")
    st.caption("現在ログインしている自分を、別の事業所にも登録するページです。")

    top_cols = st.columns([1, 1])

    with top_cols[0]:
        if st.button("← 戻る", key="back_from_other_office_register", use_container_width=True):
            st.session_state.current_page = "① 未着手の任務（掲示板）"
            st.rerun()

    with top_cols[1]:
        st.info(f"現在ログイン中: {st.session_state.get('company_name', '')} / {st.session_state.get('user', '')}")

    st.divider()

    st.write("### 他の事業所へ登録する")

    target_company_login_id = st.text_input("事業所ID", key="cross_reg_company_login_id")
    target_company_login_password = st.text_input("事業所パスワード", type="password", key="cross_reg_company_login_pw")
    entered_user_login_id = st.text_input("職員ID", key="cross_reg_user_login_id")
    entered_display_name = st.text_input("名前（ハンドルネーム）", key="cross_reg_display_name")

    if st.button("登録する", key="cross_reg_submit", use_container_width=True):
        ok, msg = register_current_user_to_other_company(
            target_company_login_id=target_company_login_id,
            target_company_login_password=target_company_login_password,
            entered_user_login_id=entered_user_login_id,
            entered_display_name=entered_display_name,
        )
        if ok:
            st.success(msg)
        else:
            st.warning(msg)

def render_company_knowbe_settings_page():
    st.title("🔐 Knowbe情報登録")
    st.caption("現在ログイン中の事業所に、Knowbeログイン情報を保存するページです。")

    current_company_id = str(st.session_state.get("company_id", "")).strip()
    current_company_name = str(st.session_state.get("company_name", "")).strip()

    saved_user, saved_pw = get_company_saved_knowbe_info(current_company_id)

    st.info(f"対象事業所: {current_company_name}")

    verify_cols = st.columns(2)
    with verify_cols[0]:
        verify_company_login_id = st.text_input("事業所ID（確認用）", key="knowbe_setting_verify_company_login_id")
    with verify_cols[1]:
        verify_company_login_password = st.text_input("事業所パスワード（確認用）", type="password", key="knowbe_setting_verify_company_login_password")

    input_cols = st.columns(2)
    with input_cols[0]:
        knowbe_login_username = st.text_input(
            "knowbeアカウント名",
            value=saved_user,
            key="knowbe_setting_login_username"
        )
    with input_cols[1]:
        knowbe_login_password = st.text_input(
            "knowbeパスワード",
            type="password",
            value=saved_pw,
            key="knowbe_setting_login_password"
        )

    st.caption(f"現在保存中のアカウント名: {mask_secret_text(saved_user)}")

    btn_cols = st.columns([1, 4])
    with btn_cols[0]:
        if st.button("登録・更新", key="save_company_knowbe_settings", use_container_width=True):
            row = authenticate_company_login(verify_company_login_id, verify_company_login_password)
            if row is None:
                st.error("事業所IDまたは事業所パスワードが違います。")
            else:
                auth_company_id = str(row.get("company_id", "")).strip()
                if auth_company_id != current_company_id:
                    st.error("現在ログイン中の事業所と一致しません。")
                elif not str(knowbe_login_username).strip() or not str(knowbe_login_password).strip():
                    st.error("knowbeアカウント名とknowbeパスワードを両方入れてください。")
                else:
                    ok = save_company_saved_knowbe_info(
                        company_id=current_company_id,
                        knowbe_login_username=knowbe_login_username,
                        knowbe_login_password=knowbe_login_password,
                    )
                    if ok:
                        create_admin_log(
                            action_type="save_company_knowbe_settings",
                            target_type="company",
                            target_id=current_company_id,
                            action_detail=f"company_name={current_company_name}"
                        )
                        st.success("Knowbe情報を保存しました！")
                        st.rerun()
                    else:
                        st.error("保存に失敗しました。")

@st.cache_data(ttl=60)
def get_company_permissions_df(company_id: str):
    df = get_user_company_permissions_df()

    # st.write("DEBUG raw permissions df =", df)

    if df is None or df.empty:
        # st.write("DEBUG permissions df is empty")
        return pd.DataFrame(columns=[
            "permission_id",
            "user_id",
            "company_id",
            "can_use",
            "is_admin",
            "status",
            "created_at",
            "updated_at",
            "memo",
        ])

    work = df.copy()

    # st.write("DEBUG input company_id =", company_id)
    # st.write("DEBUG before normalize =", work)

    work["company_id"] = work["company_id"].fillna("").astype(str).str.strip()
    work["user_id"] = work["user_id"].fillna("").astype(str).str.strip()

    # 1 / 1.0 / "1" 問題を全部吸収するです
    work["can_use"] = (
        pd.to_numeric(work["can_use"], errors="coerce")
        .fillna(0)
        .astype(int)
        .astype(str)
    )
    work["is_admin"] = (
        pd.to_numeric(work["is_admin"], errors="coerce")
        .fillna(0)
        .astype(int)
        .astype(str)
    )

    work["status"] = work["status"].fillna("").astype(str).str.strip().str.lower()

    # st.write("DEBUG company_id unique =", work["company_id"].unique())
    # st.write("DEBUG user_id unique =", work["user_id"].unique())
    # st.write("DEBUG can_use unique =", work["can_use"].unique())
    # st.write("DEBUG is_admin unique =", work["is_admin"].unique())
    # st.write("DEBUG status unique =", work["status"].unique())
    # st.write("DEBUG before filter =", work)

    work = work[
        (work["company_id"] == str(company_id).strip()) &
        (work["can_use"] == "1") &
        (work["status"] != "inactive")
    ].copy()

    # st.write("DEBUG after filter =", work)

    return work


def get_company_admin_count(company_id: str) -> int:
    perm_df = get_company_permissions_df(company_id)
    if perm_df.empty:
        return 0
    return int((perm_df["is_admin"].astype(str).str.strip() == "1").sum())


def company_has_any_admin(company_id: str) -> bool:
    return get_company_admin_count(company_id) > 0


def user_can_use_company(user_id: str, company_id: str) -> bool:
    perm_df = get_company_permissions_df(company_id)
    if perm_df.empty:
        return False

    target = perm_df[
        perm_df["user_id"].astype(str).str.strip() == str(user_id).strip()
    ]
    return not target.empty


def user_is_company_admin(user_id: str, company_id: str) -> bool:
    perm_df = get_company_permissions_df(company_id)
    if perm_df.empty:
        return False

    target = perm_df[
        (perm_df["user_id"].astype(str).str.strip() == str(user_id).strip()) &
        (perm_df["is_admin"].astype(str).str.strip() == "1")
    ]
    return not target.empty


def get_next_permission_id():
    df = get_user_company_permissions_df()
    if df is None or df.empty:
        return "P0001"

    nums = []
    for x in df["permission_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("P"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"P{next_num:04d}"

@st.cache_data(ttl=60)
def authenticate_company_login(login_id: str, login_password: str):
    df = get_companies_df()
    if df is None or df.empty:
        return None

    work = df.copy()
    work["company_login_id"] = work["company_login_id"].astype(str).str.strip()
    work["company_login_password"] = work["company_login_password"].astype(str).str.strip()
    work["status"] = work["status"].astype(str).str.strip().str.lower()

    target = work[
        (work["company_login_id"] == str(login_id).strip()) &
        (work["company_login_password"] == str(login_password).strip()) &
        (work["status"] == "active")
    ]

    if target.empty:
        return None

    return target.iloc[0].to_dict()

def get_current_company_id():
    return str(st.session_state.get("company_id", "")).strip()









def get_next_task_id(task_df=None):
    if task_df is None:
        task_df = get_tasks_df()

    if task_df is None or task_df.empty:
        return 1

    ids = pd.to_numeric(task_df["id"], errors="coerce").dropna()
    return int(ids.max()) + 1 if not ids.empty else 1


def get_diary_input_rules_df(company_id=None):
    if company_id is None:
        company_id = get_current_company_id()

    required_cols = [
        "record_id", "company_id", "date", "resident_id", "resident_name",
        "start_time", "end_time", "work_start_time", "work_end_time", "work_break_time",
        "meal_flag", "note",
        "start_memo", "end_memo", "staff_name",
        "generated_status", "generated_support", "created_at",
        "service_type", "knowbe_target", "send_status", "sent_at", "send_error",
        "record_mode"
    ]

    df = load_db("diary_input_rules")
    df = normalize_company_scoped_df(df, required_cols)
    return filter_by_company_id(df, company_id)

def get_staff_example_row(company_id: str, staff_name: str):
    df = load_db("staff_examples")

    if df is None or df.empty:
        return None

    df = df.fillna("").copy()

    for col in [
        "company_id",
        "staff_name",
        "home_start_example", "home_end_example",
        "day_start_example", "day_end_example",
        "outside_start_example", "outside_end_example",
        "updated_at",
    ]:
        if col not in df.columns:
            df[col] = ""

    df["company_id"] = df["company_id"].astype(str).str.strip()
    df["staff_name"] = df["staff_name"].astype(str).str.strip()

    hit = df[
        (df["company_id"] == str(company_id).strip()) &
        (df["staff_name"] == str(staff_name).strip())
    ].copy()

    if hit.empty:
        return None

    return hit.iloc[0].to_dict()

def get_personal_rule_row(company_id: str, staff_name: str):
    df = load_db("personal_rules")

    if df is None or df.empty:
        return None

    df = df.fillna("").copy()

    for col in [
        "company_id",
        "staff_name",
        "rule_text",
        "updated_at",
    ]:
        if col not in df.columns:
            df[col] = ""

    df["company_id"] = df["company_id"].astype(str).str.strip()
    df["staff_name"] = df["staff_name"].astype(str).str.strip()

    hit = df[
        (df["company_id"] == str(company_id).strip()) &
        (df["staff_name"] == str(staff_name).strip())
    ].copy()

    if hit.empty:
        return None

    return hit.iloc[0].to_dict()

def save_staff_examples_record(
    company_id,
    staff_name,
    home_start_example,
    home_end_example,
    day_start_example,
    day_end_example,
    outside_start_example,
    outside_end_example,
):
    df = load_db("staff_examples")
    required_cols = [
        "company_id",
        "staff_name",
        "home_start_example", "home_end_example",
        "day_start_example", "day_end_example",
        "outside_start_example", "outside_end_example",
        "updated_at"
    ]
    df = normalize_company_scoped_df(df, required_cols)

    company_id = str(company_id).strip()
    staff_name = str(staff_name).strip()
    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    mask = (
        (df["company_id"] == company_id) &
        (df["staff_name"].astype(str).str.strip() == staff_name)
    )

    new_data = {
        "company_id": company_id,
        "staff_name": staff_name,
        "home_start_example": str(home_start_example),
        "home_end_example": str(home_end_example),
        "day_start_example": str(day_start_example),
        "day_end_example": str(day_end_example),
        "outside_start_example": str(outside_start_example),
        "outside_end_example": str(outside_end_example),
        "updated_at": now_str,
    }

    if mask.any():
        for k, v in new_data.items():
            df.loc[mask, k] = v
    else:
        df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)

    save_db(df, "staff_examples")

def save_personal_rules_record(company_id, staff_name, rule_text):
    df = load_db("personal_rules")
    required_cols = [
        "company_id",
        "staff_name", "rule_text", "updated_at"
    ]
    df = normalize_company_scoped_df(df, required_cols)

    company_id = str(company_id).strip()
    staff_name = str(staff_name).strip()
    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    mask = (
        (df["company_id"] == company_id) &
        (df["staff_name"].astype(str).str.strip() == staff_name)
    )

    new_data = {
        "company_id": company_id,
        "staff_name": staff_name,
        "rule_text": str(rule_text),
        "updated_at": now_str,
    }

    if mask.any():
        for k, v in new_data.items():
            df.loc[mask, k] = v
    else:
        df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)

    save_db(df, "personal_rules")

def get_plan_row(company_id, resident_id):
    df = load_db("assistant_plans")

    if df is None or df.empty:
        return None

    df = df.fillna("").copy()

    for col in [
        "company_id",
        "resident_id",
        "long_term_goal",
        "short_term_goal",
        "updated_at",
    ]:
        if col not in df.columns:
            df[col] = ""

    df["company_id"] = df["company_id"].astype(str).str.strip()
    df["resident_id"] = df["resident_id"].astype(str).str.strip()

    hit = df[
        (df["company_id"] == str(company_id).strip()) &
        (df["resident_id"] == str(resident_id).strip())
    ].copy()

    if hit.empty:
        return None

    return hit.iloc[0].to_dict()

def get_company_row_by_id(company_id: str):
    df = get_companies_df()
    if df is None or df.empty:
        return None

    hit = df[df["company_id"].astype(str).str.strip() == str(company_id).strip()]
    if hit.empty:
        return None
    return hit.iloc[0].to_dict()


def get_current_company_row():
    current_company_id = get_current_company_id()
    if not current_company_id:
        return None
    return get_company_row_by_id(current_company_id)


def get_company_saved_knowbe_info(company_id: str):
    df = get_companies_df()
    if df is None or df.empty:
        return "", ""

    hit = df[df["company_id"].astype(str).str.strip() == str(company_id).strip()]
    if hit.empty:
        return "", ""

    row = hit.iloc[0]
    return (
        str(row.get("knowbe_login_username", "")).strip(),
        str(row.get("knowbe_login_password", "")).strip(),
    )


def save_company_saved_knowbe_info(company_id: str, knowbe_login_username: str, knowbe_login_password: str):
    df = get_companies_df()
    if df is None or df.empty:
        return False

    mask = df["company_id"].astype(str).str.strip() == str(company_id).strip()
    if not mask.any():
        return False

    df.loc[mask, "knowbe_login_username"] = str(knowbe_login_username).strip()
    df.loc[mask, "knowbe_login_password"] = str(knowbe_login_password).strip()
    df.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(df, "companies")
    return True


def reset_resident_edit_flags():
    for k in [
        "edit_basic_mode",
        "edit_schedule_mode",
        "edit_notes_mode",
        "edit_links_mode",
    ]:
        st.session_state[k] = False


def resolve_bee_company_context(
    company_login_id: str,
    company_login_password: str,
    knowbe_login_username: str,
    knowbe_login_password: str,
):
    company_login_id = str(company_login_id).strip()
    company_login_password = str(company_login_password).strip()
    knowbe_login_username = str(knowbe_login_username).strip()
    knowbe_login_password = str(knowbe_login_password).strip()

    use_manual_company = bool(company_login_id or company_login_password)

    if use_manual_company:
        if not company_login_id or not company_login_password:
            return {
                "ok": False,
                "error": "事業所IDと事業所パスワードは両方入れてください。",
            }

        row = authenticate_company_login(company_login_id, company_login_password)
        if row is None:
            return {
                "ok": False,
                "error": "事業所IDまたは事業所パスワードが違います。",
            }
    else:
        row = get_current_company_row()
        if row is None:
            return {
                "ok": False,
                "error": "現在の事業所情報が見つからありません。",
            }

    target_company_id = str(row.get("company_id", "")).strip()
    target_company_name = str(row.get("company_name", "")).strip()

    saved_knowbe_user, saved_knowbe_pw = get_company_saved_knowbe_info(target_company_id)

    final_knowbe_user = knowbe_login_username or saved_knowbe_user
    final_knowbe_pw = knowbe_login_password or saved_knowbe_pw

    return {
        "ok": True,
        "target_company_id": target_company_id,
        "target_company_name": target_company_name,
        "target_company_code": str(row.get("company_code", "")).strip(),
        "knowbe_login_username": final_knowbe_user,
        "knowbe_login_password": final_knowbe_pw,
        "using_saved_knowbe": bool((not knowbe_login_username and not knowbe_login_password) and final_knowbe_user and final_knowbe_pw),
        "has_knowbe_credentials": bool(final_knowbe_user and final_knowbe_pw),
    }

def authenticate_user_login(company_id: str, login_id: str, login_password: str):
    users_df = get_users_df()
    # st.write("DEBUG auth company_id =", company_id)
    # st.write("DEBUG auth login_id =", login_id)

    if users_df is None or users_df.empty:
        # st.write("DEBUG users empty")
        return None

    work = users_df.copy()
    work["user_login_id"] = work["user_login_id"].astype(str).str.strip()
    work["user_login_password"] = work["user_login_password"].astype(str).str.strip()
    work["status"] = work["status"].astype(str).str.strip().str.lower()

    # st.write("DEBUG users before filter =", work)

    target = work[
        (work["user_login_id"] == str(login_id).strip()) &
        (work["user_login_password"] == str(login_password).strip()) &
        (work["status"] == "active")
    ]

    # st.write("DEBUG auth matched users =", len(target))
    # st.write("DEBUG auth matched rows =", target)

    if target.empty:
        return None

    row = target.iloc[0].to_dict()
    user_id = str(row.get("user_id", "")).strip()

    # st.write("DEBUG auth user_id =", user_id)

    can_use = user_can_use_company(user_id, company_id)
    is_admin = user_is_company_admin(user_id, company_id)

    # st.write("DEBUG auth can_use =", can_use)
    # st.write("DEBUG auth is_admin =", is_admin)

    if not can_use:
        return None

    row["is_admin_resolved"] = is_admin
    return row

def get_next_user_id():
    df = get_users_df()
    if df is None or df.empty:
        return "U0001"

    nums = []
    for x in df["user_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("U"):
            num = x[1:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"U{next_num:04d}"


def create_user_with_permission(
    company_id: str,
    user_login_id: str,
    user_login_password: str,
    display_name: str,
    role_type: str = "職員",
    is_admin: str = "0",
):
    users_df = get_users_df()
    perm_df = get_user_company_permissions_df()
    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    login_id = str(user_login_id).strip()
    login_pw = str(user_login_password).strip()
    display_name = str(display_name).strip()
    role_type = str(role_type).strip()
    is_admin = "1" if str(is_admin).strip() == "1" else "0"

    if not login_id:
        return False, "IDを入れてください。"
    if not display_name:
        return False, "表示名を入れてください。"
    if not is_valid_user_password(login_pw):
        return False, "パスワードは8文字以上・英数混合・大文字必須です。"

    dup_user = users_df[
        users_df["user_login_id"].astype(str).str.strip() == login_id
    ]
    if not dup_user.empty:
        return False, "そのIDはすでに使われています。"

    user_id = get_next_user_id()

    new_user_row = pd.DataFrame([{
        "user_id": user_id,
        "company_id": str(company_id).strip(),
        "user_login_id": login_id,
        "user_login_password": login_pw,
        "display_name": display_name,
        "is_admin": is_admin,
        "role_type": role_type,
        "login_card_id": "",
        "last_login_at": "",
        "status": "active",
        "created_at": now_str,
        "updated_at": now_str,
        "memo": "",
    }])

    new_perm_row = pd.DataFrame([{
        "permission_id": get_next_permission_id(),
        "user_id": user_id,
        "company_id": str(company_id).strip(),
        "can_use": "1",
        "is_admin": is_admin,
        "status": "active",
        "created_at": now_str,
        "updated_at": now_str,
        "memo": "",
    }])

    users_df = pd.concat([users_df, new_user_row], ignore_index=True)
    perm_df = pd.concat([perm_df, new_perm_row], ignore_index=True)

    save_db(users_df, "users")
    save_db(perm_df, "user_company_permissions")

    return True, user_id

def register_current_user_to_other_company(
    target_company_login_id: str,
    target_company_login_password: str,
    entered_user_login_id: str,
    entered_display_name: str,
):
    current_company_id = str(st.session_state.get("company_id", "")).strip()
    current_user_id = str(st.session_state.get("user_id", "")).strip()

    if not current_user_id:
        return False, "個人ログイン情報が見つからありません。"

    target_company_login_id = str(target_company_login_id).strip()
    target_company_login_password = str(target_company_login_password).strip()
    entered_user_login_id = str(entered_user_login_id).strip()
    entered_display_name = str(entered_display_name).strip()

    if not target_company_login_id:
        return False, "事業所IDを入れてください。"
    if not target_company_login_password:
        return False, "事業所パスワードを入れてください。"
    if not entered_user_login_id:
        return False, "職員IDを入れてください。"
    if not entered_display_name:
        return False, "名前（ハンドルネーム）を入れてください。"

    # 入力された事業所ID/パスワードで対象事業所を認証
    company_row = authenticate_company_login(
        target_company_login_id,
        target_company_login_password
    )
    if company_row is None:
        return False, "事業所IDまたは事業所パスワードが違います。"

    target_company_id = str(company_row.get("company_id", "")).strip()
    target_company_name = str(company_row.get("company_name", "")).strip()

    if not target_company_id:
        return False, "対象事業所が見つからありません。"

    if target_company_id == current_company_id:
        return False, "現在ログインしている事業所と同じです。別事業所を入れてください。"

    users_df = get_users_df()
    my_row_df = users_df[
        users_df["user_id"].astype(str).str.strip() == current_user_id
    ].copy()

    if my_row_df.empty:
        return False, "現在の職員情報が見つからありません。"

    my_row = my_row_df.iloc[0]
    real_user_login_id = str(my_row.get("user_login_id", "")).strip()
    real_display_name = str(my_row.get("display_name", "")).strip()
    user_status = str(my_row.get("status", "")).strip().lower()

    if user_status != "active":
        return False, "この職員は現在有効ではありません。"

    if entered_user_login_id != real_user_login_id or entered_display_name != real_display_name:
        return False, "職員IDまたは名前（ハンドルネーム）が現在ログイン中の情報と一致しません。"

    perm_df = get_user_company_permissions_df()

    same_perm = perm_df[
        (perm_df["user_id"].astype(str).str.strip() == current_user_id) &
        (perm_df["company_id"].astype(str).str.strip() == target_company_id)
    ].copy()

    if not same_perm.empty:
        active_perm = same_perm[
            (same_perm["can_use"].astype(str).str.strip() == "1") &
            (same_perm["status"].astype(str).str.strip().str.lower() != "inactive")
        ]
        if not active_perm.empty:
            return False, "もうすでに登録されています"

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    new_perm_row = pd.DataFrame([{
        "permission_id": get_next_permission_id(),
        "user_id": current_user_id,
        "company_id": target_company_id,
        "can_use": "1",
        "is_admin": "0",
        "status": "active",
        "created_at": now_str,
        "updated_at": now_str,
        "memo": "cross_company_register",
    }])

    perm_df = pd.concat([perm_df, new_perm_row], ignore_index=True)
    save_db(perm_df, "user_company_permissions")

    create_admin_log(
        action_type="cross_company_register",
        target_type="company_permission",
        target_id=target_company_id,
        action_detail=f"user_id={current_user_id}, company_name={target_company_name}"
    )

    return True, f"{target_company_name} に登録できました。"

def set_company_user_status(user_id: str, company_id: str, new_status: str = "inactive"):
    users_df = get_users_df()
    perm_df = get_user_company_permissions_df()
    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    user_mask = users_df["user_id"].astype(str).str.strip() == str(user_id).strip()
    perm_mask = (
        (perm_df["user_id"].astype(str).str.strip() == str(user_id).strip()) &
        (perm_df["company_id"].astype(str).str.strip() == str(company_id).strip())
    )

    if not user_mask.any() or not perm_mask.any():
        return False, "対象スタッフが見つからありません。"

    users_df.loc[user_mask, "status"] = str(new_status).strip()
    users_df.loc[user_mask, "updated_at"] = now_str

    perm_df.loc[perm_mask, "status"] = str(new_status).strip()
    perm_df.loc[perm_mask, "updated_at"] = now_str

    save_db(users_df, "users")
    save_db(perm_df, "user_company_permissions")
    return True, "更新しました。"

def render_first_staff_register_block(key_prefix="first_staff"):
    company_id = str(st.session_state.get("company_id", "")).strip()

    st.info("この事業所にはまだ管理者がいありません。管理者が1人もいない間だけ、ここからスタッフ登録できます。")

    with st.expander("＋ スタッフ登録", expanded=False):
        display_name = st.text_input("表示名", key=f"{key_prefix}_display_name")
        user_login_id = st.text_input("ログインID", key=f"{key_prefix}_login_id")
        user_login_password = st.text_input("パスワード", type="password", key=f"{key_prefix}_login_pw")
        role_type = st.selectbox("権限", ["管理者", "職員"], key=f"{key_prefix}_role_type")

        if st.button("スタッフ登録", use_container_width=True, key=f"{key_prefix}_register_button"):
            is_admin = "1" if role_type == "管理者" else "0"

            ok, msg = create_user_with_permission(
                company_id=company_id,
                user_login_id=user_login_id,
                user_login_password=user_login_password,
                display_name=display_name,
                role_type=role_type,
                is_admin=is_admin,
            )
            if ok:
                create_admin_log(
                    action_type="first_staff_create",
                    target_type="user",
                    target_id=msg,
                    action_detail=f"display_name={display_name}, role_type={role_type}"
                )
                st.success(f"登録できました！ user_id={msg}")
                st.rerun()
            else:
                st.error(msg)


def render_admin_staff_manage_block():
    company_id = str(st.session_state.get("company_id", "")).strip()

    st.divider()
    st.subheader("👑 スタッフ管理")

    reg_tab, list_tab = st.tabs(["スタッフ登録", "スタッフ一覧"])

    with reg_tab:
        display_name = st.text_input("表示名", key="admin_staff_display_name")
        user_login_id = st.text_input("ログインID", key="admin_staff_login_id")
        user_login_password = st.text_input("パスワード", type="password", key="admin_staff_login_pw")
        role_type = st.selectbox("権限", ["管理者", "職員"], key="admin_staff_role_type")

        if st.button("登録する", use_container_width=True, key="admin_staff_register_button"):
            is_admin = "1" if role_type == "管理者" else "0"

            ok, msg = create_user_with_permission(
                company_id=company_id,
                user_login_id=user_login_id,
                user_login_password=user_login_password,
                display_name=display_name,
                role_type=role_type,
                is_admin=is_admin,
            )
            if ok:
                create_admin_log(
                    action_type="staff_create",
                    target_type="user",
                    target_id=msg,
                    action_detail=f"display_name={display_name}, role_type={role_type}"
                )
                st.success(f"登録できました！ user_id={msg}")
                st.rerun()
            else:
                st.error(msg)

    with list_tab:
        users_df = get_users_df()
        perm_df = get_company_permissions_df(company_id)

        if perm_df.empty:
            st.info("この事業所のスタッフはまだいありません。")
            return

        merged = perm_df.merge(
            users_df,
            how="left",
            on="user_id",
            suffixes=("_perm", "_user")
        )

        merged = merged.fillna("")
        try:
            merged = merged.sort_values(["is_admin", "display_name"], ascending=[False, True])
        except Exception:
            pass

        for _, row in merged.iterrows():
            user_id = str(row.get("user_id", "")).strip()
            display_name = str(row.get("display_name", "")).strip()
            user_login_id = str(row.get("user_login_id", "")).strip()
            role_type = str(row.get("role_type", "")).strip()
            is_admin = str(row.get("is_admin_perm", row.get("is_admin", ""))).strip()
            status = str(row.get("status_perm", row.get("status", ""))).strip()
            last_login_at = str(row.get("last_login_at", "")).strip()

            with st.container(border=True):
                c1, c2 = st.columns([3, 1])

                with c1:
                    st.write(f"**{display_name}**")
                    st.caption(
                        f"user_id={user_id} / login_id={user_login_id} / "
                        f"権限={'管理者' if is_admin == '1' else '職員'} / "
                        f"状態={status} / "
                        f"最終ログイン={last_login_at or '-'} / "
                        f"PW=••••••••"
                    )

                with c2:
                    if status != "inactive":
                        if st.button("無効化", key=f"deactivate_staff_{company_id}_{user_id}", use_container_width=True):
                            ok, msg = set_company_user_status(user_id, company_id, "inactive")
                            if ok:
                                create_admin_log(
                                    action_type="staff_deactivate",
                                    target_type="user",
                                    target_id=user_id,
                                    action_detail=f"display_name={display_name}"
                                )
                                st.success("無効化しました。")
                                st.rerun()
                            else:
                                st.error(msg)

def is_valid_user_password(pw: str) -> bool:
    s = str(pw)
    if len(s) < 8:
        return False
    if not any(ch.isupper() for ch in s):
        return False
    if not any(ch.islower() for ch in s):
        return False
    if not any(ch.isdigit() for ch in s):
        return False
    return True


def update_user_login_credentials(company_id: str, current_id: str, current_pw: str, new_id: str, new_pw: str):
    df = get_users_df()
    if df is None or df.empty:
        return False, "usersシートが空です"

    work = df.copy()
    work["user_login_id"] = work["user_login_id"].astype(str).str.strip()
    work["user_login_password"] = work["user_login_password"].astype(str).str.strip()
    work["user_id"] = work["user_id"].astype(str).str.strip()

    current_user_id = str(st.session_state.get("user_id", "")).strip()

    mask = (
        (work["user_id"] == current_user_id) &
        (work["user_login_id"] == str(current_id).strip()) &
        (work["user_login_password"] == str(current_pw).strip())
    )

    if not mask.any():
        return False, "現在のIDまたはパスワードが違います"

    dup_mask = (
        (work["user_login_id"] == str(new_id).strip()) &
        (work["user_id"] != current_user_id)
    )
    if dup_mask.any():
        return False, "その新しいIDはすでに使われています"

    work.loc[mask, "user_login_id"] = str(new_id).strip()
    work.loc[mask, "user_login_password"] = str(new_pw).strip()
    work.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(work, "users")
    return True, "IDとパスワードを変更しました"

def get_company_users_for_admin(company_id: str):
    users_df = get_users_df()
    perm_df = get_company_permissions_df(company_id)

    if users_df is None or users_df.empty or perm_df is None or perm_df.empty:
        return pd.DataFrame()

    merged = users_df.merge(
        perm_df,
        on="user_id",
        how="inner",
        suffixes=("", "_perm")
    )

    merged["display_name"] = merged["display_name"].fillna("").astype(str).str.strip()
    merged["status_perm"] = merged["status_perm"].fillna("").astype(str).str.strip().str.lower()
    merged["login_card_id"] = merged["login_card_id"].fillna("").astype(str).str.strip()

    merged = merged[merged["status_perm"] != "inactive"].copy()

    try:
        merged = merged.sort_values(["display_name"], ascending=[True])
    except Exception:
        pass

    return merged


def set_user_login_card_id(user_id: str, card_id: str):
    df = get_users_df()
    if df is None or df.empty:
        return False, "usersシートが空です"

    work = df.copy()
    work["user_id"] = work["user_id"].fillna("").astype(str).str.strip()

    mask = work["user_id"] == str(user_id).strip()
    if not mask.any():
        return False, "対象ユーザーが見つからありません"

    new_card_id = str(card_id).strip()
    if not new_card_id:
        return False, "カードIDが空です"

    dup_mask = (
        work["login_card_id"].fillna("").astype(str).str.strip() == new_card_id
    ) & (~mask)

    if dup_mask.any():
        dup_name = str(work.loc[dup_mask, "display_name"].iloc[0]).strip()
        return False, f"そのカードIDはすでに別ユーザーに登録済みです（{dup_name}）"

    work.loc[mask, "login_card_id"] = new_card_id
    work.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(work, "users")
    return True, "カードIDを登録しました"


def clear_user_login_card_id(user_id: str):
    df = get_users_df()
    if df is None or df.empty:
        return False, "usersシートが空です"

    work = df.copy()
    work["user_id"] = work["user_id"].fillna("").astype(str).str.strip()

    mask = work["user_id"] == str(user_id).strip()
    if not mask.any():
        return False, "対象ユーザーが見つからありません"

    work.loc[mask, "login_card_id"] = ""
    work.loc[mask, "updated_at"] = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    save_db(work, "users")
    return True, "カードIDを削除しました"


def render_ic_card_manage_page():
    if not bool(st.session_state.get("is_admin", False)):
        st.error("このページは管理者専用です。")
        return

    st.title("💳 非接触ICカード登録")
    st.caption("usersシートの login_card_id を管理するページです。")

    company_id = str(st.session_state.get("company_id", "")).strip()
    staff_df = get_company_users_for_admin(company_id)

    if staff_df is None or staff_df.empty:
        st.info("登録対象のスタッフがまだいありません。")
        return

    tab1, tab2, tab3 = st.tabs(["新規登録", "登録変更", "登録削除"])

    staff_options = [
        (
            str(row.get("user_id", "")).strip(),
            f"{str(row.get('display_name', '')).strip()} "
            f"({str(row.get('user_login_id', '')).strip()})"
        )
        for _, row in staff_df.iterrows()
    ]

    option_map = {label: user_id for user_id, label in staff_options}
    labels = list(option_map.keys())

    with tab1:
        st.subheader("新規登録")
        unregistered = staff_df[
            staff_df["login_card_id"].fillna("").astype(str).str.strip() == ""
        ].copy()

        if unregistered.empty:
            st.info("未登録スタッフはいありません。")
        else:
            new_labels = [
                f"{str(row.get('display_name', '')).strip()} ({str(row.get('user_login_id', '')).strip()})"
                for _, row in unregistered.iterrows()
            ]
            new_label_map = {
                f"{str(row.get('display_name', '')).strip()} ({str(row.get('user_login_id', '')).strip()})":
                str(row.get("user_id", "")).strip()
                for _, row in unregistered.iterrows()
            }

            selected_label = st.selectbox("登録するスタッフ", new_labels, key="ic_new_user")
            card_id_input = st.text_input(
                "カードID",
                key="ic_new_card_id",
                help="当面は手入力でもOK。あとでタッチ読取につなぎます。"
            )

            if st.button("新規登録する", key="ic_register_button", use_container_width=True):
                ok, msg = set_user_login_card_id(new_label_map[selected_label], card_id_input)
                if ok:
                    create_admin_log(
                        action_type="ic_card_register",
                        target_type="user",
                        target_id=new_label_map[selected_label],
                        action_detail=f"login_card_id={str(card_id_input).strip()}"
                    )
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    with tab2:
        st.subheader("登録変更")
        registered = staff_df[
            staff_df["login_card_id"].fillna("").astype(str).str.strip() != ""
        ].copy()

        if registered.empty:
            st.info("登録済みスタッフがいありません。")
        else:
            change_labels = [
                f"{str(row.get('display_name', '')).strip()} "
                f"({str(row.get('user_login_id', '')).strip()}) / "
                f"現在: {str(row.get('login_card_id', '')).strip()}"
                for _, row in registered.iterrows()
            ]
            change_label_map = {
                f"{str(row.get('display_name', '')).strip()} "
                f"({str(row.get('user_login_id', '')).strip()}) / "
                f"現在: {str(row.get('login_card_id', '')).strip()}":
                str(row.get("user_id", "")).strip()
                for _, row in registered.iterrows()
            }

            selected_label = st.selectbox("変更するスタッフ", change_labels, key="ic_change_user")
            new_card_id = st.text_input("新しいカードID", key="ic_change_card_id")

            if st.button("登録変更する", key="ic_change_button", use_container_width=True):
                ok, msg = set_user_login_card_id(change_label_map[selected_label], new_card_id)
                if ok:
                    create_admin_log(
                        action_type="ic_card_change",
                        target_type="user",
                        target_id=change_label_map[selected_label],
                        action_detail=f"new_login_card_id={str(new_card_id).strip()}"
                    )
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    with tab3:
        st.subheader("登録削除")
        registered = staff_df[
            staff_df["login_card_id"].fillna("").astype(str).str.strip() != ""
        ].copy()

        if registered.empty:
            st.info("削除対象の登録はありません。")
        else:
            delete_labels = [
                f"{str(row.get('display_name', '')).strip()} "
                f"({str(row.get('user_login_id', '')).strip()}) / "
                f"現在: {str(row.get('login_card_id', '')).strip()}"
                for _, row in registered.iterrows()
            ]
            delete_label_map = {
                f"{str(row.get('display_name', '')).strip()} "
                f"({str(row.get('user_login_id', '')).strip()}) / "
                f"現在: {str(row.get('login_card_id', '')).strip()}":
                str(row.get("user_id", "")).strip()
                for _, row in registered.iterrows()
            }

            selected_label = st.selectbox("削除するスタッフ", delete_labels, key="ic_delete_user")

            if st.button("登録削除する", key="ic_delete_button", use_container_width=True):
                ok, msg = clear_user_login_card_id(delete_label_map[selected_label])
                if ok:
                    create_admin_log(
                        action_type="ic_card_delete",
                        target_type="user",
                        target_id=delete_label_map[selected_label],
                        action_detail="login_card_id cleared"
                    )
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    st.divider()
    with st.expander("現在の登録状況を見る"):
        view_df = staff_df[[
            "display_name",
            "user_login_id",
            "login_card_id",
            "status_perm"
        ]].copy()
        view_df.columns = ["表示名", "ログインID", "カードID", "状態"]
        st.dataframe(view_df, use_container_width=True)

def get_document_master_df():
    return get_document_master_df_cached().copy()

def get_saved_documents_df():
    df = load_db("saved_documents")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "record_id",
            "resident_id",
            "resident_name",
            "doc_type",
            "created_at",
            "updated_at",
            "json_data"
        ])
    else:
        for col in [
            "record_id",
            "resident_id",
            "resident_name",
            "doc_type",
            "created_at",
            "updated_at",
            "json_data"
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")

@st.cache_data(ttl=60)
def get_diary_input_rules_df_cached():
    df = load_db("diary_input_rules")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "record_id", "company_id", "date", "resident_id", "resident_name",
            "start_time", "end_time", "work_start_time", "work_end_time", "work_break_time",
            "meal_flag", "note",
            "start_memo", "end_memo", "staff_name",
            "generated_status", "generated_support", "created_at",
            "service_type", "knowbe_target", "send_status", "sent_at", "send_error",
            "record_mode"
        ])
    else:
        for col in [
            "record_id", "company_id", "date", "resident_id", "resident_name",
            "start_time", "end_time", "work_start_time", "work_end_time", "work_break_time",
            "meal_flag", "note",
            "start_memo", "end_memo", "staff_name",
            "generated_status", "generated_support", "created_at",
            "service_type", "knowbe_target", "send_status", "sent_at", "send_error",
            "record_mode"
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")

def save_diary_input_record(
    date,
    resident_id,
    resident_name,
    start_time,
    end_time,
    work_start_time,
    work_end_time,
    work_break_time,
    meal_flag,
    note,
    start_memo,
    end_memo,
    staff_name,
    generated_status="",
    generated_support="",
    service_type="在宅",
    knowbe_target="",
    send_status="draft",
    sent_at="",
    send_error="",
    record_mode="gemini",
    company_id=""
):
    df = get_diary_input_rules_df()

    if df.empty:
        next_id = 1
    else:
        nums = pd.to_numeric(df["record_id"], errors="coerce").dropna()
        next_id = int(nums.max()) + 1 if not nums.empty else 1

    created_at = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    if not company_id:
        company_id = get_current_company_id()
    new_row = pd.DataFrame([{
        "record_id": next_id,
        "company_id": str(company_id),
        "date": str(date),
        "resident_id": str(resident_id),
        "resident_name": str(resident_name),
        "start_time": str(start_time),
        "end_time": str(end_time),
        "work_start_time": str(work_start_time),
        "work_end_time": str(work_end_time),
        "work_break_time": str(work_break_time),
        "meal_flag": str(meal_flag),
        "note": str(note),
        "start_memo": str(start_memo),
        "end_memo": str(end_memo),
        "staff_name": str(staff_name),
        "generated_status": str(generated_status),
        "generated_support": str(generated_support),
        "created_at": created_at,
        "service_type": str(service_type),
        "knowbe_target": str(knowbe_target),
        "send_status": str(send_status),
        "sent_at": str(sent_at),
        "send_error": str(send_error),
        "record_mode": str(record_mode),
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "diary_input_rules")
    return next_id

def update_diary_input_record_status(record_id, send_status, sent_at="", send_error=""):
    df = get_diary_input_rules_df()

    if df is None or df.empty:
        return False

    mask = df["record_id"].astype(str) == str(record_id)

    if not mask.any():
        return False

    df.loc[mask, "send_status"] = str(send_status)
    df.loc[mask, "sent_at"] = str(sent_at)
    df.loc[mask, "send_error"] = str(send_error)

    save_db(df, "diary_input_rules")
    return True

def get_diary_input_rules_df(company_id=None):
    df = get_diary_input_rules_df_cached().copy()

    if company_id is None:
        company_id = str(st.session_state.get("company_id", "")).strip()

    if "company_id" not in df.columns:
        df["company_id"] = ""

    df["company_id"] = df["company_id"].fillna("").astype(str).str.strip()
    return df[df["company_id"] == str(company_id).strip()].copy()

def validate_bee_times(
    resident_id,
    target_date,
    start_time,
    end_time,
    work_start_time,
    work_end_time,
):
    errors = []

    start_time = str(start_time).strip()
    end_time = str(end_time).strip()
    work_start_time = str(work_start_time).strip()
    work_end_time = str(work_end_time).strip()

    # 通所時間は必須
    s = _to_minutes(start_time)
    e = _to_minutes(end_time)

    if s is None or e is None:
        errors.append("時間の形式が正しくありません。HH:MM で入れてください。")
        return errors

    if s >= e:
        errors.append("開始時間と終了時間の大小が正しくありません。")

    # 作業時間が両方空なら、通所時間と同じ扱いにする
    if work_start_time == "" and work_end_time == "":
        ws = s
        we = e
        work_start_time = start_time
        work_end_time = end_time
    # 片方だけ空はエラー
    elif work_start_time == "" or work_end_time == "":
        errors.append("作業開始時間と作業終了時間は、入れるなら両方入れてください。")
        return errors
    else:
        ws = _to_minutes(work_start_time)
        we = _to_minutes(work_end_time)

        if ws is None or we is None:
            errors.append("時間の形式が正しくありません。HH:MM で入れてください。")
            return errors

    if ws >= we:
        errors.append("作業開始時間と作業終了時間の大小が正しくありません。")

    if ws < s or we > e:
        errors.append("作業時間が通所時間の範囲をはみ出しています。")

    weekday_label = _normalize_weekday_label(target_date)
    schedule_df = get_resident_schedule_df()

    if schedule_df is not None and not schedule_df.empty and weekday_label:
        work = schedule_df.copy()
        work["resident_id"] = work["resident_id"].astype(str)
        work["weekday"] = work["weekday"].astype(str).str.strip()
        work["service_type"] = work["service_type"].astype(str).str.strip()

        target_rows = work[
            (work["resident_id"] == str(resident_id)) &
            (work["weekday"] == weekday_label)
        ].copy()

        for _, row in target_rows.iterrows():
            sv = str(row.get("service_type", "")).strip()
            rs = str(row.get("start_time", "")).strip()
            re = str(row.get("end_time", "")).strip()

            if sv in ["看護", "介護"] and is_time_overlap(work_start_time, work_end_time, rs, re):
                errors.append(f"{sv}の予定（{rs}〜{re}）と作業時間が重なっています。")

    return errors

def save_document_record(resident_id, resident_name, doc_type, form_data):
    df = get_saved_documents_df()

    now_str = now_jst().strftime("%Y-%m-%d %H:%M")

    if df.empty:
        next_id = 1
    else:
        ids = pd.to_numeric(df["record_id"], errors="coerce").dropna()
        next_id = int(ids.max()) + 1 if not ids.empty else 1

    new_row = pd.DataFrame([{
        "record_id": next_id,
        "resident_id": resident_id,
        "resident_name": resident_name,
        "doc_type": doc_type,
        "created_at": now_str,
        "updated_at": now_str,
        "json_data": json.dumps(form_data, ensure_ascii=False)
    }])

    merged_df = pd.concat([df, new_row], ignore_index=True)
    save_db(merged_df, "saved_documents")
    return next_id


def update_document_record(record_id, form_data):
    df = get_saved_documents_df()
    if df.empty:
        return False

    now_str = now_jst().strftime("%Y-%m-%d %H:%M")
    mask = df["record_id"].astype(str) == str(record_id)

    if not mask.any():
        return False

    df.loc[mask, "updated_at"] = now_str
    df.loc[mask, "json_data"] = json.dumps(form_data, ensure_ascii=False)

    save_db(df, "saved_documents")
    return True


def get_document_records(doc_type, resident_id):
    df = get_saved_documents_df()
    if df.empty:
        return df

    df = df[
        (df["doc_type"].astype(str) == str(doc_type)) &
        (df["resident_id"].astype(str) == str(resident_id))
    ].copy()

    if df.empty:
        return df

    try:
        df["record_id_num"] = pd.to_numeric(df["record_id"], errors="coerce")
        df = df.sort_values(["record_id_num"], ascending=[False])
    except Exception:
        pass

    return df


def load_document_json(record_id):
    df = get_saved_documents_df()
    if df.empty:
        return None

    target = df[df["record_id"].astype(str) == str(record_id)]
    if target.empty:
        return None

    json_str = str(target.iloc[0]["json_data"]).strip()
    if not json_str:
        return None

    try:
        return json.loads(json_str)
    except Exception:
        return None

def get_gemini_api_key_from_app():
    api_key = ""

    try:
        api_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        api_key = ""

    if not api_key:
        import os
        api_key = os.environ.get("GEMINI_API_KEY", "")

    return str(api_key).strip()


def call_gemini_json(prompt: str):
    api_key = get_gemini_api_key_from_app()
    if not api_key:
        raise RuntimeError("APIキーありません")

    genai.configure(api_key=api_key)

    model_candidates = [
        "gemini-2.5-flash",
    ]

    last_error = None

    for model_name in model_candidates:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            text = str(getattr(response, "text", "")).strip()

            if not text:
                continue

            cleaned = text.replace("```json", "").replace("```", "").strip()
            return json.loads(cleaned)

        except Exception as e:
            last_error = e
            continue

    raise RuntimeError(f"Gemini全部失敗です: {last_error}")


def get_latest_saved_document(resident_id, doc_type):
    df = get_saved_documents_df()
    if df is None or df.empty:
        return None

    work = df.copy()
    work["resident_id"] = work["resident_id"].astype(str)
    work["doc_type"] = work["doc_type"].astype(str)

    work = work[
        (work["resident_id"] == str(resident_id)) &
        (work["doc_type"] == str(doc_type))
    ].copy()

    if work.empty:
        return None

    try:
        work["record_id_num"] = pd.to_numeric(work["record_id"], errors="coerce")
        work = work.sort_values(["record_id_num"], ascending=[False])
    except Exception:
        pass

    return work.iloc[0].to_dict()


def get_latest_saved_document_json(resident_id, doc_type):
    row = get_latest_saved_document(resident_id, doc_type)
    if not row:
        return None

    json_str = str(row.get("json_data", "")).strip()
    if not json_str:
        return None

    try:
        return json.loads(json_str)
    except Exception:
        return None



def build_plan_draft_generation_prompt(
    resident_name,
    source_label,
    source_data,
    new_policy="",
    new_long_goal="",
    new_short_goal="",
    new_goal_rows_policy="",
):
    source_data = source_data or {}

    prompt = f'''
就労継続支援B型の個別支援計画案を作成する。

【利用者名】
{safe_text(resident_name)}

【参照元の種類】
{safe_text(source_label)}

【参照元データ】
サービス等利用計画の総合的な方針:
{safe_text(source_data.get("policy", ""))}

長期目標:
{safe_text(source_data.get("long_goal", ""))}

短期目標:
{safe_text(source_data.get("short_goal", ""))}

具体的到達目標1:
{safe_text(source_data.get("target_1", ""))}
本人の役割1:
{safe_text(source_data.get("role_1", ""))}
支援内容1:
{safe_text(source_data.get("support_1", ""))}
支援期間1:
{safe_text(source_data.get("period_1", ""))}
担当者1:
{safe_text(source_data.get("person_1", ""))}
優先順位1:
{safe_text(source_data.get("priority_1", ""))}

具体的到達目標2:
{safe_text(source_data.get("target_2", ""))}
本人の役割2:
{safe_text(source_data.get("role_2", ""))}
支援内容2:
{safe_text(source_data.get("support_2", ""))}
支援期間2:
{safe_text(source_data.get("period_2", ""))}
担当者2:
{safe_text(source_data.get("person_2", ""))}
優先順位2:
{safe_text(source_data.get("priority_2", ""))}

具体的到達目標3:
{safe_text(source_data.get("target_3", ""))}
本人の役割3:
{safe_text(source_data.get("role_3", ""))}
支援内容3:
{safe_text(source_data.get("support_3", ""))}
支援期間3:
{safe_text(source_data.get("period_3", ""))}
担当者3:
{safe_text(source_data.get("person_3", ""))}
優先順位3:
{safe_text(source_data.get("priority_3", ""))}

【新しい方針】
サービス等利用計画の総合的な方針について:
{safe_text(new_policy)}

長期目標について:
{safe_text(new_long_goal)}

短期目標について:
{safe_text(new_short_goal)}

具体的到達目標3組全体について:
{safe_text(new_goal_rows_policy)}

【作成する項目】
- サービス等利用計画の総合的な方針
- 長期目標
- 短期目標
- 具体的到達目標・本人の役割・支援内容・支援期間・担当者・優先順位を3組

【ルール】
- 新しい方針が空でも作成すること
- 参照元データが少なくても自然に補うこと
- 就労継続支援B型の書類として自然な内容にすること
- 出力はJSONのみ
- 文章は日本語
- 余計な説明文は不要

【出力形式】
{{
  "policy": "サービス等利用計画の総合的な方針",
  "long_goal": "長期目標",
  "short_goal": "短期目標",
  "goal_rows": [
    {{
      "target": "具体的到達目標1",
      "role": "本人の役割1",
      "support": "支援内容1",
      "period": "支援期間1",
      "person": "担当者1",
      "priority": "優先順位1"
    }},
    {{
      "target": "具体的到達目標2",
      "role": "本人の役割2",
      "support": "支援内容2",
      "period": "支援期間2",
      "person": "担当者2",
      "priority": "優先順位2"
    }},
    {{
      "target": "具体的到達目標3",
      "role": "本人の役割3",
      "support": "支援内容3",
      "period": "支援期間3",
      "person": "担当者3",
      "priority": "優先順位3"
    }}
  ]
}}
'''
    return prompt

# ==========================================
# 🤫 秘密モード Gemini 自動生成まわり
# ==========================================

TEST_ALLOW_EMPTY_REFERENCE = False
# 動作確認が終わったら False にするか、
# この1行ごとコメントアウトして本番運用する。


def get_reference_doc_type_for_gemini(doc_title: str):
    return {
        "個別支援計画案": "モニタリング",
        "サービス担当者会議": "個別支援計画案",
        "個別支援計画": "サービス担当者会議",
        "モニタリング": "個別支援計画",
    }.get(str(doc_title).strip(), "")


def get_reference_json_for_gemini(resident_id, doc_title):
    ref_doc_type = get_reference_doc_type_for_gemini(doc_title)
    if not ref_doc_type:
        return None, None

    source_json = get_latest_saved_document_json(resident_id, ref_doc_type)

    if source_json:
        return ref_doc_type, source_json

    if TEST_ALLOW_EMPTY_REFERENCE:
        return f"{ref_doc_type}(参照なしテスト)", {}

    return ref_doc_type, None


def get_resident_option_map():
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        return [], {}

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    return resident_options, resident_map


def normalize_goal_rows(goal_rows, size=3):
    rows = goal_rows if isinstance(goal_rows, list) else []

    normalized = []
    for i in range(size):
        row = rows[i] if i < len(rows) and isinstance(rows[i], dict) else {}
        normalized.append({
            "target": safe_text(row.get("target", "")),
            "role": safe_text(row.get("role", "")),
            "support": safe_text(row.get("support", "")),
            "period": safe_text(row.get("period", "")),
            "person": safe_text(row.get("person", "")),
            "priority": safe_text(row.get("priority", "")),
        })

    return normalized


def build_meeting_generation_prompt(
    resident_name,
    source_label,
    source_data,
    new_policy="",
):
    source_data = source_data or {}

    prompt = f'''
就労継続支援B型のサービス担当者会議記録を作成する。

【利用者名】
{safe_text(resident_name)}

【参照元の種類】
{safe_text(source_label)}

【参照元データ】
サービス等利用計画の総合的な方針:
{safe_text(source_data.get("policy", ""))}

長期目標:
{safe_text(source_data.get("long_goal", ""))}

短期目標:
{safe_text(source_data.get("short_goal", ""))}

具体的到達目標1:
{safe_text(source_data.get("target_1", ""))}
本人の役割1:
{safe_text(source_data.get("role_1", ""))}
支援内容1:
{safe_text(source_data.get("support_1", ""))}

具体的到達目標2:
{safe_text(source_data.get("target_2", ""))}
本人の役割2:
{safe_text(source_data.get("role_2", ""))}
支援内容2:
{safe_text(source_data.get("support_2", ""))}

具体的到達目標3:
{safe_text(source_data.get("target_3", ""))}
本人の役割3:
{safe_text(source_data.get("role_3", ""))}
支援内容3:
{safe_text(source_data.get("support_3", ""))}

【新しい方針・補足】
{safe_text(new_policy)}

【作成する項目】
- 議題
- 検討内容
- 残された課題
- 結論

【ルール】
- 就労継続支援B型のサービス担当者会議として自然な内容
- 利用者の状況や支援方針がつながる内容にする
- 議題は簡潔に
- 検討内容はやや詳しく
- 出力はJSONのみ
- 文章は日本語
- 余計な説明文は不要

【出力形式】
{{
  "agenda": "議題",
  "discussion": "検討内容",
  "issues_left": "残された課題",
  "conclusion": "結論"
}}
'''
    return prompt


def build_monitoring_generation_prompt(
    resident_name,
    source_label,
    source_data,
    new_policy="",
):
    source_data = source_data or {}

    prompt = f'''
就労継続支援B型のモニタリング記録を作成する。

【利用者名】
{safe_text(resident_name)}

【参照元の種類】
{safe_text(source_label)}

【参照元データ】
サービス等利用計画の総合的な方針:
{safe_text(source_data.get("policy", ""))}

長期目標:
{safe_text(source_data.get("long_goal", ""))}

短期目標:
{safe_text(source_data.get("short_goal", ""))}

具体的到達目標1:
{safe_text(source_data.get("target_1", ""))}
本人の役割1:
{safe_text(source_data.get("role_1", ""))}
支援内容1:
{safe_text(source_data.get("support_1", ""))}
支援期間1:
{safe_text(source_data.get("period_1", ""))}
担当者1:
{safe_text(source_data.get("person_1", ""))}
優先順位1:
{safe_text(source_data.get("priority_1", ""))}

具体的到達目標2:
{safe_text(source_data.get("target_2", ""))}
本人の役割2:
{safe_text(source_data.get("role_2", ""))}
支援内容2:
{safe_text(source_data.get("support_2", ""))}
支援期間2:
{safe_text(source_data.get("period_2", ""))}
担当者2:
{safe_text(source_data.get("person_2", ""))}
優先順位2:
{safe_text(source_data.get("priority_2", ""))}

具体的到達目標3:
{safe_text(source_data.get("target_3", ""))}
本人の役割3:
{safe_text(source_data.get("role_3", ""))}
支援内容3:
{safe_text(source_data.get("support_3", ""))}
支援期間3:
{safe_text(source_data.get("period_3", ""))}
担当者3:
{safe_text(source_data.get("person_3", ""))}
優先順位3:
{safe_text(source_data.get("priority_3", ""))}

【新しい方針・補足】
{safe_text(new_policy)}

【作成する項目】
- 具体的達成目標番号1〜3の達成状況の評価
- 具体的達成目標番号1〜3の達成できている点と未達成点（要因も）
- 具体的達成目標番号1〜3の今後の対応（支援内容・方法の変更・継続・終了）

【ルール】
- 評価は「達成」「継続」「一部達成」「終了」のいずれか
- モニタリングとして自然な内容にする
- 出力はJSONのみ
- 文章は日本語
- 余計な説明文は不要

【出力形式】
{{
  "rows": [
    {{
      "status": "継続",
      "detail": "達成できている点と未達成点（要因も）1",
      "future": "今後の対応1"
    }},
    {{
      "status": "継続",
      "detail": "達成できている点と未達成点（要因も）2",
      "future": "今後の対応2"
    }},
    {{
      "status": "継続",
      "detail": "達成できている点と未達成点（要因も）3",
      "future": "今後の対応3"
    }}
  ]
}}
'''
    return prompt


def apply_generated_data_to_form(doc_title: str, generated: dict):
    generated = generated or {}

    if doc_title in ["個別支援計画案", "個別支援計画"]:
        st.session_state[f"{doc_title}_policy"] = safe_text(generated.get("policy", ""))
        st.session_state[f"{doc_title}_long_goal"] = safe_text(generated.get("long_goal", ""))
        st.session_state[f"{doc_title}_short_goal"] = safe_text(generated.get("short_goal", ""))

        rows = normalize_goal_rows(generated.get("goal_rows", []), size=3)

        for i, row in enumerate(rows, start=1):
            st.session_state[f"{doc_title}_target_{i}"] = row["target"]
            st.session_state[f"{doc_title}_role_{i}"] = row["role"]
            st.session_state[f"{doc_title}_support_{i}"] = row["support"]
            st.session_state[f"{doc_title}_period_{i}"] = row["period"]
            st.session_state[f"{doc_title}_person_{i}"] = row["person"]
            st.session_state[f"{doc_title}_priority_{i}"] = row["priority"]

    elif doc_title == "サービス担当者会議":
        st.session_state[f"{doc_title}_agenda"] = safe_text(generated.get("agenda", ""))
        st.session_state[f"{doc_title}_discussion"] = safe_text(generated.get("discussion", ""))
        st.session_state[f"{doc_title}_issues_left"] = safe_text(generated.get("issues_left", ""))
        st.session_state[f"{doc_title}_conclusion"] = safe_text(generated.get("conclusion", ""))

    elif doc_title == "モニタリング":
        rows = generated.get("rows", [])
        if not isinstance(rows, list):
            rows = []

        normalized = []
        for i in range(3):
            row = rows[i] if i < len(rows) and isinstance(rows[i], dict) else {}
            status_val = safe_text(row.get("status", ""))
            if status_val not in ["", "達成", "継続", "一部達成", "終了"]:
                status_val = "継続"

            normalized.append({
                "status": status_val,
                "detail": safe_text(row.get("detail", "")),
                "future": safe_text(row.get("future", "")),
            })

        for i, row in enumerate(normalized, start=1):
            st.session_state[f"{doc_title}_status_{i}"] = row["status"]
            st.session_state[f"{doc_title}_detail_{i}"] = row["detail"]
            st.session_state[f"{doc_title}_future_{i}"] = row["future"]


def run_secret_gemini_generation(doc_title: str, resident_id, resident_name, new_policy_text=""):
    source_label, source_data = get_reference_json_for_gemini(resident_id, doc_title)

    if source_data is None:
        raise RuntimeError(f"直近の{source_label}がありません。")

    if doc_title in ["個別支援計画案", "個別支援計画"]:
        prompt = build_plan_draft_generation_prompt(
            resident_name=resident_name,
            source_label=source_label,
            source_data=source_data,
            new_policy=new_policy_text,
            new_long_goal="",
            new_short_goal="",
            new_goal_rows_policy="",
        )
    elif doc_title == "サービス担当者会議":
        prompt = build_meeting_generation_prompt(
            resident_name=resident_name,
            source_label=source_label,
            source_data=source_data,
            new_policy=new_policy_text,
        )
    elif doc_title == "モニタリング":
        prompt = build_monitoring_generation_prompt(
            resident_name=resident_name,
            source_label=source_label,
            source_data=source_data,
            new_policy=new_policy_text,
        )
    else:
        raise RuntimeError("この書類はGemini自動生成対象ではありません。")

    return call_gemini_json(prompt)




def get_next_resident_id(master_df):
    if master_df is None or master_df.empty or "resident_id" not in master_df.columns:
        return "R0001"

    numbers = []
    for x in master_df["resident_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("R"):
            num = x[1:]
            if num.isdigit():
                numbers.append(int(num))

    next_num = max(numbers) + 1 if numbers else 1
    return f"R{next_num:04d}"

@st.cache_data(ttl=60)
def save_uploaded_document(
    category1,
    category2,
    category3,
    title,
    summary,
    memo,
    uploaded_file
):
    df = get_document_master_df()

    now_str = now_jst().strftime("%Y-%m-%d %H:%M")

    if df.empty:
        next_id = 1
    else:
        ids = pd.to_numeric(df["document_id"], errors="coerce").dropna()
        next_id = int(ids.max()) + 1 if not ids.empty else 1

    file_bytes = uploaded_file.read()
    file_data_base64 = base64.b64encode(file_bytes).decode("utf-8")

    original_filename = uploaded_file.name
    lower_name = original_filename.lower()

    if lower_name.endswith(".xlsx"):
        file_type = "xlsx"
    elif lower_name.endswith(".xls"):
        file_type = "xls"
    elif lower_name.endswith(".pdf"):
        file_type = "pdf"
    elif lower_name.endswith(".docx"):
        file_type = "docx"
    elif lower_name.endswith(".doc"):
        file_type = "doc"
    else:
        file_type = "other"

    new_row = pd.DataFrame([{
        "document_id": next_id,
        "category1": category1,
        "category2": category2,
        "category3": category3,
        "title": title,
        "file_type": file_type,
        "url": "",
        "summary": summary,
        "memo": memo,
        "status": "有効",
        "updated_at": now_str,
        "created_at": now_str,
        "original_filename": original_filename,
        "file_data_base64": file_data_base64
    }])

    merged_df = pd.concat([df, new_row], ignore_index=True)
    save_db(merged_df, "document_master")

@st.cache_data(ttl=60)
def get_document_master_df_cached():
    df = load_db("document_master")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "document_id", "category1", "category2", "category3",
            "title", "file_type", "url", "summary", "memo",
            "status", "updated_at", "created_at",
            "original_filename", "file_data_base64"
        ])
    else:
        for col in [
            "document_id", "category1", "category2", "category3",
            "title", "file_type", "url", "summary", "memo",
            "status", "updated_at", "created_at",
            "original_filename", "file_data_base64"
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")

def get_download_file_data(row):
    file_data_base64 = str(row.get("file_data_base64", "")).strip()
    original_filename = str(row.get("original_filename", "")).strip()

    if not file_data_base64 or not original_filename:
        return None, None, None

    file_bytes = base64.b64decode(file_data_base64)

    lower_name = original_filename.lower()
    if lower_name.endswith(".xlsx"):
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif lower_name.endswith(".xls"):
        mime = "application/vnd.ms-excel"
    elif lower_name.endswith(".pdf"):
        mime = "application/pdf"
    elif lower_name.endswith(".docx"):
        mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif lower_name.endswith(".doc"):
        mime = "application/msword"
    else:
        mime = "application/octet-stream"

    return file_bytes, original_filename, mime

@st.cache_data(ttl=60)
def get_external_contacts_df_cached():
    df = load_db("external_contacts")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "contact_id", "category1", "category2",
            "name", "organization", "phone", "memo"
        ])
    else:
        for col in [
            "contact_id", "category1", "category2",
            "name", "organization", "phone", "memo"
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


@st.cache_data(ttl=60)
def get_resident_links_df_cached():
    df = load_db("resident_links")
    if df is None or df.empty:
        df = pd.DataFrame(columns=["id", "resident_id", "contact_id", "role"])
    else:
        for col in ["id", "resident_id", "contact_id", "role"]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def update_active_user():
    current_user = str(st.session_state.get("user", "")).strip()
    if not current_user:
        return

    active_df = load_db("active_users")

    if active_df is None or active_df.empty:
        active_df = pd.DataFrame(columns=["user", "login_at", "last_seen"])
    else:
        for col in ["user", "login_at", "last_seen"]:
            if col not in active_df.columns:
                active_df[col] = ""

    now_str = now_jst().strftime("%Y-%m-%d %H:%M")

    keep_rows = []
    now_naive = now_jst().replace(tzinfo=None)

    for _, row in active_df.fillna("").iterrows():
        last_seen = str(row.get("last_seen", "")).strip()
        if not last_seen:
            continue
        try:
            last_dt = pd.to_datetime(last_seen).to_pydatetime()
            if (now_naive - last_dt).total_seconds() <= 15 * 60:
                keep_rows.append(row)
        except Exception:
            pass

    active_df = pd.DataFrame(keep_rows) if keep_rows else pd.DataFrame(columns=["user", "login_at", "last_seen"])

    if current_user in active_df["user"].astype(str).tolist():
        active_df.loc[active_df["user"] == current_user, "last_seen"] = now_str
    else:
        new_row = pd.DataFrame([{
            "user": current_user,
            "login_at": st.session_state.get("login_at", now_str),
            "last_seen": now_str
        }])
        active_df = pd.concat([active_df, new_row], ignore_index=True)

    save_db(active_df, "active_users")

def heartbeat_active_user():
    if "user" not in st.session_state:
        return

    now_ts = now_jst().timestamp()
    last_ping = st.session_state.get("last_active_ping", 0)

    if now_ts - last_ping >= 300:
        update_active_user()
        st.session_state["last_active_ping"] = now_ts

def sync_task_events_to_calendar(company_id=None):
    if company_id is None:
        company_id = get_current_company_id()

    task_df = get_tasks_df(company_id)
    cal_df = load_db("calendar")

    required_task_cols = get_task_required_cols()
    required_cal_cols = [
        "company_id", "id", "title", "start", "end",
        "user", "memo", "source_type", "source_task_id"
    ]

    task_df = normalize_company_scoped_df(task_df, required_task_cols)
    cal_df = normalize_company_scoped_df(cal_df, required_cal_cols)

    target_cal_df = filter_by_company_id(cal_df, company_id)

    other_cal_df = cal_df[
        cal_df["company_id"].astype(str).str.strip() != str(company_id).strip()
    ].copy()

    target_cal_df = target_cal_df[
        ~target_cal_df["source_type"].astype(str).isin(["task_deadline", "task_active"])
    ].copy()

    today = now_jst().date()
    new_events = []

    if target_cal_df.empty:
        next_id = 1
    else:
        try:
            next_id = pd.to_numeric(target_cal_df["id"], errors="coerce").max()
            next_id = 1 if pd.isna(next_id) else int(next_id) + 1
        except Exception:
            next_id = len(target_cal_df) + 1

    for _, row in task_df.iterrows():
        task_id = str(row.get("id", "")).strip()
        task_name = str(row.get("task", "")).strip()
        status = str(row.get("status", "")).strip()
        user_name = str(row.get("user", "")).strip()
        limit_str = str(row.get("limit", "")).strip()
        updated_at = str(row.get("updated_at", "")).strip()

        if limit_str:
            try:
                limit_date = pd.to_datetime(limit_str).date()
                if limit_date > today:
                    new_events.append({
                        "company_id": str(company_id).strip(),
                        "id": next_id,
                        "title": f"締切：{task_name}",
                        "start": str(limit_date),
                        "end": str(limit_date),
                        "user": user_name,
                        "memo": f"タスク期限 / 状態: {status}",
                        "source_type": "task_deadline",
                        "source_task_id": task_id,
                    })
                    next_id += 1
            except Exception:
                pass

        if status == "作業中" and updated_at:
            try:
                active_date = pd.to_datetime(updated_at).date()
                new_events.append({
                    "company_id": str(company_id).strip(),
                    "id": next_id,
                    "title": f"作業中：{task_name}",
                    "start": str(active_date),
                    "end": str(active_date),
                    "user": user_name,
                    "memo": f"現在進行中 / 着手: {updated_at}",
                    "source_type": "task_active",
                    "source_task_id": task_id,
                })
                next_id += 1
            except Exception:
                pass

    if new_events:
        add_df = pd.DataFrame(new_events)
        target_cal_df = pd.concat([target_cal_df, add_df], ignore_index=True)

    merged_cal_df = pd.concat([other_cal_df, target_cal_df], ignore_index=True)
    merged_cal_df = normalize_company_scoped_df(merged_cal_df, required_cal_cols)
    save_db(merged_cal_df, "calendar")


def start_task(task_id, company_id=None):
    if company_id is None:
        company_id = get_current_company_id()
    df = get_tasks_df(company_id)

    if df.empty:
        return

    mask = (
        (df["company_id"].astype(str).str.strip() == str(company_id).strip()) &
        (df["id"].astype(str).str.strip() == str(task_id).strip())
    )

    df.loc[mask, ["status", "user", "updated_at"]] = [
        "作業中",
        st.session_state.user,
        now_jst().strftime("%Y-%m-%d %H:%M")
    ]
    save_db(df, "task")
    sync_task_events_to_calendar(company_id)


def complete_task(task_id, company_id=None):
    if company_id is None:
        company_id = get_current_company_id()
    df = get_tasks_df(company_id)

    if df.empty:
        return

    mask = (
        (df["company_id"].astype(str).str.strip() == str(company_id).strip()) &
        (df["id"].astype(str).str.strip() == str(task_id).strip())
    )

    df.loc[mask, ["status", "updated_at"]] = [
        "完了",
        now_jst().strftime("%Y-%m-%d %H:%M")
    ]
    save_db(df, "task")
    sync_task_events_to_calendar(company_id)


def go_to_page(page_name):
    st.session_state.current_page = page_name
    st.rerun()


def render_urgent_banner():
    company_id = get_current_company_id()
    urgent_df = get_urgent_tasks_df(company_id)

    if urgent_df.empty:
        return

    urgent_count = len(urgent_df)
    critical_count = len(urgent_df[urgent_df["priority"].astype(str).str.strip() == "至急"])
    important_count = len(urgent_df[urgent_df["priority"].astype(str).str.strip() == "重要"])

    st.markdown(
        f"""
        <div style="
            background: linear-gradient(90deg, #ff7b54 0%, #ff9f43 100%);
            color: white;
            padding: 14px 18px;
            border-radius: 12px;
            margin-bottom: 16px;
            font-weight: 700;
            box-shadow: 0 2px 8px rgba(0,0,0,0.12);
        ">
            🚨 緊急タスクあり　合計 {urgent_count}件（至急 {critical_count}件 / 重要 {important_count}件）
        </div>
        """,
        unsafe_allow_html=True
    )

    col_a, col_b = st.columns([1, 5])
    with col_a:
        if st.button("緊急一覧を開く", key="open_urgent_page_button", use_container_width=True):
            go_to_page("⑧ 緊急一覧")
    with col_b:
        st.caption("クリックして、至急・重要タスクの一覧を確認できます。")

# ==========================================
# 🔑 2段階ログイン
# ==========================================
if "company_authenticated" not in st.session_state:
    st.session_state.company_authenticated = False

if "auth_mode" not in st.session_state:
    st.session_state.auth_mode = "login"   # login / change


# ---------- 事業所ログイン前 ----------
if not st.session_state.get("company_authenticated", False):
    st.markdown("<style>[data-testid='stSidebarNav'] {display: none;}</style>", unsafe_allow_html=True)
    st.warning("### 事業所ログイン💻")

    company_login_id = st.text_input("事業所ID", key="company_login_id_input")
    company_login_password = st.text_input("事業所パスワード", type="password", key="company_login_password_input")

    if st.button("事業所ログイン", use_container_width=True, key="company_login_button"):
        row = authenticate_company_login(company_login_id, company_login_password)

        if row is None:
            st.error("事業所IDまたはパスワードが違います。")
        else:
            st.session_state.company_id = str(row.get("company_id", "")).strip()
            st.session_state.company_name = str(row.get("company_name", "")).strip()
            st.session_state.company_code = str(row.get("company_code", "")).strip()
            st.session_state.company_authenticated = True
            st.rerun()

    st.stop()


# ---------- 個人ログイン前 ----------
if "user" not in st.session_state:
    st.markdown("<style>[data-testid='stSidebarNav'] {display: none;}</style>", unsafe_allow_html=True)
    st.success(f"事業所: {st.session_state.get('company_name', '')}")
    st.warning("### 個人ログイン💻")

    company_id = str(st.session_state.get("company_id", "")).strip()
    admin_exists = company_has_any_admin(company_id)

    if not admin_exists:
        render_first_staff_register_block(key_prefix="after_company_login_first_staff")
        st.divider()

    top_cols = st.columns([1, 1])
    with top_cols[0]:
        if st.button("ID・パスワード変更", use_container_width=True, key="open_change_idpw"):
            st.session_state.auth_mode = "change"
            st.rerun()

    with top_cols[1]:
        if st.button("事業所切り替え", use_container_width=True, key="back_to_company_login"):
            for k in [
                "company_authenticated", "company_id", "company_name", "company_code",
                "company_login_id", "user", "user_id", "is_admin", "login_at",
                "last_active_ping", "current_page", "bee_menu_unlocked",
                "other_office_register_unlocked", "secret_doc_mode", "heart_mode", "secret_bee_cmd"
            ]:
                if k in st.session_state:
                    del st.session_state[k]
            st.session_state.auth_mode = "login"
            st.rerun()

    st.divider()

    if st.session_state.get("auth_mode", "login") == "login":
        user_login_id = st.text_input("ID", key="user_login_id_input")
        user_login_password = st.text_input("パスワード", type="password", key="user_login_password_input")

        if st.button("個人ログイン", use_container_width=True, key="user_login_button"):
            row = authenticate_user_login(
                st.session_state.get("company_id", ""),
                user_login_id,
                user_login_password
            )

            if row is None:
                st.error("ID・パスワード、または事業所権限を確認してください。")
            else:
                st.session_state.user = str(row.get("display_name", "")).strip()
                st.session_state.user_id = str(row.get("user_id", "")).strip()
                st.session_state.is_admin = bool(row.get("is_admin_resolved", False))
                st.session_state.login_at = now_jst().strftime("%Y-%m-%d %H:%M")
                st.session_state.last_active_ping = 0
                st.session_state.auth_mode = "login"
                st.rerun()

    st.stop()

# ==========================================
# 🏠 メインメニュー
# ==========================================
heartbeat_active_user()

page_options = [
    "⓪ 検索",
    "① 未着手の任務（掲示板）",
    "② タスクの引き受け・報告",
    "③ 稼働状況・完了履歴",
    "④ チームチャット",
    "⑤ 業務マニュアル",
    "⑥ 日誌入力状況",
    "⑦ タスクカレンダー",
    "⑧ 緊急一覧",
    "⑨ 利用者情報",
    "⑩ 書類アップロード",
    "書類_個別支援計画案",
    "書類_サービス担当者会議",
    "書類_個別支援計画",
    "書類_モニタリング",
    "書類_一括書類作成",
    "書類_在宅評価シート",
    "書類_アセスメント",
    "書類_基本シート",
    "書類_就労分野シート",
    "🐝knowbe日誌入力🐝",
    "💻他事業所へ登録💻",
    "Knowbe情報登録",
    "休憩室",
    "休憩室_チャットルーム",
    "休憩室_書類アップロード",
    "休憩室_倉庫",
    "お問い合わせ",
    "内職管理",
    "スタッフ管理",
    "ICカード管理",
    "勤怠管理",
    "過去日誌照合",
    "🐝knowbe日誌一括入力🐝",
]

if "current_page" not in st.session_state or st.session_state.current_page not in page_options:
    st.session_state.current_page = "① 未着手の任務（掲示板）"

st.sidebar.markdown(
"""
<div style="display:flex;align-items:center;gap:10px;margin-left:20px;">
    <div style="font-size:36px;">&#128029;</div>
    <div>
        <div style="font-weight:bold;font-size:24px;">
        Sue for Bee
        </div>
        <div style="font-size:16px;color:gray;">
        Assistance System
        </div>
    </div>
</div>
""",
unsafe_allow_html=True
)

st.sidebar.markdown("<div style='margin-top:30px;'></div>", unsafe_allow_html=True)
st.sidebar.markdown("メニューを選択してください")

st.sidebar.markdown(
    """
    <style>
    section[data-testid="stSidebar"] .stButton {
        margin-bottom: 12px !important;
    }

    section[data-testid="stSidebar"] .stButton > button {
        width: 100% !important;
        height: 56px !important;
        min-height: 56px !important;
        border-radius: 12px !important;
        border: 1px solid #d9d9d9 !important;
        background: #ffffff !important;
        color: #1f2d3d !important;
        font-weight: 700 !important;
        padding: 0 !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
        text-align: left !important;
        box-sizing: border-box !important;
    }

    section[data-testid="stSidebar"] .stButton > button:hover {
        border-color: #ff9f43 !important;
        color: #ff7b54 !important;
        background: #fffaf5 !important;
    }

    section[data-testid="stSidebar"] .stButton > button > div {
        width: 100% !important;
        height: 56px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
        text-align: left !important;
        padding: 0 16px !important;
        box-sizing: border-box !important;
    }

    section[data-testid="stSidebar"] .stButton > button > div p,
    section[data-testid="stSidebar"] .stButton > button > div span {
        width: 100% !important;
        margin: 0 !important;
        text-align: left !important;
        justify-content: flex-start !important;
        line-height: 1.2 !important;
    }

    .menu-selected-wrap {
        width: 100%;
        margin: 0 0 12px 0;
    }

    .menu-selected-box {
        width: 100%;
        height: 56px;
        border-radius: 12px;
        border: 1px solid #ff9f43;
        background: linear-gradient(90deg, #fff1e8 0%, #fff7e6 100%);
        color: #d35400;
        font-weight: 700;
        padding: 0 16px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        text-align: center;
        display: flex;
        align-items: center;
        justify-content: center;
        box-sizing: border-box;
        line-height: 1.2;
    }

    section[data-testid="stSidebar"] div[data-testid="stMarkdown"] {
        margin-bottom: 0 !important;
    }

    section[data-testid="stSidebar"] div[data-testid="stMarkdownContainer"] {
        margin: 0 !important;
        padding: 0 !important;
    }

    section[data-testid="stSidebar"] div[data-testid="stMarkdownContainer"] p {
        margin: 0 !important;
        padding: 0 !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)



if "bee_menu_unlocked" not in st.session_state:
    st.session_state["bee_menu_unlocked"] = False
if "other_office_register_unlocked" not in st.session_state:
    st.session_state["other_office_register_unlocked"] = False
if "secret_doc_mode" not in st.session_state:
    st.session_state["secret_doc_mode"] = False
if "heart_mode" not in st.session_state:
    st.session_state["heart_mode"] = False

main_page_options = [
    "⓪ 検索",
    "① 未着手の任務（掲示板）",
    "② タスクの引き受け・報告",
    "③ 稼働状況・完了履歴",
    "④ チームチャット",
    "⑤ 業務マニュアル",
    "⑥ 日誌入力状況",
    "⑦ タスクカレンダー",
    "⑧ 緊急一覧",
    "⑨ 利用者情報",
    "⑩ 書類アップロード",
    "内職管理"
]

document_page_options = [
    ("書類_個別支援計画案", "🤫個別支援計画案🤫" if st.session_state.get("secret_doc_mode", False) else "個別支援計画案"),
    ("書類_サービス担当者会議", "🤫サービス担当者会議🤫" if st.session_state.get("secret_doc_mode", False) else "サービス担当者会議"),
    ("書類_個別支援計画", "🤫個別支援計画🤫" if st.session_state.get("secret_doc_mode", False) else "個別支援計画"),
    ("書類_モニタリング", "🤫モニタリング🤫" if st.session_state.get("secret_doc_mode", False) else "モニタリング"),
    ("書類_一括書類作成", "🤫一括書類作成🤫") if st.session_state.get("secret_doc_mode", False) else None,    
    ("書類_在宅評価シート", "🤫在宅評価シート🤫" if st.session_state.get("secret_doc_mode", False) else "在宅評価シート"),
    ("書類_アセスメント", "アセスメント"),
    ("書類_基本シート", "基本シート"),
    ("書類_就労分野シート", "就労分野シート"),
]
document_page_options = [x for x in document_page_options if x is not None] 


def process_secret_command():
    cmd = str(st.session_state.get("secret_bee_cmd", "")).strip()

    if cmd == "🐝":
        st.session_state["bee_menu_unlocked"] = True
    elif cmd == "登録💻":
        st.session_state["other_office_register_unlocked"] = True
    elif cmd == "🤫":
        st.session_state["secret_doc_mode"] = True

    st.session_state["secret_bee_cmd"] = ""


# ===== メインメニュー =====
for p in main_page_options:
    is_selected = (st.session_state.current_page == p)
    display_p = p

    if is_selected:
        st.sidebar.markdown(
            f'<div class="menu-selected-wrap"><div class="menu-selected-box">• {display_p}</div></div>',
            unsafe_allow_html=True
        )
    else:
        if st.sidebar.button(display_p, key=f"menu_{p}", use_container_width=True):
            st.session_state.current_page = p
            st.rerun()

# ===== 利用者書類 =====
st.sidebar.markdown("### 利用者書類")

for page_key, page_label in document_page_options:
    is_selected = (st.session_state.current_page == page_key)
    display_label = page_label

    if is_selected:
        st.sidebar.markdown(
            f'<div class="menu-selected-wrap"><div class="menu-selected-box">• {display_label}</div></div>',
            unsafe_allow_html=True
        )
    else:
        if st.sidebar.button(display_label, key=f"doc_menu_{page_key}", use_container_width=True):
            st.session_state.current_page = page_key
            st.rerun()


# ===== ログアウト（ここ固定） =====
if st.sidebar.button("個人ログアウト", use_container_width=True):
    for k in [
        "user", "user_id", "is_admin", "login_at", "last_active_ping",
        "current_page", "bee_menu_unlocked", "other_office_register_unlocked", "secret_doc_mode",
        "heart_mode", "secret_bee_cmd"
    ]:
        if k in st.session_state:
            del st.session_state[k]
    st.session_state.auth_mode = "login"
    st.rerun()

if st.sidebar.button("事業所切り替え", use_container_width=True):
    for k in [
        "company_authenticated", "company_id", "company_name", "company_code",
        "company_login_id", "user", "user_id", "is_admin", "login_at",
        "last_active_ping", "current_page", "bee_menu_unlocked",
        "other_office_register_unlocked", "secret_doc_mode", "heart_mode", "secret_bee_cmd"
    ]:
        if k in st.session_state:
            del st.session_state[k]
    st.session_state.auth_mode = "login"
    st.rerun()

# ===== 🐝 knowbe（条件表示） =====
if st.session_state.get("bee_menu_unlocked", False):
    knowbe_label = "🐝knowbe日誌入力🐝"
    if st.session_state.get("heart_mode", False):
        knowbe_label = "💕🐝knowbe日誌入力🐝💕"

    if st.sidebar.button("🐝knowbe日誌入力🐝", key="knowbe_single_menu_button", use_container_width=True):
        st.session_state.current_page = "🐝knowbe日誌入力🐝"
        st.rerun()

if st.session_state.get("bee_menu_unlocked", False):
    knowbe_label = "🐝knowbe日誌一括入力🐝"
    if st.session_state.get("heart_mode", False):
        knowbe_label = "💕🐝knowbe日誌一括入力🐝💕"

    if st.sidebar.button("🐝knowbe日誌一括入力🐝", key="knowbe_bulk_menu_button", use_container_width=True):
        st.session_state.current_page = "🐝knowbe日誌一括入力🐝"
        st.rerun()

# ===== 💻 他事業所へ登録（条件表示） =====
if st.session_state.get("other_office_register_unlocked", False):
    register_label = "💻他事業所へ登録💻"

    if st.sidebar.button(register_label, key="other_office_register_menu_button", use_container_width=True):
        st.session_state.current_page = "💻他事業所へ登録💻"
        st.rerun()


# ===== 入力欄 =====
st.sidebar.text_input(
    "secret command",
    key="secret_bee_cmd",
    label_visibility="collapsed",
    on_change=process_secret_command,
)

# ===== 管理者メニュー =====
if st.session_state.get("is_admin", False):

    st.sidebar.markdown("### 管理者メニュー")

    if st.sidebar.button("スタッフ登録・削除", key="menu_staff_manage", use_container_width=True):
        st.session_state.current_page = "スタッフ管理"
        st.rerun()

    if st.sidebar.button("非接触ICカード登録", key="menu_ic_card_manage", use_container_width=True):
        st.session_state.current_page = "ICカード管理"
        st.rerun()

    if st.sidebar.button("Knowbe情報登録", key="menu_knowbe_settings", use_container_width=True):
        st.session_state.current_page = "Knowbe情報登録"
        st.rerun()

    if st.sidebar.button("勤怠管理", key="menu_attendance_manage", use_container_width=True):
        st.session_state.current_page = "勤怠管理"
        st.rerun()

    if st.sidebar.button("過去日誌照合", key="menu_support_record_audit", use_container_width=True):
        st.session_state.current_page = "過去日誌照合"
        st.rerun()

# ===== 最下部 =====
st.sidebar.divider()
st.sidebar.caption("System Version 2.0")

page = st.session_state.current_page

render_urgent_banner()

TEMPLATE_FILES = {
    "個別支援計画案": "個別支援計画案.xlsx",
    "個別支援計画": "個別支援計画.xlsx",
    "サービス担当者会議": "サービス担当者会議.xlsx",
    "モニタリング": "モニタリング.xlsx",
    "在宅評価シート": "在宅評価シート.xlsx",
    "アセスメント": "アセスメントシート.xlsx",
    "基本シート": "基本シート.xlsx",
    "就労分野シート": "就労分野シート.xlsx",    
}


def create_excel_file(template_name, cell_data):

    template = TEMPLATE_FILES[template_name]

    wb = load_workbook(template)
    ws = wb.active

    try:
        if "B3" in cell_data:
            ws.title = str(cell_data["B3"])[:31]
    except Exception:
        pass

    for cell, value in cell_data.items():
        try:
            # st.write(f"DEBUG ▶ {cell} = {value}")
            ws[cell] = value
        except Exception as e:
            st.error(f"❌ エラーセル: {cell}")
            st.error(f"❌ 値: {value}")
            st.error(f"❌ エラー内容: {e}")
            raise

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer

def render_plan_form_page(doc_title: str):
    st.title(f"📄 {doc_title}")
    st.caption("まずは入力しやすい形の試作ページです。まだ保存はしません。")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    st.markdown("## 基本情報")

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_name = str(selected_row.get("resident_name", "")).strip()

    basic_cols = st.columns([7, 2, 2, 2])

    with basic_cols[0]:
        st.text_input(
            "利用者氏名",
            value=resident_name,
            key=f"{doc_title}_resident_name",
            disabled=True
        )

    with basic_cols[1]:
        year_val = st.text_input("作成年（西暦）", key=f"{doc_title}_year", placeholder="2026")
    with basic_cols[2]:
        month_val = st.text_input("月", key=f"{doc_title}_month", placeholder="3")
    with basic_cols[3]:
        day_val = st.text_input("日", key=f"{doc_title}_day", placeholder="14")

    st.divider()

    st.markdown("## 本文入力")

    policy_val = st.text_area(
        "サービス等利用計画の総合的な方針",
        key=f"{doc_title}_policy",
        height=120,
        placeholder="B8 に入る内容です"
    )

    long_goal_val = st.text_area(
        "長期目標（内容・期間等）",
        key=f"{doc_title}_long_goal",
        height=100,
        placeholder="B10 に入る内容です"
    )

    short_goal_val = st.text_area(
        "短期目標（内容・期間等）",
        key=f"{doc_title}_short_goal",
        height=100,
        placeholder="B12 に入る内容です"
    )

    st.divider()
    st.markdown("## 具体的達成目標（3行）")
    st.caption("帳票の 17〜19 行に入る部分です。")

    header_cols = st.columns([5, 3, 4, 2, 2, 2])
    with header_cols[0]:
        st.markdown("**具体的達成目標**")
    with header_cols[1]:
        st.markdown("**本人の役割**")
    with header_cols[2]:
        st.markdown("**支援内容**")
    with header_cols[3]:
        st.markdown("**支援期間**")
    with header_cols[4]:
        st.markdown("**担当者**")
    with header_cols[5]:
        st.markdown("**優先順位**")

    row_data = []

    for i in range(1, 4):
        st.markdown(f"### {i}行目")
        row_cols = st.columns([5, 3, 4, 2, 2, 2])

        with row_cols[0]:
            target_val = st.text_area(
                f"{i}行目_具体的達成目標",
                key=f"{doc_title}_target_{i}",
                height=90,
                label_visibility="collapsed",
                placeholder="C17〜C19"
            )

        with row_cols[1]:
            role_val = st.text_area(
                f"{i}行目_本人の役割",
                key=f"{doc_title}_role_{i}",
                height=90,
                label_visibility="collapsed",
                placeholder="G17〜G19"
            )

        with row_cols[2]:
            support_val = st.text_area(
                f"{i}行目_支援内容",
                key=f"{doc_title}_support_{i}",
                height=90,
                label_visibility="collapsed",
                placeholder="J17〜J19"
            )

        with row_cols[3]:
            period_val = st.text_input(
                f"{i}行目_支援期間",
                key=f"{doc_title}_period_{i}",
                label_visibility="collapsed",
                placeholder="M17〜M19"
            )

        with row_cols[4]:
            person_val = st.text_input(
                f"{i}行目_担当者",
                key=f"{doc_title}_person_{i}",
                label_visibility="collapsed",
                placeholder="O17〜O19"
            )

        with row_cols[5]:
            priority_val = st.selectbox(
                f"{i}行目_優先順位",
                ["", "1", "2", "3"],
                key=f"{doc_title}_priority_{i}",
                label_visibility="collapsed"
            )

        row_data.append({
            "target": target_val,
            "role": role_val,
            "support": support_val,
            "period": period_val,
            "person": person_val,
            "priority": priority_val,
        })

    st.divider()
    st.markdown("## 同意・担当者")

    agree_cols = st.columns([2, 2, 2, 4])

    with agree_cols[0]:
        agree_year_val = st.text_input("同意日_西暦", key=f"{doc_title}_agree_year", placeholder="2026")
    with agree_cols[1]:
        agree_month_val = st.text_input("同意日_月", key=f"{doc_title}_agree_month", placeholder="3")
    with agree_cols[2]:
        agree_day_val = st.text_input("同意日_日", key=f"{doc_title}_agree_day", placeholder="14")
    with agree_cols[3]:
        manager_val = st.text_input("サービス担当責任者", key=f"{doc_title}_manager", placeholder="N21")

    st.divider()

    with st.expander("入力内容確認"):
        st.write(f"利用者氏名: {resident_name}")
        st.write(f"作成年月日: {year_val} / {month_val} / {day_val}")
        st.write(f"総合的な方針: {policy_val}")
        st.write(f"長期目標: {long_goal_val}")
        st.write(f"短期目標: {short_goal_val}")
        st.write(f"同意書日付: {agree_year_val} / {agree_month_val} / {agree_day_val}")
        st.write(f"サービス担当責任者: {manager_val}")

        for idx, item in enumerate(row_data, start=1):
            st.markdown(f"**{idx}行目**")
            st.write(item)

    st.divider()
    st.markdown("### Excel出力")

    cell_data = {
        # 基本情報
        "E5": resident_name,
        "M5": year_val,
        "O5": month_val,
        "Q5": day_val,

        # 総合方針・目標
        "B8": policy_val,
        "B10": long_goal_val,
        "B12": short_goal_val,

        # ===== 1行目 =====
        "C17": row_data[0]["target"],
        "G17": row_data[0]["role"],
        "J17": row_data[0]["support"],
        "M17": row_data[0]["period"],
        "O17": row_data[0]["person"],
        "Q17": row_data[0]["priority"],

        # ===== 2行目 =====
        "C18": row_data[1]["target"],
        "G18": row_data[1]["role"],
        "J18": row_data[1]["support"],
        "M18": row_data[1]["period"],
        "O18": row_data[1]["person"],
        "Q18": row_data[1]["priority"],

        # ===== 3行目 =====
        "C19": row_data[2]["target"],
        "G19": row_data[2]["role"],
        "J19": row_data[2]["support"],
        "M19": row_data[2]["period"],
        "O19": row_data[2]["person"],
        "Q19": row_data[2]["priority"],

        # 同意・担当者
        "B21": agree_year_val,
        "E21": agree_month_val,
        "H21": agree_day_val,
        "N21": manager_val
    }

    template_name = doc_title
    file_name = f"{doc_title}_{year_val}.{month_val}.{day_val}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )

def render_meeting_form_page(doc_title: str):
    st.title(f"📄 {doc_title}")
    st.caption("サービス担当者会議の入力UIです。まだ保存はしません。")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    st.markdown("## 基本情報")

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_name = str(selected_row.get("resident_name", "")).strip()

    # -----------------------------
    # 作成年月日
    # M3 / O3 / Q3
    # -----------------------------
    st.markdown("### 作成年月日")
    create_cols = st.columns([2, 2, 2, 6])

    with create_cols[0]:
        create_year = st.text_input("作成年（西暦）", key=f"{doc_title}_create_year", placeholder="2026")
    with create_cols[1]:
        create_month = st.text_input("月", key=f"{doc_title}_create_month", placeholder="3")
    with create_cols[2]:
        create_day = st.text_input("日", key=f"{doc_title}_create_day", placeholder="14")

    # -----------------------------
    # 利用者名 / 作成者
    # C4 / M4
    # -----------------------------
    base_cols = st.columns([6, 4])

    with base_cols[0]:
        st.text_input(
            "利用者名",
            value=resident_name,
            key=f"{doc_title}_resident_name",
            disabled=True
        )
    with base_cols[1]:
        creator_name = st.text_input(
            "作成者（担当者）名",
            value=st.session_state.user,
            key=f"{doc_title}_creator_name"
        )

    st.divider()

    # -----------------------------
    # 開催日時 / 開催場所
    # C5 E5 G5 / M5
    # -----------------------------
    st.markdown("## 開催情報")

    meeting_cols = st.columns([2, 2, 2, 6])

    with meeting_cols[0]:
        meeting_year = st.text_input("開催年（西暦）", key=f"{doc_title}_meeting_year", placeholder="2026")
    with meeting_cols[1]:
        meeting_month = st.text_input("開催月", key=f"{doc_title}_meeting_month", placeholder="3")
    with meeting_cols[2]:
        meeting_day = st.text_input("開催日", key=f"{doc_title}_meeting_day", placeholder="14")
    with meeting_cols[3]:
        meeting_place = st.text_input("開催場所", key=f"{doc_title}_meeting_place", placeholder="事業所相談室")

    st.divider()

    # -----------------------------
    # 会議出席者
    # E8 J8 O8 / E9 J9 O9 / E10 J10 O10
    # -----------------------------
    st.markdown("## 会議出席者")

    header_cols = st.columns(3)
    with header_cols[0]:
        st.markdown("**左列**")
    with header_cols[1]:
        st.markdown("**中央列**")
    with header_cols[2]:
        st.markdown("**右列**")

    row1 = st.columns(3)
    with row1[0]:
        manager_name = st.text_input("管理者名", key=f"{doc_title}_manager_name")
    with row1[1]:
        staff_name = st.text_input("支援員名", key=f"{doc_title}_staff_name")
    with row1[2]:
        attendee_user_name = st.text_input(
            "利用者名",
            value=resident_name,
            key=f"{doc_title}_attendee_user_name"
        )

    row2 = st.columns(3)
    with row2[0]:
        care_manager_name = st.text_input("ケアマネ", key=f"{doc_title}_care_manager_name")
    with row2[1]:
        nurse_name = st.text_input("看護師", key=f"{doc_title}_nurse_name")
    with row2[2]:
        family_name = st.text_input("親族", key=f"{doc_title}_family_name")

    row3 = st.columns(3)
    with row3[0]:
        service_manager_name = st.text_input("サービス管理責任者", key=f"{doc_title}_service_manager_name")
    with row3[1]:
        consultant_name = st.text_input("相談員", key=f"{doc_title}_consultant_name")
    with row3[2]:
        keyperson_name = st.text_input("キーパーソン", key=f"{doc_title}_keyperson_name")

    st.divider()

    # -----------------------------
    # 本文
    # C11〜C14
    # -----------------------------
    st.markdown("## 会議内容")

    agenda = st.text_area(
        "議題",
        key=f"{doc_title}_agenda",
        height=90,
        placeholder="C11"
    )

    discussion = st.text_area(
        "検討内容",
        key=f"{doc_title}_discussion",
        height=140,
        placeholder="C12"
    )

    issues_left = st.text_area(
        "残された課題",
        key=f"{doc_title}_issues_left",
        height=100,
        placeholder="C13"
    )

    conclusion = st.text_area(
        "結論",
        key=f"{doc_title}_conclusion",
        height=100,
        placeholder="C14"
    )

    st.divider()

    with st.expander("入力内容確認"):
        st.write(f"利用者名: {resident_name}")
        st.write(f"作成年月日: {create_year} / {create_month} / {create_day}")
        st.write(f"作成者: {creator_name}")
        st.write(f"開催日時: {meeting_year} / {meeting_month} / {meeting_day}")
        st.write(f"開催場所: {meeting_place}")

        st.markdown("**会議出席者**")
        st.write({
            "管理者名": manager_name,
            "支援員名": staff_name,
            "利用者名": attendee_user_name,
            "ケアマネ": care_manager_name,
            "看護師": nurse_name,
            "親族": family_name,
            "サービス管理責任者": service_manager_name,
            "相談員": consultant_name,
            "キーパーソン": keyperson_name,
        })

        st.write(f"議題: {agenda}")
        st.write(f"検討内容: {discussion}")
        st.write(f"残された課題: {issues_left}")
        st.write(f"結論: {conclusion}")

    st.divider()
    st.markdown("### Excel出力")

    cell_data = {
            "M3": create_year,
            "O3": create_month,
            "Q3": create_day,
            "C4": resident_name,
            "M4": creator_name,
            "C5": meeting_year,
            "E5": meeting_month,
            "G5": meeting_day,
            "M5": meeting_place,
            "E8": manager_name,
            "J8": staff_name,
            "O8": attendee_user_name,
            "E9": care_manager_name,
            "J9": nurse_name,
            "O9": family_name,
            "E10": service_manager_name,
            "J10": consultant_name,
            "O10": keyperson_name,
            "C11": agenda,
            "C12": discussion,
            "C13": issues_left,
            "C14": conclusion,
    }

    template_name = doc_title
    file_name = f"{doc_title}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )


def render_monitoring_form_page(doc_title: str):
    st.title(f"📄 {doc_title}")
    st.caption("モニタリングの入力UIです。まだ保存はしません。")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    st.markdown("## 基本情報")

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_name = str(selected_row.get("resident_name", "")).strip()

    base_cols = st.columns([6, 2, 2, 2])

    with base_cols[0]:
        st.text_input(
            "利用者名",
            value=resident_name,
            key=f"{doc_title}_resident_name",
            disabled=True
        )

    with base_cols[1]:
        year_val = st.text_input("実施年（西暦）", key=f"{doc_title}_year", placeholder="2026")
    with base_cols[2]:
        month_val = st.text_input("月", key=f"{doc_title}_month", placeholder="3")
    with base_cols[3]:
        day_val = st.text_input("日", key=f"{doc_title}_day", placeholder="14")

    st.divider()

    st.markdown("## モニタリング内容")
    st.caption("具体的達成目標番号ごとに入力してください。")

    row_data = []

    for i in range(1, 4):
        st.markdown(f"### 具体的達成目標番号{i}")

        row_cols = st.columns([2, 5, 5])

        with row_cols[0]:
            status_val = st.selectbox(
                f"達成状況の評価_{i}",
                ["", "達成", "継続", "一部達成", "終了"],
                key=f"{doc_title}_status_{i}"
            )

        with row_cols[1]:
            detail_val = st.text_area(
                f"達成できている点と未達成点（要因も）_{i}",
                key=f"{doc_title}_detail_{i}",
                height=120,
                placeholder=f"D{7+i}"
            )

        with row_cols[2]:
            future_val = st.text_area(
                f"今後の対応（支援内容・方法の変更・継続・終了）_{i}",
                key=f"{doc_title}_future_{i}",
                height=120,
                placeholder=f"E{7+i}"
            )

        row_data.append({
            "status": status_val,
            "detail": detail_val,
            "future": future_val,
        })

        st.divider()

    with st.expander("入力内容確認"):
        st.write(f"利用者名: {resident_name}")
        st.write(f"実施年月日: {year_val} / {month_val} / {day_val}")

        for idx, item in enumerate(row_data, start=1):
            st.markdown(f"**具体的達成目標番号{idx}**")
            st.write(item)

    st.divider()
    st.markdown("### Excel出力")

    cell_data = {
            "C5": resident_name,
            "E5": year_val,
            "G5": month_val,
            "I5": day_val,

            "C8": row_data[0]["status"],
            "D8": row_data[0]["detail"],
            "E8": row_data[0]["future"],

            "C9": row_data[1]["status"],
            "D9": row_data[1]["detail"],
            "E9": row_data[1]["future"],

            "C10": row_data[2]["status"],
            "D10": row_data[2]["detail"],
            "E10": row_data[2]["future"],
    }

    template_name = doc_title
    file_name = f"{doc_title}_{year_val}.{month_val}.{day_val}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )

def render_home_evaluation_form_page(doc_title: str):
    st.title(f"📄 {doc_title}")
    st.caption("在宅評価シートの入力UIです。まだ保存はしません。")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    st.markdown("## 基本情報")

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_name = str(selected_row.get("resident_name", "")).strip()

    base_cols = st.columns([6, 2, 2])

    with base_cols[0]:
        st.text_input(
            "利用者名",
            value=resident_name,
            key=f"{doc_title}_resident_name",
            disabled=True
        )

    with base_cols[1]:
        year_val = st.text_input("年（西暦）", key=f"{doc_title}_year", placeholder="2026")
    with base_cols[2]:
        month_val = st.text_input("月", key=f"{doc_title}_month", placeholder="3")

    st.divider()

    st.markdown("## 目標")
    goal1 = st.text_area("目標1", key=f"{doc_title}_goal1", height=80, placeholder="B7")
    goal2 = st.text_area("目標2", key=f"{doc_title}_goal2", height=80, placeholder="B8")
    goal3 = st.text_area("目標3", key=f"{doc_title}_goal3", height=80, placeholder="B9")

    st.divider()

    service_manager = st.text_input(
        "サービス管理責任者",
        value=st.session_state.user,
        key=f"{doc_title}_service_manager",
        placeholder="H11"
    )

    st.divider()

    st.markdown("## 月間評価")
    monthly1 = st.text_area("月間評価1", key=f"{doc_title}_monthly1", height=90, placeholder="B12")
    monthly2 = st.text_area("月間評価2", key=f"{doc_title}_monthly2", height=90, placeholder="B13")
    monthly3 = st.text_area("月間評価3", key=f"{doc_title}_monthly3", height=90, placeholder="B14")

    st.divider()

    st.markdown("## 週ごとの評価（週報）")
    st.caption("土曜日の日付は、入力した年・月から自動表示します。")

    ym_key = f"{doc_title}_auto_year_month"
    current_ym = f"{year_val}-{month_val}"

    # 先に土曜日一覧を計算する
    try:
        y = int(str(year_val).strip())
        m = int(str(month_val).strip())
        if 1 <= m <= 12:
            saturday_dates = get_saturday_dates_for_month(y, m)
        else:
            saturday_dates = []
    except Exception:
        saturday_dates = []

    # 年月が変わったときだけ自動反映
    if st.session_state.get(ym_key) != current_ym:
        for i in range(1, 6):
            sat_key = f"{doc_title}_sat_{i}"
            sat_text = ""
            if len(saturday_dates) >= i:
                sat_text = saturday_dates[i - 1].strftime("%Y-%m-%d")
            st.session_state[sat_key] = sat_text

        st.session_state[ym_key] = current_ym
    try:
        y = int(str(year_val).strip())
        m = int(str(month_val).strip())
        if 1 <= m <= 12:
            saturday_dates = get_saturday_dates_for_month(y, m)
    except Exception:
        saturday_dates = []

    week_rows = []

    for i in range(1, 6):
        st.markdown(f"### 第{i}週")

        sat_key = f"{doc_title}_sat_{i}"

        row1 = st.columns([2, 8])
        with row1[0]:
            sat_input = st.text_input(
                f"第{i}週 土曜日日付",
                key=sat_key
            )
        with row1[1]:
            weekly_report = st.text_area(
                f"第{i}週 週報",
                key=f"{doc_title}_weekly_report_{i}",
                height=80,
                placeholder=f"C{17 + i * 2}"
            )

        row2 = st.columns([6, 4])
        with row2[1]:
            visit_manager = st.text_input(
                f"第{i}週 訪問したサービス管理責任者名",
                key=f"{doc_title}_visit_manager_{i}",
                placeholder=f"J{18 + i * 2}"
            )

        week_rows.append({
            "sat_date": sat_input,
            "weekly_report": weekly_report,
            "visit_manager": visit_manager,
        })

        st.divider()

    with st.expander("入力内容確認"):
        st.write(f"利用者名: {resident_name}")
        st.write(f"対象年月: {year_val}年 / {month_val}月")
        st.write(f"サービス管理責任者: {service_manager}")

        st.markdown("**目標**")
        st.write({
            "目標1": goal1,
            "目標2": goal2,
            "目標3": goal3,
        })

        st.markdown("**月間評価**")
        st.write({
            "月間評価1": monthly1,
            "月間評価2": monthly2,
            "月間評価3": monthly3,
        })

        for idx, item in enumerate(week_rows, start=1):
            st.markdown(f"**第{idx}週**")
            st.write(item)

    st.divider()
    st.markdown("### Excel出力")

    cell_data = {
            "B3": resident_name,
            "J3": year_val,
            "L3": month_val,

            "B7": goal1,
            "B8": goal2,
            "B9": goal3,

            "H11": service_manager,

            "B12": monthly1,
            "B13": monthly2,
            "B14": monthly3,

            "A19": week_rows[0]["sat_date"],
            "C19": week_rows[0]["weekly_report"],
            "J20": week_rows[0]["visit_manager"],

            "A21": week_rows[1]["sat_date"],
            "C21": week_rows[1]["weekly_report"],
            "J22": week_rows[1]["visit_manager"],

            "A23": week_rows[2]["sat_date"],
            "C23": week_rows[2]["weekly_report"],
            "J24": week_rows[2]["visit_manager"],

            "A25": week_rows[3]["sat_date"],
            "C25": week_rows[3]["weekly_report"],
            "J26": week_rows[3]["visit_manager"],

            "A27": week_rows[4]["sat_date"],
            "C27": week_rows[4]["weekly_report"],
            "J28": week_rows[4]["visit_manager"],
    }

    template_name = doc_title
    file_name = f"{doc_title}_{year_val}.{month_val}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )

# ==========================================
# 書類_アセスメント（フェイスシート・試作UI）
# ==========================================
def render_assessment_form_page(doc_title: str):
    st.title("📋 アセスメントシート")
    st.caption("フェイスシート入力ページです。入力とExcel出力までつなぐです。")

    st.markdown("## 基本情報")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_name = str(selected_row.get("resident_name", "")).strip()

    # -----------------------------
    # 聴き取り情報
    # P1 / Y1 / AD1 / AG1
    # -----------------------------
    st.markdown("### 聴き取り情報")

    hear_cols = st.columns([3, 2, 1, 1])
    with hear_cols[0]:
        interviewer_name = st.text_input("聴き取り者名", key=f"{doc_title}_interviewer_name")
    with hear_cols[1]:
        hear_year = st.text_input("聞き取り日（西暦）", key=f"{doc_title}_hear_year", placeholder="2026")
    with hear_cols[2]:
        hear_month = st.text_input("月", key=f"{doc_title}_hear_month", placeholder="3")
    with hear_cols[3]:
        hear_day = st.text_input("日", key=f"{doc_title}_hear_day", placeholder="15")

    st.divider()

    # -----------------------------
    # 氏名
    # F4 / F5
    # -----------------------------
    st.markdown("### 氏名")

    name_cols = st.columns([3, 4])
    with name_cols[0]:
        furigana = st.text_input("フリガナ", key=f"{doc_title}_furigana")
    with name_cols[1]:
        full_name = st.text_input(
            "氏名",
            value=resident_name,
            key=f"{doc_title}_full_name"
        )

    # -----------------------------
    # 生年月日
    # U4 / AA4 / AC4 / AG4
    # -----------------------------
    st.markdown("### 生年月日")

    birth_cols = st.columns([2, 1, 1, 1])
    with birth_cols[0]:
        birth_year = st.text_input("西暦", key=f"{doc_title}_birth_year", placeholder="1980")
    with birth_cols[1]:
        birth_month = st.text_input("月", key=f"{doc_title}_birth_month", placeholder="1")
    with birth_cols[2]:
        birth_day = st.text_input("日", key=f"{doc_title}_birth_day", placeholder="1")
    with birth_cols[3]:
        age = st.text_input("年齢", key=f"{doc_title}_age", placeholder="45")

    st.divider()

    # -----------------------------
    # 現住所
    # H7 / K7 / F8 / AA7 / AA9
    # -----------------------------
    st.markdown("## 現住所")

    current_top_cols = st.columns([1, 1, 2, 2])
    with current_top_cols[0]:
        current_zip_1 = st.text_input("郵便番号（上3桁）", key=f"{doc_title}_current_zip_1", placeholder="557")
    with current_top_cols[1]:
        current_zip_2 = st.text_input("郵便番号（下4桁）", key=f"{doc_title}_current_zip_2", placeholder="0000")
    with current_top_cols[2]:
        current_phone = st.text_input("連絡先電話番号", key=f"{doc_title}_current_phone")
    with current_top_cols[3]:
        nearest_station = st.text_input("最寄り駅", key=f"{doc_title}_nearest_station")

    current_address = st.text_area(
        "住所",
        key=f"{doc_title}_current_address",
        height=100
    )

    st.divider()

    # -----------------------------
    # 緊急連絡先
    # H11 / K11 / F12 / AA11 / AA13
    # -----------------------------
    st.markdown("## 緊急連絡先")

    emergency_top_cols = st.columns([1, 1, 2, 2])
    with emergency_top_cols[0]:
        emergency_zip_1 = st.text_input("郵便番号（上3桁）", key=f"{doc_title}_emergency_zip_1", placeholder="557")
    with emergency_top_cols[1]:
        emergency_zip_2 = st.text_input("郵便番号（下4桁）", key=f"{doc_title}_emergency_zip_2", placeholder="0000")
    with emergency_top_cols[2]:
        emergency_relation = st.text_input("続柄", key=f"{doc_title}_emergency_relation")
    with emergency_top_cols[3]:
        emergency_phone_fax = st.text_input("電話 / FAX", key=f"{doc_title}_emergency_phone_fax")

    emergency_address = st.text_area(
        "住所",
        key=f"{doc_title}_emergency_address",
        height=100
    )

    st.divider()

    # -----------------------------
    # 援護実施機関
    # F15 / J15 / X15 / AF15
    # -----------------------------
    st.markdown("## 援護実施機関")

    support_cols = st.columns([2, 1, 3, 2])
    with support_cols[0]:
        support_city = st.text_input("市区名", key=f"{doc_title}_support_city")
    with support_cols[1]:
        support_city_type = st.selectbox(
            "区 / 市",
            ["区", "市"],
            key=f"{doc_title}_support_city_type"
        )
    with support_cols[2]:
        support_office = st.text_input("相談支援事業所", key=f"{doc_title}_support_office")
    with support_cols[3]:
        support_worker = st.text_input("担当ワーカー", key=f"{doc_title}_support_worker")

    st.divider()

    # -----------------------------
    # 手帳・福祉制度
    # -----------------------------
    st.markdown("## 手帳・福祉制度")

    handbook_cols = st.columns([2, 2, 1, 1])
    with handbook_cols[0]:
        handbook_grade = st.text_input("手帳に記入されている級", key=f"{doc_title}_handbook_grade")
    with handbook_cols[1]:
        handbook_year = st.text_input("手帳取得年（西暦）", key=f"{doc_title}_handbook_year", placeholder="2020")
    with handbook_cols[2]:
        handbook_month = st.text_input("月", key=f"{doc_title}_handbook_month", placeholder="1")
    with handbook_cols[3]:
        handbook_day = st.text_input("日", key=f"{doc_title}_handbook_day", placeholder="1")

    disability_summary = st.text_area(
        "障害状況の概要、および手帳取得経緯",
        key=f"{doc_title}_disability_summary",
        height=120
    )

    welfare_cols1 = st.columns(2)
    with welfare_cols1[0]:
        support_level = st.selectbox(
            "障害支援区分",
            ["区分無し", "区分1", "区分2", "区分3", "区分4", "区分5", "区分6"],
            key=f"{doc_title}_support_level"
        )
    with welfare_cols1[1]:
        guardian_status = st.selectbox(
            "成年後見人の有無",
            ["なし", "あり"],
            key=f"{doc_title}_guardian_status"
        )

    guardian_name = st.text_input(
        "成年後見人の氏名（ありの場合）",
        key=f"{doc_title}_guardian_name"
    )

    welfare_cols2 = st.columns(2)
    with welfare_cols2[0]:
        pension_status = st.selectbox(
            "年金（種別）の有無",
            ["なし", "あり"],
            key=f"{doc_title}_pension_status"
        )
    with welfare_cols2[1]:
        allowance_status = st.selectbox(
            "特別児童扶養手当その他の有無",
            ["なし", "あり"],
            key=f"{doc_title}_allowance_status"
        )

    pension_detail = st.text_input(
        "年金詳細（何級・月額いくら等）",
        key=f"{doc_title}_pension_detail"
    )
    allowance_detail = st.text_input(
        "特別児童扶養手当その他詳細（何級・月額いくら等）",
        key=f"{doc_title}_allowance_detail"
    )

    welfare_cols3 = st.columns(3)
    with welfare_cols3[0]:
        transport_pass = st.selectbox(
            "大阪市交通局優待乗車証",
            ["無料", "半額", "なし"],
            key=f"{doc_title}_transport_pass"
        )
    with welfare_cols3[1]:
        welfare_status = st.selectbox(
            "生活保護の受給状況",
            ["あり", "なし"],
            key=f"{doc_title}_welfare_status"
        )
    with welfare_cols3[2]:
        public_support = st.text_input(
            "公費医療・福祉用具等の利用",
            key=f"{doc_title}_public_support"
        )

    st.divider()

    # -----------------------------
    # 家族の状況
    # -----------------------------
    st.markdown("## 家族の状況")

    family_rows = []
    for i in range(4):
        st.markdown(f"### 家族{i+1}")
        cols = st.columns([2, 1, 1, 2, 1, 2])

        with cols[0]:
            fam_name = st.text_input("氏名", key=f"{doc_title}_fam_name_{i}")
        with cols[1]:
            fam_relation = st.text_input("続柄", key=f"{doc_title}_fam_relation_{i}")
        with cols[2]:
            fam_age = st.text_input("年齢", key=f"{doc_title}_fam_age_{i}")
        with cols[3]:
            fam_job = st.text_input("職業等", key=f"{doc_title}_fam_job_{i}")
        with cols[4]:
            fam_live = st.selectbox("同居・別居", ["同居", "別居", ""], key=f"{doc_title}_fam_live_{i}")
        with cols[5]:
            fam_note = st.text_input("特記事項", key=f"{doc_title}_fam_note_{i}")

        family_rows.append({
            "name": fam_name,
            "relation": fam_relation,
            "age": fam_age,
            "job": fam_job,
            "live": fam_live,
            "note": fam_note
        })

    st.divider()

    # -----------------------------
    # 住居環境等
    # -----------------------------
    st.markdown("## 住居環境等")

    housing_cols = st.columns(2)
    with housing_cols[0]:
        housing_transport = st.selectbox(
            "住居の状況（交通手段）",
            ["電車", "バス", "自転車", "その他"],
            key=f"{doc_title}_housing_transport"
        )
    with housing_cols[1]:
        housing_transport_other = st.text_input(
            "住居の状況（その他内容）",
            key=f"{doc_title}_housing_transport_other"
        )

    housing_use_status = st.selectbox(
        "住居の状況 利用する場合具体的な状況",
        ["単独利用", "家族等の付き添い", "その他"],
        key=f"{doc_title}_housing_use_status"
    )
    housing_use_status_other = st.text_input(
        "住居の状況 利用状況（その他内容）",
        key=f"{doc_title}_housing_use_status_other"
    )

    transport_cols = st.columns(2)
    with transport_cols[0]:
        available_transport = st.selectbox(
            "利用可能な交通手段",
            ["電車", "バス", "自転車", "その他"],
            key=f"{doc_title}_available_transport"
        )
    with transport_cols[1]:
        available_transport_other = st.text_input(
            "利用可能な交通手段（その他内容）",
            key=f"{doc_title}_available_transport_other"
        )

    available_use_status = st.selectbox(
        "利用可能交通手段 利用する場合具体的な状況",
        ["単独利用", "家族等の付き添い", "その他"],
        key=f"{doc_title}_available_use_status"
    )
    available_use_status_other = st.text_input(
        "利用可能交通手段 利用状況（その他内容）",
        key=f"{doc_title}_available_use_status_other"
    )

    st.divider()

    # -----------------------------
    # 生活歴
    # -----------------------------
    st.markdown("## 生活歴")

    life_rows = []
    for i in range(3):
        st.markdown(f"### 生活歴{i+1}")
        cols = st.columns([1, 3])
        with cols[0]:
            life_date = st.text_input("西暦年月日", key=f"{doc_title}_life_date_{i}", placeholder="2000/04")
        with cols[1]:
            life_history = st.text_area(
                "生活歴（学歴や転居等の経緯）",
                key=f"{doc_title}_life_history_{i}",
                height=80
            )
        life_rows.append({
            "date": life_date,
            "history": life_history
        })

    st.divider()

    # -----------------------------
    # 医療機関の受診状況等
    # -----------------------------
    st.markdown("## 医療機関の受診状況等")

    medical_cols1 = st.columns([2, 3])
    with medical_cols1[0]:
        disease_name = st.text_input("病名（複数入力可）", key=f"{doc_title}_disease_name")
    with medical_cols1[1]:
        disease_symptom = st.text_input("症状（複数入力可）", key=f"{doc_title}_disease_symptom")

    st.markdown("### 医療機関")
    medical_cols2 = st.columns([2, 2, 2, 1, 2])
    with medical_cols2[0]:
        hospital_name = st.text_input("病院名", key=f"{doc_title}_hospital_name")
    with medical_cols2[1]:
        doctor_name = st.text_input("医師名", key=f"{doc_title}_doctor_name")
    with medical_cols2[2]:
        hospital_contact = st.text_input("連絡先", key=f"{doc_title}_hospital_contact")
    with medical_cols2[3]:
        visit_frequency = st.text_input("通院頻度", key=f"{doc_title}_visit_frequency")
    with medical_cols2[4]:
        medication_status = st.text_input("服薬状況等", key=f"{doc_title}_medication_status")

    st.divider()

    # -----------------------------
    # 心身状況等
    # -----------------------------
    st.markdown("## 心身状況等")

    mind_rows = []
    for i in range(2):
        st.markdown(f"### 心身状況{i+1}")
        cols = st.columns([2, 3, 3])
        with cols[0]:
            mind_disease = st.text_input("障害名・病名", key=f"{doc_title}_mind_disease_{i}")
        with cols[1]:
            mind_symptom = st.text_input("症状など", key=f"{doc_title}_mind_symptom_{i}")
        with cols[2]:
            mind_support = st.text_input("必要な支援の内容", key=f"{doc_title}_mind_support_{i}")

        mind_rows.append({
            "disease": mind_disease,
            "symptom": mind_symptom,
            "support": mind_support
        })

    st.divider()

    # -----------------------------
    # 障害福祉サービスなどの利用状況
    # -----------------------------
    st.markdown("## 障害福祉サービスなどの利用状況")

    service_rows = []
    for i in range(3):
        st.markdown(f"### サービス利用{i+1}")
        cols = st.columns([1, 2, 1, 2])
        with cols[0]:
            service_date = st.text_input("利用開始日（時期）", key=f"{doc_title}_service_date_{i}")
        with cols[1]:
            service_name = st.text_input("サービス名", key=f"{doc_title}_service_name_{i}")
        with cols[2]:
            service_amount = st.text_input("利用量/月", key=f"{doc_title}_service_amount_{i}")
        with cols[3]:
            service_office = st.text_input("事業所名", key=f"{doc_title}_service_office_{i}")

        service_rows.append({
            "date": service_date,
            "name": service_name,
            "amount": service_amount,
            "office": service_office
        })

    st.divider()

    # -----------------------------
    # 生活の流れ等
    # -----------------------------
    st.markdown("## 生活の流れ等")

    day_flow = st.text_area(
        "標準的な１日の生活の流れ（起床から就寝まで）",
        key=f"{doc_title}_day_flow",
        height=120
    )
    special_note = st.text_area(
        "特記事項（１週間の過ごし方やいきがい、趣味、特技など）",
        key=f"{doc_title}_special_note",
        height=120
    )

    st.divider()

    # -----------------------------
    # 総合所見
    # -----------------------------
    st.markdown("## 総合所見")

    wish_user = st.text_area(
        "当事業所のサービス利用に対する本人の希望",
        key=f"{doc_title}_wish_user",
        height=100
    )
    wish_family = st.text_area(
        "当事業所のサービス利用に対する保護者・関係者の希望する方向性",
        key=f"{doc_title}_wish_family",
        height=100
    )
    future_direction = st.text_area(
        "フェイスシートからみる課題や今後の方向性",
        key=f"{doc_title}_future_direction",
        height=120
    )

    st.divider()

    with st.expander("入力内容確認"):
        st.write({
            "聴き取り者名": interviewer_name,
            "聞き取り日": f"{hear_year}/{hear_month}/{hear_day}",
            "フリガナ": furigana,
            "氏名": full_name,
            "生年月日": f"{birth_year}/{birth_month}/{birth_day}",
            "年齢": age,
            "現住所_郵便番号上3桁": current_zip_1,
            "現住所_郵便番号下4桁": current_zip_2,
            "現住所": current_address,
            "現住所_電話番号": current_phone,
            "現住所_最寄り駅": nearest_station,
            "緊急連絡先_郵便番号上3桁": emergency_zip_1,
            "緊急連絡先_郵便番号下4桁": emergency_zip_2,
            "緊急連絡先_住所": emergency_address,
            "緊急連絡先_続柄": emergency_relation,
            "緊急連絡先_電話FAX": emergency_phone_fax,
            "援護実施機関_市区名": support_city,
            "援護実施機関_区市": support_city_type,
            "援護実施機関_相談支援事業所": support_office,
            "援護実施機関_担当ワーカー": support_worker,
            "手帳級": handbook_grade,
            "手帳取得日": f"{handbook_year}/{handbook_month}/{handbook_day}",
            "障害概要・取得経緯": disability_summary,
            "障害支援区分": support_level,
            "成年後見人": guardian_status,
            "成年後見人名": guardian_name,
            "年金": pension_status,
            "年金詳細": pension_detail,
            "特別児童扶養手当等": allowance_status,
            "特別児童扶養手当等詳細": allowance_detail,
            "大阪市交通局優待乗車証": transport_pass,
            "生活保護": welfare_status,
            "公費医療・福祉用具等": public_support,
            "家族状況": family_rows,
            "住居の状況": [housing_transport, housing_transport_other, housing_use_status, housing_use_status_other],
            "利用可能交通手段": [available_transport, available_transport_other, available_use_status, available_use_status_other],
            "生活歴": life_rows,
            "病名": disease_name,
            "症状": disease_symptom,
            "医療機関": [hospital_name, doctor_name, hospital_contact, visit_frequency, medication_status],
            "心身状況": mind_rows,
            "福祉サービス利用": service_rows,
            "1日の生活": day_flow,
            "特記事項": special_note,
            "本人の希望": wish_user,
            "家族・関係者の希望": wish_family,
            "課題・今後方向": future_direction,
        })

    # -----------------------------
    # 保存用データ作成
    # -----------------------------
    form_data = {
        "interviewer_name": interviewer_name,
        "hear_year": hear_year,
        "hear_month": hear_month,
        "hear_day": hear_day,
        "furigana": furigana,
        "full_name": full_name,
        "birth_year": birth_year,
        "birth_month": birth_month,
        "birth_day": birth_day,
        "age": age,
        "current_zip_1": current_zip_1,
        "current_zip_2": current_zip_2,
        "current_phone": current_phone,
        "nearest_station": nearest_station,
        "current_address": current_address,
        "emergency_zip_1": emergency_zip_1,
        "emergency_zip_2": emergency_zip_2,
        "emergency_relation": emergency_relation,
        "emergency_phone_fax": emergency_phone_fax,
        "emergency_address": emergency_address,
        "support_city": support_city,
        "support_city_type": support_city_type,
        "support_office": support_office,
        "support_worker": support_worker,
        "handbook_grade": handbook_grade,
        "handbook_year": handbook_year,
        "handbook_month": handbook_month,
        "handbook_day": handbook_day,
        "disability_summary": disability_summary,
        "support_level": support_level,
        "guardian_status": guardian_status,
        "guardian_name": guardian_name,
        "pension_status": pension_status,
        "allowance_status": allowance_status,
        "pension_detail": pension_detail,
        "allowance_detail": allowance_detail,
        "transport_pass": transport_pass,
        "welfare_status": welfare_status,
        "public_support": public_support,
        "family_rows": family_rows,
        "housing_transport": housing_transport,
        "housing_transport_other": housing_transport_other,
        "housing_use_status": housing_use_status,
        "housing_use_status_other": housing_use_status_other,
        "available_transport": available_transport,
        "available_transport_other": available_transport_other,
        "available_use_status": available_use_status,
        "available_use_status_other": available_use_status_other,
        "life_rows": life_rows,
        "disease_name": disease_name,
        "disease_symptom": disease_symptom,
        "hospital_name": hospital_name,
        "doctor_name": doctor_name,
        "hospital_contact": hospital_contact,
        "visit_frequency": visit_frequency,
        "medication_status": medication_status,
        "mind_rows": mind_rows,
        "service_rows": service_rows,
        "day_flow": day_flow,
        "special_note": special_note,
        "wish_user": wish_user,
        "wish_family": wish_family,
        "future_direction": future_direction,
    }

    st.divider()
    st.markdown("### Excel出力")

    guardian_value = "なし" if guardian_status == "なし" else f"あり：{guardian_name}" if guardian_name.strip() else "あり"
    pension_value = "なし" if pension_status == "なし" else f"あり：{pension_detail}" if pension_detail.strip() else "あり"
    allowance_value = "なし" if allowance_status == "なし" else f"あり：{allowance_detail}" if allowance_detail.strip() else "あり"

    housing_transport_value = housing_transport_other.strip() if housing_transport == "その他" and housing_transport_other.strip() else housing_transport
    housing_status_value = housing_use_status_other.strip() if housing_use_status == "その他" and housing_use_status_other.strip() else housing_use_status

    available_transport_value = available_transport_other.strip() if available_transport == "その他" and available_transport_other.strip() else available_transport
    available_status_value = available_use_status_other.strip() if available_use_status == "その他" and available_use_status_other.strip() else available_use_status

    cell_data = {
            "P1": interviewer_name,
            "Y1": hear_year,
            "AD1": hear_month,
            "AG1": hear_day,

            "F4": furigana,
            "F5": full_name,

            "U4": birth_year,
            "AA4": birth_month,
            "AC4": birth_day,
            "AG4": age,

            "H7": current_zip_1,
            "K7": current_zip_2,
            "F8": current_address,
            "AA7": current_phone,
            "AA9": nearest_station,

            "H11": emergency_zip_1,
            "K11": emergency_zip_2,
            "F12": emergency_address,
            "AA11": emergency_relation,
            "AA13": emergency_phone_fax,

            "F15": support_city,
            "J15": support_city_type,
            "X15": support_office,
            "AF15": support_worker,

            "G19": handbook_grade,
            "Y19": handbook_year,
            "AD19": handbook_month,
            "AG19": handbook_day,
            "G21": disability_summary,
            "G24": support_level,
            "X24": guardian_value,
            "G26": pension_value,
            "X26": allowance_value,
            "G28": transport_pass,
            "Q28": welfare_status,
            "Z28": public_support,

            "G33": family_rows[0]["name"],
            "N33": family_rows[0]["relation"],
            "Q33": family_rows[0]["age"],
            "T33": family_rows[0]["job"],
            "X33": family_rows[0]["live"],
            "AB33": family_rows[0]["note"],

            "G35": family_rows[1]["name"],
            "N35": family_rows[1]["relation"],
            "Q35": family_rows[1]["age"],
            "T35": family_rows[1]["job"],
            "X35": family_rows[1]["live"],
            "AB35": family_rows[1]["note"],

            "G37": family_rows[2]["name"],
            "N37": family_rows[2]["relation"],
            "Q37": family_rows[2]["age"],
            "T37": family_rows[2]["job"],
            "X37": family_rows[2]["live"],
            "AB37": family_rows[2]["note"],

            "G39": family_rows[3]["name"],
            "N39": family_rows[3]["relation"],
            "Q39": family_rows[3]["age"],
            "T39": family_rows[3]["job"],
            "X39": family_rows[3]["live"],
            "AB39": family_rows[3]["note"],

            "G43": housing_transport_value,
            "G45": available_transport_value,

            "A50": life_rows[0]["date"],
            "G50": life_rows[0]["history"],
            "A52": life_rows[1]["date"],
            "G52": life_rows[1]["history"],
            "A54": life_rows[2]["date"],
            "G54": life_rows[2]["history"],

            "A60": disease_name,
            "M60": disease_symptom,
            "C64": hospital_name,
            "C65": doctor_name,
            "C66": hospital_contact,
            "M64": visit_frequency,
            "R64": medication_status,

            "A70": mind_rows[0]["disease"],
            "L70": mind_rows[0]["symptom"],
            "X70": mind_rows[0]["support"],
            "A73": mind_rows[1]["disease"],
            "L73": mind_rows[1]["symptom"],
            "X73": mind_rows[1]["support"],

            "A79": service_rows[0]["date"],
            "G79": service_rows[0]["name"],
            "Q79": service_rows[0]["amount"],
            "Y79": service_rows[0]["office"],

            "A81": service_rows[1]["date"],
            "G81": service_rows[1]["name"],
            "Q81": service_rows[1]["amount"],
            "Y81": service_rows[1]["office"],

            "A83": service_rows[2]["date"],
            "G83": service_rows[2]["name"],
            "Q83": service_rows[2]["amount"],
            "Y83": service_rows[2]["office"],

            "A88": day_flow,
            "A94": special_note,

            "M100": wish_user,
            "M103": wish_family,
            "M107": future_direction,
    }

    template_name = doc_title
    file_name = f"{doc_title}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )

    resident_id = str(selected_row.get("resident_id", "")).strip()

    saved_df = get_document_records("アセスメント", resident_id)

    saved_options = ["新規作成"]
    saved_map = {"新規作成": None}

    if saved_df is not None and not saved_df.empty:
        for _, row in saved_df.iterrows():
            rid = str(row.get("record_id", "")).strip()
            updated_at = str(row.get("updated_at", "")).strip()
            label = f"{rid} / {updated_at}"
            saved_options.append(label)
            saved_map[label] = rid

    basic_cols = st.columns([7, 2, 2, 2])

    with basic_cols[0]:
        st.text_input(
            "利用者氏名",
            value=resident_name,
            key=f"{doc_title}_resident_name",
            disabled=True
        )

    with basic_cols[1]:
        year_val = st.text_input("作成年（西暦）", key=f"{doc_title}_year", placeholder="2026")
    with basic_cols[2]:
        month_val = st.text_input("月", key=f"{doc_title}_month", placeholder="3")
    with basic_cols[3]:
        day_val = st.text_input("日", key=f"{doc_title}_day", placeholder="14")

    st.divider()

    st.markdown("## 本文入力")

    policy_val = st.text_area(
        "サービス等利用計画の総合的な方針",
        key=f"{doc_title}_policy",
        height=120,
        placeholder="B8 に入る内容です"
    )

    long_goal_val = st.text_area(
        "長期目標（内容・期間等）",
        key=f"{doc_title}_long_goal",
        height=100,
        placeholder="B10 に入る内容です"
    )

    short_goal_val = st.text_area(
        "短期目標（内容・期間等）",
        key=f"{doc_title}_short_goal",
        height=100,
        placeholder="B12 に入る内容です"
    )

    st.divider()
    st.markdown("## 具体的達成目標（3行）")
    st.caption("帳票の 17〜19 行に入る部分です。")

    header_cols = st.columns([5, 3, 4, 2, 2, 2])
    with header_cols[0]:
        st.markdown("**具体的達成目標**")
    with header_cols[1]:
        st.markdown("**本人の役割**")
    with header_cols[2]:
        st.markdown("**支援内容**")
    with header_cols[3]:
        st.markdown("**支援期間**")
    with header_cols[4]:
        st.markdown("**担当者**")
    with header_cols[5]:
        st.markdown("**優先順位**")

    row_data = []

    for i in range(1, 4):
        st.markdown(f"### {i}行目")
        row_cols = st.columns([5, 3, 4, 2, 2, 2])

        with row_cols[0]:
            target_val = st.text_area(
                f"{i}行目_具体的達成目標",
                key=f"{doc_title}_target_{i}",
                height=90,
                label_visibility="collapsed",
                placeholder="C17〜C19"
            )

        with row_cols[1]:
            role_val = st.text_area(
                f"{i}行目_本人の役割",
                key=f"{doc_title}_role_{i}",
                height=90,
                label_visibility="collapsed",
                placeholder="G17〜G19"
            )

        with row_cols[2]:
            support_val = st.text_area(
                f"{i}行目_支援内容",
                key=f"{doc_title}_support_{i}",
                height=90,
                label_visibility="collapsed",
                placeholder="J17〜J19"
            )

        with row_cols[3]:
            period_val = st.text_input(
                f"{i}行目_支援期間",
                key=f"{doc_title}_period_{i}",
                label_visibility="collapsed",
                placeholder="M17〜M19"
            )

        with row_cols[4]:
            person_val = st.text_input(
                f"{i}行目_担当者",
                key=f"{doc_title}_person_{i}",
                label_visibility="collapsed",
                placeholder="O17〜O19"
            )

        with row_cols[5]:
            priority_val = st.selectbox(
                f"{i}行目_優先順位",
                ["", "1", "2", "3"],
                key=f"{doc_title}_priority_{i}",
                label_visibility="collapsed"
            )

        row_data.append({
            "target": target_val,
            "role": role_val,
            "support": support_val,
            "period": period_val,
            "person": person_val,
            "priority": priority_val,
        })

    st.divider()
    st.markdown("## 同意・担当者")

    agree_cols = st.columns([2, 2, 2, 4])

    with agree_cols[0]:
        agree_year_val = st.text_input("同意日_西暦", key=f"{doc_title}_agree_year", placeholder="2026")
    with agree_cols[1]:
        agree_month_val = st.text_input("同意日_月", key=f"{doc_title}_agree_month", placeholder="3")
    with agree_cols[2]:
        agree_day_val = st.text_input("同意日_日", key=f"{doc_title}_agree_day", placeholder="14")
    with agree_cols[3]:
        manager_val = st.text_input("サービス担当責任者", key=f"{doc_title}_manager", placeholder="N21")

    st.divider()

    form_data = {
        "interviewer_name": interviewer_name,
        "hear_year": hear_year,
        "hear_month": hear_month,
        "hear_day": hear_day,
        "furigana": furigana,
        "full_name": full_name,
        "birth_year": birth_year,
        "birth_month": birth_month,
        "birth_day": birth_day,
        "age": age,
        "current_zip_1": current_zip_1,
        "current_zip_2": current_zip_2,
        "current_phone": current_phone,
        "nearest_station": nearest_station,
        "current_address": current_address,
        "emergency_zip_1": emergency_zip_1,
        "emergency_zip_2": emergency_zip_2,
        "emergency_relation": emergency_relation,
        "emergency_phone_fax": emergency_phone_fax,
        "emergency_address": emergency_address,
        "support_city": support_city,
        "support_city_type": support_city_type,
        "support_office": support_office,
        "support_worker": support_worker,
        "handbook_grade": handbook_grade,
        "handbook_year": handbook_year,
        "handbook_month": handbook_month,
        "handbook_day": handbook_day,
        "disability_summary": disability_summary,
        "support_level": support_level,
        "guardian_status": guardian_status,
        "guardian_name": guardian_name,
        "pension_status": pension_status,
        "allowance_status": allowance_status,
        "pension_detail": pension_detail,
        "allowance_detail": allowance_detail,
        "transport_pass": transport_pass,
        "welfare_status": welfare_status,
        "public_support": public_support,
        "family_rows": family_rows,
        "housing_transport": housing_transport,
        "housing_transport_other": housing_transport_other,
        "housing_use_status": housing_use_status,
        "housing_use_status_other": housing_use_status_other,
        "available_transport": available_transport,
        "available_transport_other": available_transport_other,
        "available_use_status": available_use_status,
        "available_use_status_other": available_use_status_other,
        "life_rows": life_rows,
        "disease_name": disease_name,
        "disease_symptom": disease_symptom,
        "hospital_name": hospital_name,
        "doctor_name": doctor_name,
        "hospital_contact": hospital_contact,
        "visit_frequency": visit_frequency,
        "medication_status": medication_status,
        "mind_rows": mind_rows,
        "service_rows": service_rows,
        "day_flow": day_flow,
        "special_note": special_note,
        "wish_user": wish_user,
        "wish_family": wish_family,
        "future_direction": future_direction,
    }

    with st.expander("入力内容確認"):
        st.write(f"利用者氏名: {resident_name}")
        st.write(f"作成年月日: {year_val} / {month_val} / {day_val}")
        st.write(f"総合的な方針: {policy_val}")
        st.write(f"長期目標: {long_goal_val}")
        st.write(f"短期目標: {short_goal_val}")
        st.write(f"同意書日付: {agree_year_val} / {agree_month_val} / {agree_day_val}")
        st.write(f"サービス担当責任者: {manager_val}")

        for idx, item in enumerate(row_data, start=1):
            st.markdown(f"**{idx}行目**")
            st.write(item)

    st.markdown("### 保存済みデータ呼び出し")

    selected_saved_label = st.selectbox(
        "保存済みデータ",
        saved_options,
        key=f"{doc_title}_saved_record_select"
    )

    selected_record_id = saved_map[selected_saved_label]

    if st.button("保存済みを読み込む", key=f"{doc_title}_load_saved"):
        if selected_record_id is not None:
            saved_json = load_document_json(selected_record_id)

            if saved_json:
                for k, v in saved_json.items():
                    st.session_state[f"{doc_title}_{k}"] = v

                st.session_state[f"{doc_title}_loaded_record_id"] = selected_record_id
                st.success("保存済みデータを読み込んだです！")
                st.rerun()
            else:
                st.warning("保存データが見つからありません。")
        else:
            st.info("まだ保存済みデータがないので、新規保存してください。")

    st.markdown("### 保存")

    loaded_record_id = st.session_state.get(f"{doc_title}_loaded_record_id")

    save_cols = st.columns([1, 1, 4])

    with save_cols[0]:
        if st.button("新規保存", key=f"{doc_title}_save_new"):
            new_id = save_document_record(
                resident_id=resident_id,
                resident_name=resident_name,
                doc_type="アセスメント",
                form_data=form_data
            )
            sync_resident_master_from_assessment(
                resident_id=selected_row.get("resident_id", ""),
                welfare_status=welfare_status
            )            
            st.session_state[f"{doc_title}_loaded_record_id"] = new_id
            st.success(f"新規保存しました！ record_id = {new_id}")

    with save_cols[1]:
        if st.button("上書き保存", key=f"{doc_title}_save_update"):
            if loaded_record_id:
                ok = update_document_record(loaded_record_id, form_data)
                if ok:
                    st.success(f"上書き保存しました！ record_id = {loaded_record_id}")
                else:
                    st.warning("上書き対象が見つからありません。")
            else:
                st.warning("先に保存済みデータを読み込むか、新規保存してください。")

def render_basic_sheet_form_page(doc_title: str):
    st.title("📋 基本シート")
    st.caption("基本シート入力ページです。入力と保存とExcel出力までつなぐです。")

    st.markdown("## 基本情報")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_id = str(selected_row.get("resident_id", "")).strip()
    resident_name = str(selected_row.get("resident_name", "")).strip()

    # -----------------------------
    # 保存済み一覧の準備
    # -----------------------------
    saved_df = get_document_records("基本シート", resident_id)

    saved_options = ["新規作成"]
    saved_map = {"新規作成": None}

    if saved_df is not None and not saved_df.empty:
        for _, row in saved_df.iterrows():
            rid = str(row.get("record_id", "")).strip()
            updated_at = str(row.get("updated_at", "")).strip()
            label = f"{rid} / {updated_at}"
            saved_options.append(label)
            saved_map[label] = rid

    # -----------------------------
    # 聞き取り情報
    # P1 / AB1 / AG1 / AJ1
    # -----------------------------
    st.markdown("### 聞き取り情報")

    hear_cols = st.columns([4, 2, 1, 1])
    with hear_cols[0]:
        interviewer_name = st.text_input("聞き取り者名", key=f"{doc_title}_interviewer_name")
    with hear_cols[1]:
        hear_year = st.text_input("聞き取り日（西暦）", key=f"{doc_title}_hear_year", placeholder="2026")
    with hear_cols[2]:
        hear_month = st.text_input("月", key=f"{doc_title}_hear_month", placeholder="3")
    with hear_cols[3]:
        hear_day = st.text_input("日", key=f"{doc_title}_hear_day", placeholder="15")

    st.divider()

    # -----------------------------
    # 1. 生活と家族に関すること
    # A10
    # -----------------------------
    st.markdown("## １．生活と家族に関すること")
    life_family = st.text_area(
        "経済環境、住環境、日常の意思決定、家庭内での立ち位置や役割など",
        key=f"{doc_title}_life_family",
        height=220
    )

    st.divider()

    # -----------------------------
    # 2. 健康維持に関すること
    # A29
    # -----------------------------
    st.markdown("## ２．健康維持に関すること")
    health = st.text_area(
        "服薬管理、食事管理、睡眠、病気への注意、疾病への認識、通院、具合が悪くなった時の対応、医学的管理、必要な体力の維持など",
        key=f"{doc_title}_health",
        height=260
    )

    st.divider()

    # -----------------------------
    # 3. 社会生活に関すること
    # A46
    # -----------------------------
    st.markdown("## ３．社会生活に関すること")
    social = st.text_area(
        "キーパーソンの存在、人付き合い、社会参加など",
        key=f"{doc_title}_social",
        height=220
    )

    st.divider()

    # -----------------------------
    # 4. その他
    # A63
    # -----------------------------
    st.markdown("## ４．その他")
    other = st.text_area(
        "日常生活動作、日常生活への支障や課題など",
        key=f"{doc_title}_other",
        height=220
    )

    st.divider()

    # -----------------------------
    # 5. 総合所見
    # I74 / I80
    # -----------------------------
    st.markdown("## ５．総合所見")

    opinion = st.text_area(
        "聞き取り者所見",
        key=f"{doc_title}_opinion",
        height=140
    )

    direction = st.text_area(
        "支援の方針・方向性など",
        key=f"{doc_title}_direction",
        height=140
    )

    st.divider()

    # -----------------------------
    # 保存用データ作成
    # -----------------------------
    form_data = {
        "interviewer_name": interviewer_name,
        "hear_year": hear_year,
        "hear_month": hear_month,
        "hear_day": hear_day,
        "life_family": life_family,
        "health": health,
        "social": social,
        "other": other,
        "opinion": opinion,
        "direction": direction,
    }

    with st.expander("入力内容確認"):
        st.write({
            "利用者名": resident_name,
            "聞き取り者名": interviewer_name,
            "聞き取り日": f"{hear_year}/{hear_month}/{hear_day}",
            "生活と家族に関すること": life_family,
            "健康維持に関すること": health,
            "社会生活に関すること": social,
            "その他": other,
            "聞き取り者所見": opinion,
            "支援の方針・方向性など": direction,
        })

    st.markdown("### 保存済みデータ呼び出し")

    selected_saved_label = st.selectbox(
        "保存済みデータ",
        saved_options,
        key=f"{doc_title}_saved_record_select"
    )

    selected_record_id = saved_map[selected_saved_label]

    if st.button("保存済みを読み込む", key=f"{doc_title}_load_saved"):
        if selected_record_id is not None:
            saved_json = load_document_json(selected_record_id)

            if saved_json:
                for k, v in saved_json.items():
                    st.session_state[f"{doc_title}_{k}"] = v

                st.session_state[f"{doc_title}_loaded_record_id"] = selected_record_id
                st.success("保存済みデータを読み込んだです！")
                st.rerun()
            else:
                st.warning("保存データが見つからありません。")
        else:
            st.info("まだ保存済みデータがないので、新規保存してください。")

    st.markdown("### 保存")

    loaded_record_id = st.session_state.get(f"{doc_title}_loaded_record_id")

    save_cols = st.columns([1, 1, 4])

    with save_cols[0]:
        if st.button("新規保存", key=f"{doc_title}_save_new"):
            new_id = save_document_record(
                resident_id=resident_id,
                resident_name=resident_name,
                doc_type="基本シート",
                form_data=form_data
            )
            st.session_state[f"{doc_title}_loaded_record_id"] = new_id
            st.success(f"新規保存しました！ record_id = {new_id}")

    with save_cols[1]:
        if st.button("上書き保存", key=f"{doc_title}_save_update"):
            if loaded_record_id:
                ok = update_document_record(loaded_record_id, form_data)
                if ok:
                    st.success(f"上書き保存しました！ record_id = {loaded_record_id}")
                else:
                    st.warning("上書き対象が見つからありません。")
            else:
                st.warning("先に保存済みデータを読み込むか、新規保存してください。")

    st.divider()
    st.markdown("### Excel出力")

    cell_data = {
            "P1": interviewer_name,
            "AB1": hear_year,
            "AG1": hear_month,
            "AJ1": hear_day,
            "A10": life_family,
            "A29": health,
            "A46": social,
            "A63": other,
            "I74": opinion,
            "I80": direction,
    }

    template_name = doc_title
    file_name = f"{doc_title}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )

def render_work_field_form_page(doc_title: str):
    st.title(f"📄 {doc_title}")
    st.info("就労分野シートはまだ作成中です。")

def render_work_sheet_form_page(doc_title: str):
    st.title("📋 就労分野シート")
    st.caption("就労分野シート入力ページです。入力・保存・呼び出し・Excel出力まで対応版です。")

    st.markdown("## 基本情報")
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if master_df is None or master_df.empty:
        st.warning("利用者情報がまだ登録されていません。先に⑨ 利用者情報から利用者を登録してください。")
        return

    master_df = master_df.fillna("").copy()

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"
        if status:
            label += f" / {status}"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    selected_label = st.selectbox(
        "誰の書類を入力するか",
        resident_options,
        key=f"{doc_title}_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_id = str(selected_row.get("resident_id", "")).strip()
    resident_name = str(selected_row.get("resident_name", "")).strip()

    # ---------------------------------
    # 保存済みデータ一覧
    # ---------------------------------
    saved_df = get_document_records("就労分野シート", resident_id)

    saved_options = ["新規作成"]
    saved_map = {"新規作成": None}

    if saved_df is not None and not saved_df.empty:
        for _, row in saved_df.iterrows():
            rid = str(row.get("record_id", "")).strip()
            updated_at = str(row.get("updated_at", "")).strip()
            label = f"{rid} / {updated_at}"
            saved_options.append(label)
            saved_map[label] = rid

    # -----------------------------
    # 聞き取り情報
    # P1 / Y1 / AD1 / AG1
    # -----------------------------
    st.markdown("### 聞き取り情報")

    hear_cols = st.columns([4, 2, 1, 1])
    with hear_cols[0]:
        interviewer_name = st.text_input("聞き取り者名", key=f"{doc_title}_interviewer_name")
    with hear_cols[1]:
        hear_year = st.text_input("聞き取り日（西暦）", key=f"{doc_title}_hear_year", placeholder="2026")
    with hear_cols[2]:
        hear_month = st.text_input("月", key=f"{doc_title}_hear_month", placeholder="3")
    with hear_cols[3]:
        hear_day = st.text_input("日", key=f"{doc_title}_hear_day", placeholder="15")

    st.divider()

    # -----------------------------
    # 4〜8行目 表
    # -----------------------------
    st.markdown("### 評価基準")

    import pandas as pd

    criteria_df = pd.DataFrame({
        "段階": ["1", "2", "3", "4"],
        "評価基準": [
            "できる",
            "少し支援が必要",
            "支援が必要",
            "できない"
        ],
        "目安": [
            "本人の力で達成できる場合",
            "声かけやプログラムにより達成できる場合",
            "個別の支援や配慮など工夫が必要な場合",
            "障がい特性上、達成が困難な場合"
        ]
    })

    st.dataframe(criteria_df, use_container_width=True, hide_index=True)

    # -----------------------------
    # 1. 健康管理
    # -----------------------------
    st.markdown("## １．健康管理")

    health_item1 = st.selectbox("①体調管理　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_health_item1")
    health_item2 = st.selectbox("②服薬管理　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_health_item2")
    health_item3 = st.selectbox("③食事管理　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_health_item3")

    health_point = st.text_area(
        "支援ポイント等記載欄",
        key=f"{doc_title}_health_point",
        height=160
    )

    st.divider()

    # -----------------------------
    # 2. 日常生活管理
    # -----------------------------
    st.markdown("## ２．日常生活管理")

    daily_item1 = st.selectbox("①生活リズム　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_daily_item1")
    daily_item2 = st.selectbox("②みだしなみ　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_daily_item2")
    daily_item3 = st.selectbox("③清潔保持　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_daily_item3")
    daily_item4 = st.selectbox("④金銭管理　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_daily_item4")
    daily_item5 = st.selectbox("⑤移動　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_daily_item5")

    daily_point = st.text_area(
        "支援ポイント等記載欄",
        key=f"{doc_title}_daily_point",
        height=180
    )

    st.divider()

    # -----------------------------
    # 3. 就労に関すること
    # -----------------------------
    st.markdown("## ３．就労に関すること")

    work_item1 = st.selectbox("①就労理解　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_work_item1")
    work_item2 = st.selectbox("②就労意識　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_work_item2")
    work_item3 = st.selectbox("③就労意欲　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_work_item3")
    work_item4 = st.selectbox("④体力、精神力　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_work_item4")
    work_item5 = st.selectbox("⑤家族の支援　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_work_item5")

    work_point = st.text_area(
        "支援ポイント等記載欄",
        key=f"{doc_title}_work_point",
        height=180
    )

    st.divider()

    # -----------------------------
    # 5. 基本的労働習慣
    # -----------------------------
    st.markdown("## ５．基本的労働習慣")

    labor_item1 = st.selectbox("①あいさつ　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_labor_item1")
    labor_item2 = st.selectbox("②報告・連絡・相談　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_labor_item2")
    labor_item3 = st.selectbox("③毎日の通所　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_labor_item3")
    labor_item4 = st.selectbox("④規則の順守　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_labor_item4")

    labor_point = st.text_area(
        "支援ポイント等記載欄",
        key=f"{doc_title}_labor_point",
        height=160
    )

    st.divider()

    # -----------------------------
    # 6. 職業適性
    # -----------------------------
    st.markdown("## ６．職業適性")

    aptitude_item1 = st.selectbox("①指示の理解　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_aptitude_item1")
    aptitude_item2 = st.selectbox("②持続力、集中力　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_aptitude_item2")
    aptitude_item3 = st.selectbox("③正確性　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_aptitude_item3")
    aptitude_item4 = st.selectbox("④責任感　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_aptitude_item4")
    aptitude_item5 = st.selectbox("⑤自己理解①　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_aptitude_item5")
    aptitude_item6 = st.selectbox("⑥自己理解②　支援の必要性", ["1", "2", "3", "4"], key=f"{doc_title}_aptitude_item6")

    aptitude_point = st.text_area(
        "支援ポイント等記載欄",
        key=f"{doc_title}_aptitude_point",
        height=200
    )

    st.divider()

    # -----------------------------
    # 7. 利用経過及び就労歴
    # -----------------------------
    st.markdown("## ７．現状に至る障がい福祉サービスの利用経過及び就労歴")

    history_rows = []

    for i in range(3):
        st.markdown(f"### 就労歴 {i + 1}")

        col1 = st.columns([1, 2])
        with col1[0]:
            period = st.text_input("期間（時期）", key=f"{doc_title}_history_period_{i}")
        with col1[1]:
            company = st.text_input("企業・事業所名", key=f"{doc_title}_history_company_{i}")

        col2 = st.columns([2, 2])
        with col2[0]:
            business = st.text_area(
                "事業内容（具体的に）",
                key=f"{doc_title}_history_business_{i}",
                height=100
            )
        with col2[1]:
            condition = st.text_area(
                "労働条件（給与、労働時間、休日など）",
                key=f"{doc_title}_history_condition_{i}",
                height=100
            )

        note = st.text_area(
            "特記事項（職場でのエピソード、離職理由等）",
            key=f"{doc_title}_history_note_{i}",
            height=120
        )

        history_rows.append({
            "period": period,
            "company": company,
            "business": business,
            "condition": condition,
            "note": note,
        })

    hope_service = st.text_area(
        "当事業所のサービス利用に対する本人の希望",
        key=f"{doc_title}_hope_service",
        height=120
    )

    st.divider()

    # -----------------------------
    # 8. 総合所見
    # -----------------------------
    st.markdown("## ８．総合所見")

    hope_work = st.text_area(
        "本人が希望している就労内容",
        key=f"{doc_title}_hope_work",
        height=120
    )

    readiness = st.text_area(
        "本人が自覚している就労準備性の状況",
        key=f"{doc_title}_readiness",
        height=120
    )

    suitable_work = st.text_area(
        "本人に向いていると判断される就労内容",
        key=f"{doc_title}_suitable_work",
        height=120
    )

    opinion = st.text_area(
        "聞き取り者所見",
        key=f"{doc_title}_opinion",
        height=120
    )

    direction = st.text_area(
        "支援の方向性や方針",
        key=f"{doc_title}_direction",
        height=120
    )

    st.divider()

    # ---------------------------------
    # 保存用データ
    # ---------------------------------
    form_data = {
        "interviewer_name": interviewer_name,
        "hear_year": hear_year,
        "hear_month": hear_month,
        "hear_day": hear_day,
        "health_item1": health_item1,
        "health_item2": health_item2,
        "health_item3": health_item3,
        "health_point": health_point,
        "daily_item1": daily_item1,
        "daily_item2": daily_item2,
        "daily_item3": daily_item3,
        "daily_item4": daily_item4,
        "daily_item5": daily_item5,
        "daily_point": daily_point,
        "work_item1": work_item1,
        "work_item2": work_item2,
        "work_item3": work_item3,
        "work_item4": work_item4,
        "work_item5": work_item5,
        "work_point": work_point,
        "labor_item1": labor_item1,
        "labor_item2": labor_item2,
        "labor_item3": labor_item3,
        "labor_item4": labor_item4,
        "labor_point": labor_point,
        "aptitude_item1": aptitude_item1,
        "aptitude_item2": aptitude_item2,
        "aptitude_item3": aptitude_item3,
        "aptitude_item4": aptitude_item4,
        "aptitude_item5": aptitude_item5,
        "aptitude_item6": aptitude_item6,
        "aptitude_point": aptitude_point,
        "history_rows": history_rows,
        "hope_service": hope_service,
        "hope_work": hope_work,
        "readiness": readiness,
        "suitable_work": suitable_work,
        "opinion": opinion,
        "direction": direction,
    }

    with st.expander("入力内容確認"):
        st.write({
            "利用者名": resident_name,
            "聞き取り者名": interviewer_name,
            "聞き取り日": f"{hear_year}/{hear_month}/{hear_day}",
            "健康管理": [health_item1, health_item2, health_item3, health_point],
            "日常生活管理": [daily_item1, daily_item2, daily_item3, daily_item4, daily_item5, daily_point],
            "就労に関すること": [work_item1, work_item2, work_item3, work_item4, work_item5, work_point],
            "基本的労働習慣": [labor_item1, labor_item2, labor_item3, labor_item4, labor_point],
            "職業適性": [aptitude_item1, aptitude_item2, aptitude_item3, aptitude_item4, aptitude_item5, aptitude_item6, aptitude_point],
            "就労歴": history_rows,
            "サービス利用希望": hope_service,
            "総合所見": [hope_work, readiness, suitable_work, opinion, direction],
        })

    st.markdown("### 保存済みデータ呼び出し")

    selected_saved_label = st.selectbox(
        "保存済みデータ",
        saved_options,
        key=f"{doc_title}_saved_record_select"
    )

    selected_record_id = saved_map[selected_saved_label]

    if st.button("保存済みを読み込む", key=f"{doc_title}_load_saved"):
        if selected_record_id is not None:
            saved_json = load_document_json(selected_record_id)

            if saved_json:
                for k, v in saved_json.items():
                    st.session_state[f"{doc_title}_{k}"] = v

                st.session_state[f"{doc_title}_loaded_record_id"] = selected_record_id
                st.success("保存済みデータを読み込みました！")
                st.rerun()
            else:
                st.warning("保存データが見つからありません。")
        else:
            st.info("まだ保存済みデータがないので、新規保存してください。")

    st.markdown("### 保存")

    loaded_record_id = st.session_state.get(f"{doc_title}_loaded_record_id")

    save_cols = st.columns([1, 1, 4])

    with save_cols[0]:
        if st.button("新規保存", key=f"{doc_title}_save_new"):
            new_id = save_document_record(
                resident_id=resident_id,
                resident_name=resident_name,
                doc_type="就労分野シート",
                form_data=form_data
            )
            st.session_state[f"{doc_title}_loaded_record_id"] = new_id
            st.success(f"新規保存しました！ record_id = {new_id}")

    with save_cols[1]:
        if st.button("上書き保存", key=f"{doc_title}_save_update"):
            if loaded_record_id:
                ok = update_document_record(loaded_record_id, form_data)
                if ok:
                    st.success(f"上書き保存しました！ record_id = {loaded_record_id}")
                else:
                    st.warning("上書き対象が見つからありません。")
            else:
                st.warning("先に保存済みデータを読み込むか、新規保存してください。")

    st.divider()
    st.markdown("### Excel出力")

    cell_data = {
            "P1": interviewer_name,
            "Y1": hear_year,
            "AD1": hear_month,
            "AG1": hear_day,

            # 1. 健康管理
            "M14": health_item1,
            "M16": health_item2,
            "M18": health_item3,
            "P14": health_point,

            # 2. 日常生活管理
            "M22": daily_item1,
            "M24": daily_item2,
            "M26": daily_item3,
            "M28": daily_item4,
            "M30": daily_item5,
            "P22": daily_point,

            # 3. 就労に関すること
            "M44": work_item1,
            "M46": work_item2,
            "M48": work_item3,
            "M50": work_item4,
            "M52": work_item5,
            "P44": work_point,

            # 5. 基本的労働習慣
            "M58": labor_item1,
            "M60": labor_item2,
            "M62": labor_item3,
            "M64": labor_item4,
            "P58": labor_point,

            # 6. 職業適性
            "M68": aptitude_item1,
            "M70": aptitude_item2,
            "M72": aptitude_item3,
            "M74": aptitude_item4,
            "M76": aptitude_item5,
            "M78": aptitude_item6,
            "P68": aptitude_point,

            # 7. 利用経過及び就労歴
            "A84": history_rows[0]["period"],
            "A85": history_rows[0]["company"],
            "J84": history_rows[0]["business"],
            "W84": history_rows[0]["condition"],
            "A87": history_rows[0]["note"],

            "A91": history_rows[1]["period"],
            "A92": history_rows[1]["company"],
            "J91": history_rows[1]["business"],
            "W91": history_rows[1]["condition"],
            "A94": history_rows[1]["note"],

            "A98": history_rows[2]["period"],
            "A99": history_rows[2]["company"],
            "J98": history_rows[2]["business"],
            "W98": history_rows[2]["condition"],
            "A101": history_rows[2]["note"],

            "J99": hope_service,

            # 8. 総合所見
            "J105": hope_work,
            "J109": readiness,
            "J112": suitable_work,
            "J116": opinion,
            "J121": direction,
    }

    template_name = doc_title
    file_name = f"{doc_title}.xlsx"

    render_excel_download_block(
        doc_title=doc_title,
        file_name=file_name,
        template_name=template_name,
        cell_data=cell_data
    )

@st.cache_data(ttl=60)
def get_staff_examples_df_cached():
    df = load_db("staff_examples")
    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "staff_name",
            "home_start_example", "home_end_example",
            "day_start_example", "day_end_example",
            "outside_start_example", "outside_end_example",
            "updated_at"
        ])
    else:
        for col in [
            "staff_name",
            "home_start_example", "home_end_example",
            "day_start_example", "day_end_example",
            "outside_start_example", "outside_end_example",
            "updated_at"
        ]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def get_staff_examples_df():
    return get_staff_examples_df_cached().copy()


@st.cache_data(ttl=60)
def get_personal_rules_df_cached():
    df = load_db("personal_rules")
    if df is None or df.empty:
        df = pd.DataFrame(columns=["staff_name", "rule_text", "updated_at"])
    else:
        for col in ["staff_name", "rule_text", "updated_at"]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def get_personal_rules_df():
    return get_personal_rules_df_cached().copy()


@st.cache_data(ttl=60)
def get_assistant_plans_df_cached():
    df = load_db("assistant_plans")
    if df is None or df.empty:
        df = pd.DataFrame(columns=["resident_id", "long_term_goal", "short_term_goal", "updated_at"])
    else:
        for col in ["resident_id", "long_term_goal", "short_term_goal", "updated_at"]:
            if col not in df.columns:
                df[col] = ""
    return df.fillna("")


def get_assistant_plans_df():
    df = load_db("resident_schedule")

    if df is None or df.empty:
        return pd.DataFrame()

    df = df.fillna("").copy()

    if "company_id" not in df.columns:
        df["company_id"] = ""

    company_id = str(st.session_state.get("company_id", "")).strip()

    df["company_id"] = df["company_id"].astype(str).str.strip()
    return df[df["company_id"] == company_id]


def save_staff_examples_record(
    company_id,
    staff_name,
    home_start_example,
    home_end_example,
    day_start_example,
    day_end_example,
    outside_start_example,
    outside_end_example,
):
    df = load_db("staff_examples")
    required_cols = [
        "company_id",
        "staff_name",
        "home_start_example", "home_end_example",
        "day_start_example", "day_end_example",
        "outside_start_example", "outside_end_example",
        "updated_at"
    ]
    df = normalize_company_scoped_df(df, required_cols)

    company_id = str(company_id).strip()
    staff_name = str(staff_name).strip()
    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    mask = (
        (df["company_id"] == company_id) &
        (df["staff_name"].astype(str).str.strip() == staff_name)
    )

    new_data = {
        "company_id": company_id,
        "staff_name": staff_name,
        "home_start_example": str(home_start_example),
        "home_end_example": str(home_end_example),
        "day_start_example": str(day_start_example),
        "day_end_example": str(day_end_example),
        "outside_start_example": str(outside_start_example),
        "outside_end_example": str(outside_end_example),
        "updated_at": now_str,
    }

    if mask.any():
        for k, v in new_data.items():
            df.loc[mask, k] = v
    else:
        df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)

    save_db(df, "staff_examples")


def save_personal_rules_record(company_id, staff_name, rule_text):
    df = load_db("personal_rules")

    if df is None or df.empty:
        df = pd.DataFrame(columns=[
            "company_id",
            "staff_name",
            "rule_text",
            "updated_at",
        ])
    else:
        df = df.fillna("").copy()
        for col in [
            "company_id",
            "staff_name",
            "rule_text",
            "updated_at",
        ]:
            if col not in df.columns:
                df[col] = ""

    company_id = str(company_id).strip()
    staff_name = str(staff_name).strip()
    updated_at = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    df["company_id"] = df["company_id"].astype(str).str.strip()
    df["staff_name"] = df["staff_name"].astype(str).str.strip()

    hit_idx = df.index[
        (df["company_id"] == company_id) &
        (df["staff_name"] == staff_name)
    ].tolist()

    row_data = {
        "company_id": company_id,
        "staff_name": staff_name,
        "rule_text": str(rule_text),
        "updated_at": updated_at,
    }

    if hit_idx:
        idx = hit_idx[0]
        for k, v in row_data.items():
            df.at[idx, k] = v
    else:
        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)

    save_db(df, "personal_rules")
    return True

def build_examples_text(service_type, example_row):
    if not example_row:
        return ""

    if service_type == "在宅":
        return (
            f"在宅作業開始例文:\n{example_row.get('home_start_example', '')}\n\n"
            f"在宅作業終了例文:\n{example_row.get('home_end_example', '')}"
        )
    elif service_type == "通所":
        return (
            f"通所作業開始例文:\n{example_row.get('day_start_example', '')}\n\n"
            f"通所作業終了例文:\n{example_row.get('day_end_example', '')}"
        )
    else:
        return (
            f"施設外作業開始例文:\n{example_row.get('outside_start_example', '')}\n\n"
            f"施設外作業終了例文:\n{example_row.get('outside_end_example', '')}"
        )


def generate_status_support_with_gemini(
    service_type,
    meal_flag,
    note,
    start_memo,
    end_memo,
    examples_text,
    rule_text,
    plan_text=""
):
    api_key = get_gemini_api_key_from_app()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY が取得できませんでした。")

    prompt = f"""
あなたは障害福祉サービスのKnowbe日誌入力を補助するアシスタントです。

以下の条件で、必ずJSON形式のみで返してください。
キーは "generated_status" と "generated_support" の2つです。

【前提】
- service_type: {service_type}
- meal_flag: {meal_flag}
- 備考: {note}

【利用者状態メモ（そのままの意味を尊重）】
{start_memo}

【職員考察メモ（そのままの意味を尊重）】
{end_memo}

【スタッフ例文】
{examples_text}

【個人ルール】
{rule_text}

【支援計画】
{plan_text}

【絶対ルール】
- generated_status は「利用者状態」に入れる文章
- generated_support は「職員考察」に入れる文章
- start_memo の内容は generated_status に反映
- end_memo の内容は generated_support に反映
- 事実を勝手に増やさない
- 余計な見出しはつけない
- JSON以外は返さない
"""

    client = get_genai_client(api_key)
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
    )
    text = (response.text or "").strip()

    if not text:
        raise RuntimeError("Geminiの応答が空です")

    text = text.replace("```json", "").replace("```", "").strip()
    data = json.loads(text)

    return (
        str(data.get("generated_status", "")).strip(),
        str(data.get("generated_support", "")).strip(),
    )

def get_gemini_api_key_from_app():
    api_key = ""

    try:
        api_key = st.secrets.get("GEMINI_API_KEY", "")
    except Exception:
        api_key = ""

    if not api_key:
        try:
            if "gemini" in st.secrets and "api_key" in st.secrets["gemini"]:
                api_key = st.secrets["gemini"]["api_key"]
        except Exception:
            pass

    if not api_key:
        import os
        api_key = os.environ.get("GEMINI_API_KEY", "")

    return str(api_key).strip()


def generate_bee_texts(
    resident_name,
    service_type,
    meal_flag="",
    note="",
    start_memo="",
    end_memo="",
    examples_text="",
    rule_text="",
    plan_text="",
    note_text=None,
    staff_name=None,
):
    if (not str(note).strip()) and note_text is not None:
        note = note_text

    meal_flag = "" if meal_flag is None else str(meal_flag).strip()
    note = "" if note is None else str(note).strip()
    start_memo = "" if start_memo is None else str(start_memo).strip()
    end_memo = "" if end_memo is None else str(end_memo).strip()
    examples_text = "" if examples_text is None else str(examples_text).strip()
    rule_text = "" if rule_text is None else str(rule_text).strip()
    plan_text = "" if plan_text is None else str(plan_text).strip()

    api_key = get_gemini_api_key_from_app()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY が取得できなかったです")

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    prompt = f"""
あなたは就労継続支援B型の支援記録作成アシスタントです。
以下の情報をもとに、Knowbeへそのまま貼り付けられる
「利用者状態」と「職員考察」を作ってください。

【利用者名】
{resident_name}

【サービス種別】
{service_type}

【食事】
{meal_flag}

【備考】
{note}

【利用者状態メモ】
{start_memo}

【職員考察メモ】
{end_memo}

【スタッフ例文】
{examples_text}

【個人ルール】
{rule_text}

【支援計画】
{plan_text}

【ルール】
- 出力はJSONのみ
- generated_status は「利用者状態」
- generated_support は「職員考察」
- 事実を勝手に増やさない
- 見出しや箇条書きは不要
- 支援記録として自然で丁寧な文
- 3文程度
- 100〜150文字程度目安
- Knowbeへそのまま貼れる長さ

【出力形式】
{{
  "generated_status": "ここに利用者状態",
  "generated_support": "ここに職員考察"
}}
"""
    


    response = model.generate_content(prompt)
    result_text = (response.text or "").strip()

    if not result_text:
        raise RuntimeError("Geminiの応答が空です")

    cleaned = result_text.replace("```json", "").replace("```", "").strip()

    try:
        data = json.loads(cleaned)
        generated_status = str(data.get("generated_status", "")).strip()
        generated_support = str(data.get("generated_support", "")).strip()
    except Exception:
        raise RuntimeError(f"Gemini出力の解析に失敗です: {cleaned}")

    if not generated_status or not generated_support:
        raise RuntimeError(f"Gemini出力の解析に失敗です: {cleaned}")

    return generated_status, generated_support

def get_knowbe_credentials_from_app(company_id=None):
    if company_id is None:
        company_id = get_current_company_id()

    username = ""
    password = ""

    try:
        company_row = get_company_row_by_id(company_id)
    except Exception:
        company_row = {}

    if company_row:
        username = str(company_row.get("knowbe_login_username", "")).strip()
        password = str(company_row.get("knowbe_login_password", "")).strip()

    return username, password

def send_to_knowbe_from_bee(
    record_id=None,
    company_id="",
    target_date="",
    resident_name="",
    service_type="",
    start_time="",
    end_time="",
    meal_flag="",
    note_text="",
    generated_status="",
    generated_support="",
    staff_name="",
    knowbe_target="",
    work_start_time="",
    work_end_time="",
    work_break_time="",
    work_memo="",
    login_username="",
    login_password="",
    send_user_status=True,
    send_staff_comment=True,
):
    import traceback

    login_username = str(login_username).strip()
    login_password = str(login_password).strip()

    if not login_username or not login_password:
        raise RuntimeError("Knowbeアカウント名またはKnowbeパスワードが未設定です。")

    try:
        from run_assistance import send_one_record_from_app  # type: ignore

        ok = send_one_record_from_app(
            target_date=str(target_date).strip(),
            resident_name=str(resident_name).strip(),
            service_type=str(service_type).strip(),
            start_time=str(start_time).strip(),
            end_time=str(end_time).strip(),
            meal_flag=str(meal_flag).strip(),
            note_text=str(note_text).strip(),
            generated_status=str(generated_status).strip(),
            generated_support=str(generated_support).strip(),
            staff_name=str(staff_name).strip(),
            knowbe_target=str(knowbe_target).strip(),
            login_username=login_username,
            login_password=login_password,
            work_start_time=str(work_start_time).strip(),
            work_end_time=str(work_end_time).strip(),
            work_break_time=str(work_break_time).strip(),
            work_memo=str(work_memo).strip(),
            send_user_status=bool(send_user_status),
            send_staff_comment=bool(send_staff_comment),
        )

    except Exception:
        st.code(traceback.format_exc())
        raise

    if not ok:
        raise RuntimeError("run_assistance.send_one_record_from_app が False を返しました")

    return True

import streamlit as st
import pandas as pd
from datetime import date

def render_bulk_knowbe_diary_page():
    st.title("🐝 knowbe日誌入力（一括）")

    current_company_id = st.session_state.get("company_id", "")
    login_staff_name = st.session_state.get("display_name", "")

    if not current_company_id:
        st.error("事業所情報が取得できません。")
        return

    # ===== Knowbeログイン情報解決（単発ページに寄せる）=====
    ctx = resolve_bee_company_context(
        company_login_id="",
        company_login_password="",
        knowbe_login_username="",
        knowbe_login_password="",
    )

    if not ctx.get("ok", False):
        st.error(ctx.get("error", "事業所情報の確認に失敗しました。"))
        return

    resolved_knowbe_user = ctx["knowbe_login_username"]
    resolved_knowbe_pw = ctx["knowbe_login_password"]

    if not str(resolved_knowbe_user).strip() or not str(resolved_knowbe_pw).strip():
        st.error("Knowbe情報が未登録です。送信するには登録が必要です。")
        return

    # ===== 利用者一覧取得 =====
    residents_df = get_resident_master_df(current_company_id)
    if residents_df.empty:
        st.warning("利用者データがありません。")
        return

    if "status" in residents_df.columns:
        residents_df = residents_df[residents_df["status"].fillna("利用中") == "利用中"]

    residents_df = residents_df.reset_index(drop=True)

    if residents_df.empty:
        st.warning("この事業所の利用中利用者がいません。")
        return

    # ===== 共通日付 =====
    bulk_target_date = st.date_input(
        "対象日（全員共通）",
        value=date.today(),
        key="bulk_target_date"
    )

    st.markdown("---")

    send_targets = []

    for idx, r in residents_df.iterrows():
        resident_id = str(r.get("resident_id", "")).strip()
        resident_name = str(r.get("resident_name", "")).strip()

        if not resident_id or not resident_name:
            continue

        block_key = f"bulk_{resident_id}"

        with st.container():
            st.subheader(f"{resident_name}")

            enabled = st.checkbox(
                "この利用者を送信対象にする",
                value=True,
                key=f"{block_key}_enabled"
            )

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                start_time = st.text_input("開始時間", value="10:00", key=f"{block_key}_start_time")
            with c2:
                end_time = st.text_input("終了時間", value="10:50", key=f"{block_key}_end_time")
            with c3:
                meal_flag = st.selectbox("食事提供", ["なし", "あり"], key=f"{block_key}_meal_flag")
            with c4:
                service_type = st.selectbox(
                    "サービス種別",
                    ["在宅", "通所", "施設外就労"],
                    key=f"{block_key}_service_type"
                )

            c5, c6, c7 = st.columns(3)
            with c5:
                work_start_time = st.text_input("作業開始時間", value="", key=f"{block_key}_work_start_time")
            with c6:
                work_end_time = st.text_input("作業終了時間", value="", key=f"{block_key}_work_end_time")
            with c7:
                break_time = st.text_input("休憩時間", value="", key=f"{block_key}_break_time")

            diary_input_staff = st.text_input(
                "日誌入力者",
                value=login_staff_name,
                key=f"{block_key}_staff_name"
            )

            remark_mode = st.radio(
                "備考の入力方法",
                ["候補から選ぶ", "直接入力"],
                horizontal=True,
                key=f"{block_key}_remark_mode"
            )

            if remark_mode == "候補から選ぶ":
                remark_candidates = ["在宅利用", "食事摂取量 10/10", "施設外就労(実施報告書等添付)", "入院", ""]
                remark_text = st.selectbox(
                    "備考",
                    remark_candidates,
                    key=f"{block_key}_remark_text_select"
                )
            else:
                remark_text = st.text_input("備考", value="", key=f"{block_key}_remark_text")

            c8, c9 = st.columns(2)
            with c8:
                start_memo = st.text_area("開始メモ", height=140, key=f"{block_key}_start_memo")
            with c9:
                end_memo = st.text_area("終了メモ", height=140, key=f"{block_key}_end_memo")

            send_mode = st.radio(
                "送信方式",
                ["Geminiで編集して送信", "入力文のまま送信"],
                horizontal=True,
                key=f"{block_key}_send_mode"
            )

            st.markdown("---")

            send_targets.append({
                "enabled": enabled,
                "resident_id": resident_id,
                "resident_name": resident_name,
                "target_date": bulk_target_date.strftime("%Y-%m-%d"),
                "start_time": start_time,
                "end_time": end_time,
                "meal_flag": meal_flag,
                "service_type": service_type,
                "work_start_time": work_start_time,
                "work_end_time": work_end_time,
                "break_time": break_time,
                "staff_name": diary_input_staff,
                "remark_text": remark_text,
                "start_memo": start_memo,
                "end_memo": end_memo,
                "send_mode": "gemini" if send_mode == "Geminiで編集して送信" else "raw",
            })

    if st.button("knowbeへ送信する", type="primary", use_container_width=True):
        active_targets = [x for x in send_targets if x["enabled"]]

        valid_targets = []
        invalid_messages = []

        for item in active_targets:
            if not str(item.get("staff_name", "")).strip():
                invalid_messages.append(f"{item['resident_name']}：日誌入力者が未入力")
                continue

            valid_targets.append(item)

        if invalid_messages:
            st.warning("一部送信されない利用者があります：")
            for msg in invalid_messages:
                st.warning(msg)

        if not valid_targets:
            st.warning("送信できる利用者が1人もいません。")
            return

        success_count = 0
        error_count = 0

        progress = st.progress(0)
        status_box = st.empty()

        for i, item in enumerate(valid_targets, start=1):
            try:
                status_box.info(f"{i}/{len(valid_targets)} 送信中: {item['resident_name']}")

                start_memo_to_send = item["start_memo"]
                end_memo_to_send = item["end_memo"]

                if item["send_mode"] == "gemini":
                    start_memo_to_send = edit_text_with_gemini_for_start_memo(
                        resident_name=item["resident_name"],
                        original_text=item["start_memo"],
                        remark_text=item["remark_text"],
                    )
                    end_memo_to_send = edit_text_with_gemini_for_end_memo(
                        resident_name=item["resident_name"],
                        original_text=item["end_memo"],
                        remark_text=item["remark_text"],
                    )

                ok = send_to_knowbe_from_bee(
                    company_id=current_company_id,
                    target_date=item["target_date"],
                    resident_name=item["resident_name"],
                    service_type=item["service_type"],
                    start_time=item["start_time"],
                    end_time=item["end_time"],
                    meal_flag=item["meal_flag"],
                    note_text=item["remark_text"],
                    generated_status=start_memo_to_send,
                    generated_support=end_memo_to_send,
                    staff_name=item["staff_name"],
                    knowbe_target="bulk_gemini" if item["send_mode"] == "gemini" else "bulk_raw",
                    work_start_time=item["work_start_time"],
                    work_end_time=item["work_end_time"],
                    work_break_time=item.get("break_time", ""),
                    work_memo="",
                    login_username=resolved_knowbe_user,
                    login_password=resolved_knowbe_pw,
                    send_user_status=True,
                    send_staff_comment=True,
                )

                if ok:
                    success_count += 1
                else:
                    error_count += 1
                    st.error(f"{item['resident_name']} の送信で失敗しました（Knowbe未反映）")

            except Exception as e:
                error_count += 1
                st.error(f"{item['resident_name']} の送信でエラー: {e}")

            progress.progress(i / len(valid_targets))

        status_box.success(f"送信完了：成功 {success_count}件 / エラー {error_count}件")

def render_bee_journal_page():
    st.title("🐝knowbe日誌入力🐝")
    st.caption("Sue for Bee Assistance 専用の裏メニューです。")

    current_company_name = str(st.session_state.get("company_name", "")).strip()
    current_staff_name = str(st.session_state.get("user", "")).strip()

    st.markdown("## 事業所選択")

    company_cols = st.columns(2)
    with company_cols[0]:
        bee_company_login_id = st.text_input(
            "事業所ID",
            key="bee_company_login_id",
            placeholder="空欄なら現在ログイン中の事業所を使います"
        )
    with company_cols[1]:
        bee_company_login_password = st.text_input(
            "事業所パスワード",
            type="password",
            key="bee_company_login_password",
            placeholder="空欄なら現在ログイン中の事業所を使います"
        )

    st.markdown("## knowbeアカウント情報入力")

    knowbe_cols = st.columns(2)
    with knowbe_cols[0]:
        bee_knowbe_login_username = st.text_input(
            "knowbeアカウント名",
            key="bee_knowbe_login_username"
        )
    with knowbe_cols[1]:
        bee_knowbe_login_password = st.text_input(
            "knowbeパスワード",
            type="password",
            key="bee_knowbe_login_password"
        )

    ctx = resolve_bee_company_context(
        company_login_id=bee_company_login_id,
        company_login_password=bee_company_login_password,
        knowbe_login_username=bee_knowbe_login_username,
        knowbe_login_password=bee_knowbe_login_password,
    )

    if not ctx.get("ok", False):
        st.error(ctx.get("error", "事業所情報の確認に失敗しました。"))
        return

    target_company_id = ctx["target_company_id"]
    target_company_name = ctx["target_company_name"]
    resolved_knowbe_user = ctx["knowbe_login_username"]
    resolved_knowbe_pw = ctx["knowbe_login_password"]

    info_cols = st.columns(2)
    with info_cols[0]:
        st.info(f"対象事業所: {target_company_name or current_company_name}")

    with info_cols[1]:
        if ctx.get("using_saved_knowbe", False):
            st.success(f"保存済みKnowbe情報を使用する：{mask_secret_text(resolved_knowbe_user)}")

        elif ctx.get("has_knowbe_credentials", False):
            st.success(f"入力されたKnowbe情報を使用する：{mask_secret_text(resolved_knowbe_user)}")

        else:
            st.error("Knowbe情報が未登録です。送信するには登録が必要です。")

            if bool(st.session_state.get("is_admin", False)):
                if st.button("Knowbe情報登録ページへ", key="go_knowbe_settings_from_bee", use_container_width=True):
                    st.session_state.current_page = "⑨管理者"
                    st.rerun()
            else:
                st.warning("管理者以外は登録できません。管理者へ報告してください。")
    st.markdown("## 利用者選択")

    master_df = get_resident_master_df(target_company_id)

    if master_df.empty:
        st.warning("この事業所に所属する利用者がまだ登録されていません。")
        return

    resident_options = []
    resident_map = {}

    for _, row in master_df.iterrows():
        rid = str(row.get("resident_id", "")).strip()
        rname = str(row.get("resident_name", "")).strip()

        if not rname:
            continue

        label = f"{rname}"
        if rid:
            label += f" ({rid})"

        resident_options.append(label)
        resident_map[label] = row.to_dict()

    if not resident_options:
        st.warning("この事業所に所属する利用者がまだ登録されていません。")
        return

    selected_label = st.selectbox(
        "利用者を選ぶ",
        resident_options,
        key="bee_resident_select"
    )

    selected_row = resident_map[selected_label]
    resident_id = str(selected_row.get("resident_id", "")).strip()
    resident_name = str(selected_row.get("resident_name", "")).strip()

    st.markdown("### 保存データ呼び出し")

    diary_df = get_diary_input_rules_df()

    if diary_df is not None and not diary_df.empty:
        df_user = diary_df[
            diary_df["resident_name"].astype(str) == str(resident_name)
        ].copy()

        if not df_user.empty:
            try:
                df_user["record_id_num"] = pd.to_numeric(df_user["record_id"], errors="coerce")
                df_user = df_user.sort_values("record_id_num", ascending=False)
            except Exception:
                pass

            load_options = []
            load_map = {}

            for _, r in df_user.head(10).iterrows():
                label = f"{r.get('date', '')} / {r.get('start_time', '')}〜{r.get('end_time', '')} / ID:{r.get('record_id', '')}"
                load_options.append(label)
                load_map[label] = r.to_dict()

            load_col1, load_col2 = st.columns([5, 1])

            with load_col1:
                selected_saved_label = st.selectbox(
                    "過去の保存データ",
                    [""] + load_options,
                    key="bee_saved_record_select"
                )

            with load_col2:
                st.write("")
                st.write("")
                if st.button("呼び出す", key="bee_load_saved_record", use_container_width=True):
                    if selected_saved_label:
                        rec = load_map[selected_saved_label]

                        st.session_state["bee_target_date"] = (
                            pd.to_datetime(rec.get("date", "")).date()
                            if str(rec.get("date", "")).strip()
                            else now_jst().date()
                        )
                        st.session_state["start_time"] = str(rec.get("start_time", ""))
                        st.session_state["end_time"] = str(rec.get("end_time", ""))
                        st.session_state["bee_meal_flag"] = str(rec.get("meal_flag", "なし"))
                        st.session_state["bee_note_text"] = str(rec.get("note", ""))
                        st.session_state["bee_start_memo"] = str(rec.get("start_memo", ""))
                        st.session_state["bee_end_memo"] = str(rec.get("end_memo", ""))
                        st.session_state["bee_staff_name"] = str(rec.get("staff_name", current_staff_name))
                        st.session_state["bee_service_type"] = str(rec.get("service_type", "在宅"))
                        st.success("保存データを呼び出しました！")
                        st.rerun()

    st.divider()
    st.markdown("## 日誌入力")

    target_date = st.date_input(
        "対象日",
        value=st.session_state.get("bee_target_date", now_jst().date()),
        key="bee_target_date"
    )

    input_cols = st.columns([1, 1, 1, 1])

    with input_cols[0]:
        start_time = st.text_input(
            "開始時間",
            value=st.session_state.get("start_time", ""),
            key="start_time",
            placeholder="10:00"
        )

    with input_cols[1]:
        end_time = st.text_input(
            "終了時間",
            value=st.session_state.get("end_time", ""),
            key="end_time",
            placeholder="10:50"
        )

    with input_cols[2]:
        meal_flag = st.selectbox(
            "食事提供",
            ["なし", "あり"],
            index=0 if st.session_state.get("bee_meal_flag", "なし") == "なし" else 1,
            key="bee_meal_flag"
        )

    with input_cols[3]:
        service_type = st.selectbox(
            "サービス種別",
            ["在宅", "通所", "施設外就労"],
            index=["在宅", "通所", "施設外就労"].index(
                st.session_state.get("bee_service_type", "在宅")
                if st.session_state.get("bee_service_type", "在宅") in ["在宅", "通所", "施設外就労"]
                else "在宅"
            ),
            key="bee_service_type"
        )

    work_time_col1, work_time_col2, work_time_col3 = st.columns(3)

    with work_time_col1:
        work_start_time = st.text_input(
            "作業開始時間",
            value=st.session_state.get("bee_work_start_time", start_time),
            key="bee_work_start_time"
        )

    with work_time_col2:
        work_end_time = st.text_input(
            "作業終了時間",
            value=st.session_state.get("bee_work_end_time", end_time),
            key="bee_work_end_time"
        )

    with work_time_col3:
        work_break_time = st.text_input(
            "休憩時間",
            value=st.session_state.get("bee_work_break_time", ""),
            key="bee_work_break_time"
        )

    staff_name = st.text_input(
        "日誌入力者",
        value=st.session_state.get("bee_staff_name", current_staff_name),
        key="bee_staff_name"
    )

    note_mode = st.radio(
        "備考の入力方法",
        ["候補から選ぶ", "直接入力"],
        horizontal=True,
        key="bee_note_mode"
    )

    if note_mode == "候補から選ぶ":
        note_candidates = ["在宅利用", "食事摂取量 10/10", "施設外就労(実施報告書等添付)", "入院", ""]
        default_note = st.session_state.get("bee_note_text", "")
        default_index = note_candidates.index(default_note) if default_note in note_candidates else 0

        note = st.selectbox(
            "備考",
            note_candidates,
            index=default_index,
            key="bee_note_select"
        )
    else:
        note = st.text_area(
            "備考",
            value=st.session_state.get("bee_note_text", ""),
            key="bee_note_text",
            height=80
        )

    st.markdown("""
    <style>
    div.stButton > button {
        white-space: normal !important;
        height: auto !important;
        min-height: 3.8rem !important;
        line-height: 1.4 !important;
        padding-top: 0.6rem !important;
        padding-bottom: 0.6rem !important;
    }
    div.stButton > button p {
        white-space: pre-line !important;
        line-height: 1.4 !important;
        margin: 0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    if "knowbe_bee_use_plan" not in st.session_state:
        st.session_state["knowbe_bee_use_plan"] = False

    use_plan = st.session_state.get("knowbe_bee_use_plan", False)

    plan_row = get_plan_row(target_company_id, resident_id)
    plan_text = ""
    if use_plan and plan_row:
        plan_text = (
            f"長期目標: {plan_row.get('long_term_goal', '')}\n"
            f"短期目標: {plan_row.get('short_term_goal', '')}"
        )

    example_row = get_staff_example_row(target_company_id, staff_name)
    rule_row = get_personal_rule_row(target_company_id, staff_name)

    examples_text = build_examples_text(service_type, example_row)
    loaded_rule_text = rule_row.get("rule_text", "") if rule_row else ""

    st.caption(f"DEBUG 個人ルール読込: {loaded_rule_text!r}")

    preview_note = note if note_mode == "候補から選ぶ" else st.session_state.get("bee_note_text", "")

    memo_cols = st.columns(2)

    with memo_cols[0]:
        start_memo = st.text_area(
            "開始メモ",
            value=st.session_state.get("bee_start_memo", ""),
            key="bee_start_memo",
            height=140
        )

    with memo_cols[1]:
        end_memo = st.text_area(
            "終了メモ",
            value=st.session_state.get("bee_end_memo", ""),
            key="bee_end_memo",
            height=140
        )

    send_memo_cols = st.columns(4)

    with send_memo_cols[0]:
        start_send_raw = st.button(
            "<利用者状態欄>\n開始メモをだけを入力した\n文章のままknowbeへ送信",
            key="bee_send_start_raw",
            use_container_width=True
        )

    with send_memo_cols[1]:
        start_send_gemini = st.button(
            "<職員考察欄>\n開始メモをだけをGeminiで\n文を編集してknowbeへ送信",
            key="bee_send_start_gemini",
            use_container_width=True
        )

    with send_memo_cols[2]:
        end_send_raw = st.button(
            "<利用者状態欄>\n終了メモをだけを入力した\n文章のままknowbeへ送信",
            key="bee_send_end_raw",
            use_container_width=True
        )

    with send_memo_cols[3]:
        end_send_gemini = st.button(
            "<職員考察欄>\n終了メモをだけをGeminiで\n文を編集してknowbeへ送信",
            key="bee_send_end_gemini",
            use_container_width=True
        )

    save_payload = {
        "company_id": target_company_id,
        "company_name": target_company_name,
        "date": str(target_date),
        "resident_id": resident_id,
        "resident_name": resident_name,
        "start_time": start_time,
        "end_time": end_time,
        "meal_flag": meal_flag,
        "note": preview_note,
        "start_memo": start_memo,
        "end_memo": end_memo,
        "staff_name": staff_name,
        "service_type": service_type,
        "knowbe_user": st.session_state.get("bee_knownbe_user_name", "未登録"),
        "use_plan": st.session_state.get("knowbe_bee_use_plan", False),
    }

    st.divider()
    st.markdown("## 送信")

    show_time_errors = False
    if (
        str(start_time).strip()
        or str(end_time).strip()
        or str(work_start_time).strip()
        or str(work_end_time).strip()
    ):
        show_time_errors = True

    if show_time_errors:
        time_errors = validate_bee_times(
            resident_id=resident_id,
            target_date=target_date,
            start_time=start_time,
            end_time=end_time,
            work_start_time=work_start_time,
            work_end_time=work_end_time,
        )
    else:
        time_errors = []

    if show_time_errors and time_errors:
        for err in time_errors:
            st.error(err)

    # st.caption(f"DEBUG 個人ルール読込: {loaded_rule_text!r}")
    # st.caption(f"DEBUG start_memo: {start_memo!r}")
    # st.caption(f"DEBUG end_memo: {end_memo!r}")

    send_cols = st.columns([1, 1])

    with send_cols[0]:
        bulk_send_gemini = st.button(
            "開始メモをGeminiで編集して利用者状態欄へ\n終了メモをGeminiで編集して職員考察欄へ\nまとめて送信",
            key="bee_bulk_send_gemini",
            width="stretch",
            disabled=bool(time_errors)
        )

    with send_cols[1]:
        bulk_send_raw = st.button(
            "開始メモを入力した文のまま編集なしで利用者状態欄へ\n終了メモを入力した文のまま編集なしで職員考察欄へ\nまとめて送信",
            key="bee_bulk_send_raw",
            width="stretch",
            disabled=bool(time_errors)
        )

    if bulk_send_gemini:
        try:
            st.session_state.pop("bee_generated_status", None)
            st.session_state.pop("bee_generated_support", None)

            generated_status, generated_support = generate_bee_texts(
                resident_name=resident_name,
                service_type=service_type,
                meal_flag=meal_flag,
                note_text=preview_note,
                start_memo=start_memo,
                end_memo=end_memo,
                staff_name=staff_name,
                plan_text=plan_text,
                examples_text=examples_text,
                rule_text=loaded_rule_text,
            )

            st.session_state["bee_generated_status"] = generated_status
            st.session_state["bee_generated_support"] = generated_support

            ok = send_to_knowbe_from_bee(
                company_id=target_company_id,
                target_date=str(target_date),
                resident_name=resident_name,
                service_type=service_type,
                start_time=start_time,
                end_time=end_time,
                meal_flag=meal_flag,
                note_text=preview_note,
                generated_status=generated_status,
                generated_support=generated_support,
                staff_name=staff_name,
                knowbe_target="bulk_gemini",
                work_start_time=work_start_time,
                work_end_time=work_end_time,
                work_break_time=work_break_time,
                work_memo="",
                login_username=resolved_knowbe_user,
                login_password=resolved_knowbe_pw,
                send_user_status=True,
                send_staff_comment=True,
            )

            if ok:
                st.success("Gemini文で一気送信できました！")
                st.session_state.pop("bee_generated_status", None)
                st.session_state.pop("bee_generated_support", None)

        except Exception as e:
            st.error(f"Gemini一気送信失敗です: {e}")

    elif bulk_send_raw:
        try:
            ok = send_to_knowbe_from_bee(
                company_id=target_company_id,
                target_date=str(target_date),
                resident_name=resident_name,
                service_type=service_type,
                start_time=start_time,
                end_time=end_time,
                meal_flag=meal_flag,
                note_text=preview_note,
                generated_status=start_memo,
                generated_support=end_memo,
                staff_name=staff_name,
                knowbe_target="bulk_raw",
                work_start_time=work_start_time,
                work_end_time=work_end_time,
                work_break_time=work_break_time,
                work_memo="",
                login_username=resolved_knowbe_user,
                login_password=resolved_knowbe_pw,
                send_user_status=True,
                send_staff_comment=True,
            )

            if ok:
                st.success("Gemini編集なしでそのままKnowbeへ入力できました！")

        except Exception as e:
            st.error(f"編集なし送信失敗です: {e}")

    st.warning(
        "Knowbe送信を使うには、画面上でKnowbe情報を入力するか、管理者メニューの『Knowbe情報登録』で保存してください。"
    )

    st.divider()
    st.markdown("## 補助設定")

    if "knowbe_bee_use_plan" not in st.session_state:
        st.session_state["knowbe_bee_use_plan"] = False

    use_plan = st.checkbox(
        "個別支援計画を参照する",
        key="knowbe_bee_use_plan"
    )

    st.divider()
    st.markdown("## スタッフ例文・個人ルール")

    if not example_row:
        st.warning("この入力者のスタッフ例文は未登録です。")
    if not rule_row:
        st.warning("この入力者の個人ルールは未登録です。")

    if "bee_rule_edit_open" not in st.session_state:
        st.session_state["bee_rule_edit_open"] = False

    is_editing = st.session_state["bee_rule_edit_open"]

    ex_cols1 = st.columns(2)
    with ex_cols1[0]:
        home_start_value = st.text_area(
            "在宅作業開始例文",
            value=example_row.get("home_start_example", "") if example_row else "",
            key="bee_home_start_example_unified",
            height=100,
            disabled=not is_editing
        )
    with ex_cols1[1]:
        home_end_value = st.text_area(
            "在宅作業終了例文",
            value=example_row.get("home_end_example", "") if example_row else "",
            key="bee_home_end_example_unified",
            height=100,
            disabled=not is_editing
        )

    ex_cols2 = st.columns(2)
    with ex_cols2[0]:
        day_start_value = st.text_area(
            "通所作業開始例文",
            value=example_row.get("day_start_example", "") if example_row else "",
            key="bee_day_start_example_unified",
            height=100,
            disabled=not is_editing
        )
    with ex_cols2[1]:
        day_end_value = st.text_area(
            "通所作業終了例文",
            value=example_row.get("day_end_example", "") if example_row else "",
            key="bee_day_end_example_unified",
            height=100,
            disabled=not is_editing
        )

    ex_cols3 = st.columns(2)
    with ex_cols3[0]:
        outside_start_value = st.text_area(
            "施設外作業開始例文",
            value=example_row.get("outside_start_example", "") if example_row else "",
            key="bee_outside_start_example_unified",
            height=100,
            disabled=not is_editing
        )
    with ex_cols3[1]:
        outside_end_value = st.text_area(
            "施設外作業終了例文",
            value=example_row.get("outside_end_example", "") if example_row else "",
            key="bee_outside_end_example_unified",
            height=100,
            disabled=not is_editing
        )

    bottom_cols = st.columns([5, 1])
    with bottom_cols[0]:
        rule_text_value = st.text_area(
            "個人ルール",
            value=rule_row.get("rule_text", "") if rule_row else "未登録です",
            key="bee_rule_text_unified",
            height=160,
            disabled=not is_editing,
            placeholder="未登録です"
        )

    with bottom_cols[1]:
        st.write("")
        st.write("")
        if not is_editing:
            if st.button("編集", key="bee_rule_edit_open_btn", use_container_width=True):
                st.session_state["bee_rule_edit_open"] = True
                st.rerun()
        else:
            if st.button("登録", key="bee_save_examples_rules", use_container_width=True):
                save_staff_examples_record(
                    company_id=target_company_id,
                    staff_name=staff_name,
                    home_start_example=home_start_value,
                    home_end_example=home_end_value,
                    day_start_example=day_start_value,
                    day_end_example=day_end_value,
                    outside_start_example=outside_start_value,
                    outside_end_example=outside_end_value,
                )
                save_personal_rules_record(
                    company_id=target_company_id,
                    staff_name=staff_name,
                    rule_text="" if rule_text_value == "未登録です" else rule_text_value,
                )
                st.success("スタッフ例文・個人ルールを登録しました！")
                st.session_state["bee_rule_edit_open"] = False
                st.rerun()

    st.divider()
    st.markdown("## 入力内容確認")
    st.json(save_payload)




def get_external_contacts_df():
    return get_external_contacts_df_cached().copy()


def get_resident_links_df():
    return get_resident_links_df_cached().copy()


def get_next_contact_id(contact_df):
    if contact_df is None or contact_df.empty or "contact_id" not in contact_df.columns:
        return "C001"

    numbers = []
    for x in contact_df["contact_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("C"):
            num = x[1:]
            if num.isdigit():
                numbers.append(int(num))

    next_num = max(numbers) + 1 if numbers else 1
    return f"C{next_num:03d}"


def get_resident_contact_cards(resident_id):
    links_df = get_resident_links_df()
    contacts_df = get_external_contacts_df()

    if links_df.empty or contacts_df.empty:
        return pd.DataFrame(columns=[
            "contact_id", "category1", "category2",
            "name", "organization", "phone", "memo", "role"
        ])

    target_links = links_df[
        links_df["resident_id"].astype(str) == str(resident_id)
    ].copy()

    if target_links.empty:
        return pd.DataFrame(columns=[
            "contact_id", "category1", "category2",
            "name", "organization", "phone", "memo", "role"
        ])

    merged = target_links.merge(
        contacts_df,
        how="left",
        on="contact_id"
    )

    for col in ["category1", "category2", "name", "organization", "phone", "memo", "role"]:
        if col not in merged.columns:
            merged[col] = ""

    merged = merged.fillna("")
    return merged

def get_contact_residents(contact_id, links_df=None, master_df=None):
    if links_df is None:
        links_df = get_resident_links_df()
    if master_df is None:
        company_id = get_current_company_id()
        master_df = get_resident_master_df(company_id)

    if links_df.empty or master_df.empty:
        return pd.DataFrame(columns=["resident_id", "resident_name", "status", "role"])

    target_links = links_df[
        links_df["contact_id"].astype(str) == str(contact_id)
    ].copy()

    if target_links.empty:
        return pd.DataFrame(columns=["resident_id", "resident_name", "status", "role"])

    merged = target_links.merge(
        master_df[["resident_id", "resident_name", "status"]],
        how="left",
        on="resident_id"
    )

    for col in ["resident_id", "resident_name", "status", "role"]:
        if col not in merged.columns:
            merged[col] = ""

    return merged.fillna("")


def get_schedule_slot_num(row):
    memo_val = str(row.get("memo", "")).strip()
    if memo_val.startswith("slot:"):
        try:
            num = int(memo_val.replace("slot:", ""))
            return num if 1 <= num <= 4 else 0
        except Exception:
            return 0
    return 0

def build_schedule_form_base(schedule_df, resident_id):
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    service_types = ["病院", "看護", "介護"]

    base = {
        service: {
            "place": "",
            "phone": "",
            "person_in_charge": "",
            "days": {wd: ["", "", "", ""] for wd in weekdays}
        }
        for service in service_types
    }

    if schedule_df is None or schedule_df.empty:
        return base

    view = schedule_df[
        (schedule_df["resident_id"].astype(str) == str(resident_id)) &
        (schedule_df["service_type"].astype(str).isin(service_types))
    ].copy()

    if view.empty:
        return base

    view["_slot_num"] = view.apply(get_schedule_slot_num, axis=1)
    view = view.sort_values(["service_type", "weekday", "_slot_num", "start_time"])

    for _, hit in view.iterrows():
        service_type = str(hit.get("service_type", "")).strip()
        weekday = str(hit.get("weekday", "")).strip()

        if service_type not in base or weekday not in base[service_type]["days"]:
            continue

        if not base[service_type]["place"]:
            base[service_type]["place"] = str(hit.get("place", "")).strip()
        if not base[service_type]["phone"]:
            base[service_type]["phone"] = str(hit.get("phone", "")).strip()
        if not base[service_type]["person_in_charge"]:
            base[service_type]["person_in_charge"] = str(hit.get("person_in_charge", "")).strip()

        slot_num = get_schedule_slot_num(hit)
        if 1 <= slot_num <= 4:
            start_time = str(hit.get("start_time", "")).strip()
            end_time = str(hit.get("end_time", "")).strip()

            time_text = start_time
            if end_time:
                time_text += f"〜{end_time}"

            base[service_type]["days"][weekday][slot_num - 1] = time_text

    return base

def render_resident_schedule_html(schedule_view):
    schedule_view = schedule_view.copy()

    color_map = {
        "病院": "#f6c90e",
        "看護": "#9fd3e6",
        "介護": "#b7e20f",
    }
    col_order = ["月", "火", "水", "木", "金", "土", "日"]

    legend_html = ""
    for name, color in color_map.items():
        legend_html += f'''
        <div style="
            display:inline-block;
            background:{color};
            color:#111;
            font-weight:700;
            padding:8px 28px;
            border:2px solid #111;
            margin-right:10px;
            margin-bottom:10px;
            min-width:120px;
            text-align:center;
        ">{name}</div>
        '''

    table_html = '<table style="width:100%; border-collapse:collapse; table-layout:fixed;">'
    table_html += '<tr>'
    for wd in col_order:
        table_html += f'''
        <th style="
            border:2px solid #111;
            padding:8px;
            text-align:center;
            background:#fafafa;
            width:14.28%;
        ">{wd}</th>
        '''
    table_html += '</tr><tr>'

    for wd in col_order:
        day_df = schedule_view[schedule_view["weekday"].astype(str) == wd].copy()
        items = []

        if not day_df.empty:
            day_df["_slot_num"] = day_df.apply(get_schedule_slot_num, axis=1)
            day_df = day_df.sort_values(["_slot_num", "service_type", "start_time"])

            for _, hit in day_df.iterrows():
                service_type = str(hit.get("service_type", "")).strip()
                bg = color_map.get(service_type, "#eeeeee")

                start_time = str(hit.get("start_time", "")).strip()
                end_time = str(hit.get("end_time", "")).strip()

                time_text = start_time
                if end_time:
                    time_text += f"〜{end_time}"

                items.append(
                    f'''
                    <div style="
                        background:{bg};
                        border:2px solid #222;
                        padding:6px 8px;
                        margin:6px 0;
                        font-weight:700;
                        text-align:left;
                        min-height:34px;
                        box-sizing:border-box;
                        overflow:hidden;
                    ">{time_text}</div>
                    '''
                )

        table_html += f'''
        <td style="
            border:2px solid #111;
            vertical-align:top;
            padding:6px;
            height:190px;
            width:14.28%;
            box-sizing:border-box;
        ">
            {''.join(items)}
        </td>
        '''

    table_html += "</tr></table>"

    components.html(legend_html + table_html, height=320)

def go_resident_detail(resident_id):
    st.session_state.selected_resident_id = resident_id
    st.rerun()


def back_to_resident_list():
    st.session_state.selected_resident_id = ""
    st.rerun()

# ログイン中メンバー表示
try:
    active_df = load_db("active_users")
except Exception:
    active_df = pd.DataFrame(columns=["user", "login_at", "last_seen"])

st.sidebar.markdown(f"### 👤 ログイン中")

if active_df is None or active_df.empty:
    st.sidebar.write("現在ログイン中の人はいありません。")
else:
    active_df = active_df.fillna("")
    now_dt = now_jst().replace(tzinfo=None)
    visible_rows = []

    for _, row in active_df.iterrows():
        last_seen = str(row.get("last_seen", "")).strip()
        if not last_seen:
            continue
        try:
            last_dt = pd.to_datetime(last_seen).to_pydatetime()
            if (now_dt - last_dt).total_seconds() <= 15 * 60:
                visible_rows.append(row)
        except Exception:
            pass

    if visible_rows:
        for row in visible_rows:
            user_name = str(row.get("user", "")).strip()
            login_at = str(row.get("login_at", "")).strip()
            st.sidebar.write(f"**{user_name}**")
            st.sidebar.caption(f"ログイン: {login_at}")
    else:
        st.sidebar.write("現在ログイン中の人はいありません。")

## マイ状況
try:
    company_id = get_current_company_id()
    task_df = get_tasks_df(company_id)
except Exception:
    task_df = pd.DataFrame(columns=["id", "task", "status", "user", "limit", "priority", "updated_at"])

my_active = task_df[
    (task_df["status"].astype(str).str.strip() == "作業中") &
    (task_df["user"].astype(str).str.strip() == st.session_state.user)
]

my_todo = task_df[
    (task_df["status"].astype(str).str.strip() == "未着手")
]

st.sidebar.divider()
st.sidebar.markdown("### 📌 マイ状況")

active_count = len(my_active)
todo_count = len(my_todo)

if st.sidebar.button(f"作業中: {active_count}件", key="go_my_active", use_container_width=True):
    st.session_state.current_page = "② タスクの引き受け・報告"
    st.rerun()

if st.sidebar.button(f"未着手全体: {todo_count}件", key="go_todo_board", use_container_width=True):
    st.session_state.current_page = "① 未着手の任務（掲示板）"
    st.rerun()

# ここを追加するです
page = st.session_state.get("current_page", "① 未着手の任務（掲示板）")

st.sidebar.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

is_break_selected = (st.session_state.current_page == "休憩室")

if is_break_selected:
    st.sidebar.markdown(
        '<div class="menu-selected-wrap"><div class="menu-selected-box">● 🍵休憩室🍵</div></div>',
        unsafe_allow_html=True
    )
else:
    if st.sidebar.button("🍵休憩室🍵", key="menu_break_room_fixed", use_container_width=True):
        st.session_state.current_page = "休憩室"
        st.rerun()

is_contact_selected = (st.session_state.current_page == "お問い合わせ")

if is_contact_selected:
    st.sidebar.markdown(
        '<div class="menu-selected-wrap"><div class="menu-selected-box">● 📩 お問い合わせ</div></div>',
        unsafe_allow_html=True
    )
else:
    if st.sidebar.button("📩 お問い合わせ", key="menu_contact_fixed", use_container_width=True):
        st.session_state.current_page = "お問い合わせ"
        st.rerun()

# ==========================================
# ① 未着手の任務（掲示板）
# ==========================================
if page == "① 未着手の任務（掲示板）":
    def show_task_board_page():
        st.title("📋 未着手タスク一覧")
        st.write("現在、依頼されている業務の一覧です。新しいタスクを登録することも可能です。")

        current_company_id = get_current_company_id()

        with st.expander("➕ 新規タスクを登録する"):
            with st.form("task_form"):
                bulk_tasks = st.text_area(
                    "タスク名（1行に1件。まとめて貼り付けOK）",
                    placeholder="請求書の確認\n支援記録の見直し\n送迎表の作成"
                )
                t_prio = st.select_slider("緊急度", options=["通常", "重要", "至急"])
                t_limit = st.date_input("完了期限", now_jst().date())

                if st.form_submit_button("タスクを登録"):
                    lines = [x.strip() for x in str(bulk_tasks).replace("\r\n", "\n").split("\n") if x.strip()]

                    if lines:
                        company_id = get_current_company_id()
                        df = get_tasks_df(company_id)

                        next_id = get_next_task_id(df)
                        new_rows = []
                        now_str = now_jst().strftime("%Y-%m-%d %H:%M")

                        for task_name in lines:
                            new_rows.append({
                                "company_id": current_company_id,
                                "id": next_id,
                                "task": task_name,
                                "status": "未着手",
                                "user": "",
                                "limit": str(t_limit),
                                "priority": t_prio,
                                "updated_at": now_str
                            })
                            next_id += 1

                        add_df = pd.DataFrame(new_rows)
                        merged_df = pd.concat([df, add_df], ignore_index=True)

                        save_db(merged_df, "task")
                        sync_task_events_to_calendar(current_company_id)

                        st.success(f"{len(new_rows)}件のタスクを登録しました！")
                        st.rerun()
                    else:
                        st.error("タスクを1件以上入力してください。")

        df = get_tasks_df(current_company_id)
        todo = df[df["status"].astype(str).str.strip() == "未着手"].copy()

        if not todo.empty:
            prio_map = {"至急": 0, "重要": 1, "通常": 2}
            todo["p_val"] = todo["priority"].map(prio_map)
            st.dataframe(
                todo.sort_values(["p_val", "limit"])[["priority", "limit", "task"]],
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("現在、未着手のタスクはありません。")

    show_task_board_page()
    
# ==========================================
# ② タスクの引き受け・報告
# ==========================================
elif page == "② タスクの引き受け・報告":
    def show_my_tasks_page():
        st.title("🎯 タスク管理")

        current_company_id = get_current_company_id()
        df = get_tasks_df(current_company_id)

        st.subheader("📦 新しくタスクを引き受ける")
        todo = df[df["status"].astype(str).str.strip() == "未着手"].copy()

        if todo.empty:
            st.write("引き受け可能なタスクはありません。")

        for _, row in todo.iterrows():
            p_symbol = "🔴 [至急]" if row["priority"] == "至急" else "🟡 [重要]" if row["priority"] == "重要" else "⚪ [通常]"
            if st.button(
                f"{p_symbol} {row['task']} (期限:{row['limit']}) を開始する",
                key=f"get_{current_company_id}_{row['id']}"
            ):
                start_task(row["id"], current_company_id)
                st.rerun()

        st.divider()
        st.subheader("⚡ 現在対応中のタスク")

        my_tasks = df[
            (df["status"].astype(str).str.strip() == "作業中") &
            (df["user"].astype(str).str.strip() == str(st.session_state.user).strip())
        ].copy()

        if my_tasks.empty:
            st.write("現在、対応中のタスクはありません。")

        for _, row in my_tasks.iterrows():
            if st.button(
                f"✅ {row['task']} の完了を報告する",
                key=f"done_{current_company_id}_{row['id']}",
                type="primary"
            ):
                complete_task(row["id"], current_company_id)
                st.rerun()

    show_my_tasks_page()
# ==========================================
# ③ 稼働状況・完了履歴
# ==========================================
elif page == "③ 稼働状況・完了履歴":
    @st.fragment(run_every=180)
    def show_status_page():
        st.title("📊 チーム稼働状況")

        current_company_id = get_current_company_id()
        df = get_tasks_df(current_company_id)
        df = normalize_company_scoped_df(df, get_task_required_cols())

        # ------------------------------------------
        # 期限アラート
        # ------------------------------------------
        st.subheader("⏰ 期限アラート")

        today = now_jst().date()
        tomorrow = today + timedelta(days=1)

        deadline_alerts = []

        for _, row in df.iterrows():
            status = str(row.get("status", "")).strip()
            task_name = str(row.get("task", "")).strip()
            limit_str = str(row.get("limit", "")).strip()

            if status == "完了":
                continue

            if limit_str:
                try:
                    limit_date = pd.to_datetime(limit_str).date()

                    if limit_date < today:
                        deadline_alerts.append(("期限切れ", task_name, str(limit_date)))
                    elif limit_date == today:
                        deadline_alerts.append(("今日期限", task_name, str(limit_date)))
                    elif limit_date == tomorrow:
                        deadline_alerts.append(("明日期限", task_name, str(limit_date)))
                except Exception:
                    pass

        if deadline_alerts:
            for kind, task_name, d in deadline_alerts:
                if kind == "期限切れ":
                    st.error(f"🔴 {kind}: {task_name}（{d}）")
                elif kind == "今日期限":
                    st.warning(f"🟠 {kind}: {task_name}（{d}）")
                else:
                    st.info(f"🟡 {kind}: {task_name}（{d}）")
        else:
            st.write("期限アラートはありません。")

        st.divider()

        # ------------------------------------------
        # 現在稼働中のメンバー
        # ------------------------------------------
        st.subheader("🏃 現在稼働中のメンバー")

        active = df[df["status"].astype(str).str.strip() == "作業中"].copy()

        if active.empty:
            st.write("現在稼働中のメンバーはいません。")
        else:
            for _, row in active.iterrows():
                st.info(f"👤 **{row['user']}**： {row['task']} （着手時刻：{row['updated_at']}）")

        st.divider()

        # ------------------------------------------
        # 完了済みのタスク（過去7日間）
        # ------------------------------------------
        st.subheader("✅ 完了済みのタスク（過去7日間）")

        one_week_ago = now_jst() - timedelta(days=7)

        done_rows = []

        for _, row in df.iterrows():
            status = str(row.get("status", "")).strip()
            updated_at = str(row.get("updated_at", "")).strip()

            if status != "完了" or not updated_at:
                continue

            try:
                done_dt = pd.to_datetime(updated_at)
                if done_dt >= one_week_ago:
                    done_rows.append(row)
            except Exception:
                pass

        if done_rows:
            done_df = pd.DataFrame(done_rows)
            done_df = done_df.sort_values("updated_at", ascending=False)
            st.table(done_df[["updated_at", "user", "task"]])
        else:
            st.write("直近1週間以内の完了履歴はありません。")

    show_status_page()

# ==========================================
# ④ チームチャット（画像添付対応です！）
# ==========================================
elif page == "④ チームチャット":
    @st.fragment(run_every=15)
    def show_chat_page():
        st.title("💬 業務連絡チャット")
        chat_df = load_db("chat")
        today = now_jst().strftime("%Y-%m-%d")

        # 必要列を保証
        if chat_df is None or chat_df.empty:
            chat_df = pd.DataFrame(columns=["date", "time", "user", "message", "image_data"])
        else:
            for col in ["date", "time", "user", "message", "image_data"]:
                if col not in chat_df.columns:
                    chat_df[col] = ""

        mode = st.radio("表示切り替え", ["本日の連絡", "ログ倉庫（過去分）"], horizontal=True)

        if mode == "本日の連絡":
            with st.form("chat_form", clear_on_submit=True):
                msg = st.text_area("業務連絡を入力してください")
                uploaded_image = st.file_uploader(
                    "画像を添付（ドラッグ＆ドロップ可）",
                    type=["png", "jpg", "jpeg"],
                    key="chat_image_uploader"
                )

                if uploaded_image is not None:
                    st.image(uploaded_image, caption="送信前プレビュー", use_container_width=True)

                if st.form_submit_button("送信"):
                    if msg or uploaded_image is not None:
                        image_data = ""

                        if uploaded_image is not None:
                            file_bytes = uploaded_image.read()
                            image_data = base64.b64encode(file_bytes).decode("utf-8")

                        new_msg = pd.DataFrame([{
                            "date": today,
                            "time": datetime.now().strftime("%H:%M"),
                            "user": st.session_state.user,
                            "message": msg,
                            "image_data": image_data
                        }])

                        save_db(pd.concat([chat_df, new_msg], ignore_index=True), "chat")
                        st.success("送信しました！")
                        st.rerun()
                    else:
                        st.error("メッセージか画像のどちらかを入れてください。")

            st.divider()

            display_df = chat_df[chat_df["date"].astype(str) == today].copy()

            if display_df.empty:
                st.write("本日の連絡事項はありません。")

            display_df = display_df.fillna("")

            for _, row in display_df.sort_values("time", ascending=False).iterrows():
                user_name = str(row.get("user", ""))
                msg_text = str(row.get("message", ""))
                time_text = str(row.get("time", ""))
                image_data = str(row.get("image_data", ""))

                st.markdown(f"**{user_name}** ({time_text})")

                if msg_text.strip():
                    st.write(msg_text)

                if image_data.strip():
                    try:
                        image_bytes = base64.b64decode(image_data)
                        st.image(image_bytes, caption="添付画像", use_container_width=True)
                    except Exception:
                        st.warning("画像の読み込みに失敗しました。")

                st.divider()

        else:
            st.subheader("📚 過去の連絡ログ")

            if chat_df.empty:
                st.write("過去ログはありません。")
                return

            dates = sorted(
                [str(d).strip() for d in chat_df["date"].dropna().tolist() if str(d).strip()],
                reverse=True
            )

            if not dates:
                st.write("過去ログはありません。")
                return

            target_date = st.selectbox("参照する日付を選択してください", dates)

            if target_date:
                log_df = chat_df[chat_df["date"].astype(str) == target_date].copy().fillna("")

                for _, row in log_df.sort_values("time", ascending=False).iterrows():
                    user_name = str(row.get("user", ""))
                    msg_text = str(row.get("message", ""))
                    time_text = str(row.get("time", ""))
                    image_data = str(row.get("image_data", ""))

                    st.markdown(f"**{user_name}** ({time_text})")

                    if msg_text.strip():
                        st.write(msg_text)

                    if image_data.strip():
                        try:
                            image_bytes = base64.b64decode(image_data)
                            st.image(image_bytes, caption="添付画像", use_container_width=True)
                        except Exception:
                            st.warning("画像の読み込みに失敗しました。")

                    st.divider()

    show_chat_page()

# ==========================================
# ⑤ 業務マニュアル ver2 (画像D&D + 削除対応)
# ==========================================
elif page == "⑤ 業務マニュアル":
    @st.fragment(run_every=180)
    def show_manual_page():
        st.title("📚 業務マニュアル")

        m_df = load_db("manual")

        if m_df is None or m_df.empty:
            m_df = pd.DataFrame(columns=["id", "title", "content", "image_data", "created_at"])
        else:
            for col in ["id", "title", "content", "image_data", "created_at"]:
                if col not in m_df.columns:
                    m_df[col] = ""

        with st.expander("📝 新しいマニュアルを作成する"):
            with st.form("manual_form"):
                title = st.text_input("マニュアルのタイトル")
                content = st.text_area("手順の説明")

                uploaded_file = st.file_uploader(
                    "写真をドラッグ&ドロップ",
                    type=["png", "jpg", "jpeg"]
                )

                if uploaded_file is not None:
                    st.image(
                        uploaded_file,
                        caption="アップロード画像プレビュー",
                        use_container_width=True
                    )

                if st.form_submit_button("保存する"):
                    if title and content:
                        if m_df.empty:
                            next_id = 1
                        else:
                            ids = pd.to_numeric(m_df["id"], errors="coerce").dropna()
                            next_id = int(ids.max()) + 1 if not ids.empty else 1

                        image_data = ""
                        if uploaded_file is not None:
                            bytes_data = uploaded_file.getvalue()
                            image_data = base64.b64encode(bytes_data).decode("utf-8")

                        new_m = pd.DataFrame([{
                            "id": next_id,
                            "title": title,
                            "content": content,
                            "image_data": image_data,
                            "created_at": now_jst().strftime("%Y-%m-%d %H:%M")
                        }])

                        save_db(pd.concat([m_df, new_m], ignore_index=True), "manual")
                        st.success("マニュアルを保存しました！")
                        st.rerun()
                    else:
                        st.error("タイトルと説明は必須です。")

        st.divider()

        m_df = load_db("manual")

        if m_df is None or m_df.empty:
            st.info("マニュアルはまだ登録されていません。")
            return

        for col in ["id", "title", "content", "image_data", "created_at"]:
            if col not in m_df.columns:
                m_df[col] = ""

        m_df = m_df.fillna("")

        try:
            display_df = m_df.sort_values("created_at", ascending=False)
        except Exception:
            display_df = m_df.copy()

        for _, row in display_df.iterrows():
            manual_id = row.get("id", "")
            title = str(row.get("title", "")).strip()
            content = str(row.get("content", "")).strip()
            created_at = str(row.get("created_at", "")).strip()
            image_data = str(row.get("image_data", "")).strip()

            with st.expander(f"📖 {title} (作成日: {created_at})"):
                if content:
                    st.write(content)

                if image_data:
                    try:
                        image_bytes = base64.b64decode(image_data)
                        st.image(
                            image_bytes,
                            caption="添付画像",
                            use_container_width=True
                        )
                    except Exception:
                        st.warning("画像データの読み込みに失敗しました。")

                if st.button("🗑️ このマニュアルを削除する", key=f"delete_manual_{manual_id}"):
                    new_df = m_df[m_df["id"].astype(str) != str(manual_id)].copy()
                    save_db(new_df, "manual")
                    st.success("削除しました。")
                    st.rerun()

    show_manual_page()

# ==========================================
# ⑥ 日誌入力状況（年つき横表・Excel風です！）
# ==========================================
elif page == "⑥ 日誌入力状況":
    @st.fragment(run_every=180)
    def show_record_status_page():
        st.title("📝 日誌入力状況管理")

        # 表示したい年の範囲
        start_year = 2025
        end_year = 2026

        # 年月列を作るです
        month_cols = []
        for year in range(start_year, end_year + 1):
            for month in range(1, 13):
                month_cols.append(f"{year}年{month}月")

        required_cols = ["resident_name"] + month_cols

        r_df = load_db("record_status")

        # 空でも最低限の形に整えるです
        if r_df is None or r_df.empty:
            r_df = pd.DataFrame(columns=required_cols)
        else:
            for col in required_cols:
                if col not in r_df.columns:
                    r_df[col] = ""
            r_df = r_df[required_cols].copy()

        # ==========================================
        # 利用者一括追加
        # ==========================================
        with st.expander("👤 利用者をまとめて追加する"):
            bulk_names = st.text_area(
                "名前を入力（改行ごとに1人。Excelの縦列コピーもOK）",
                placeholder="荒木 和也\n石田 愛子\n岩城 勝也"
            )

            if st.button("利用者を追加する", key="add_residents_button"):
                names = [
                    n.strip()
                    for n in str(bulk_names).replace("\r\n", "\n").split("\n")
                    if n.strip()
                ]

                if not names:
                    st.error("名前を1人以上入力してください。")
                else:
                    existing_names = set(
                        str(x).strip()
                        for x in r_df["resident_name"].dropna().tolist()
                        if str(x).strip()
                    )

                    new_rows = []
                    skipped = []

                    for name in names:
                        if name in existing_names:
                            skipped.append(name)
                        else:
                            row = {"resident_name": name}
                            for m in month_cols:
                                row[m] = ""
                            new_rows.append(row)

                    if new_rows:
                        add_df = pd.DataFrame(new_rows)
                        r_df = pd.concat([r_df, add_df], ignore_index=True)
                        save_db(r_df, "record_status")

                        if skipped:
                            st.success(
                                f"{len(new_rows)}人追加しました。重複スキップ: {', '.join(skipped)}"
                            )
                        else:
                            st.success(f"{len(new_rows)}人追加しました。")
                        st.rerun()
                    else:
                        st.warning("全員すでに登録済みです。")

        st.divider()
        st.caption("各セルに「未入力」「15日まで」「完了」など自由に入力できます。")

        # data_editorで落ちないように、全部文字列にそろえるです
        r_df = r_df.fillna("")

        for col in required_cols:
            if col not in r_df.columns:
                r_df[col] = ""
            r_df[col] = r_df[col].astype(str)

        # 列設定を自動生成
        column_config = {
            "resident_name": st.column_config.TextColumn("氏名", width="medium")
        }
        for col in month_cols:
            column_config[col] = st.column_config.TextColumn(col, width="small")

        # 表示用
        edited_df = st.data_editor(
            r_df,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config=column_config,
            key="record_status_editor"
        )

        # 保存ボタン
        col1, col2 = st.columns([1, 4])
        with col1:
            if st.button("💾 表を保存する", type="primary", key="save_record_status_button"):
                edited_df = edited_df.fillna("")

                # 氏名空欄の行は除外
                edited_df = edited_df[
                    edited_df["resident_name"].astype(str).str.strip() != ""
                ].copy()

                # 必要列だけにそろえる
                for col in required_cols:
                    if col not in edited_df.columns:
                        edited_df[col] = ""
                    edited_df[col] = edited_df[col].astype(str)

                edited_df = edited_df[required_cols]

                save_db(edited_df, "record_status")
                st.success("保存しました！")
                st.rerun()

    show_record_status_page()

# ==========================================
# ⑦ 勤務カレンダー
# ==========================================
elif page == "⑦ タスクカレンダー":
    sync_task_events_to_calendar()

    @st.fragment(run_every=180)
    def show_calendar_page():
        st.title("📅 タスクカレンダー")

        try:
            cal_df = load_db("calendar")
            company_id = get_current_company_id()
            task_df = get_tasks_df(company_id)

        except Exception:
            st.warning("Googleスプレッドシートとの通信が一時的に不安定です。少し待って再読み込みしてください。")
            return

        # calendarシートの最低列を保証
        if cal_df is None or cal_df.empty:
            cal_df = pd.DataFrame(columns=["id", "title", "start", "end", "user", "memo", "source_type", "source_task_id"])
        else:
            for col in ["id", "title", "start", "end", "user", "memo", "source_type", "source_task_id"]:
                if col not in cal_df.columns:
                    cal_df[col] = ""

        # taskシートの最低列を保証
        if task_df is None or task_df.empty:
            task_df = pd.DataFrame(columns=["id", "task", "status", "user", "limit", "priority", "updated_at"])
        else:
            for col in ["id", "task", "status", "user", "limit", "priority", "updated_at"]:
                if col not in task_df.columns:
                    task_df[col] = ""

        with st.expander("➕ 予定を追加する"):
            with st.form("calendar_form"):
                title = st.text_input("予定名")
                start_date = st.date_input("開始日")
                end_date = st.date_input("終了日")
                user_name = st.text_input("担当者", value=st.session_state.user)
                memo = st.text_area("メモ")

                if st.form_submit_button("予定を保存する"):
                    if title:
                        if cal_df.empty:
                            next_id = 1
                        else:
                            try:
                                next_id = pd.to_numeric(cal_df["id"], errors="coerce").max()
                                next_id = 1 if pd.isna(next_id) else int(next_id) + 1
                            except Exception:
                                next_id = len(cal_df) + 1

                        new_row = pd.DataFrame([{
                            "id": next_id,
                            "title": title,
                            "start": str(start_date),
                            "end": str(end_date),
                            "user": user_name,
                            "memo": memo,
                            "source_type": "manual",
                            "source_task_id": ""
                        }])

                        save_db(pd.concat([cal_df, new_row], ignore_index=True), "calendar")
                        st.success("予定を保存しました！")
                        st.rerun()
                    else:
                        st.error("予定名を入れてください。")

        st.divider()

        events = []
        event_ids = set()

        # ------------------------------------------
        # ① 手入力の予定（calendarシート）
        # ------------------------------------------
        display_cal_df = cal_df.fillna("")

        for _, row in display_cal_df.iterrows():
            title = str(row.get("title", "")).strip()
            start = str(row.get("start", "")).strip()
            end = str(row.get("end", "")).strip()
            user_name = str(row.get("user", "")).strip()
            memo = str(row.get("memo", "")).strip()
            source_type = str(row.get("source_type", "")).strip()

            # task由来イベントは下で task_df から作るので、ここでは表示しません
            if source_type in ["task_deadline", "task_active"]:
                continue

            if title and start:
                short_title = title if len(title) <= 10 else title[:10] + "…"
                event_title = short_title if not user_name else f"{user_name}：{short_title}"
                event_id = f"manual_{row.get('id', '')}"

                if event_id not in event_ids:
                    events.append({
                        "id": event_id,
                        "title": event_title,
                        "start": start,
                        "end": end if end else start,
                        "color": "#3788d8",
                        "extendedProps": {
                            "memo": memo,
                            "user": user_name,
                            "source_type": "manual",
                            "full_title": title,
                        }
                    })
                    event_ids.add(event_id)

        # ------------------------------------------
        # ② taskシートから締切イベントを自動生成
        # ------------------------------------------
        display_task_df = task_df.fillna("")
        today = now_jst().date()

        for _, row in display_task_df.iterrows():
            task_id = str(row.get("id", "")).strip()
            task_name = str(row.get("task", "")).strip()
            status = str(row.get("status", "")).strip()
            user_name = str(row.get("user", "")).strip()
            priority = str(row.get("priority", "")).strip()
            limit_str = str(row.get("limit", "")).strip()
            updated_at = str(row.get("updated_at", "")).strip()

            # 締切が今日以降なら出す
            if limit_str:
                try:
                    limit_date = pd.to_datetime(limit_str).date()
                    if limit_date >= today:
                        event_id = f"deadline_{task_id}"
                        short_title = task_name if len(task_name) <= 10 else task_name[:10] + "…"

                        if event_id not in event_ids:
                            events.append({
                                "id": event_id,
                                "title": f"締切：{short_title}",
                                "start": str(limit_date),
                                "end": str(limit_date),
                                "color": "#e74c3c",
                                "extendedProps": {
                                    "memo": f"優先度: {priority} / 状態: {status}",
                                    "user": user_name,
                                    "source_type": "task_deadline",
                                    "full_title": task_name,
                                }
                            })
                            event_ids.add(event_id)
                except Exception:
                    pass

            # 作業中なら開始日で出す
            if status == "作業中" and updated_at:
                try:
                    active_date = pd.to_datetime(updated_at).date()
                    event_id = f"active_{task_id}"
                    short_title = task_name if len(task_name) <= 10 else task_name[:10] + "…"

                    if event_id not in event_ids:
                        events.append({
                            "id": event_id,
                            "title": f"作業中：{short_title}",
                            "start": str(active_date),
                            "end": str(active_date),
                            "color": "#f39c12",
                            "extendedProps": {
                                "memo": f"着手: {updated_at}",
                                "user": user_name,
                                "source_type": "task_active",
                                "full_title": task_name,
                            }
                        })
                        event_ids.add(event_id)
                except Exception:
                    pass

        calendar_options = {
            "initialView": "dayGridMonth",
            "locale": "ja",
            "height": "auto",
            "dayMaxEvents": 3,
            "moreLinkClick": "popover",
            "displayEventTime": False,
            "eventDisplay": "block",
            "headerToolbar": {
                "left": "prev,next today",
                "center": "title",
                "right": "dayGridMonth,timeGridWeek,listWeek"
            }
        }

        state = st_calendar(
            events=events,
            options=calendar_options,
            key="work_calendar"
        )



        st.divider()
        st.subheader("📋 登録済み予定一覧")

        if not display_cal_df.empty:
            st.dataframe(
                display_cal_df[["start", "end", "user", "title", "memo"]],
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("手入力の予定はまだありません。")

        # ------------------------------------------
        # 選択日の予定一覧
        # ------------------------------------------
        st.divider()
        st.subheader("🗓️ 選択日の予定一覧")

        selected_date = None

        if state and state.get("dateClick"):
            selected_date = str(state["dateClick"].get("date", ""))[:10]
        elif state and state.get("eventClick"):
            selected_date = str(state["eventClick"]["event"].get("start", ""))[:10]

        if selected_date:
            day_events = []

            for ev in events:
                ev_start = str(ev.get("start", ""))[:10]
                if ev_start == selected_date:
                    ext = ev.get("extendedProps", {})
                    source_type = ext.get("source_type", "")

                    type_label = "手入力予定"
                    if source_type == "task_deadline":
                        type_label = "締切タスク"
                    elif source_type == "task_active":
                        type_label = "作業中タスク"

                    day_events.append({
                        "種類": type_label,
                        "予定名": ext.get("full_title", ev.get("title", "")),
                        "担当者": ext.get("user", ""),
                        "メモ": ext.get("memo", "")
                    })

            if day_events:
                st.dataframe(
                    pd.DataFrame(day_events),
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.info(f"{selected_date} の予定はありません。")
        else:
            st.caption("日付か予定をクリックすると、その日の一覧を下に表示します。")

        # ------------------------------------------
        # クリックした予定の詳細
        # ------------------------------------------
        if state and state.get("eventClick"):
            clicked = state["eventClick"]["event"]
            ext = clicked.get("extendedProps", {})

            st.divider()
            st.subheader("🔍 選択中の予定")

            full_title = ext.get("full_title", "") or clicked.get("title", "")
            source_type = ext.get("source_type", "")

            type_label = "手入力予定"
            if source_type == "task_deadline":
                type_label = "締切タスク"
            elif source_type == "task_active":
                type_label = "作業中タスク"

            st.write(f"**種類**: {type_label}")
            st.write(f"**予定名**: {full_title}")
            st.write(f"**開始**: {clicked.get('start', '')}")
            st.write(f"**終了**: {clicked.get('end', '')}")

            if ext.get("user"):
                st.write(f"**担当者**: {ext.get('user')}")

            if ext.get("memo"):
                st.write(f"**メモ**: {ext.get('memo')}")

    show_calendar_page()

# ==========================================
# ⑧ 緊急一覧
# ==========================================
elif page == "⑧ 緊急一覧":
    def show_urgent_page():
        st.title("🚨 緊急一覧")

        company_id = get_current_company_id()
        urgent_df = get_urgent_tasks_df(company_id)

        if urgent_df.empty:
            st.success("現在、至急・重要タスクはありません。")
            return

        col1, col2, col3, col4, col5 = st.columns([1, 1, 1, 1, 3])

        with col1:
            if st.button("全部", key="urgent_filter_all", use_container_width=True):
                st.session_state.urgent_filter = "全部"
        with col2:
            if st.button("至急", key="urgent_filter_critical", use_container_width=True):
                st.session_state.urgent_filter = "至急"
        with col3:
            if st.button("重要", key="urgent_filter_important", use_container_width=True):
                st.session_state.urgent_filter = "重要"
        with col4:
            if st.button("未着手", key="urgent_filter_todo", use_container_width=True):
                st.session_state.urgent_filter = "未着手"

        current_filter = st.session_state.get("urgent_filter", "全部")

        if current_filter == "至急":
            urgent_df = urgent_df[urgent_df["priority"].astype(str).str.strip() == "至急"]
        elif current_filter == "重要":
            urgent_df = urgent_df[urgent_df["priority"].astype(str).str.strip() == "重要"]
        elif current_filter == "未着手":
            urgent_df = urgent_df[urgent_df["status"].astype(str).str.strip() == "未着手"]

        st.caption(f"現在の表示: {current_filter}")

        if urgent_df.empty:
            st.info("条件に合う緊急タスクはありません。")
            return

        # ===============================
        # 緊急カード表示（2列・期限色対応）
        # ===============================
        cols = st.columns(2)

        for i, (_, row) in enumerate(urgent_df.iterrows()):
            with cols[i % 2]:
                task_id = str(row.get("id", "")).strip()
                task_name = str(row.get("task", "")).strip()
                priority = str(row.get("priority", "")).strip()
                status = str(row.get("status", "")).strip()
                user_name = str(row.get("user", "")).strip()
                limit_str = str(row.get("limit", "")).strip()
                updated_at = str(row.get("updated_at", "")).strip()

                limit_date = None
                try:
                    limit_date = pd.to_datetime(limit_str, errors="coerce").date()
                except:
                    pass

                today = now_jst().date()

                if limit_date and limit_date < today:
                    icon = "🩸"
                    border_color = "#d63031"
                    bg_color = "#ffeaea"
                elif priority == "至急":
                    icon = "🚨"
                    border_color = "#ff4d4f"
                    bg_color = "#fff1f0"
                else:
                    icon = "⚠️"
                    border_color = "#ff9f43"
                    bg_color = "#fff7e6"

                assignee = user_name if user_name else "未割当"

                with st.container(border=True):
                    st.markdown(
                        f"""
                        <div style="
                            border-left: 8px solid {border_color};
                            background-color: {bg_color};
                            padding: 16px;
                            border-radius: 12px;
                            margin-bottom: 12px;
                            min-height: 160px;
                        ">
                        <div style="font-size:20px; font-weight:700; margin-bottom:6px;">
                        {icon} {priority}
                        </div>
                        <div style="font-size:22px; font-weight:700; margin-bottom:10px;">
                        {task_name}
                        </div>
                        <div style="line-height:1.9;">
                        <b>状態:</b> {status}<br>
                        <b>担当:</b> {assignee}<br>
                        <b>期限:</b> {limit_str}<br>
                        <b>更新:</b> {updated_at}
                        </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                    action_cols = st.columns([1, 1, 4])

                    if status == "未着手":
                        with action_cols[0]:
                            if st.button("開始する", key=f"urgent_start_{task_id}", use_container_width=True):
                                start_task(task_id)
                                st.success("タスクを開始しました！")
                                st.rerun()

                    elif status == "作業中":
                        if user_name == st.session_state.user:
                            with action_cols[1]:
                                if st.button("完了する", key=f"urgent_done_{task_id}", use_container_width=True):
                                    complete_task(task_id)
                                    st.success("タスクを完了しました！")
                                    st.rerun()
                        else:
                            with action_cols[2]:
                                st.caption(f"現在 {user_name} さんが対応中です。")

    show_urgent_page()

# ==========================================
# ⑨ 利用者情報（軽量化版 + 至急アラート連動）
# ==========================================
elif page == "⑨ 利用者情報":
    def show_resident_page():
        if "resident_mode" not in st.session_state:
            st.session_state["resident_mode"] = "利用者一覧"

        st.title("👤 利用者情報")

        current_company_id = str(st.session_state.get("company_id", "")).strip()

        if "selected_resident_id" not in st.session_state:
            st.session_state["selected_resident_id"] = ""

        master_df = get_resident_master_df(current_company_id)

        # ------------------------------------------
        # 一覧モード
        # ------------------------------------------
        if not st.session_state.get("selected_resident_id"):
            reset_resident_edit_flags()

            top_cols = st.columns([1, 1, 3])

            with top_cols[0]:
                if st.button("利用者一覧", use_container_width=True):
                    st.session_state["resident_mode"] = "利用中"
                    st.rerun()

            with top_cols[1]:
                if st.button("退所者一覧", use_container_width=True):
                    st.session_state["resident_mode"] = "退所"
                    st.rerun()

            with top_cols[2]:
                st.caption(f"現在表示: {st.session_state.get('resident_mode', '利用者一覧')}")

            st.divider()

            with st.expander("➕ 新しい利用者を追加する"):
                weekdays = ["月", "火", "水", "木", "金", "土", "日"]

                service_defs = [
                    ("病院", "#F8E7A1"),
                    ("看護", "#CFEAF6"),
                    ("介護", "#DDEDB7"),
                ]

                slot_placeholders = [
                    "例 10:00〜11:00",
                    "2つ目があれば入力",
                    "3つ目があれば入力",
                    "4つ目があれば入力",
                ]

                def parse_add_time_range(raw_text: str):
                    raw = str(raw_text).strip()
                    if not raw:
                        return "", ""

                    raw = raw.replace("～", "〜").replace("~", "〜").replace("-", "〜")
                    if "〜" in raw:
                        start_time, end_time = [x.strip() for x in raw.split("〜", 1)]
                        return start_time, end_time

                    return raw, ""

                st.markdown(
                    """
                    <style>
                    .add-svc-title {
                        display:inline-block;
                        min-width:140px;
                        text-align:center;
                        font-weight:700;
                        font-size:20px;
                        color:#111;
                        border:3px solid #111;
                        padding:8px 18px;
                        margin-top:10px;
                        margin-bottom:12px;
                    }
                    .add-week-head {
                        text-align:center;
                        font-weight:700;
                        border:2px solid #111;
                        padding:7px 0;
                        background:#fafafa;
                        margin-bottom:4px;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True
                )

                with st.form("resident_add_form"):
                    st.markdown("### 基本情報")

                    # 1行目
                    basic1 = st.columns(3)
                    with basic1[0]:
                        resident_name = st.text_input("利用者名", key="add_resident_name")
                    with basic1[1]:
                        status = st.selectbox("状態", ["利用中", "退所"], key="add_status")
 
                    # 2行目
                    basic2 = st.columns(2)
                    with basic2[0]:
                        disability_type = st.selectbox("障害区分", ["精神", "身体"], key="add_disability")
                    with basic2[1]:
                        public_assistance = st.selectbox("生活保護受給", ["あり", "なし"], key="add_pa")

                    # 3行目
                    basic3 = st.columns(2)
                    with basic3[0]:
                        consultant = st.text_input("相談員", key="add_consultant")
                    with basic3[1]:
                        consultant_phone = st.text_input("相談員電話", key="add_consultant_phone")

                    # 4行目
                    basic4 = st.columns(2)
                    with basic4[0]:
                        caseworker = st.text_input("ケースワーカー", key="add_caseworker")
                    with basic4[1]:
                        caseworker_phone = st.text_input("ケースワーカー電話", key="add_caseworker_phone")

                    st.markdown("### 病院・看護・介護の週間予定")
                    st.caption("同じ曜日に2回以上ある場合は、同じ曜日の下の2つ目・3つ目・4つ目にもそのまま入力してください。Enterは不要です。")

                    weekly_inputs = {}

                    for service_name, service_color in service_defs:
                        st.markdown(
                            f'<div class="add-svc-title" style="background:{service_color};">{service_name}</div>',
                            unsafe_allow_html=True
                        )

                        top_info = st.columns([4, 4, 3])
                        with top_info[0]:
                            place_val = st.text_input(
                                f"{service_name}名",
                                key=f"add_{service_name}_place",
                                placeholder=f"{service_name}名"
                            )
                        with top_info[1]:
                            person_val = st.text_input(
                                f"{service_name}担当",
                                key=f"add_{service_name}_person",
                                placeholder="担当"
                            )
                        with top_info[2]:
                            phone_val = st.text_input(
                                f"{service_name}電話",
                                key=f"add_{service_name}_phone",
                                placeholder="電話"
                            )

                        day_cols = st.columns(7)

                        day_values = {}

                        for i, wd in enumerate(weekdays):
                            with day_cols[i]:
                                st.markdown(f'<div class="add-week-head">{wd}</div>', unsafe_allow_html=True)

                                slots = []
                                for slot_idx in range(4):
                                    slot_val = st.text_input(
                                        f"{service_name}_{wd}_{slot_idx+1}",
                                        key=f"add_{service_name}_{wd}_{slot_idx+1}",
                                        label_visibility="collapsed",
                                        placeholder=slot_placeholders[slot_idx]
                                    )
                                    slots.append(slot_val)

                                day_values[wd] = slots

                        weekly_inputs[service_name] = {
                            "place": str(place_val).strip(),
                            "person_in_charge": str(person_val).strip(),
                            "phone": str(phone_val).strip(),
                            "days": day_values
                        }

                    submit_cols = st.columns(2)
                    with submit_cols[0]:
                        add_resident = st.form_submit_button("利用者を登録する", use_container_width=True)
                    with submit_cols[1]:
                        cancel_add = st.form_submit_button("入力をやめる", use_container_width=True)

                    if add_resident:
                        if resident_name.strip():
                            next_resident_id = get_next_resident_id(master_df)
                            now_str = now_jst().strftime("%Y-%m-%d %H:%M")

                            new_master_row = pd.DataFrame([{
                                "company_id": current_company_id,
                                "resident_id": next_resident_id,
                                "resident_name": resident_name.strip(),
                                "status": status,
                                "disability_type": disability_type,
                                "public_assistance": public_assistance,
                                "consultant": consultant.strip(),
                                "consultant_phone": consultant_phone.strip(),
                                "caseworker": caseworker.strip(),
                                "caseworker_phone": caseworker_phone.strip(),
                                "hospital": weekly_inputs["病院"]["place"],
                                "hospital_phone": weekly_inputs["病院"]["phone"],
                                "nurse": weekly_inputs["看護"]["place"],
                                "nurse_phone": weekly_inputs["看護"]["phone"],
                                "care": weekly_inputs["介護"]["place"],
                                "care_phone": weekly_inputs["介護"]["phone"],
                                "created_at": now_str,
                                "updated_at": now_str
                            }])
                        
                            all_master_df = load_db("resident_master")
                            if all_master_df is None or all_master_df.empty:
                                all_master_df = pd.DataFrame(columns=[
                                    "company_id",
                                    "resident_id", "resident_name", "status",
                                    "disability_type",
                                    "public_assistance",
                                    "consultant", "consultant_phone",
                                    "caseworker", "caseworker_phone",
                                    "hospital", "hospital_phone",
                                    "nurse", "nurse_phone",
                                    "care", "care_phone",
                                    "created_at", "updated_at"
                                ])
                            else:
                                all_master_df = all_master_df.fillna("").copy()
                                for col in [
                                    "company_id",
                                    "resident_id", "resident_name", "status",
                                    "disability_type",
                                    "public_assistance",
                                    "consultant", "consultant_phone",
                                    "caseworker", "caseworker_phone",
                                    "hospital", "hospital_phone",
                                    "nurse", "nurse_phone",
                                    "care", "care_phone",
                                    "created_at", "updated_at"
                                ]:
                                    if col not in all_master_df.columns:
                                        all_master_df[col] = ""

                            new_master_df = pd.concat([all_master_df, new_master_row], ignore_index=True)
                            save_db(new_master_df, "resident_master")

                            schedule_df_add = load_db("resident_schedule")
                            if schedule_df_add is None or schedule_df_add.empty:
                                schedule_df_add = pd.DataFrame(columns=[
                                    "company_id",
                                    "id", "resident_id", "weekday", "service_type",
                                    "start_time", "end_time", "place", "phone", "person_in_charge", "memo"
                                ])
                            else:
                                schedule_df_add = schedule_df_add.fillna("").copy()
                                for col in [
                                    "company_id",
                                    "id", "resident_id", "weekday", "service_type",
                                    "start_time", "end_time", "place", "phone", "person_in_charge", "memo"
                                ]:
                                    if col not in schedule_df_add.columns:
                                        schedule_df_add[col] = ""

                            next_schedule_id = get_next_numeric_id(schedule_df_add, "id", 1)
                            new_schedule_rows = []

                            for service_type, data in weekly_inputs.items():
                                place_name = str(data["place"]).strip()
                                phone_text = str(data["phone"]).strip()
                                person_text = str(data["person_in_charge"]).strip()

                                for wd, slot_list in data["days"].items():
                                    for slot_index, raw_time in enumerate(slot_list, start=1):
                                        start_time, end_time = parse_add_time_range(raw_time)
                                        if not start_time and not end_time:
                                            continue

                                        new_schedule_rows.append({
                                            "company_id": current_company_id,
                                            "id": next_schedule_id,
                                            "resident_id": next_resident_id,
                                            "weekday": wd,
                                            "service_type": service_type,
                                            "start_time": start_time,
                                            "end_time": end_time,
                                            "place": place_name,
                                            "phone": phone_text,
                                            "person_in_charge": person_text,
                                            "memo": f"slot:{slot_index}"
                                        })
                                        next_schedule_id += 1

                            if new_schedule_rows:
                                add_schedule_df = pd.DataFrame(new_schedule_rows)
                                merged_schedule_df = pd.concat([schedule_df_add, add_schedule_df], ignore_index=True)
                                save_db(merged_schedule_df.fillna(""), "resident_schedule")

                            st.success("利用者を登録しました！")
                            st.rerun()
                        else:
                            st.error("利用者名を入力してください。")

                    if cancel_add:
                        st.rerun()

            st.divider()

            search_word = st.text_input("名前検索", placeholder="利用者名を入力")

            list_df = master_df.copy()
            list_df = list_df[
                list_df["status"].astype(str).str.strip() == st.session_state.get("resident_mode", "利用者一覧")
            ].copy()

            if search_word.strip():
                list_df = list_df[
                    list_df["resident_name"].astype(str).str.contains(search_word.strip(), case=False, na=False)
                ].copy()

            if not list_df.empty:
                list_df = list_df.sort_values("resident_name")

            if list_df.empty:
                st.info("該当する利用者はいありません。")
                return

            cols = st.columns(2)

            for i, (_, row) in enumerate(list_df.iterrows()):
                with cols[i % 2]:
                    company_id = str(row.get("company_id", "")).strip()
                    resident_id = str(row.get("resident_id", "")).strip()
                    resident_name = str(row.get("resident_name", "")).strip()
                    status = str(row.get("status", "")).strip()
                    consultant = str(row.get("consultant", "")).strip()

                    if status == "利用中":
                        status_label = "🟢 利用中"
                        bg = "#eefbf0"
                        border = "#2ecc71"
                    else:
                        status_label = "⚫ 退所"
                        bg = "#f3f4f6"
                        border = "#7f8c8d"

                    with st.container(border=True):
                        st.markdown(
                            f"""
                            <div style="
                                border-left: 8px solid {border};
                                background-color: {bg};
                                padding: 14px;
                                border-radius: 10px;
                                min-height: 140px;
                                margin-bottom: 10px;
                            ">
                                <div style="font-size: 20px; font-weight: 700; margin-bottom: 8px;">
                                    {resident_name}
                                </div>
                                <div style="margin-bottom: 6px;">{status_label}</div>
                                <div>相談員: {consultant if consultant else '未登録'}</div>
                                <div>ID: {resident_id}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        if st.button(
                            "詳細を見る",
                            key=f"open_resident_{company_id}_{resident_id}_{i}",
                            use_container_width=True
                        ):
                            st.session_state.selected_resident_id = resident_id
                            reset_resident_edit_flags()
                            st.rerun()
            return

        # ------------------------------------------
        # 個人詳細モード
        # ------------------------------------------
        selected_id = str(st.session_state.selected_resident_id).strip()

        detail_df = master_df[master_df["resident_id"].astype(str) == selected_id].copy()
        if detail_df.empty:
            st.warning("利用者情報が見つからありません。")
            if st.button("一覧に戻る", use_container_width=True):
                st.session_state.selected_resident_id = ""
                reset_resident_edit_flags()
                st.rerun()
            return

        row = detail_df.iloc[0]
        resident_name = str(row.get("resident_name", "")).strip()
        status = str(row.get("status", "")).strip()

        schedule_df = get_resident_schedule_df()
        notes_df = get_resident_notes_df()

        company_id = get_current_company_id()
        task_df_detail = get_tasks_df(company_id)

        back_cols = st.columns([1, 5])
        with back_cols[0]:
            if st.button("← 一覧に戻る", use_container_width=True):
                st.session_state.selected_resident_id = ""
                reset_resident_edit_flags()
                st.rerun()

        status_label = "🟢 利用中" if status == "利用中" else "⚫ 退所"

        st.subheader(f"{resident_name} 様")
        st.caption(f"{status_label} / ID: {selected_id}")

        st.divider()
        st.markdown("### ☎️ 関係者・外部連携")

        contact_cards_df = get_resident_contact_cards(selected_id)

        if contact_cards_df.empty:
            st.info("この利用者に紐づく関係者はまだ登録されていません。")
        else:
            category_icon_map = {
                "医療": "🏥",
                "外部連携": "🤝",
                "家族": "👪",
                "業者": "🛠️",
            }

            display_order = ["医療", "外部連携", "家族", "業者"]

            for cat1 in display_order:
                cat_df = contact_cards_df[
                    contact_cards_df["category1"].astype(str).str.strip() == cat1
                ].copy()

                if cat_df.empty:
                    continue

                icon = category_icon_map.get(cat1, "📌")
                st.markdown(f"#### {icon} {cat1}")

                for _, crow in cat_df.iterrows():
                    role = str(crow.get("role", "")).strip()
                    category2 = str(crow.get("category2", "")).strip()
                    name = str(crow.get("name", "")).strip()
                    org = str(crow.get("organization", "")).strip()
                    phone = str(crow.get("phone", "")).strip()
                    memo = str(crow.get("memo", "")).strip()
                    contact_id = str(crow.get("contact_id", "")).strip()

                    title_text = role if role else category2

                    with st.container(border=True):
                        st.markdown(
                            f"""
                            <div style="
                                border-left: 6px solid #9aa5b1;
                                background:#fafafa;
                                padding:12px 14px;
                                border-radius:10px;
                                margin-bottom:10px;
                            ">
                                <div style="font-size:18px; font-weight:700; margin-bottom:6px;">
                                    {title_text}
                                </div>
                                <div style="line-height:1.8;">
                                    <b>氏名:</b> {name}<br>
                                    <b>事業所:</b> {org}<br>
                                    <b>電話:</b> {phone}<br>
                                    <b>分類:</b> {category2}<br>
                                    <b>ID:</b> {contact_id}
                                </div>
                                <div style="margin-top:8px; color:#555;">
                                    {memo}
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )        

        # ------------------------------------------
        # この人の至急アラート
        # ------------------------------------------
        st.markdown("### 🚨 この人のために今すぐやること")

        urgent_person_df = task_df_detail[
            task_df_detail["priority"].astype(str).str.strip().isin(["至急", "重要"]) &
            (task_df_detail["status"].astype(str).str.strip() != "完了") &
            task_df_detail["task"].astype(str).str.contains(resident_name, case=False, na=False)
        ].copy()

        if not urgent_person_df.empty:
            prio_map = {"至急": 0, "重要": 1}
            urgent_person_df["prio_sort"] = urgent_person_df["priority"].map(prio_map).fillna(9)
            urgent_person_df["limit_sort"] = pd.to_datetime(urgent_person_df["limit"], errors="coerce")
            urgent_person_df = urgent_person_df.sort_values(
                ["prio_sort", "limit_sort", "updated_at"],
                ascending=[True, True, False]
            )

            for _, trow in urgent_person_df.iterrows():
                t_id = str(trow.get("id", "")).strip()
                t_name = str(trow.get("task", "")).strip()
                t_priority = str(trow.get("priority", "")).strip()
                t_status = str(trow.get("status", "")).strip()
                t_user = str(trow.get("user", "")).strip()
                t_limit = str(trow.get("limit", "")).strip()
                t_updated = str(trow.get("updated_at", "")).strip()

                limit_date = pd.to_datetime(t_limit, errors="coerce")
                today = now_jst().date()

                if pd.notna(limit_date) and limit_date.date() < today:
                    icon = "🩸"
                    border_color = "#d63031"
                    bg_color = "#ffeaea"
                elif t_priority == "至急":
                    icon = "🚨"
                    border_color = "#ff4d4f"
                    bg_color = "#fff1f0"
                else:
                    icon = "⚠️"
                    border_color = "#ff9f43"
                    bg_color = "#fff7e6"

                assignee = t_user if t_user else "未割当"

                with st.container(border=True):
                    st.markdown(
                        f"""
                        <div style="
                            border-left: 8px solid {border_color};
                            background-color: {bg_color};
                            padding: 14px;
                            border-radius: 10px;
                            margin-bottom: 10px;
                        ">
                            <div style="font-size:18px; font-weight:700; margin-bottom:6px;">
                                {icon} {t_priority} - {t_name}
                            </div>
                            <div style="line-height:1.8;">
                                <b>状態:</b> {t_status}<br>
                                <b>担当:</b> {assignee}<br>
                                <b>期限:</b> {t_limit}<br>
                                <b>更新:</b> {t_updated}
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                    btn_cols = st.columns([1, 1, 4])

                    if t_status == "未着手":
                        with btn_cols[0]:
                            if st.button("開始する", key=f"resident_urgent_start_{t_id}", use_container_width=True):
                                start_task(t_id)
                                st.success("タスクを開始しました！")
                                st.rerun()

                    elif t_status == "作業中":
                        if t_user == st.session_state.user:
                            with btn_cols[1]:
                                if st.button("完了する", key=f"resident_urgent_done_{t_id}", use_container_width=True):
                                    complete_task(t_id)
                                    st.success("タスクを完了しました！")
                                    st.rerun()
                        else:
                            with btn_cols[2]:
                                st.caption(f"現在 {t_user} さんが対応中です。")
        else:
            st.info("この利用者に連動した至急・重要タスクは今のところありません。")

        st.divider()
        st.markdown("### 基本情報")

        info_cols = st.columns(2)
        with info_cols[0]:
            st.write(f"**相談員**: {row.get('consultant', '')}")
            st.write(f"**相談員電話**: {row.get('consultant_phone', '')}")
            st.write(f"**ケースワーカー**: {row.get('caseworker', '')}")
            st.write(f"**ケースワーカー電話**: {row.get('caseworker_phone', '')}")
            st.write(f"**病院**: {row.get('hospital', '')}")
            st.write(f"**病院電話**: {row.get('hospital_phone', '')}")

        with info_cols[1]:
            st.write(f"**看護**: {row.get('nurse', '')}")
            st.write(f"**看護電話**: {row.get('nurse_phone', '')}")
            st.write(f"**介護**: {row.get('care', '')}")
            st.write(f"**介護電話**: {row.get('care_phone', '')}")
            st.write(f"**登録日**: {row.get('created_at', '')}")
            st.write(f"**更新日**: {row.get('updated_at', '')}")

        st.divider()
        st.markdown("### 週間予定")

        schedule_view = schedule_df[schedule_df["resident_id"].astype(str) == selected_id].copy()
        schedule_view = schedule_view[
            schedule_view["service_type"].astype(str).isin(["病院", "看護", "介護"])
        ].copy()

        if schedule_view.empty:
            st.info("週間予定はまだ登録されていません。")
        else:
            render_resident_schedule_html(schedule_view)

        st.divider()
        st.markdown("### 共有メモ")

        notes_view = notes_df[notes_df["resident_id"].astype(str) == selected_id].copy()
        if not notes_view.empty:
            try:
                notes_view = notes_view.sort_values("date", ascending=False)
            except Exception:
                pass

            for _, note_row in notes_view.iterrows():
                with st.container(border=True):
                    st.write(f"**{note_row.get('date', '')}**  {note_row.get('user', '')}")
                    st.write(note_row.get("note", ""))
        else:
            st.info("共有メモはまだありません。")


        st.divider()
        st.markdown("### ➕ 関係者を登録する")

        with st.expander("新しい関係者を追加"):
            with st.form(f"resident_contact_add_form_{selected_id}"):
                form_col1, form_col2 = st.columns(2)

                with form_col1:
                    category1 = st.selectbox(
                        "大分類",
                        ["医療", "外部連携", "家族", "業者"],
                        key=f"contact_cat1_{selected_id}"
                    )

                    category2_map = {
                        "医療": ["主治医", "訪問看護", "薬局", "行政", "その他"],
                        "外部連携": ["ケアマネ", "相談員", "行政", "その他"],
                        "家族": ["家族", "成年後見人", "身元引受人", "その他"],
                        "業者": ["配食", "福祉用具", "修理", "その他"],
                    }

                    category2 = st.selectbox(
                        "分類",
                        category2_map.get(category1, ["その他"]),
                        key=f"contact_cat2_{selected_id}"
                    )

                    role = st.text_input(
                        "この利用者に対する役割",
                        value=category2,
                        key=f"contact_role_{selected_id}"
                    )

                    name = st.text_input("氏名", key=f"contact_name_{selected_id}")

                with form_col2:
                    organization = st.text_input("事業所名", key=f"contact_org_{selected_id}")
                    phone = st.text_input("電話番号", key=f"contact_phone_{selected_id}")
                    memo = st.text_area("メモ", key=f"contact_memo_{selected_id}")

                save_contact = st.form_submit_button("関係者を登録する", use_container_width=True)

                if save_contact:
                    if name.strip() or organization.strip():
                        contacts_df = get_external_contacts_df()
                        links_df = get_resident_links_df()

                        next_contact_id = get_next_contact_id(contacts_df)
                        next_link_id = get_next_numeric_id(links_df, "id", 1)

                        new_contact_row = pd.DataFrame([{
                            "contact_id": next_contact_id,
                            "category1": category1,
                            "category2": category2,
                            "name": name.strip(),
                            "organization": organization.strip(),
                            "phone": phone.strip(),
                            "memo": memo.strip(),
                        }])

                        new_contacts_df = pd.concat([contacts_df, new_contact_row], ignore_index=True)
                        save_db(new_contacts_df, "external_contacts")

                        new_link_row = pd.DataFrame([{
                            "id": next_link_id,
                            "resident_id": selected_id,
                            "contact_id": next_contact_id,
                            "role": role.strip(),
                        }])

                        new_links_df = pd.concat([links_df, new_link_row], ignore_index=True)
                        save_db(new_links_df, "resident_links")

                        st.success("関係者を登録しました！")
                        st.rerun()
                    else:
                        st.error("氏名か事業所名のどちらかは入れてください。")


        st.divider()
        st.markdown("### 編集・追加")

        edit_cols = st.columns(3)
        with edit_cols[0]:
            if st.button("基本情報を編集", key=f"edit_basic_{selected_id}", use_container_width=True):
                st.session_state.edit_resident_basic = True
                st.session_state.edit_resident_schedule = False
                st.session_state.edit_resident_note = False
                st.rerun()

        with edit_cols[1]:
            if st.button("予定を追加", key=f"edit_schedule_{selected_id}", use_container_width=True):
                st.session_state.edit_resident_basic = False
                st.session_state.edit_resident_schedule = True
                st.session_state.edit_resident_note = False
                st.rerun()

        with edit_cols[2]:
            if st.button("メモを追加", key=f"edit_note_{selected_id}", use_container_width=True):
                st.session_state.edit_resident_basic = False
                st.session_state.edit_resident_schedule = False
                st.session_state.edit_resident_note = True
                st.rerun()

        # ------------------------------------------
        # 基本情報編集
        # ------------------------------------------
        if st.session_state.get("edit_resident_basic", False):
            st.divider()
            st.markdown("#### 基本情報を編集")

            with st.form(f"resident_basic_form_{selected_id}"):
                resident_name = st.text_input("利用者名", value=str(row.get("resident_name", "")))
                status = st.selectbox(
                    "状態",
                    ["利用中", "退所"],
                    index=0 if str(row.get("status", "")).strip() == "利用中" else 1
                )
                consultant = st.text_input("相談員", value=str(row.get("consultant", "")))
                consultant_phone = st.text_input("相談員電話", value=str(row.get("consultant_phone", "")))
                caseworker = st.text_input("ケースワーカー", value=str(row.get("caseworker", "")))
                caseworker_phone = st.text_input("ケースワーカー電話", value=str(row.get("caseworker_phone", "")))
                hospital = st.text_input("病院", value=str(row.get("hospital", "")))
                hospital_phone = st.text_input("病院電話", value=str(row.get("hospital_phone", "")))
                nurse = st.text_input("看護", value=str(row.get("nurse", "")))
                nurse_phone = st.text_input("看護電話", value=str(row.get("nurse_phone", "")))
                care = st.text_input("介護", value=str(row.get("care", "")))
                care_phone = st.text_input("介護電話", value=str(row.get("care_phone", "")))

                save_col1, save_col2 = st.columns(2)
                with save_col1:
                    save_basic = st.form_submit_button("基本情報を保存する", use_container_width=True)
                with save_col2:
                    cancel_basic = st.form_submit_button("キャンセル", use_container_width=True)

                if save_basic:
                    update_df = master_df.copy()
                    now_str = now_jst().strftime("%Y-%m-%d %H:%M")

                    target_mask = update_df["resident_id"].astype(str) == selected_id
                    update_df.loc[target_mask, "resident_name"] = resident_name.strip()
                    update_df.loc[target_mask, "status"] = status
                    update_df.loc[target_mask, "consultant"] = consultant.strip()
                    update_df.loc[target_mask, "consultant_phone"] = consultant_phone.strip()
                    update_df.loc[target_mask, "caseworker"] = caseworker.strip()
                    update_df.loc[target_mask, "caseworker_phone"] = caseworker_phone.strip()
                    update_df.loc[target_mask, "hospital"] = hospital.strip()
                    update_df.loc[target_mask, "hospital_phone"] = hospital_phone.strip()
                    update_df.loc[target_mask, "nurse"] = nurse.strip()
                    update_df.loc[target_mask, "nurse_phone"] = nurse_phone.strip()
                    update_df.loc[target_mask, "care"] = care.strip()
                    update_df.loc[target_mask, "care_phone"] = care_phone.strip()
                    update_df.loc[target_mask, "updated_at"] = now_str

                    save_db(update_df, "resident_master")
                    st.session_state.edit_resident_basic = False
                    st.success("基本情報を保存しました！")
                    st.rerun()

                if cancel_basic:
                    st.session_state.edit_resident_basic = False
                    st.rerun()

        # ------------------------------------------
        # 週間予定編集
        # ------------------------------------------
        if st.session_state.get("edit_resident_schedule", False):
            st.divider()
            st.markdown("#### 病院・看護・介護の週間予定を編集")
            st.caption("同じ曜日に2回以上ある場合は、同じ曜日の下の2つ目・3つ目・4つ目にもそのまま入力してください。Enterは不要です。")

            schedule_base = build_schedule_form_base(schedule_df, selected_id)
            weekdays = ["月", "火", "水", "木", "金", "土", "日"]
            service_types = ["病院", "看護", "介護"]

            color_map = {
                "病院": "#F8E7A1",
                "看護": "#CFEAF6",
                "介護": "#DDEDB7",
            }

            slot_placeholders = [
                "例 10:00〜11:00",
                "2つ目があれば入力",
                "3つ目があれば入力",
                "4つ目があれば入力",
            ]

            with st.form(f"resident_schedule_form_{selected_id}"):
                weekly_inputs = {}

                for service in service_types:
                    st.markdown(
                        f"""
                        <div style="
                            background:{color_map[service]};
                            border:2px solid #111;
                            padding:8px 16px;
                            font-weight:700;
                            display:inline-block;
                            min-width:120px;
                            text-align:center;
                            margin-top:8px;
                            margin-bottom:10px;
                        ">{service}</div>
                        """,
                        unsafe_allow_html=True
                    )

                    top_cols = st.columns([3, 3, 3])
                    with top_cols[0]:
                        place_val = st.text_input(
                            f"{service}名",
                            value=schedule_base[service]["place"],
                            key=f"{selected_id}_{service}_place"
                        )
                    with top_cols[1]:
                        phone_val = st.text_input(
                            f"{service}電話",
                            value=schedule_base[service]["phone"],
                            key=f"{selected_id}_{service}_phone"
                        )
                    with top_cols[2]:
                        person_val = st.text_input(
                            f"{service}担当",
                            value=schedule_base[service]["person_in_charge"],
                            key=f"{selected_id}_{service}_person"
                        )

                    st.markdown(
                        """
                        <style>
                        .weekly-head {
                            text-align:center;
                            font-weight:700;
                            padding:6px 0;
                            border:2px solid #111;
                            background:#fafafa;
                            margin-bottom:4px;
                        }
                        </style>
                        """,
                        unsafe_allow_html=True
                    )

                    day_values = {}
                    day_cols = st.columns(7)

                    for i, wd in enumerate(weekdays):
                        with day_cols[i]:
                            st.markdown(f'<div class="weekly-head">{wd}</div>', unsafe_allow_html=True)

                            slots = []
                            for slot_idx in range(4):
                                slot_val = schedule_base[service]["days"][wd][slot_idx]
                                new_val = st.text_input(
                                    f"{wd}{slot_idx+1}枠",
                                    value=slot_val,
                                    key=f"{selected_id}_{service}_{wd}_{slot_idx+1}",
                                    label_visibility="collapsed",
                                    placeholder=slot_placeholders[slot_idx]
                                )
                                slots.append(new_val)

                            day_values[wd] = slots

                    weekly_inputs[service] = {
                        "place": str(place_val).strip(),
                        "phone": str(phone_val).strip(),
                        "person_in_charge": str(person_val).strip(),
                        "days": day_values
                    }

                    st.markdown("<br>", unsafe_allow_html=True)

                save_col1, save_col2 = st.columns(2)
                with save_col1:
                    save_weekly = st.form_submit_button("週間予定を保存する", use_container_width=True)
                with save_col2:
                    cancel_weekly = st.form_submit_button("キャンセル", use_container_width=True)

                if save_weekly:
                    keep_df = schedule_df[
                        ~(
                            (schedule_df["resident_id"].astype(str) == selected_id) &
                            (schedule_df["service_type"].astype(str).isin(["病院", "看護", "介護"]))
                        )
                    ].copy()

                    next_id = get_next_numeric_id(schedule_df, "id", 1)
                    new_rows = []

                    for service_type, data in weekly_inputs.items():
                        place_name = data["place"]
                        phone_text = data["phone"]
                        person_text = data["person_in_charge"]

                        for wd, slot_list in data["days"].items():
                            for slot_index, time_range in enumerate(slot_list, start=1):
                                start_time, end_time = parse_time_range(time_range)
                                if not start_time and not end_time:
                                    continue

                                new_rows.append({
                                    "id": next_id,
                                    "resident_id": selected_id,
                                    "weekday": wd,
                                    "service_type": service_type,
                                    "start_time": start_time,
                                    "end_time": end_time,
                                    "place": place_name,
                                    "phone": phone_text,
                                    "person_in_charge": person_text,
                                    "memo": f"slot:{slot_index}"
                                })
                                next_id += 1

                    if new_rows:
                        add_df = pd.DataFrame(new_rows)
                        save_df = pd.concat([keep_df, add_df], ignore_index=True)
                    else:
                        save_df = keep_df.copy()

                    save_df = save_df.fillna("")
                    save_db(save_df, "resident_schedule")
                    st.session_state.edit_resident_schedule = False
                    st.success("週間予定を保存しました！")
                    st.rerun()

                if cancel_weekly:
                    st.session_state.edit_resident_schedule = False
                    st.rerun()

            st.markdown("#### 現在の週間予定")

            current_view = schedule_df[schedule_df["resident_id"].astype(str) == selected_id].copy()
            current_view = current_view[
                current_view["service_type"].astype(str).isin(["病院", "看護", "介護"])
            ].copy()

            if current_view.empty:
                st.info("週間予定はまだ登録されていません。")
            else:
                render_resident_schedule_html(current_view)

            delete_target = schedule_df[
                (schedule_df["resident_id"].astype(str) == selected_id) &
                (schedule_df["service_type"].astype(str).isin(["病院", "看護", "介護"]))
            ].copy()

            if not delete_target.empty:
                if st.button("週間予定をすべて削除", key=f"delete_weekly_schedule_{selected_id}", use_container_width=True):
                    new_schedule_df = schedule_df[
                        ~(
                            (schedule_df["resident_id"].astype(str) == selected_id) &
                            (schedule_df["service_type"].astype(str).isin(["病院", "看護", "介護"]))
                        )
                    ].copy()
                    save_db(new_schedule_df, "resident_schedule")
                    st.success("週間予定を削除しました。")
                    st.rerun()

        # ------------------------------------------
        # メモ追加
        # ------------------------------------------
        if st.session_state.get("edit_resident_note", False):
            st.divider()
            st.markdown("#### 共有メモを追加")

            with st.form(f"resident_note_form_{selected_id}"):
                note_date = st.date_input("日付", value=now_jst().date())
                note_text = st.text_area("共有メモ")

                save_col1, save_col2 = st.columns(2)
                with save_col1:
                    add_note = st.form_submit_button("メモを追加する", use_container_width=True)
                with save_col2:
                    cancel_note = st.form_submit_button("キャンセル", use_container_width=True)

                if add_note:
                    if note_text.strip():
                        next_id = get_next_numeric_id(notes_df, "id", 1)
                        new_row = pd.DataFrame([{
                            "id": next_id,
                            "resident_id": selected_id,
                            "date": str(note_date),
                            "user": str(st.session_state.get("user", "")),
                            "note": note_text.strip()
                        }])

                        new_notes_df = pd.concat([notes_df, new_row], ignore_index=True)
                        save_db(new_notes_df, "resident_notes")
                        st.session_state.edit_resident_note = False
                        st.success("共有メモを追加しました！")
                        st.rerun()
                    else:
                        st.error("メモ内容を入力してください。")

                if cancel_note:
                    st.session_state.edit_resident_note = False
                    st.rerun()

            notes_delete_df = notes_df[
                notes_df["resident_id"].astype(str) == selected_id
            ].copy()

            if not notes_delete_df.empty:
                try:
                    notes_delete_df = notes_delete_df.sort_values("date", ascending=False)
                except Exception:
                    pass

                st.caption("登録済みメモを削除する場合は下から選んでください。")

                for _, nrow in notes_delete_df.iterrows():
                    nid = str(nrow.get("id", "")).strip()

                    with st.container(border=True):
                        st.write(f"**{nrow.get('date', '')}**  {nrow.get('user', '')}")
                        st.write(nrow.get("note", ""))

                        if st.button(
                            "このメモを削除",
                            key=f"delete_note_{selected_id}_{nid}",
                            use_container_width=True
                        ):
                            latest_notes_df = get_resident_notes_df()
                            new_notes_df = latest_notes_df[
                                latest_notes_df["id"].astype(str) != nid
                            ].copy()

                            save_db(new_notes_df, "resident_notes")
                            st.success("メモを削除しました。")
                            st.rerun()

                            links_df = get_resident_links_df()
                            new_links_df = links_df[
                                ~(
                                    (links_df["resident_id"].astype(str) == str(selected_id)) &
                                    (links_df["contact_id"].astype(str) == str(contact_id))
                                )
                            ].copy()
                            save_db(new_links_df, "resident_links")
                            st.success("この利用者との紐づきを削除しました。")
                            st.rerun()

    show_resident_page()
                            

elif page == "⓪ 検索":

    st.title("🔍 検索")
    st.write("利用者・関係者・資料をまとめて探せるページです。")

    # ------------------------------------------
    # 書類検索
    # ------------------------------------------
    st.markdown("## 📁 書類検索")

    CATEGORY1_OPTIONS = [
        "全部",
        "運営関連",
        "外部連携",
        "その他",
    ]

    CATEGORY2_MAP = {
        "全部": ["全部"],
        "運営関連": [
            "全部",
            "マニュアル",
            "帳票",
            "研修",
            "行政提出",
            "その他",
        ],
        "外部連携": [
            "全部",
            "病院",
            "訪問看護",
            "ケアマネ",
            "薬局",
            "行政",
            "その他",
        ],
        "その他": [
            "全部",
            "その他",
        ],
    }

    doc_df = load_db("document_master")

    if doc_df is None or doc_df.empty:
        doc_df = pd.DataFrame(columns=[
            "document_id", "category1", "category2", "category3",
            "title", "file_type", "url", "summary", "memo",
            "status", "updated_at", "created_at"
        ])
    else:
        for col in [
            "document_id", "category1", "category2", "category3",
            "title", "file_type", "url", "summary", "memo",
            "status", "updated_at", "created_at"
        ]:
            if col not in doc_df.columns:
                doc_df[col] = ""

    doc_df = doc_df.fillna("").copy()

    # 書類アップロード・倉庫の簡易検索
    shared_keyword = st.text_input(
        "書類アップロード・倉庫を検索",
        key="shared_doc_search"
    )

    if shared_keyword.strip():
        result_df = search_shared_documents(shared_keyword)

        if result_df.empty:
            st.info("該当する資料がありません。")
        else:
            st.markdown(f"### 倉庫・アップロード検索結果（{len(result_df)}件）")
            for _, row in result_df.iterrows():
                st.markdown("---")
                st.markdown(f"**{row['title']}**")
                st.caption(f"{row['source']} / {row['id']} / {row['file_name']}")
                st.write(row["description"])
                st.caption(f"カテゴリ: {row['category_main']} / {row['category_sub']}")
                if row["source"] == "倉庫":
                    st.caption(f"公開設定: {row['visibility_type']}")

    search_cols = st.columns([2, 2, 2, 3])

    with search_cols[0]:
        cat1 = st.selectbox("カテゴリ1", CATEGORY1_OPTIONS, key="doc_search_cat1")

    with search_cols[1]:
        cat2_options = CATEGORY2_MAP.get(cat1, ["全部"])
        cat2 = st.selectbox("カテゴリ2", cat2_options, key="doc_search_cat2")

    with search_cols[2]:
        status_candidates = ["全部"]
        if not doc_df.empty:
            status_values = sorted([
                x for x in doc_df["status"].astype(str).unique().tolist()
                if str(x).strip()
            ])
            status_candidates += status_values
        status_filter = st.selectbox("状態", status_candidates, key="doc_search_status")

    with search_cols[3]:
        kw = st.text_input(
            "キーワード",
            key="doc_search_kw",
            placeholder="タイトル・概要・メモで検索"
        )

    view_df = doc_df.copy()

    if cat1 != "全部":
        view_df = view_df[view_df["category1"].astype(str) == cat1]

    if cat2 != "全部":
        view_df = view_df[view_df["category2"].astype(str) == cat2]

    if status_filter != "全部":
        view_df = view_df[view_df["status"].astype(str) == status_filter]

    if kw.strip():
        kw_l = kw.strip().lower()
        view_df = view_df[
            view_df.apply(
                lambda row:
                    kw_l in str(row.get("title", "")).lower()
                    or kw_l in str(row.get("summary", "")).lower()
                    or kw_l in str(row.get("memo", "")).lower()
                    or kw_l in str(row.get("category1", "")).lower()
                    or kw_l in str(row.get("category2", "")).lower()
                    or kw_l in str(row.get("category3", "")).lower(),
                axis=1
            )
        ]

    if view_df.empty:
        st.info("条件に合う資料はありません。")
    else:
        try:
            view_df = view_df.sort_values("updated_at", ascending=False)
        except Exception:
            pass

        for _, row in view_df.iterrows():
            document_id = str(row.get("document_id", "")).strip()
            category1 = str(row.get("category1", "")).strip()
            category2 = str(row.get("category2", "")).strip()
            category3 = str(row.get("category3", "")).strip()
            title = str(row.get("title", "")).strip()
            file_type = str(row.get("file_type", "")).strip()
            summary = str(row.get("summary", "")).strip()
            memo = str(row.get("memo", "")).strip()
            status = str(row.get("status", "")).strip()

            with st.container(border=True):
                st.markdown(f"**{title if title else '無題資料'}**")
                st.write(f"{category1} / {category2} / {category3}")
                st.write(f"種類: {file_type} / 状態: {status}")
                if summary:
                    st.write(f"概要: {summary}")
                if memo:
                    st.write(f"メモ: {memo}")

                file_bytes, filename, mime = get_download_file_data(row)
                if file_bytes:
                    st.download_button(
                        "ダウンロード",
                        data=file_bytes,
                        file_name=filename,
                        mime=mime,
                        key=f"download_doc_{document_id}",
                        use_container_width=True
                    )

    st.divider()

    # ------------------------------------------
    # 関係者検索
    # ------------------------------------------
    st.markdown("## ☎️ 関係者検索")

    contacts_df = get_external_contacts_df()
    links_df = get_resident_links_df()
    company_id = get_current_company_id()
    master_df = get_resident_master_df(company_id)

    if contacts_df.empty:
        st.info("関係者データはまだ登録されていません。")
    else:
        for col in ["contact_id", "category1", "category2", "name", "organization", "phone", "memo"]:
            if col not in contacts_df.columns:
                contacts_df[col] = ""

        contacts_df = contacts_df.fillna("").copy()

        search_cols = st.columns([2, 2, 3])

        with search_cols[0]:
            contact_cat1 = st.selectbox(
                "大分類で絞る",
                ["全部", "医療", "外部連携", "家族", "業者"],
                key="contact_search_cat1"
            )

        with search_cols[1]:
            contact_cat2 = st.selectbox(
                "小分類で絞る",
                ["全部", "病院", "訪問看護", "ケアマネ", "相談員", "家族", "その他"],
                key="contact_search_cat2"
            )

        with search_cols[2]:
            contact_kw = st.text_input(
                "関係者キーワード",
                key="contact_search_kw",
                placeholder="氏名・所属・電話番号など"
            )

        contact_view_df = contacts_df.copy()

        if contact_cat1 != "全部":
            contact_view_df = contact_view_df[
                contact_view_df["category1"].astype(str) == contact_cat1
            ]

        if contact_cat2 != "全部":
            contact_view_df = contact_view_df[
                contact_view_df["category2"].astype(str) == contact_cat2
            ]

        if contact_kw.strip():
            kw2 = contact_kw.strip().lower()
            contact_view_df = contact_view_df[
                contact_view_df.apply(
                    lambda row:
                        kw2 in str(row.get("name", "")).lower()
                        or kw2 in str(row.get("organization", "")).lower()
                        or kw2 in str(row.get("phone", "")).lower()
                        or kw2 in str(row.get("memo", "")).lower()
                        or kw2 in str(row.get("category1", "")).lower()
                        or kw2 in str(row.get("category2", "")).lower(),
                    axis=1
                )
            ]

        if contact_view_df.empty:
            st.info("該当する関係者はありません。")
        else:
            for _, row in contact_view_df.iterrows():
                contact_id = str(row.get("contact_id", "")).strip()
                name = str(row.get("name", "")).strip()
                organization = str(row.get("organization", "")).strip()
                phone = str(row.get("phone", "")).strip()
                memo = str(row.get("memo", "")).strip()
                category1 = str(row.get("category1", "")).strip()
                category2 = str(row.get("category2", "")).strip()

                linked_names = []
                if not links_df.empty and not master_df.empty:
                    target_links = links_df[links_df["contact_id"].astype(str) == contact_id]
                    for _, lrow in target_links.iterrows():
                        rid = str(lrow.get("resident_id", "")).strip()
                        hit = master_df[master_df["resident_id"].astype(str) == rid]
                        if not hit.empty:
                            linked_name = str(hit.iloc[0].get("resident_name", "")).strip()
                            if linked_name:
                                linked_names.append(linked_name)

                with st.container(border=True):
                    st.markdown(f"**{name if name else '名称未設定'}**")
                    st.write(f"{organization} / {category1} / {category2}")
                    if phone:
                        st.write(f"電話: {phone}")
                    if memo:
                        st.write(f"メモ: {memo}")
                    if linked_names:
                        st.write("担当利用者: " + "、".join(linked_names))

    st.divider()

    # ------------------------------------------
    # 利用者検索
    # ------------------------------------------
    st.markdown("## 👤 利用者検索")

    current_company_id = get_current_company_id()
    resident_df = get_resident_master_df(current_company_id)

    resident_search_cols = st.columns([2, 2, 2, 3])

    with resident_search_cols[0]:
        resident_status = st.selectbox(
            "利用状態",
            ["全部", "利用中", "停止中", "終了"],
            key="resident_search_status"
        )

    with resident_search_cols[1]:
        disability_type_filter = st.selectbox(
            "障害区分",
            ["全部", "精神", "身体"],
            key="resident_search_disability_type"
        )

    with resident_search_cols[2]:
        public_assistance_filter = st.selectbox(
            "生活保護受給",
            ["全部", "あり", "なし"],
            key="resident_search_public_assistance"
        )

    with resident_search_cols[3]:
        resident_kw = st.text_input(
            "キーワード",
            key="resident_search_kw",
            placeholder="利用者名・利用者ID・相談員・病院名などで検索"
        )

    resident_view_df = resident_df.copy()

    if disability_type_filter != "全部":
        resident_view_df = resident_view_df[
            resident_view_df["disability_type"].astype(str).str.strip() == disability_type_filter
        ].copy()

    if public_assistance_filter != "全部":
        resident_view_df = resident_view_df[
            resident_view_df["public_assistance"].astype(str).str.strip() == public_assistance_filter
        ].copy()

    if resident_kw.strip():
        kw = resident_kw.strip().lower()
        resident_view_df = resident_view_df[
            resident_view_df.apply(
                lambda row:
                    kw in str(row.get("resident_id", "")).lower()
                    or kw in str(row.get("resident_name", "")).lower()
                    or kw in str(row.get("consultant", "")).lower()
                    or kw in str(row.get("caseworker", "")).lower()
                    or kw in str(row.get("hospital", "")).lower()
                    or kw in str(row.get("nurse", "")).lower()
                    or kw in str(row.get("care", "")).lower(),
                axis=1
            )
        ]

    if resident_view_df.empty:
        st.info("条件に合う利用者はいありません。")
    else:
        try:
            resident_view_df = resident_view_df.sort_values(
                ["status", "resident_id"],
                ascending=[True, True]
            )
        except Exception:
            pass

        for _, row in resident_view_df.iterrows():
            resident_id = str(row.get("resident_id", "")).strip()
            resident_name = str(row.get("resident_name", "")).strip()
            status = str(row.get("status", "")).strip()
            public_assistance = str(row.get("public_assistance", "")).strip()
            consultant = str(row.get("consultant", "")).strip()
            caseworker = str(row.get("caseworker", "")).strip()
            hospital = str(row.get("hospital", "")).strip()
            nurse = str(row.get("nurse", "")).strip()
            care = str(row.get("care", "")).strip()

            with st.container(border=True):
                st.markdown(f"**{resident_name if resident_name else '氏名未設定'}**")
                st.write(f"利用者ID: {resident_id}")
                st.write(f"状態: {status} / 生活保護受給: {public_assistance}")
                st.write(f"相談員: {consultant} / ケースワーカー: {caseworker}")
                st.write(f"病院: {hospital} / 看護: {nurse} / 介護: {care}")

def get_piecework_master_df(company_id=None):
    import pandas as pd

    if company_id is None:
        company_id = str(st.session_state.get("company_id", "")).strip()

    cols = [
        "company_id",
        "piecework_id",
        "piecework_name",
        "client_name",
        "arrival_date",
        "delivery_date",
        "quantity",
        "unit_price",
        "purchase_price",
        "defect_quantity",
        "final_delivery_quantity",
        "income",
        "unit",
        "note",
        "status",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    try:
        df = load_db("piecework_master")

        if df is None or df.empty:
            return pd.DataFrame(columns=cols)

        df = df.copy()

        for c in cols:
            if c not in df.columns:
                df[c] = ""

        df = df[cols].fillna("").copy()

        if company_id:
            df = df[
                df["company_id"].astype(str).str.strip() == str(company_id).strip()
            ].copy()

        return df

    except Exception:
        return pd.DataFrame(columns=cols)

def get_piecework_entries_df(company_id=None):
    import pandas as pd

    if company_id is None:
        company_id = str(st.session_state.get("company_id", "")).strip()

    cols = [
        "company_id",
        "entry_id",
        "piecework_id",
        "entry_date",
        "entry_type",
        "item_name",
        "quantity",
        "unit_price",
        "amount",
        "partner",
        "note",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    try:
        df = load_db("piecework_entries")

        if df is None or df.empty:
            return pd.DataFrame(columns=cols)

        df = df.copy()

        for c in cols:
            if c not in df.columns:
                df[c] = ""

        df = df[cols].fillna("").copy()

        if company_id:
            df = df[
                df["company_id"].astype(str).str.strip() == str(company_id).strip()
            ].copy()

        return df

    except Exception:
        return pd.DataFrame(columns=cols)

def save_piecework_production(piecework_id, user_id, user_name, quantity):
    import pandas as pd

    cols = [
        "company_id",
        "production_id",
        "piecework_id",
        "work_date",
        "user_id",
        "user_name",
        "quantity",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    df = get_piecework_production_df()

    if df is None or df.empty:
        df = pd.DataFrame(columns=cols)

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    now = now_jst()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")

    new_row = pd.DataFrame([{
        "company_id": str(st.session_state.get("company_id", "")).strip(),
        "production_id": f"PR{now.strftime('%Y%m%d%H%M%S%f')}",
        "piecework_id": str(piecework_id).strip(),
        "work_date": now.strftime("%Y-%m-%d"),
        "user_id": str(user_id).strip(),
        "user_name": str(user_name).strip(),
        "quantity": int(quantity),
        "created_at": now_str,
        "updated_at": now_str,
        "is_deleted": "0",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "piecework_production")

def get_piecework_production_df(company_id=None):
    import pandas as pd

    if company_id is None:
        company_id = str(st.session_state.get("company_id", "")).strip()

    cols = [
        "company_id",
        "production_id",
        "piecework_id",
        "work_date",
        "user_id",
        "user_name",
        "quantity",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    try:
        df = load_db("piecework_production")

        if df is None or df.empty:
            return pd.DataFrame(columns=cols)

        df = df.copy()

        for c in cols:
            if c not in df.columns:
                df[c] = ""

        df = df[cols].fillna("").copy()

        if company_id:
            df = df[
                df["company_id"].astype(str).str.strip() == str(company_id).strip()
            ].copy()

        return df

    except Exception:
        return pd.DataFrame(columns=cols)

def get_next_piecework_id():
    df = get_piecework_master_df()

    if df is None or df.empty:
        return "PW0001"

    nums = []
    for x in df["piecework_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("PW"):
            num = x[2:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"PW{next_num:04d}"

def save_piecework_master(
    piecework_name,
    client_name,
    arrival_date,
    delivery_date,
    quantity,
    unit_price,
    purchase_price,
    defect_quantity,
    final_delivery_quantity,
    income,
    unit,
    note,
    status="active",
):
    import pandas as pd

    df = get_piecework_master_df(company_id=None)

    cols = [
        "company_id",
        "piecework_id",
        "piecework_name",
        "client_name",
        "arrival_date",
        "delivery_date",
        "quantity",
        "unit_price",
        "purchase_price",
        "defect_quantity",
        "final_delivery_quantity",
        "income",
        "unit",
        "note",
        "status",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    if df is None or df.empty:
        df = pd.DataFrame(columns=cols)

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    new_row = pd.DataFrame([{
        "company_id": str(st.session_state.get("company_id", "")).strip(),
        "piecework_id": get_next_piecework_id(),
        "piecework_name": str(piecework_name).strip(),
        "client_name": str(client_name).strip(),
        "arrival_date": str(arrival_date).strip(),
        "delivery_date": str(delivery_date).strip(),
        "quantity": str(quantity).strip(),
        "unit_price": str(unit_price).strip(),
        "purchase_price": str(purchase_price).strip(),
        "defect_quantity": str(defect_quantity).strip(),
        "final_delivery_quantity": str(final_delivery_quantity).strip(),
        "income": str(income).strip(),
        "unit": str(unit).strip(),
        "note": str(note).strip(),
        "status": str(status).strip(),
        "created_at": now_str,
        "updated_at": now_str,
        "is_deleted": "0",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "piecework_master")

def get_next_piecework_client_id():
    df = get_piecework_clients_df()

    if df is None or df.empty:
        return "PC0001"

    nums = []
    for x in df["client_id"].fillna("").astype(str):
        x = x.strip().upper()
        if x.startswith("PC"):
            num = x[2:]
            if num.isdigit():
                nums.append(int(num))

    next_num = max(nums) + 1 if nums else 1
    return f"PC{next_num:04d}"

def save_piecework_client(
    client_name,
    client_address,
    contact_1,
    contact_person_1,
    contact_2,
    contact_person_2,
):
    import pandas as pd

    cols = [
        "company_id",
        "client_id",
        "client_name",
        "client_address",
        "contact_1",
        "contact_person_1",
        "contact_2",
        "contact_person_2",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    df = get_piecework_clients_df(company_id=None)

    if df is None or df.empty:
        df = pd.DataFrame(columns=cols)

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    now_str = now_jst().strftime("%Y-%m-%d %H:%M:%S")

    new_row = pd.DataFrame([{
        "company_id": str(st.session_state.get("company_id", "")).strip(),
        "client_id": get_next_piecework_client_id(),
        "client_name": str(client_name).strip(),
        "client_address": str(client_address).strip(),
        "contact_1": str(contact_1).strip(),
        "contact_person_1": str(contact_person_1).strip(),
        "contact_2": str(contact_2).strip(),
        "contact_person_2": str(contact_person_2).strip(),
        "created_at": now_str,
        "updated_at": now_str,
        "is_deleted": "0",
    }])

    df = pd.concat([df, new_row], ignore_index=True)
    save_db(df, "piecework_clients")

def get_piecework_clients_df(company_id=None):
    import pandas as pd

    if company_id is None:
        company_id = str(st.session_state.get("company_id", "")).strip()

    cols = [
        "company_id",
        "client_id",
        "client_name",
        "client_address",
        "contact_1",
        "contact_person_1",
        "contact_2",
        "contact_person_2",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    try:
        df = load_db("piecework_clients")

        if df is None or df.empty:
            return pd.DataFrame(columns=cols)

        df = df.copy()

        for c in cols:
            if c not in df.columns:
                df[c] = ""

        df = df[cols].fillna("").copy()

        if company_id:
            df = df[
                df["company_id"].astype(str).str.strip() == str(company_id).strip()
            ].copy()

        return df

    except Exception:
        return pd.DataFrame(columns=cols)

def render_piecework_page():
    import pandas as pd
    import streamlit as st

    st.title("内職管理")
    st.caption("内職案件の一覧・検索・詳細確認ができます。")

    df = get_piecework_master_df()

    clients_df = get_piecework_clients_df()

    if clients_df is None:
        clients_df = pd.DataFrame()

    if not clients_df.empty:
        clients_df = clients_df[
            clients_df["is_deleted"].astype(str).str.strip() != "1"
        ].copy()

        client_name_list = clients_df["client_name"].fillna("").astype(str).str.strip().tolist()
        client_name_list = [x for x in client_name_list if x]
    else:
        client_name_list = []

    if not client_name_list:
        client_name_list = ["企業未登録"]

    if df is None:
        df = pd.DataFrame()

    if "selected_piecework_id" not in st.session_state:
        st.session_state.selected_piecework_id = ""

    top_cols = st.columns([1, 1])

    with top_cols[0]:
        if st.button(
            "← 休憩室へ移動",
            key="back_from_piecework",
            width="stretch",
            type="secondary",
        ):
            st.session_state.current_page = "休憩室"
            st.rerun()

    with top_cols[1]:
        if st.button(
            "選択中の内職を解除",
            key="clear_selected_piecework",
            width="stretch",
            type="secondary",
        ):
            st.session_state.selected_piecework_id = ""
            st.rerun()

    st.divider()

    with st.expander("＋ 企業を登録する"):
        client_name = st.text_input("企業名", key="client_name")
        client_address = st.text_input("住所", key="client_address")

        row1 = st.columns(2)
        with row1[0]:
            contact_1 = st.text_input("連絡先(1)", key="contact_1")
        with row1[1]:
            contact_person_1 = st.text_input("担当者", key="contact_person_1")

        row2 = st.columns(2)
        with row2[0]:
            contact_2 = st.text_input("連絡先(2)", key="contact_2")
        with row2[1]:
            contact_person_2 = st.text_input("担当者", key="contact_person_2")

            if st.button("企業を登録", key="save_piecework_client", width="stretch", type="secondary"):
                if not str(client_name).strip():
                    st.error("企業名を入れてください。")
                else:
                    save_piecework_client(
                        client_name=client_name,
                        client_address=client_address,
                        contact_1=contact_1,
                        contact_person_1=contact_person_1,
                        contact_2=contact_2,
                        contact_person_2=contact_person_2,
                    )
                    st.success("企業を登録しました。")
                    st.rerun()

    with st.expander("＋ 内職を登録する"):
        row1 = st.columns([3, 2])
        with row1[0]:
            new_piecework_name = st.text_input("内職名", key="new_piecework_name")
        with row1[1]:
            selected_client = st.selectbox("企業名", client_name_list, key="selected_client")

        row2 = st.columns(2)
        with row2[0]:
            arrival_date = st.date_input("納入日", key="arrival_date")
        with row2[1]:
            delivery_date = st.date_input("納品日", key="delivery_date")

        row3 = st.columns(3)
        with row3[0]:
            quantity = st.number_input("数量", min_value=0, step=1, key="new_piecework_quantity")
        with row3[1]:
            unit_price = st.number_input("単価（円）", min_value=0, step=1, key="new_piecework_unit_price")
        with row3[2]:
            purchase_price = st.number_input("購入価格（円）", min_value=0, step=1, key="purchase_price")

        note = st.text_area("備考", key="new_piecework_note", height=80)

        if st.button("内職を登録", key="save_new_piecework", width="stretch", type="secondary"):
            if not str(new_piecework_name).strip():
                st.error("内職名を入れてください。")
            elif client_name_list == ["企業未登録"]:
                st.error("先に企業を登録してください。")
            else:
                save_piecework_master(
                    piecework_name=new_piecework_name,
                    client_name=selected_client,
                    arrival_date=arrival_date,
                    delivery_date=delivery_date,
                    quantity=quantity,
                    unit_price=unit_price,
                    purchase_price=purchase_price,
                    defect_quantity=0,
                    final_delivery_quantity=0,
                    income=0,
                    unit="個",
                    note=note,
                    status="active",
                )
                st.success("内職案件を登録しました。")
                st.rerun()

    filter_cols = st.columns([2, 1, 1])

    with filter_cols[0]:
        search_text = st.text_input(
            "内職名検索",
            key="piecework_search_text",
            placeholder="内職名を入力",
        )

    with filter_cols[1]:
        status_filter = st.selectbox(
            "状態",
            ["すべて", "進行中", "停止中", "終了"],
            key="piecework_status_filter",
        )

    with filter_cols[2]:
        sort_order = st.selectbox(
            "並び順",
            ["更新日が新しい順", "作成日が新しい順", "名前順"],
            key="piecework_sort_order",
        )

    st.divider()

    if df.empty:
        st.info("まだ内職案件がありません。")
        return

    work = df.copy()

    if "is_deleted" in work.columns:
        work = work[
            work["is_deleted"].astype(str).str.strip() != "1"
        ].copy()

    if "company_id" in work.columns:
        current_company_id = str(st.session_state.get("company_id", "")).strip()
        if current_company_id:
            work = work[
                work["company_id"].astype(str).str.strip() == current_company_id
            ].copy()

    for c in ["piecework_name", "status", "unit", "note", "piecework_id", "created_at", "updated_at"]:
        if c not in work.columns:
            work[c] = ""
        work[c] = work[c].fillna("").astype(str)

    if str(search_text).strip():
        keyword = str(search_text).strip().lower()
        work = work[
            work["piecework_name"].str.lower().str.contains(keyword, na=False)
        ].copy()

    status_map = {
        "進行中": "active",
        "停止中": "stopped",
        "終了": "closed",
    }

    if status_filter != "すべて":
        target_status = status_map.get(status_filter, "")
        work = work[
            work["status"].astype(str).str.strip().str.lower() == target_status
        ].copy()

    try:
        work["updated_at_dt"] = pd.to_datetime(work["updated_at"], errors="coerce")
        work["created_at_dt"] = pd.to_datetime(work["created_at"], errors="coerce")
    except Exception:
        work["updated_at_dt"] = pd.NaT
        work["created_at_dt"] = pd.NaT

    if sort_order == "更新日が新しい順":
        work = work.sort_values(["updated_at_dt"], ascending=[False], na_position="last")
    elif sort_order == "作成日が新しい順":
        work = work.sort_values(["created_at_dt"], ascending=[False], na_position="last")
    else:
        work = work.sort_values(["piecework_name"], ascending=[True], na_position="last")

    if work.empty:
        st.info("条件に一致する内職案件がありません。")
        return

    st.caption(f"件数: {len(work)}件")

    cards = work.to_dict("records")
    for i in range(0, len(cards), 2):
        row_cols = st.columns(2)

        for j in range(2):
            idx = i + j
            if idx >= len(cards):
                continue

            row = cards[idx]

            piecework_id = str(row.get("piecework_id", "")).strip()
            piecework_name = str(row.get("piecework_name", "")).strip() or "名称なし"
            status_raw = str(row.get("status", "")).strip().lower()
            unit = str(row.get("unit", "")).strip()
            updated_at = str(row.get("updated_at", "")).strip()

            is_selected = (
                str(st.session_state.get("selected_piecework_id", "")).strip() == piecework_id
            )

            if status_raw == "active":
                status_label = "進行中"
                status_color = "#ECFDF5"
            elif status_raw == "stopped":
                status_label = "停止中"
                status_color = "#FEFCE8"
            elif status_raw == "closed":
                status_label = "終了"
                status_color = "#F3F4F6"
            else:
                status_label = "未設定"
                status_color = "#F9FAFB"

            with row_cols[j]:
                st.markdown(
                    f"""
                    <div style="
                        background:{status_color};
                        border:1px solid #E5E7EB;
                        border-radius:16px;
                        padding:16px;
                        margin-bottom:12px;
                    ">
                        <div style="font-size:20px; font-weight:700; margin-bottom:8px;">
                            {piecework_name}{' ✓' if is_selected else ''}
                        </div>
                        <div style="font-size:13px; color:#6B7280; margin-bottom:4px;">
                            状態：{status_label}
                        </div>
                        <div style="font-size:13px; color:#6B7280; margin-bottom:4px;">
                            単位：{unit if unit else '-'}
                        </div>
                        <div style="font-size:13px; color:#6B7280; margin-bottom:4px;">
                            最終更新日：{updated_at if updated_at else '-'}
                        </div>
                        <div style="font-size:12px; color:#9CA3AF;">
                            ID：{piecework_id if piecework_id else '-'}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                if st.button(
                    "詳細を見る",
                    key=f"piecework_detail_{piecework_id}",
                    width="stretch",
                    type="primary" if is_selected else "secondary",
                ):
                    st.session_state.selected_piecework_id = piecework_id
                    st.rerun()

    selected_piecework_id = str(st.session_state.get("selected_piecework_id", "")).strip()

    if selected_piecework_id:
        st.divider()
        st.markdown("## 内職詳細")

        selected_df = work[
            work["piecework_id"].astype(str).str.strip() == selected_piecework_id
        ].copy()

        if selected_df.empty:
            st.warning("選択中の内職案件が見つかりません。")
            return

        row = selected_df.iloc[0]

        piecework_name = str(row.get("piecework_name", "")).strip()
        status_raw = str(row.get("status", "")).strip().lower()
        unit = str(row.get("unit", "")).strip()
        note = str(row.get("note", "")).strip()

        if status_raw == "active":
            status_label = "進行中"
        elif status_raw == "stopped":
            status_label = "停止中"
        elif status_raw == "closed":
            status_label = "終了"
        else:
            status_label = "未設定"

        st.markdown(f"### {piecework_name if piecework_name else '名称なし'}")
        st.caption(f"状態: {status_label}")

        info_cols = st.columns(3)
        with info_cols[0]:
            st.write(f"**単位**: {unit if unit else '-'}")
        with info_cols[1]:
            st.write(f"**ID**: {selected_piecework_id}")
        with info_cols[2]:
            st.write(f"**状態**: {status_label}")

        st.write(f"**備考**: {note if note else '-'}")

        st.markdown("### 📥 実績入力")

        st.markdown("### 👤 利用者ごとの作成数登録")

        company_id = get_current_company_id()
        master_df = get_resident_master_df(company_id)
        if resident_df is None:
            resident_df = pd.DataFrame()

        resident_options = []
        resident_map = {}

        if not resident_df.empty:
            if "company_id" in resident_df.columns:
                resident_df = resident_df[
                    resident_df["company_id"].astype(str).str.strip() == str(st.session_state.get("company_id", "")).strip()
                ].copy()

            if "status" in resident_df.columns:
                resident_df = resident_df[
                    resident_df["status"].astype(str).str.strip() != "退所"
                ].copy()

            for _, r in resident_df.iterrows():
                rid = str(r.get("resident_id", "")).strip()
                rname = str(r.get("resident_name", "")).strip()
                if rid and rname:
                    label = f"{rname}（{rid}）"
                    resident_options.append(label)
                    resident_map[label] = {
                        "resident_id": rid,
                        "resident_name": rname,
                    }

        prod_cols = st.columns([1, 1, 1])

        with prod_cols[0]:
            st.text_input(
                "内職名",
                value=piecework_name if piecework_name else "",
                key=f"production_piecework_name_{selected_piecework_id}",
                disabled=True,
            )

        with prod_cols[1]:
            selected_resident_label = st.selectbox(
                "利用者",
                resident_options if resident_options else ["利用者未登録"],
                key=f"production_resident_{selected_piecework_id}",
            )

        with prod_cols[2]:
            production_quantity_input = st.number_input(
                "数量",
                min_value=0,
                step=1,
                key=f"production_quantity_{selected_piecework_id}",
            )

        if st.button("登録", key=f"save_piecework_production_{selected_piecework_id}", width="stretch"):
            if not resident_options:
                st.error("利用者が登録されていません。")
            elif production_quantity_input <= 0:
                st.error("数量を入れてください。")
            else:
                picked = resident_map.get(selected_resident_label, {})
                save_piecework_production(
                    piecework_id=selected_piecework_id,
                    user_id=picked.get("resident_id", ""),
                    user_name=picked.get("resident_name", ""),
                    quantity=production_quantity_input,
                )
                st.success("作成数を登録しました。")
                st.rerun()

        row1 = st.columns(3)

        with row1[0]:
            defect_quantity_input = st.number_input(
                "不良/欠品数",
                min_value=0,
                step=1,
                key=f"detail_defect_quantity_{selected_piecework_id}"
            )

        with row1[1]:
            final_delivery_quantity_input = st.number_input(
                "最終納品数",
                min_value=0,
                step=1,
                key=f"detail_final_delivery_quantity_{selected_piecework_id}"
            )

        with row1[2]:
            income_input = st.number_input(
                "収入（円）",
                min_value=0,
                step=1,
                key=f"detail_income_{selected_piecework_id}"
            )

        if st.button("実績を登録", key=f"save_piecework_entry_{selected_piecework_id}", width="stretch"):
            save_piecework_entry(
                piecework_id=selected_piecework_id,
                defect_quantity=defect_quantity_input,
                final_delivery_quantity=final_delivery_quantity_input,
                income=income_input,
            )
            st.success("実績を登録しました")
            st.rerun()


        st.divider()
        st.markdown("### 📅 表示年月")

        now_dt = datetime.now()
        year_key = f"piecework_year_{selected_piecework_id}"
        month_key = f"piecework_month_{selected_piecework_id}"

        if year_key not in st.session_state:
            st.session_state[year_key] = now_dt.year
        if month_key not in st.session_state:
            st.session_state[month_key] = now_dt.month

        ym_cols = st.columns([1, 1, 1])

        with ym_cols[0]:
            selected_year = st.number_input(
                "年",
                min_value=2020,
                max_value=2100,
                value=int(st.session_state[year_key]),
                step=1,
                key=f"{year_key}_input",
            )

        with ym_cols[1]:
            selected_month = st.number_input(
                "月",
                min_value=1,
                max_value=12,
                value=int(st.session_state[month_key]),
                step=1,
                key=f"{month_key}_input",
            )

        with ym_cols[2]:
            st.write("")
            st.write("")
            if st.button(
                "表示",
                key=f"apply_piecework_ym_{selected_piecework_id}",
                width="stretch",
                type="secondary",
            ):
                st.session_state[year_key] = int(selected_year)
                st.session_state[month_key] = int(selected_month)
                st.rerun()

        view_year = int(st.session_state[year_key])
        view_month = int(st.session_state[month_key])

        st.caption(f"現在表示：{view_year}年 {view_month}月")

        entries_df = get_piecework_entries_df()
        production_df = get_piecework_production_df()

        if entries_df is None:
            entries_df = pd.DataFrame()
        if production_df is None:
            production_df = pd.DataFrame()

        month_entries = pd.DataFrame()
        if not entries_df.empty:
            month_entries = entries_df[
                entries_df["piecework_id"].astype(str).str.strip() == selected_piecework_id
            ].copy()

            if "is_deleted" in month_entries.columns:
                month_entries = month_entries[
                    month_entries["is_deleted"].astype(str).str.strip() != "1"
                ].copy()

            month_entries["entry_date_dt"] = pd.to_datetime(
                month_entries["entry_date"], errors="coerce"
            )

            month_entries = month_entries[
                (month_entries["entry_date_dt"].dt.year == view_year) &
                (month_entries["entry_date_dt"].dt.month == view_month)
            ].copy()

        month_production = pd.DataFrame()
        if not production_df.empty:
            month_production = production_df[
                production_df["piecework_id"].astype(str).str.strip() == selected_piecework_id
            ].copy()

            if "is_deleted" in month_production.columns:
                month_production = month_production[
                    month_production["is_deleted"].astype(str).str.strip() != "1"
                ].copy()

            month_production["work_date_dt"] = pd.to_datetime(
                month_production["work_date"], errors="coerce"
            )

            month_production = month_production[
                (month_production["work_date_dt"].dt.year == view_year) &
                (month_production["work_date_dt"].dt.month == view_month)
            ].copy()

        # 数値化
        if not month_entries.empty:
            month_entries["defect_quantity_num"] = pd.to_numeric(
                month_entries["defect_quantity"], errors="coerce"
            ).fillna(0)
            month_entries["final_delivery_quantity_num"] = pd.to_numeric(
                month_entries["final_delivery_quantity"], errors="coerce"
            ).fillna(0)
            month_entries["income_num"] = pd.to_numeric(
                month_entries["income"], errors="coerce"
            ).fillna(0)

        if not month_production.empty:
            month_production["quantity_num"] = pd.to_numeric(
                month_production["quantity"], errors="coerce"
            ).fillna(0)

        defect_total = 0
        final_delivery_total = 0
        sales_total = 0

        if not month_entries.empty:
            defect_total = month_entries["defect_quantity_num"].sum()
            final_delivery_total = month_entries["final_delivery_quantity_num"].sum()
            sales_total = month_entries["income_num"].sum()

        production_total = 0
        if not month_production.empty:
            production_total = month_production["quantity_num"].sum()

        expense_total = 0
        try:
            expense_total = float(row.get("purchase_price", 0) or 0)
        except Exception:
            expense_total = 0

        profit_total = sales_total - expense_total

        st.divider()
        st.markdown("### 📊 月間サマリー")

        cols = st.columns(6)

        with cols[0]:
            st.metric("作成数", f"{int(production_total)}個")

        with cols[1]:
            st.metric("不良数", f"{int(defect_total)}個")

        with cols[2]:
            st.metric("最終納品数", f"{int(final_delivery_total)}個")

        with cols[3]:
            st.metric("支出合計", f"¥{int(expense_total):,}")

        with cols[4]:
            st.metric("売上合計", f"¥{int(sales_total):,}")

        with cols[5]:
            st.metric("差引", f"¥{int(profit_total):,}")

        st.divider()
        st.markdown("### 📋 利用者ごとの作成記録")

        if month_production.empty:
            st.info("この月の利用者別作成記録はありません。")
        else:
            show_user_prod = month_production.copy()

            unit_price_val = 0
            try:
                unit_price_val = float(row.get("unit_price", 0) or 0)
            except Exception:
                unit_price_val = 0

            show_user_prod["amount"] = (
                show_user_prod["quantity_num"] * unit_price_val
            ).astype(int)

            show_user_prod = show_user_prod[[
                "work_date",
                "user_name",
                "quantity",
                "amount",
            ]].copy()

            show_user_prod.columns = ["日付", "利用者名", "数量", "金額"]

            st.dataframe(show_user_prod, width="stretch", height=220)

        st.divider()
        st.markdown("### 履歴プレビュー")

        prev_cols = st.columns(2)

        with prev_cols[0]:
            st.markdown("#### 入出金")
            if month_entries.empty:
                st.info("この月の入出金記録はありません。")
            else:
                show_entries = month_entries.copy()
                show_entries = show_entries[[
                    "entry_date",
                    "defect_quantity",
                    "final_delivery_quantity",
                    "income",
                ]].copy()
                st.dataframe(show_entries, width="stretch", height=250)

        with prev_cols[1]:
            st.markdown("#### 作成記録")
            if month_production.empty:
                st.info("この月の作成記録はありません。")
            else:
                show_prod = month_production.copy()
                show_prod = show_prod[[
                    "work_date",
                    "user_name",
                    "quantity",
                ]].copy()
                st.dataframe(show_prod, width="stretch", height=250)

        st.info("次でここに『支出登録・売上登録・作成数登録』を追加していくある。")

def save_piecework_entry(
    piecework_id,
    defect_quantity,
    final_delivery_quantity,
    income,
):
    import pandas as pd

    cols = [
        "company_id",
        "entry_id",
        "piecework_id",
        "entry_date",
        "defect_quantity",
        "final_delivery_quantity",
        "income",
        "created_at",
        "updated_at",
        "is_deleted",
    ]

    df = get_piecework_entries_df()

    if df is None or df.empty:
        df = pd.DataFrame(columns=cols)

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    now = now_jst()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")

    new_row = pd.DataFrame([{
        "company_id": str(st.session_state.get("company_id", "")),
        "entry_id": f"PE{now.strftime('%Y%m%d%H%M%S')}",
        "piecework_id": piecework_id,
        "entry_date": now.strftime("%Y-%m-%d"),
        "defect_quantity": defect_quantity,
        "final_delivery_quantity": final_delivery_quantity,
        "income": income,
        "created_at": now_str,
        "updated_at": now_str,
        "is_deleted": "0",
    }])

    df = pd.concat([df, new_row], ignore_index=True)

    save_db(df, "piecework_entries")

def render_bulk_documents_page():
    st.title("🤫一括書類作成🤫")
    st.caption("個別支援計画案・サービス担当者会議・個別支援計画をまとめて作成するページです。")

    resident_options, resident_map = get_resident_option_map()

    if not resident_options:
        st.warning("利用者情報がまだありません。")
        return

    st.markdown("## 利用者選択")
    selected_label = st.selectbox(
        "利用者を選択",
        resident_options,
        key="bulk_docs_resident_select"
    )

    selected_row = resident_map.get(selected_label, {})
    resident_id = str(selected_row.get("resident_id", "")).strip()
    resident_name = str(selected_row.get("resident_name", "")).strip()

    st.write(f"利用者ID: {resident_id}")
    st.write(f"利用者名: {resident_name}")

    st.divider()
    st.markdown("## 個別支援計画案")
    plan_draft_cols = st.columns([2, 2, 2, 4])

    with plan_draft_cols[0]:
        st.text_input("計画案_年", key="bulk_plan_draft_year", placeholder="2026")
    with plan_draft_cols[1]:
        st.text_input("計画案_月", key="bulk_plan_draft_month", placeholder="3")
    with plan_draft_cols[2]:
        st.text_input("計画案_日", key="bulk_plan_draft_day", placeholder="29")
    with plan_draft_cols[3]:
        st.text_input("計画案_サービス管理責任者", key="bulk_plan_draft_manager", placeholder="サービス管理責任者")

    st.markdown("#### 計画案_短期目標ごとの入力")
    for i in range(1, 4):
        row_cols = st.columns([3, 3])
        with row_cols[0]:
            st.text_input(
                f"計画案_支援期間{i}",
                key=f"bulk_plan_draft_period_{i}",
                placeholder=f"{i}つ目の具体的到達目標の支援期間"
            )
        with row_cols[1]:
            st.text_input(
                f"計画案_担当者{i}",
                key=f"bulk_plan_draft_person_{i}",
                value="全職員"
            )

    st.divider()
    st.markdown("## サービス担当者会議")

    st.markdown("### サ会議_作成年月日")
    meeting_create_cols = st.columns([2, 2, 2, 4])

    with meeting_create_cols[0]:
        st.text_input("サ会議_作成年", key="bulk_meeting_create_year", placeholder="2026")
    with meeting_create_cols[1]:
        st.text_input("サ会議_作成月", key="bulk_meeting_create_month", placeholder="3")
    with meeting_create_cols[2]:
        st.text_input("サ会議_作成日", key="bulk_meeting_create_day", placeholder="29")
    with meeting_create_cols[3]:
        st.text_input("サ会議_作成者", key="bulk_meeting_creator", placeholder="作成者名")

    st.markdown("### サ会議_開催日時")
    meeting_hold_cols = st.columns([2, 2, 2])

    with meeting_hold_cols[0]:
        st.text_input("サ会議_開催年", key="bulk_meeting_year", placeholder="2026")
    with meeting_hold_cols[1]:
        st.text_input("サ会議_開催月", key="bulk_meeting_month", placeholder="3")
    with meeting_hold_cols[2]:
        st.text_input("サ会議_開催日", key="bulk_meeting_day", placeholder="29")

    st.text_input("サ会議_開催情報", key="bulk_meeting_info")

    st.markdown("### 会議出席者")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.text_input("管理者", key="bulk_meeting_admin")
        st.text_input("ケアマネ", key="bulk_meeting_caremanager")
        st.text_input("サービス管理責任者", key="bulk_meeting_manager")

    with col2:
        st.text_input("支援員", key="bulk_meeting_staff")
        st.text_input("看護師", key="bulk_meeting_nurse")
        st.text_input("相談員", key="bulk_meeting_consultant")

    with col3:
        st.text_input("利用者", key="bulk_meeting_user", value=resident_name)
        st.text_input("親族", key="bulk_meeting_family")
        st.text_input("キーパーソン", key="bulk_meeting_keyperson")

    st.divider()
    st.markdown("## 個別支援計画")
    plan_cols = st.columns([2, 2, 2, 4])

    with plan_cols[0]:
        st.text_input("本計画_年", key="bulk_plan_year", placeholder="2026")
    with plan_cols[1]:
        st.text_input("本計画_月", key="bulk_plan_month", placeholder="3")
    with plan_cols[2]:
        st.text_input("本計画_日", key="bulk_plan_day", placeholder="29")
    with plan_cols[3]:
        st.text_input("本計画_サービス管理責任者", key="bulk_plan_manager", placeholder="サービス管理責任者")

    st.markdown("#### 本計画_短期目標ごとの入力")
    for i in range(1, 4):
        row_cols = st.columns([3, 3])
        with row_cols[0]:
            st.text_input(
                f"本計画_支援期間{i}",
                key=f"bulk_plan_period_{i}",
                placeholder=f"{i}つ目の具体的到達目標の支援期間"
            )
        with row_cols[1]:
            st.text_input(
                f"本計画_担当者{i}",
                key=f"bulk_plan_person_{i}",
                value="全職員"
            )

    st.divider()
    st.markdown("## 確認")
    with st.expander("入力内容確認"):
        st.write(f"利用者: {resident_name} ({resident_id})")

        st.write(
            f"計画案日付: "
            f"{st.session_state.get('bulk_plan_draft_year', '')}/"
            f"{st.session_state.get('bulk_plan_draft_month', '')}/"
            f"{st.session_state.get('bulk_plan_draft_day', '')}"
        )
        st.write(f"計画案サビ管: {st.session_state.get('bulk_plan_draft_manager', '')}")

        for i in range(1, 4):
            st.write(
                f"計画案 {i}行目 / 期間: {st.session_state.get(f'bulk_plan_draft_period_{i}', '')} "
                f"/ 担当者: {st.session_state.get(f'bulk_plan_draft_person_{i}', '')}"
            )

        st.write(
            f"サ会議作成年月日: "
            f"{st.session_state.get('bulk_meeting_create_year', '')}/"
            f"{st.session_state.get('bulk_meeting_create_month', '')}/"
            f"{st.session_state.get('bulk_meeting_create_day', '')}"
        )

        st.write(
            f"サ会議開催年月日: "
            f"{st.session_state.get('bulk_meeting_year', '')}/"
            f"{st.session_state.get('bulk_meeting_month', '')}/"
            f"{st.session_state.get('bulk_meeting_day', '')}"
        )
        st.write(f"サ会議作成者: {st.session_state.get('bulk_meeting_creator', '')}")
        st.write(f"サ会議開催情報: {st.session_state.get('bulk_meeting_info', '')}")

        st.markdown("#### 👥 会議出席者")
        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            st.write(f"管理者: {st.session_state.get('bulk_meeting_admin', '')}")
            st.write(f"ケアマネ: {st.session_state.get('bulk_meeting_caremanager', '')}")
            st.write(f"サービス管理責任者: {st.session_state.get('bulk_meeting_manager', '')}")
        with cc2:
            st.write(f"支援員: {st.session_state.get('bulk_meeting_staff', '')}")
            st.write(f"看護師: {st.session_state.get('bulk_meeting_nurse', '')}")
            st.write(f"相談員: {st.session_state.get('bulk_meeting_consultant', '')}")
        with cc3:
            st.write(f"利用者: {st.session_state.get('bulk_meeting_user', '')}")
            st.write(f"親族: {st.session_state.get('bulk_meeting_family', '')}")
            st.write(f"キーパーソン: {st.session_state.get('bulk_meeting_keyperson', '')}")

        st.write(
            f"本計画日付: "
            f"{st.session_state.get('bulk_plan_year', '')}/"
            f"{st.session_state.get('bulk_plan_month', '')}/"
            f"{st.session_state.get('bulk_plan_day', '')}"
        )
        st.write(f"本計画サビ管: {st.session_state.get('bulk_plan_manager', '')}")

        for i in range(1, 4):
            st.write(
                f"本計画 {i}行目 / 期間: {st.session_state.get(f'bulk_plan_period_{i}', '')} "
                f"/ 担当者: {st.session_state.get(f'bulk_plan_person_{i}', '')}"
            )

    st.divider()

    bulk_generate_key = f"bulk_generate_all_{resident_id}"

    if st.button("🚀 3枚まとめて作成", key=bulk_generate_key):
        resident_name = str(resident_name).strip()

        latest_monitoring_json = get_latest_saved_document_json(resident_id, "モニタリング")
        if not latest_monitoring_json:
            st.error("直近のモニタリングがありません。先にモニタリングを保存してください。")
            return

        draft_periods = [
            st.session_state.get("bulk_plan_draft_period_1", ""),
            st.session_state.get("bulk_plan_draft_period_2", ""),
            st.session_state.get("bulk_plan_draft_period_3", ""),
        ]
        draft_persons = [
            st.session_state.get("bulk_plan_draft_person_1", "全職員"),
            st.session_state.get("bulk_plan_draft_person_2", "全職員"),
            st.session_state.get("bulk_plan_draft_person_3", "全職員"),
        ]

        final_periods = [
            st.session_state.get("bulk_plan_period_1", ""),
            st.session_state.get("bulk_plan_period_2", ""),
            st.session_state.get("bulk_plan_period_3", ""),
        ]
        final_persons = [
            st.session_state.get("bulk_plan_person_1", "全職員"),
            st.session_state.get("bulk_plan_person_2", "全職員"),
            st.session_state.get("bulk_plan_person_3", "全職員"),
        ]

        meeting_info = st.session_state.get("bulk_meeting_info", "")

        attendees_dict = {
            "admin": st.session_state.get("bulk_meeting_admin", ""),
            "caremanager": st.session_state.get("bulk_meeting_caremanager", ""),
            "manager": st.session_state.get("bulk_meeting_manager", ""),
            "staff": st.session_state.get("bulk_meeting_staff", ""),
            "nurse": st.session_state.get("bulk_meeting_nurse", ""),
            "consultant": st.session_state.get("bulk_meeting_consultant", ""),
            "user": st.session_state.get("bulk_meeting_user", ""),
            "family": st.session_state.get("bulk_meeting_family", ""),
            "keyperson": st.session_state.get("bulk_meeting_keyperson", ""),
        }

        attendees_text = f"""
管理者: {attendees_dict["admin"]}
ケアマネ: {attendees_dict["caremanager"]}
サービス管理責任者: {attendees_dict["manager"]}
支援員: {attendees_dict["staff"]}
看護師: {attendees_dict["nurse"]}
相談員: {attendees_dict["consultant"]}
利用者: {attendees_dict["user"]}
親族: {attendees_dict["family"]}
キーパーソン: {attendees_dict["keyperson"]}
""".strip()

        try:
            plan_draft_json = generate_json_with_gemini(
                build_bulk_plan_from_monitoring_prompt(
                    resident_name=resident_name,
                    monitoring_json=latest_monitoring_json,
                )
            )
            plan_draft_json = apply_bulk_plan_overrides(
                plan_draft_json,
                draft_periods,
                draft_persons,
            )

            meeting_json = generate_json_with_gemini(
                build_bulk_meeting_prompt(
                    resident_name=resident_name,
                    plan_json=plan_draft_json,
                    meeting_info=meeting_info,
                    attendees_text=attendees_text,
                )
            )

            plan_final_json = generate_json_with_gemini(
                build_bulk_final_plan_prompt(
                    resident_name=resident_name,
                    draft_plan_json=plan_draft_json,
                    meeting_json=meeting_json,
                )
            )
            plan_final_json = apply_bulk_plan_overrides(
                plan_final_json,
                final_periods,
                final_persons,
            )

            st.session_state["bulk_plan_draft_json"] = plan_draft_json
            st.session_state["bulk_meeting_json"] = meeting_json
            st.session_state["bulk_plan_final_json"] = plan_final_json

            st.success("3枚すべて生成完了です。")

        except Exception as e:
            st.error(f"一括生成エラー: {e}")

    if st.session_state.get("bulk_plan_draft_json"):
        st.markdown("### 📄 計画案（生成結果）")
        st.text_area(
            "計画案",
            json.dumps(st.session_state["bulk_plan_draft_json"], ensure_ascii=False, indent=2),
            height=220,
            key="bulk_plan_draft_json_view"
        )

    if st.session_state.get("bulk_meeting_json"):
        st.markdown("### 📄 サ会議（生成結果）")
        st.text_area(
            "サ会議",
            json.dumps(st.session_state["bulk_meeting_json"], ensure_ascii=False, indent=2),
            height=220,
            key="bulk_meeting_json_view"
        )

    if st.session_state.get("bulk_plan_final_json"):
        st.markdown("### 📄 本計画（生成結果）")
        st.text_area(
            "本計画",
            json.dumps(st.session_state["bulk_plan_final_json"], ensure_ascii=False, indent=2),
            height=220,
            key="bulk_plan_final_json_view"
        )

    if (
        st.session_state.get("bulk_plan_draft_json")
        and st.session_state.get("bulk_meeting_json")
        and st.session_state.get("bulk_plan_final_json")
    ):
        import io
        import zipfile

        st.divider()
        st.markdown("### 📦 一括ダウンロード")

        if st.button("📥 3枚まとめてダウンロード", key="bulk_download_zip"):
            try:
                zip_buffer = io.BytesIO()

                draft_year = st.session_state.get("bulk_plan_draft_year", "")
                draft_month = st.session_state.get("bulk_plan_draft_month", "")
                draft_day = st.session_state.get("bulk_plan_draft_day", "")
                draft_manager = st.session_state.get("bulk_plan_draft_manager", "")

                create_year = st.session_state.get("bulk_meeting_create_year", "")
                create_month = st.session_state.get("bulk_meeting_create_month", "")
                create_day = st.session_state.get("bulk_meeting_create_day", "")
                meeting_year = st.session_state.get("bulk_meeting_year", "")
                meeting_month = st.session_state.get("bulk_meeting_month", "")
                meeting_day = st.session_state.get("bulk_meeting_day", "")
                meeting_info = st.session_state.get("bulk_meeting_info", "")

                attendees_dict = {
                    "admin": st.session_state.get("bulk_meeting_admin", ""),
                    "caremanager": st.session_state.get("bulk_meeting_caremanager", ""),
                    "manager": st.session_state.get("bulk_meeting_manager", ""),
                    "staff": st.session_state.get("bulk_meeting_staff", ""),
                    "nurse": st.session_state.get("bulk_meeting_nurse", ""),
                    "consultant": st.session_state.get("bulk_meeting_consultant", ""),
                    "user": st.session_state.get("bulk_meeting_user", ""),
                    "family": st.session_state.get("bulk_meeting_family", ""),
                    "keyperson": st.session_state.get("bulk_meeting_keyperson", ""),
                }

                plan_year = st.session_state.get("bulk_plan_year", "")
                plan_month = st.session_state.get("bulk_plan_month", "")
                plan_day = st.session_state.get("bulk_plan_day", "")
                plan_manager = st.session_state.get("bulk_plan_manager", "")

                with zipfile.ZipFile(zip_buffer, "w") as zf:
                    cell_data_plan = build_plan_cell_data_from_json(
                        st.session_state["bulk_plan_draft_json"],
                        resident_name,
                        draft_year,
                        draft_month,
                        draft_day,
                        draft_manager,
                    )
                    file_plan = create_excel_file("個別支援計画案", cell_data_plan)
                    zf.writestr(f"{resident_name}_計画案.xlsx", file_plan.getvalue())

                    cell_data_meeting = build_meeting_cell_data_from_json(
                        st.session_state["bulk_meeting_json"],
                        resident_name,
                        create_year,
                        create_month,
                        create_day,
                        meeting_year,
                        meeting_month,
                        meeting_day,
                        meeting_info,
                        attendees_dict,
                        st.session_state.get("bulk_meeting_creator", ""),
                    )
                    file_meeting = create_excel_file("サービス担当者会議", cell_data_meeting)
                    zf.writestr(f"{resident_name}_サ会議.xlsx", file_meeting.getvalue())

                    cell_data_final = build_plan_cell_data_from_json(
                        st.session_state["bulk_plan_final_json"],
                        resident_name,
                        plan_year,
                        plan_month,
                        plan_day,
                        plan_manager,
                    )
                    file_final = create_excel_file("個別支援計画", cell_data_final)
                    zf.writestr(f"{resident_name}_本計画.xlsx", file_final.getvalue())

                zip_buffer.seek(0)

                st.download_button(
                    label="📦 ZIPダウンロード",
                    data=zip_buffer,
                    file_name=f"{resident_name}_書類一式.zip",
                    mime="application/zip",
                    key="bulk_zip_download_button"
                )

            except Exception as e:
                st.error(f"ZIP作成エラー: {e}")

from copy import copy
from io import BytesIO
from openpyxl import load_workbook, Workbook


def sanitize_excel_sheet_name(name: str) -> str:
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    safe = str(name or "").strip()

    for ch in invalid_chars:
        safe = safe.replace(ch, " ")

    safe = safe.strip()
    if not safe:
        safe = "sheet"

    return safe[:31]


def build_home_eval_multi_workbook(resident_files: list[tuple[str, bytes]]) -> bytes:
    """
    resident_files:
      [
        ("荒木和也", <xlsx bytes>),
        ("石田愛子", <xlsx bytes>),
        ...
      ]
    """
    if not resident_files:
        raise RuntimeError("結合する在宅評価シートがありません。")

    master_name, master_bytes = resident_files[0]
    master_wb = load_workbook(BytesIO(master_bytes))
    master_ws = master_wb.active
    master_ws.title = sanitize_excel_sheet_name(master_name)

    for resident_name, file_bytes in resident_files[1:]:
        src_wb = load_workbook(BytesIO(file_bytes))
        src_ws = src_wb.active

        new_ws = master_wb.create_sheet(title=sanitize_excel_sheet_name(resident_name))

        # 列幅コピー
        for col_key, dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_key].width = dim.width

        # 行高さコピー
        for row_key, dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_key].height = dim.height

        # セル値・書式コピー
        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                new_cell.value = cell.value

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)

        # 結合セルコピー
        for merged_range in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

        # シート表示設定
        new_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
        new_ws.freeze_panes = src_ws.freeze_panes

    out = BytesIO()
    master_wb.save(out)
    out.seek(0)
    return out.getvalue()

def render_secret_home_eval_auto_page():
    st.title("🤫在宅評価シート🤫")
    st.caption("Knowbeの支援記録を読み込み、在宅評価シートを自動作成する裏ページです。")

    resident_options, resident_map = get_resident_option_map()

    if not resident_options:
        st.warning("利用者情報がまだありません。")
        return

    st.markdown("## 基本情報")

    base_cols = st.columns([2, 2, 3])

    with base_cols[0]:
        create_year = st.text_input("作成年", key="secret_home_eval_year", placeholder="2026")
    with base_cols[1]:
        create_month = st.text_input("作成月", key="secret_home_eval_month", placeholder="3")
    with base_cols[2]:
        manager_name = st.text_input(
            "月間評価のサビ管名",
            key="secret_home_eval_manager_name",
            placeholder="サービス管理責任者名"
        )

    st.markdown("## 訪問者（職員者名）")
    visit_cols_1 = st.columns(2)
    visit_cols_2 = st.columns(2)
    visit_cols_3 = st.columns(2)

    with visit_cols_1[0]:
        visit_1 = st.text_input("第1週 訪問者（職員者名）", key="secret_home_eval_visit_1")
    with visit_cols_1[1]:
        visit_2 = st.text_input("第2週 訪問者（職員者名）", key="secret_home_eval_visit_2")
    with visit_cols_2[0]:
        visit_3 = st.text_input("第3週 訪問者（職員者名）", key="secret_home_eval_visit_3")
    with visit_cols_2[1]:
        visit_4 = st.text_input("第4週 訪問者（職員者名）", key="secret_home_eval_visit_4")
    with visit_cols_3[0]:
        visit_5 = st.text_input("第5週 訪問者（職員者名）", key="secret_home_eval_visit_5")

    weekly_dates = build_home_eval_week_ranges(create_year, create_month)

    with st.expander("週の日付確認"):
        st.write(f"第1週: {weekly_dates.get('1', '')}")
        st.write(f"第2週: {weekly_dates.get('2', '')}")
        st.write(f"第3週: {weekly_dates.get('3', '')}")
        st.write(f"第4週: {weekly_dates.get('4', '')}")
        st.write(f"第5週: {weekly_dates.get('5', '')}")

    st.divider()

    if st.button("🚀 Knowbeから在宅評価シートを全員分自動作成", key="secret_home_eval_auto_generate_all"):
        if not str(create_year).strip() or not str(create_month).strip():
            st.error("作成年と作成月を入力してください。")
            return

        if not str(manager_name).strip():
            st.error("月間評価のサビ管名を入力してください。")
            return

        weekly_visits = {
            "1": visit_1,
            "2": visit_2,
            "3": visit_3,
            "4": visit_4,
            "5": visit_5,
        }

        driver = None
        created_files = []
        created_names = []
        failed_names = []
        support_record_map = {}
        home_eval_json_map = {}

        try:
            login_username, login_password = get_knowbe_login_credentials()
            if not login_username or not login_password:
                st.error("Knowbeログイン情報が取得できませんでした。")
                return

            with st.spinner("Knowbeへ接続中..."):
                driver = build_chrome_driver()
                driver.get("https://mgr.knowbe.jp/v2/")
                time.sleep(1.0)
                manual_login_wait(driver, login_username, login_password)

            progress = st.progress(0)
            status_box = st.empty()

            total_count = len(resident_options)

            for idx, selected_label in enumerate(resident_options, start=1):
                selected_row = resident_map.get(selected_label, {})
                resident_name = str(selected_row.get("resident_name", "")).strip()

                if not resident_name:
                    failed_names.append(f"{selected_label}：利用者名取得失敗")
                    progress.progress(idx / total_count)
                    continue

                try:
                    status_box.info(f"{idx}/{total_count} 作成中: {resident_name}")

                    support_record_text = fetch_support_record_text_for_month(
                        driver=driver,
                        resident_name=resident_name,
                        year=int(str(create_year).strip()),
                        month=int(str(create_month).strip()),
                    )

                    support_record_map[resident_name] = support_record_text

                    home_eval_json = generate_json_with_gemini(
                        build_home_eval_from_support_record_prompt(
                            resident_name=resident_name,
                            year_val=create_year,
                            month_val=create_month,
                            support_record_text=support_record_text,
                        )
                    )

                    home_eval_json_map[resident_name] = home_eval_json

                    goals = home_eval_json.get("goals", [])
                    monthly_evaluations = home_eval_json.get("monthly_evaluations", [])
                    weekly_reports = home_eval_json.get("weekly_reports", {})

                    cell_data = build_home_eval_cell_data(
                        resident_name=resident_name,
                        create_year=create_year,
                        create_month=create_month,
                        manager_name=manager_name,
                        goals=goals,
                        monthly_evaluations=monthly_evaluations,
                        weekly_dates=weekly_dates,
                        weekly_reports=weekly_reports,
                        weekly_visits=weekly_visits,
                    )

                    excel_buffer = create_excel_file("在宅評価シート", cell_data)

                    if hasattr(excel_buffer, "getvalue"):
                        excel_bytes = excel_buffer.getvalue()
                    else:
                        excel_bytes = excel_buffer

                    created_files.append((resident_name, excel_bytes))
                    created_names.append(resident_name)

                except Exception as e:
                    failed_names.append(f"{resident_name}：{e}")

                progress.progress(idx / total_count)

            if not created_files:
                st.error("全員分の在宅評価シート作成に失敗しました。")
                if failed_names:
                    with st.expander("失敗一覧"):
                        for msg in failed_names:
                            st.write(msg)
                return

            merged_book_bytes = build_home_eval_multi_workbook(created_files)

            st.session_state["secret_home_eval_all_book"] = merged_book_bytes
            st.session_state["secret_home_eval_all_book_name"] = f"{create_year}年{create_month}月_在宅評価シート_全員分.xlsx"
            st.session_state["secret_home_eval_all_created_names"] = created_names
            st.session_state["secret_home_eval_all_failed_names"] = failed_names
            st.session_state["secret_home_eval_support_record_map"] = support_record_map
            st.session_state["secret_home_eval_json_map"] = home_eval_json_map

            status_box.success(f"全員分作成完了：成功 {len(created_names)}名 / 失敗 {len(failed_names)}名")

        except Exception as e:
            st.error(f"在宅評価シート全員分自動作成エラー: {e}")

        finally:
            try:
                if driver is not None:
                    driver.quit()
            except Exception:
                pass

    created_names = st.session_state.get("secret_home_eval_all_created_names", [])
    failed_names = st.session_state.get("secret_home_eval_all_failed_names", [])

    if created_names:
        st.markdown("### 作成できた利用者")
        for name in created_names:
            st.write(f"・{name}")

    if failed_names:
        with st.expander("作成できなかった利用者"):
            for msg in failed_names:
                st.write(f"・{msg}")

    if st.session_state.get("secret_home_eval_all_book") is not None:
        st.download_button(
            label="📥 在宅評価シート（全員分1ブック）をダウンロード",
            data=st.session_state["secret_home_eval_all_book"],
            file_name=st.session_state.get("secret_home_eval_all_book_name", "在宅評価シート_全員分.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="secret_home_eval_download_all"
        )

# ==========================================
# 利用者書類
# ==========================================
def render_secret_generation_panel(doc_title: str):
    st.markdown("## 🤫 Gemini自動作成")
    st.caption("直近の前段書類を参照して、入力欄へ自動反映します。")

    resident_options, resident_map = get_resident_option_map()

    if not resident_options:
        st.warning("利用者情報がまだ登録されていません。")
        return

    selected_label = st.session_state.get(f"{doc_title}_resident_select", "")
    selected_row = resident_map.get(selected_label, {})

    resident_id = str(selected_row.get("resident_id", "")).strip()
    resident_name = str(selected_row.get("resident_name", "")).strip()

    ref_doc_type = get_reference_doc_type_for_gemini(doc_title)

    info_cols = st.columns([3, 2])

    with info_cols[0]:
        st.info(
            f"参照元: {ref_doc_type if ref_doc_type else 'なし'}"
            + (" / テスト時は参照なしでも作成可" if TEST_ALLOW_EMPTY_REFERENCE else "")
        )

    with info_cols[1]:
        st.caption("先に下の利用者選択を触っておくと確実です。")

    new_policy_text = st.text_area(
        "新しい方針・補足（任意）",
        key=f"{doc_title}_secret_new_policy",
        height=100,
        placeholder="Geminiに追加で考慮させたい内容があれば入れる",
    )

    if st.button(f"🤖 {doc_title}をGeminiで作成", key=f"{doc_title}_secret_generate"):
        if not selected_label or not resident_id or not resident_name:
            st.warning("先に下の利用者選択をしてから押してください。")
            return

        try:
            generated = run_secret_gemini_generation(
                doc_title=doc_title,
                resident_id=resident_id,
                resident_name=resident_name,
                new_policy_text=new_policy_text,
            )

            apply_generated_data_to_form(doc_title, generated)
            st.success("Geminiで作成して入力欄へ反映しました。")
            st.rerun()

        except Exception as e:
            st.error(f"Gemini作成でエラーです: {e}")


def render_secret_page(doc_title: str):
    st.title(f"🤫 {doc_title}")

    if doc_title in ["個別支援計画案", "サービス担当者会議", "個別支援計画", "モニタリング"]:
        render_secret_generation_panel(doc_title)
        st.divider()

    if doc_title == "サービス担当者会議":
        render_meeting_form_page(doc_title)
    elif doc_title == "モニタリング":
        render_monitoring_form_page(doc_title)
    elif doc_title == "在宅評価シート":
        render_home_evaluation_form_page(doc_title)
    elif doc_title == "アセスメント":
        render_assessment_form_page(doc_title)
    elif doc_title == "基本シート":
        render_basic_sheet_form_page(doc_title)
    elif doc_title == "就労分野シート":
        render_work_sheet_form_page(doc_title)
    else:
        render_plan_form_page(doc_title)

import traceback

def run_page_debug(page_name, fn):
    try:
        st.info(f"DEBUG: {page_name} に入りました")
        fn()
    except Exception as e:
        st.error(f"{page_name} でエラーです: {e}")
        st.code(traceback.format_exc())

if page == "書類_個別支援計画案":
    if st.session_state.get("secret_doc_mode", False):
        run_page_debug("書類_個別支援計画案", lambda: render_secret_page("個別支援計画案"))
    else:
        run_page_debug("書類_個別支援計画案", lambda: render_plan_form_page("個別支援計画案"))

elif page == "書類_サービス担当者会議":
    if st.session_state.get("secret_doc_mode", False):
        run_page_debug("書類_サービス担当者会議", lambda: render_secret_page("サービス担当者会議"))
    else:
        run_page_debug("書類_サービス担当者会議", lambda: render_meeting_form_page("サービス担当者会議"))

elif page == "書類_個別支援計画":
    if st.session_state.get("secret_doc_mode", False):
        run_page_debug("書類_個別支援計画", lambda: render_secret_page("個別支援計画"))
    else:
        run_page_debug("書類_個別支援計画", lambda: render_plan_form_page("個別支援計画"))

elif page == "書類_モニタリング":
    if st.session_state.get("secret_doc_mode", False):
        run_page_debug("書類_モニタリング", lambda: render_secret_page("モニタリング"))
    else:
        run_page_debug("書類_モニタリング", lambda: render_monitoring_form_page("モニタリング"))

elif page == "書類_在宅評価シート":
    if st.session_state.get("secret_doc_mode", False):
        run_page_debug("書類_在宅評価シート", render_secret_home_eval_auto_page)
    else:
        run_page_debug("書類_在宅評価シート", lambda: render_home_evaluation_form_page("在宅評価シート"))

elif page == "書類_アセスメント":
    run_page_debug("書類_アセスメント", lambda: render_assessment_form_page("アセスメント"))

elif page == "書類_基本シート":
    run_page_debug("書類_基本シート", lambda: render_basic_sheet_form_page("基本シート"))

elif page == "書類_就労分野シート":
    run_page_debug("書類_就労分野シート", lambda: render_work_sheet_form_page("就労分野シート"))

elif page == "🐝knowbe日誌入力🐝":
    run_page_debug("🐝knowbe日誌入力🐝", render_bee_journal_page)

elif page == "🐝knowbe日誌一括入力🐝":
    run_page_debug("🐝knowbe日誌一括入力🐝", render_bulk_knowbe_diary_page)

elif page == "💻他事業所へ登録💻":
    run_page_debug("💻他事業所へ登録💻", render_other_office_register_page)

elif page == "休憩室":
    run_page_debug("休憩室", render_break_room_page)

elif page == "休憩室_チャットルーム":
    run_page_debug("休憩室_チャットルーム", render_chat_room_page)

elif page == "休憩室_書類アップロード":
    run_page_debug("休憩室_書類アップロード", render_archive_page)

elif page == "⑩ 書類アップロード":
    run_page_debug("⑩ 書類アップロード", render_archive_page)

elif page == "休憩室_倉庫":
    run_page_debug("休憩室_倉庫", render_warehouse_page)

elif page == "内職管理":
    run_page_debug("内職管理", render_piecework_page)

elif page == "スタッフ管理":
    if not st.session_state.get("is_admin", False):
        st.error("このページは管理者専用です。")
    else:
        run_page_debug("スタッフ管理", render_admin_staff_manage_block)

elif page == "ICカード管理":
    run_page_debug("ICカード管理", render_ic_card_manage_page)

elif page == "Knowbe情報登録":
    run_page_debug("Knowbe情報登録", render_company_knowbe_settings_page)

elif page == "お問い合わせ":
    run_page_debug("お問い合わせ", render_contact_page)

elif page == "書類_一括書類作成":
    run_page_debug("書類_一括書類作成", render_bulk_documents_page)

elif page == "勤怠管理":
    if not st.session_state.get("is_admin", False):
        st.error("このページは管理者専用です。")
    else:
        run_page_debug("勤怠管理", render_attendance_page)

elif page == "過去日誌照合":
    if not st.session_state.get("is_admin", False):
        st.error("このページは管理者専用です。")
    else:
        run_page_debug("過去日誌照合", render_support_record_audit_page)