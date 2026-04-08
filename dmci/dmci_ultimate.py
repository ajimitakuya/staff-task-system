import json
import re
import time
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any

import pandas as pd
import gspread

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import ChromeOptions, EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC


# =========================================================
# 設定読込
# =========================================================
CONFIG_PATH = Path(__file__).parent / "config_local.json"


def load_config() -> Dict[str, Any]:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError("config_local.json が見つからないある")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


CFG = load_config()

BASE_URL = CFG["base_url"]
USER_ID = CFG["user_id"]
PASSWORD = CFG["password"]
BROWSER = CFG.get("browser", "chrome").lower()
TASKS = CFG["tasks"]

TODAY_STR = datetime.now().strftime("%Y%m%d")
STAMP_STR = datetime.now().strftime("%Y%m%d_%H%M%S")

DESKTOP = Path.home() / "Desktop"
WORK_DIR = DESKTOP / f"{CFG.get('desktop_output_dir_name_prefix', 'DMCI_TABLE_AUTORUN')}_{STAMP_STR}"
XLSX_DIR = WORK_DIR / "xlsx"
DEBUG_DIR = WORK_DIR / "debug"
LOG_DIR = WORK_DIR / "logs"
HTML_DIR = DEBUG_DIR / "html"
SHOT_DIR = DEBUG_DIR / "screenshots"
JSON_DIR = DEBUG_DIR / "json"

RUN_LOG = LOG_DIR / "run.log"
RESULT_JSON = JSON_DIR / "result_summary.json"
FINAL_BOOK_PATH = WORK_DIR / f"{TODAY_STR}.xlsx"


# =========================================================
# 基本
# =========================================================
def ensure_dirs():
    for p in [XLSX_DIR, DEBUG_DIR, LOG_DIR, HTML_DIR, SHOT_DIR, JSON_DIR]:
        p.mkdir(parents=True, exist_ok=True)


def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    with open(RUN_LOG, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def dump_json(obj, path: Path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2, default=str)


def safe_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(s))


def now_tag() -> str:
    return datetime.now().strftime("%H%M%S")


def save_screenshot(driver, name: str):
    path = SHOT_DIR / f"{safe_filename(name)}_{now_tag()}.png"
    try:
        driver.save_screenshot(str(path))
    except Exception:
        pass


def save_html(driver, name: str):
    path = HTML_DIR / f"{safe_filename(name)}_{now_tag()}.html"
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(driver.page_source)
    except Exception:
        pass


def debug_dump(driver, name: str):
    save_screenshot(driver, name)
    save_html(driver, name)
    try:
        meta = {
            "url": driver.current_url,
            "title": driver.title,
            "timestamp": datetime.now().isoformat(),
            "window_handles": driver.window_handles,
        }
        dump_json(meta, JSON_DIR / f"{safe_filename(name)}_{now_tag()}.json")
    except Exception:
        pass


def retry(func, retries=3, wait_sec=3, label="retry_target"):
    last_err = None
    for i in range(1, retries + 1):
        try:
            log(f"{label}: {i}/{retries} 回目")
            return func()
        except Exception as e:
            last_err = e
            log(f"{label}: 失敗 / {e}")
            time.sleep(wait_sec)
    raise last_err


# =========================================================
# Google Sheets
# =========================================================
def get_gspread_client():
    sa_file = Path(__file__).parent / CFG["google_service_account_file"]
    gc = gspread.service_account(filename=str(sa_file))
    return gc


def get_spreadsheet():
    gc = get_gspread_client()
    ss = gc.open_by_key(CFG["google_spreadsheet_id"])
    return ss


def ensure_worksheet(ss, title: str, rows=3000, cols=40):
    try:
        return ss.worksheet(title)
    except Exception:
        return ss.add_worksheet(title=title, rows=rows, cols=cols)


def write_df_to_gsheet(ss, sheet_name: str, df: pd.DataFrame):
    ws_name = CFG["google_sheet_names"].get(sheet_name, sheet_name)
    ws = ensure_worksheet(ss, ws_name, rows=max(len(df) + 100, 3000), cols=max(len(df.columns) + 5, 20))
    ws.clear()
    values = [df.columns.tolist()] + df.fillna("").astype(str).values.tolist()
    ws.update("A1", values)
    log(f"Google Sheets書込完了: {ws_name} / {len(df)}件")


def append_run_log(ss, level: str, message: str):
    ws = ensure_worksheet(ss, "run_log", rows=5000, cols=10)
    ws.append_row(
        [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), level, message],
        value_input_option="USER_ENTERED"
    )


def write_run_summary(ss, summary: Dict[str, Any]):
    ws = ensure_worksheet(ss, "run_summary", rows=2000, cols=30)
    ws.clear()

    rows = [["key", "value"]]
    for k, v in summary.items():
        if isinstance(v, (dict, list)):
            rows.append([k, json.dumps(v, ensure_ascii=False)])
        else:
            rows.append([k, str(v)])
    ws.update("A1", rows)


# =========================================================
# ブラウザ
# =========================================================
def build_driver():
    if BROWSER == "edge":
        options = EdgeOptions()
    else:
        options = ChromeOptions()

    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")
    options.add_argument("--lang=en-US")

    if BROWSER == "edge":
        driver = webdriver.Edge(options=options)
    else:
        driver = webdriver.Chrome(options=options)

    driver.set_page_load_timeout(120)
    return driver


# =========================================================
# Selenium共通
# =========================================================
def wait_visible(driver, by, value, timeout=30):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, value))
    )


def wait_present(driver, by, value, timeout=30):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )


def js_click(driver, elem):
    driver.execute_script("arguments[0].click();", elem)


def robust_click(driver, elem):
    try:
        elem.click()
        return
    except Exception:
        pass
    js_click(driver, elem)


def robust_click_xpath(driver, xpaths: List[str], timeout=10):
    last_err = None
    for xp in xpaths:
        try:
            elem = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xp))
            )
            robust_click(driver, elem)
            return True
        except Exception as e:
            last_err = e
    if last_err:
        raise last_err
    return False


def hover(driver, elem):
    ActionChains(driver).move_to_element(elem).pause(1.0).perform()


# =========================================================
# ログインと遷移
# =========================================================
def open_login(driver):
    driver.get(BASE_URL)
    debug_dump(driver, "login_page")

    uid = wait_visible(driver, By.ID, "txtUserId", timeout=120)
    pw = wait_visible(driver, By.ID, "txtPassword", timeout=120)

    uid.clear()
    uid.send_keys(USER_ID)
    pw.clear()
    pw.send_keys(PASSWORD)
    pw.send_keys(Keys.ENTER)

    time.sleep(3)
    debug_dump(driver, "after_login_submit")


def click_entity_code(driver, entity_code: str):
    xpaths = [
        f"//td[normalize-space()='{entity_code}']",
        f"//td[contains(normalize-space(), '{entity_code}')]",
    ]
    robust_click_xpath(driver, xpaths, timeout=60)
    time.sleep(3)
    debug_dump(driver, f"after_entity_{entity_code}")


def open_unit_availability(driver):
    prop_xpaths = [
        "//div[@id='sidebar-menu-title' and contains(., 'Properties')]",
        "//*[normalize-space()='Properties']",
        "//nav//*[contains(., 'Properties')]",
    ]

    moved = False
    for xp in prop_xpaths:
        try:
            prop = wait_present(driver, By.XPATH, xp, timeout=10)
            hover(driver, prop)
            time.sleep(1.5)
            robust_click_xpath(
                driver,
                [
                    "//a[contains(., 'Unit Availability')]",
                    "//*[normalize-space()='Unit Availability']",
                    "//*[contains(@href, 'propertyunit')]",
                ],
                timeout=8,
            )
            moved = True
            break
        except Exception:
            continue

    if not moved:
        driver.get("https://seller.dmcihomes.com/propertyunit?filtercategory=0&filterstatus=1")

    time.sleep(3)
    debug_dump(driver, "unit_availability_opened")


# =========================================================
# 検索条件
# =========================================================
def select_by_value_or_text(driver, select_id: str, value: Optional[str], text_hint: Optional[str]):
    sel = Select(wait_visible(driver, By.ID, select_id, timeout=20))

    if value:
        try:
            sel.select_by_value(value)
            return
        except Exception:
            pass

    if text_hint:
        for opt in sel.options:
            if text_hint.lower() in (opt.text or "").lower():
                sel.select_by_visible_text(opt.text)
                return

    options = [{"value": o.get_attribute("value"), "text": o.text} for o in sel.options]
    dump_json(options, JSON_DIR / f"{select_id}_options_{now_tag()}.json")
    raise RuntimeError(f"{select_id} の選択ができなかったある")


def run_search_if_needed(driver, task: Dict):
    if not task.get("need_search", False):
        return

    select_by_value_or_text(driver, "project", task.get("project_value"), task.get("project_text_hint"))
    select_by_value_or_text(driver, "category", task.get("category_value"), task.get("category_text_hint"))
    debug_dump(driver, f"after_select_{task['sheet_name']}")

    robust_click_xpath(
        driver,
        [
            "//button[@id='btnSearch']",
            "//button[contains(., 'Search')]",
        ],
        timeout=15,
    )
    time.sleep(5)
    debug_dump(driver, f"after_search_{task['sheet_name']}")


def set_view_by_100(driver):
    sel = Select(wait_visible(driver, By.ID, "selectedcount", timeout=20))
    sel.select_by_value("100")
    time.sleep(4)
    debug_dump(driver, "after_view_by_100")

def normalize_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def score_row_text(text: str) -> int:
    """
    Unit Availabilityの1行っぽさをざっくり点数化する。
    """
    t = normalize_text(text)
    if not t:
        return 0

    score = 0

    keywords = [
        "avail", "onhold", "reserved", "sold",
        "condo", "parking", "service area",
        "php",
    ]
    for kw in keywords:
        if kw in t.lower():
            score += 2

    # 価格っぽい
    if re.search(r"php\s*[\d,]+(?:\.\d{2})?", t, re.I):
        score += 3

    # 日付っぽい
    if re.search(r"\b\d{1,2}/\d{1,2}/\d{4}\b", t):
        score += 2

    # 面積っぽい
    if re.search(r"\b\d+(?:\.\d+)?\b", t):
        score += 1

    # 長すぎるのはむしろ行っぽい
    if len(t) >= 30:
        score += 1

    return score


def split_cells_from_text(text: str) -> List[str]:
    """
    divベースの行を雑にセル分割する。
    まず複数スペースで区切り、だめなら単語列のまま返す。
    """
    t = normalize_text(text)
    if not t:
        return []

    # 2個以上スペース区切り優先
    parts = [p.strip() for p in re.split(r"\s{2,}", t) if p.strip()]
    if len(parts) >= 4:
        return parts

    # fallback: 単一スペース区切り
    parts = [p.strip() for p in t.split(" ") if p.strip()]
    return parts

# =========================================================
# 一覧表直読み
# =========================================================
def parse_html_table(driver) -> List[Dict[str, Any]]:
    """
    table固定ではなく、
    - table/tr
    - role=row
    - divの繰り返し行
    を総当たりで探す。
    """

    # -----------------------------------------------------
    # 1) まず table を探す
    # -----------------------------------------------------
    tables = driver.find_elements(By.XPATH, "//table")
    best_records: List[Dict[str, Any]] = []
    best_meta: Dict[str, Any] = {"mode": None, "headers": [], "count": 0}

    for idx, table in enumerate(tables, start=1):
        try:
            rows = table.find_elements(By.XPATH, ".//tr")
            if len(rows) < 2:
                continue

            header_cells = rows[0].find_elements(By.XPATH, ".//th|.//td")
            headers = [normalize_text(c.text) for c in header_cells]

            if len(headers) < 4:
                continue

            records = []
            for tr in rows[1:]:
                cells = tr.find_elements(By.XPATH, ".//td")
                vals = [normalize_text(c.text) for c in cells]
                if not vals:
                    continue

                joined = normalize_text(" ".join(vals))
                if score_row_text(joined) < 2 and len(vals) < 4:
                    continue

                row = {}
                for i, h in enumerate(headers):
                    key = h if h else f"col_{i+1}"
                    row[key] = vals[i] if i < len(vals) else ""
                records.append(row)

            if len(records) > len(best_records):
                best_records = records
                best_meta = {
                    "mode": "table",
                    "headers": headers,
                    "count": len(records),
                    "table_index": idx,
                }
        except Exception:
            continue

    if best_records:
        dump_json(best_meta, JSON_DIR / f"table_detect_meta_{now_tag()}.json")
        return best_records

    # -----------------------------------------------------
    # 2) role=row を探す
    # -----------------------------------------------------
    role_rows = driver.find_elements(By.XPATH, "//*[@role='row']")
    role_records = []

    for i, row_elem in enumerate(role_rows, start=1):
        try:
            txt = normalize_text(row_elem.text)
            if score_row_text(txt) < 3:
                continue

            cells = row_elem.find_elements(By.XPATH, ".//*[@role='cell'] | .//div | .//span")
            vals = [normalize_text(c.text) for c in cells if normalize_text(c.text)]
            if len(vals) < 3:
                vals = split_cells_from_text(txt)

            if len(vals) < 3:
                continue

            role_records.append({
                "raw_text": txt,
                **{f"col_{j+1}": v for j, v in enumerate(vals)}
            })
        except Exception:
            continue

    if role_records:
        dump_json(
            {
                "mode": "role_row",
                "count": len(role_records),
                "sample": role_records[:5],
            },
            JSON_DIR / f"table_detect_meta_{now_tag()}.json"
        )
        return role_records

    # -----------------------------------------------------
    # 3) divベースの繰り返しブロックを探す
    # -----------------------------------------------------
    candidate_xpaths = [
        "//div[contains(@class,'row')]",
        "//div[contains(@class,'grid')]//div",
        "//div[contains(@class,'item')]",
        "//div[contains(@class,'data')]//div",
        "//div[contains(@class,'list')]//div",
        "//*[contains(@class,'datatable')]//*[self::div or self::li]",
        "//*[contains(@id,'datatable')]//*[self::div or self::li]",
    ]

    all_candidates = []
    seen_texts = set()

    for xp in candidate_xpaths:
        elems = driver.find_elements(By.XPATH, xp)
        for e in elems:
            try:
                txt = normalize_text(e.text)
                if not txt:
                    continue
                if txt in seen_texts:
                    continue
                seen_texts.add(txt)

                score = score_row_text(txt)
                if score < 4:
                    continue

                vals = split_cells_from_text(txt)
                if len(vals) < 3:
                    continue

                all_candidates.append({
                    "score": score,
                    "raw_text": txt,
                    "vals": vals,
                    "xpath_source": xp,
                })
            except Exception:
                continue

    # スコア高い順に並べる
    all_candidates = sorted(all_candidates, key=lambda x: (-x["score"], -len(x["vals"])))

    div_records = []
    for item in all_candidates:
        div_records.append({
            "raw_text": item["raw_text"],
            **{f"col_{i+1}": v for i, v in enumerate(item["vals"])}
        })

    if div_records:
        dump_json(
            {
                "mode": "div_rows",
                "count": len(div_records),
                "sample": div_records[:5],
            },
            JSON_DIR / f"table_detect_meta_{now_tag()}.json"
        )
        return div_records

    # -----------------------------------------------------
    # 4) 全滅ならデバッグ保存して落とす
    # -----------------------------------------------------
    debug_dump(driver, f"no_table_like_structure_{now_tag()}")
    raise RuntimeError("table/div行構造が見つからないある")


def find_next_button(driver):
    candidates = driver.find_elements(
        By.XPATH,
        "//a | //button | //span"
    )

    for e in candidates:
        try:
            txt = (e.text or "").strip().lower()
            cls = (e.get_attribute("class") or "").lower()
            aria = (e.get_attribute("aria-label") or "").strip().lower()

            if txt in ("next", ">", "next ›", "›") or "next" in aria:
                if "disabled" in cls:
                    continue
                if e.is_displayed():
                    return e
        except Exception:
            continue

    # DataTablesっぽい next
    xpaths = [
        "//a[contains(@id, '_next')]",
        "//li[contains(@class,'next')]/a",
        "//a[contains(., 'Next')]",
        "//button[contains(., 'Next')]",
    ]
    for xp in xpaths:
        elems = driver.find_elements(By.XPATH, xp)
        for e in elems:
            try:
                cls = (e.get_attribute("class") or "").lower()
                if "disabled" in cls:
                    continue
                if e.is_displayed():
                    return e
            except Exception:
                continue

    return None


def get_page_signature(records: List[Dict[str, Any]]) -> str:
    if not records:
        return "EMPTY"
    sample = records[:3]
    return json.dumps(sample, ensure_ascii=False, sort_keys=True)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    取れた列名をなるべく統一する。
    """
    rename_map = {}

    for c in df.columns:
        lc = str(c).strip().lower()

        if lc == "property":
            rename_map[c] = "Property"
        elif lc == "building":
            rename_map[c] = "Building"
        elif lc == "unit":
            rename_map[c] = "Unit"
        elif lc == "tower":
            rename_map[c] = "Tower"
        elif lc == "floor":
            rename_map[c] = "Floor"
        elif lc == "status":
            rename_map[c] = "Status"
        elif lc == "category":
            rename_map[c] = "Category"
        elif lc == "type":
            rename_map[c] = "Type"
        elif "gross" in lc and "area" in lc:
            rename_map[c] = "GrossArea"
        elif lc == "location":
            rename_map[c] = "Location"
        elif "property unit" in lc:
            rename_map[c] = "PropertyUnit"
        elif "rfo" in lc and "date" in lc:
            rename_map[c] = "RFODate"
        elif "tandem" in lc or "package" in lc:
            rename_map[c] = "TandemPackage"
        elif "list" in lc and "price" in lc:
            rename_map[c] = "ListPrice"

    df = df.rename(columns=rename_map)

    if "Property" in df.columns:
        df = df.drop(columns=["Property"])

    if "GrossArea" in df.columns:
        df["GrossArea"] = pd.to_numeric(
            df["GrossArea"].astype(str).str.replace(",", "", regex=False),
            errors="coerce"
        )

    if "RFODate" in df.columns:
        df["RFODate"] = pd.to_datetime(df["RFODate"], errors="coerce")

    if "ListPrice" in df.columns:
        df["ListPrice"] = (
            df["ListPrice"]
            .astype(str)
            .str.replace("Php", "", regex=False)
            .str.replace(",", "", regex=False)
            .str.strip()
        )
        df["ListPrice"] = pd.to_numeric(df["ListPrice"], errors="coerce")

    ordered_cols = [
        "Building",
        "Unit",
        "Tower",
        "Floor",
        "Status",
        "Category",
        "Type",
        "GrossArea",
        "Location",
        "PropertyUnit",
        "RFODate",
        "TandemPackage",
        "ListPrice",
    ]
    ordered_existing = [c for c in ordered_cols if c in df.columns]
    other_cols = [c for c in df.columns if c not in ordered_existing]

    return df[ordered_existing + other_cols]


def collect_all_pages_table(driver, logical_name: str) -> pd.DataFrame:
    all_rows: List[Dict[str, Any]] = []
    seen_signatures = set()
    page_no = 1
    max_pages = 50  # 暴走防止

    while page_no <= max_pages:
        time.sleep(2)
        debug_dump(driver, f"table_page_{logical_name}_{page_no}")

        rows = parse_html_table(driver)
        sig = get_page_signature(rows)

        if sig in seen_signatures:
            log(f"{logical_name}: 同じページ署名を検出、巡回終了")
            break

        seen_signatures.add(sig)

        for r in rows:
            r["_page_no"] = page_no
        all_rows.extend(rows)

        next_btn = find_next_button(driver)
        if not next_btn:
            log(f"{logical_name}: Nextボタンなし、巡回終了")
            break

        before_url = driver.current_url
        before_sig = sig

        try:
            robust_click(driver, next_btn)
            time.sleep(4)

            # ページ変化待ち
            changed = False
            for _ in range(10):
                time.sleep(1)
                try:
                    new_rows = parse_html_table(driver)
                    new_sig = get_page_signature(new_rows)
                    if new_sig != before_sig:
                        changed = True
                        break
                except Exception:
                    pass

            if not changed:
                log(f"{logical_name}: Nextクリック後に内容変化なし、巡回終了")
                break

            page_no += 1

        except Exception as e:
            log(f"{logical_name}: Nextクリック失敗、巡回終了 / {e}")
            break

    if not all_rows:
        raise RuntimeError(f"一覧データ取得が0件だったある: {logical_name}")

    df = pd.DataFrame(all_rows)

    # 列の型を雑に整える
    if "raw_text" in df.columns and len(df.columns) <= 4:
        # 完全に生テキストしか取れてない場合も、そのまま保存できるようにする
        log(f"{logical_name}: 生テキスト主体で保存するある / rows={len(df)}")
        return df

    df = normalize_columns(df)
    log(f"{logical_name}: 一覧直読み完了 / {len(df)}件 / {page_no}ページ")
    return df


# =========================================================
# Excel
# =========================================================
def format_excel(path: Path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        header_map = {}
        for c in ws[1]:
            header_map[c.value] = c.column_letter

        if "RFODate" in header_map:
            col = header_map["RFODate"]
            for cell in ws[col][1:]:
                cell.number_format = 'yyyy"年"mm"月"dd"日"'

        if "ListPrice" in header_map:
            col = header_map["ListPrice"]
            for cell in ws[col][1:]:
                cell.number_format = '#,##0.00'

        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                val = "" if cell.value is None else str(cell.value)
                widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(val))

        for col, width in widths.items():
            ws.column_dimensions[col].width = min(max(width + 2, 12), 42)

    wb.save(path)


def save_single_excel(df: pd.DataFrame, sheet_name: str, path: Path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    format_excel(path)


def save_final_excel(items: List[Dict[str, pd.DataFrame]], out_path: Path):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for item in items:
            item["df"].to_excel(writer, sheet_name=item["sheet_name"], index=False)
    format_excel(out_path)


# =========================================================
# 1タスク
# =========================================================
def process_task(driver, task: Dict, ss) -> Dict[str, Any]:
    name = task["sheet_name"]
    result = {
        "sheet_name": name,
        "entity_code": task["entity_code"],
        "status": "running",
        "row_count": 0,
        "source_mode": "html_table",
        "xlsx_path": None,
        "error": None,
    }

    try:
        append_run_log(ss, "INFO", f"開始: {name}")

        retry(lambda: open_login(driver), retries=3, wait_sec=5, label=f"{name}_open_login")
        retry(lambda: click_entity_code(driver, task["entity_code"]), retries=3, wait_sec=3, label=f"{name}_entity")
        retry(lambda: open_unit_availability(driver), retries=3, wait_sec=3, label=f"{name}_unit_availability")
        retry(lambda: run_search_if_needed(driver, task), retries=3, wait_sec=3, label=f"{name}_run_search")
        retry(lambda: set_view_by_100(driver), retries=3, wait_sec=3, label=f"{name}_viewby")

        df = retry(
            lambda: collect_all_pages_table(driver, name),
            retries=2,
            wait_sec=5,
            label=f"{name}_table_collect"
        )

        result["row_count"] = int(len(df))

        xlsx_path = XLSX_DIR / f"{safe_filename(name)}.xlsx"
        save_single_excel(df, name, xlsx_path)
        result["xlsx_path"] = str(xlsx_path)

        write_df_to_gsheet(ss, name, df)

        result["status"] = "success"
        append_run_log(ss, "INFO", f"成功: {name} / {len(df)}件 / mode=html_table")
        return {"result": result, "df": df}

    except Exception as e:
        result["status"] = "failed"
        result["error"] = str(e)
        debug_dump(driver, f"task_failed_{name}")
        append_run_log(ss, "ERROR", f"失敗: {name} / {e}")
        traceback.print_exc()
        return {"result": result, "df": None}


# =========================================================
# main
# =========================================================
def main():
    ensure_dirs()
    log(f"作業開始: {WORK_DIR}")

    summary = {
        "started_at": datetime.now().isoformat(),
        "work_dir": str(WORK_DIR),
        "browser": BROWSER,
        "tasks": [],
        "final_excel": None,
    }

    ss = get_spreadsheet()
    append_run_log(ss, "INFO", "全体開始")

    driver = None
    final_items = []

    try:
        driver = build_driver()

        for task in TASKS:
            out = process_task(driver, task, ss)
            summary["tasks"].append(out["result"])

            if out["df"] is not None:
                final_items.append({
                    "sheet_name": task["sheet_name"],
                    "df": out["df"],
                })

            dump_json(summary, RESULT_JSON)
            write_run_summary(ss, summary)

        if final_items:
            save_final_excel(final_items, FINAL_BOOK_PATH)
            summary["final_excel"] = str(FINAL_BOOK_PATH)

        summary["finished_at"] = datetime.now().isoformat()
        dump_json(summary, RESULT_JSON)
        write_run_summary(ss, summary)
        append_run_log(ss, "INFO", "全体終了")

        log("全処理完了ある")
        if summary["final_excel"]:
            log(f"最終Excel: {FINAL_BOOK_PATH}")

    except Exception as e:
        summary["fatal_error"] = str(e)
        summary["finished_at"] = datetime.now().isoformat()
        dump_json(summary, RESULT_JSON)
        try:
            write_run_summary(ss, summary)
            append_run_log(ss, "ERROR", f"main異常終了: {e}")
        except Exception:
            pass
        traceback.print_exc()

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        log("終了ある")


if __name__ == "__main__":
    main()