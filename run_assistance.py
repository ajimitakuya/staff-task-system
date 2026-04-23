# -*- coding: utf-8 -*-
"""
knowbe 日誌入力スクリプト（安定統合版・2026-02-23 / dropdown強化fix）
- Excelは必ずTEMPにコピーしてから openpyxl で読む（Y:直読み破損対策）
- 手動ログイン（2FA/Googleログイン変動に強い）
- ログイン後、必ず「利用実績（日ごと）」で指定日付へ移動
- 日付変更後はテーブル安定待ち
- 編集(鉛筆)は「行の右端セル button/svg」優先
- 条約準拠：
    * B列が「通所」「施設外就労」のみ処理（それ以外は触らない）
    * 「通所」は E列=提供なし → 在宅扱い(備考=在宅利用), E列=提供あり → 来所扱い(備考=食事摂取量などExcel優先)
    * 実績入力は B〜F（サービス/開始/終了/食事/備考）を入れて保存まで完了
- API：キーが無いなら絶対に進めない（条約）
"""

import os
import streamlit as st
import time
import datetime
import re
import math
import shutil
import tempfile
from data_access import get_companies_df
from dataclasses import dataclass
from typing import List, Tuple, Optional

import random
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.action_chains import ActionChains

from journal_rewrite import generate_journal_from_memo

print("[DEBUG] RUN_ASSISTANCE __file__ =", __file__, flush=True)

# =========================
# 設定
# =========================
def get_knowbe_login_credentials(company_id=None):
    """
    1) まず company_id で companies シートを参照
    2) 見つからなければ旧 office_key ベースへフォールバック
    """
    username = ""
    password = ""

    # ---------------------------------
    # ① company_id 優先
    # ---------------------------------
    if company_id is None:
        company_id = st.session_state.get("company_id", "")

    company_id = str(company_id or "").strip()

    if company_id:
        try:
            df = get_companies_df()
        except Exception:
            df = None

        if df is not None and not df.empty:
            work = df.fillna("").copy()

            if "company_id" in work.columns:
                work["company_id"] = work["company_id"].astype(str).str.strip()

                hit = work[work["company_id"] == company_id]
                if not hit.empty:
                    row = hit.iloc[0]
                    username = str(row.get("knowbe_login_username", "")).strip()
                    password = str(row.get("knowbe_login_password", "")).strip()

                    print(f"[SECRETS CHECK NOW] company_id={company_id}", flush=True)
                    print(f"[SECRETS CHECK NOW] company-based username exists={bool(username)}", flush=True)
                    print(f"[SECRETS CHECK NOW] company-based password exists={bool(password)}", flush=True)

                    if username and password:
                        return username, password

    # ---------------------------------
    # ② 旧 office_key フォールバック
    # ---------------------------------
    office_key = str(st.session_state.get("office_key", "support")).strip().lower()
    if office_key not in ("support", "home"):
        office_key = "support"

    secret_user_key = f"KB_LOGIN_USERNAME_{office_key.upper()}"
    secret_pass_key = f"KB_LOGIN_PASSWORD_{office_key.upper()}"

    try:
        username = st.secrets.get(secret_user_key, "")
        password = st.secrets.get(secret_pass_key, "")
    except Exception:
        username = ""
        password = ""

    if not username:
        username = os.environ.get(secret_user_key, "")
    if not password:
        password = os.environ.get(secret_pass_key, "")

    print(f"[SECRETS CHECK NOW] office={office_key}", flush=True)
    print(f"[SECRETS CHECK NOW] username_key={secret_user_key}", flush=True)
    print(f"[SECRETS CHECK NOW] password_key={secret_pass_key}", flush=True)
    print(f"[SECRETS CHECK NOW] LOGIN_USERNAME exists={bool(username)}", flush=True)
    print(f"[SECRETS CHECK NOW] LOGIN_PASSWORD exists={bool(password)}", flush=True)

    return username, password

def build_chrome_driver():
    import platform
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service as ChromeService

    options = webdriver.ChromeOptions()
    system_name = platform.system().lower()

    # ===== Linux / Streamlit Cloud / GitHub側実行想定 =====
    if system_name == "linux":
        options.binary_location = "/usr/bin/chromium"
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1400,900")
        options.add_argument("--remote-debugging-port=9222")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--disable-extensions")

        driver = webdriver.Chrome(
            service=ChromeService("/usr/bin/chromedriver"),
            options=options
        )

    # ===== Windows / ローカル実行 =====
    else:
        # options.add_argument("--headless=new")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-features=RendererCodeIntegrity")
        options.add_argument("--disable-backgrounding-occluded-windows")

        driver = webdriver.Chrome(options=options)

    driver.set_page_load_timeout(60)
    driver.set_script_timeout(60)
    driver.implicitly_wait(2)
    return driver

BASE_URL = "https://mgr.knowbe.jp/v2/"
REPORT_DAILY_URL = "https://mgr.knowbe.jp/v2/#/report/daily"
EXCEL_PATH_DEFAULT = os.path.join(os.path.dirname(__file__), "日誌基本情報サポート.xlsx")

SHEET_MAIN = "基本情報"
SHEET_TREATY = "条約"

WAIT_SHORT = 5
WAIT_NORMAL = 20

# 鉛筆SVG path(d)（フォールバック）
PENCIL_D_PREFIX = "M3 17.25V21"


# =========================
# データ構造
# =========================
@dataclass
class PersonItem:
    name: str
    service: str      # B列（通所/施設外就労/その他）
    start: str        # C列
    end: str          # D列
    meal: str         # E列（提供あり/提供なし）
    note: str         # F列（在宅利用/食事摂取量 x/10/空欄など）
    user_state: str   # H列（今回未使用）
    staff_note: str   # I列（今回未使用）
    staff_name: str   # K列（今回未使用）
    staff_mark: str   # L列（今回未使用）
    work_start: str = ""
    work_end: str = ""
    work_break: str = "0"
    work_memo: str = ""   


# =========================
# ログ/デバッグ
# =========================
def log(msg: str):
    print(msg, flush=True)

def dump_debug(driver, tag: str):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    outdir = os.path.join(os.path.dirname(__file__), "dbg")
    os.makedirs(outdir, exist_ok=True)
    png = os.path.join(outdir, f"debug_{tag}_{ts}.png")
    html = os.path.join(outdir, f"debug_{tag}_{ts}.html")
    try:
        driver.save_screenshot(png)
    except Exception:
        pass
    try:
        with open(html, "w", encoding="utf-8") as f:
            f.write(driver.page_source)
    except Exception:
        pass


# =========================
# Excel: 安定読み取り（TEMPへコピー）
# =========================
def stage_excel_local(src_path: str) -> str:
    src_path = os.path.abspath(src_path)
    base = os.path.basename(src_path)
    if base.startswith("~$"):
        raise RuntimeError("[FATAL] Excelの一時ファイル(~$)を読もうとしてるある。元のxlsxを指定してくれある")
    if not os.path.exists(src_path):
        raise RuntimeError(f"[FATAL] Excelが見つからないある: {src_path}")

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    staged = os.path.join(tempfile.gettempdir(), f"knowbe_staged_{ts}.xlsx")
    shutil.copy2(src_path, staged)

    if os.path.getsize(staged) < 1000:
        raise RuntimeError(f"[FATAL] staged excel が小さすぎるある: {staged}")
    return staged

def norm(v) -> str:
    return "" if v is None else str(v).strip()

def read_treaty_check(excel_path: str):
    staged = stage_excel_local(excel_path)
    wb = openpyxl.load_workbook(staged, data_only=True)
    if SHEET_TREATY not in wb.sheetnames:
        raise RuntimeError(f"[FATAL] 条約シートが無いある: {SHEET_TREATY}")
    ws = wb[SHEET_TREATY]
    a1 = (ws["A1"].value or "")
    b1 = (ws["B1"].value or "")
    c1 = (ws["C1"].value or "")
    wb.close()

def normalize_hhmm(v) -> str:
    """
    Excel由来の時刻が
      - datetime.time / datetime.datetime
      - '10:01:00' / '10:01'
    どれでも最終的に 'HH:MM' にする
    """
    if v is None:
        return ""
    if isinstance(v, datetime.time):
        return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, datetime.datetime):
        return f"{v.hour:02d}:{v.minute:02d}"

    s = str(v).strip()
    m = re.match(r"^\s*(\d{1,2}):(\d{2})(?::\d{2})?\s*$", s)
    if m:
        hh = int(m.group(1)); mm = int(m.group(2))
        return f"{hh:02d}:{mm:02d}"
    return ""  # ★None返し事故を防ぐ

def read_sheet_data(excel_path: str) -> Tuple[int, int, int, List[PersonItem]]:
    staged = stage_excel_local(excel_path)
    wb = openpyxl.load_workbook(staged, data_only=True)

    if SHEET_MAIN not in wb.sheetnames:
        raise RuntimeError(f"[FATAL] 基本情報シートが無いある: {SHEET_MAIN}")
    ws = wb[SHEET_MAIN]

    y = int(ws["A1"].value)
    m = int(ws["B1"].value)
    d = int(ws["C1"].value)

    items: List[PersonItem] = []
    row = 3
    while True:
        name = norm(ws[f"A{row}"].value)
        if not name:
            break

        start = normalize_hhmm(ws[f"C{row}"].value)
        end   = normalize_hhmm(ws[f"D{row}"].value)

        items.append(PersonItem(
            name=name,
            service=norm(ws[f"B{row}"].value),
            start=start,   # ← 空欄なら空欄のまま
            end=end,       # ← 空欄なら空欄のまま
            meal=norm(ws[f"E{row}"].value),
            note=norm(ws[f"F{row}"].value),
            user_state=norm(ws[f"H{row}"].value),
            staff_note=norm(ws[f"I{row}"].value),
            staff_name=norm(ws[f"K{row}"].value),
            staff_mark=norm(ws[f"L{row}"].value),
        ))
        row += 1

    wb.close()
    return y, m, d, items

# =========================
# Selenium ユーティリティ
# =========================
def safe_click(driver, el) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.05)
    except Exception:
        pass
    try:
        el.click()
        return True
    except Exception:
        pass
    try:
        ActionChains(driver).move_to_element(el).pause(0.05).click(el).perform()
        return True
    except Exception:
        pass
    try:
        driver.execute_script("arguments[0].click();", el)
        return True
    except Exception:
        return False

def manual_login_wait(driver, login_username, login_password):
    login_username = "" if login_username is None else str(login_username).strip()
    login_password = "" if login_password is None else str(login_password).strip()

    if not login_username or not login_password:
        raise RuntimeError("[FATAL] login_username / login_password が空ある")

    last_error = ""
    start_ts = time.time()

    while time.time() - start_ts < 30:
        url = driver.current_url or ""

        if "login" not in url and "mgr.knowbe.jp" in url:
            return

        try:
            user_el = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.ID, "username"))
            )
            pass_el = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.ID, "password"))
            )

            set_input_value(driver, user_el, login_username)
            time.sleep(0.2)
            set_input_value(driver, pass_el, login_password)
            time.sleep(0.3)

            span = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//span[contains(normalize-space(.),'上記に同意してログイン')]")
                )
            )
            btn = span.find_element(By.XPATH, "./ancestor::button[1]")

            if not safe_click(driver, btn):
                driver.execute_script("arguments[0].click();", btn)

            t0 = time.time()
            while time.time() - t0 < 10:
                cur = driver.current_url or ""
                if "login" not in cur and "mgr.knowbe.jp" in cur:
                    time.sleep(1.0)
                    return
                time.sleep(0.3)

        except Exception as e:
            last_error = str(e)

        time.sleep(1)

    raise RuntimeError(
        f"[FATAL] 自動ログインが30秒でタイムアウトしたある。"
        f" current_url={driver.current_url!r} last_error={last_error!r}"
    )

def get_top_dialog(driver):
    ds = driver.find_elements(By.CSS_SELECTOR, "[role='dialog']")
    return ds[-1] if ds else None

def close_dialog_if_open(driver):
    dlg = get_top_dialog(driver)
    if not dlg:
        return
    for xp in [
        ".//button[contains(.,'キャンセル')]",
        ".//button[contains(.,'閉じる')]",
        ".//button[contains(.,'戻る')]",
    ]:
        try:
            b = dlg.find_element(By.XPATH, xp)
            if safe_click(driver, b):
                try:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element(dlg))
                except Exception:
                    pass
                return
        except Exception:
            pass
    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    except Exception:
        pass

def wait_table_stable_after_date_change(driver, timeout=20):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
    )
    time.sleep(0.8)
    t0 = time.time()
    while time.time() - t0 < 10:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        if rows:
            for r in rows[:6]:
                try:
                    tds = r.find_elements(By.TAG_NAME, "td")
                    if not tds:
                        continue
                    last = tds[-1]
                    if last.find_elements(By.TAG_NAME, "button") or last.find_elements(By.TAG_NAME, "svg"):
                        return
                except Exception:
                    continue
        time.sleep(0.25)


# =========================
# 日付移動（ピッカー）
# =========================
def goto_report_daily(driver):
    driver.get("https://mgr.knowbe.jp/v2/#/report/daily")
    time.sleep(3.0)

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re


def _click_by_visible_text(driver, tag_name: str, visible_text: str, timeout: int = 10) -> bool:
    """
    画面上の表示文字で要素を探してクリックする
    例: <p>記録</p> / <span>支援記録</span>
    """
    xpath = f"//{tag_name}[normalize-space(.)='{visible_text}']"

    try:
        el = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.2)
        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)
        time.sleep(1.0)
        return True
    except Exception:
        return False

def parse_header_date_text(s: str) -> Optional[Tuple[int, int, int]]:
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3))

def get_current_header_date(driver) -> Optional[Tuple[int, int, int]]:
    try:
        header = driver.find_element(By.CSS_SELECTOR, "#reportDailyHeader")
        txt = (header.text or "").replace("\n", " ").strip()
        got = parse_header_date_text(txt)
        if got:
            return got
    except Exception:
        pass
    return None

def open_date_picker(driver) -> bool:
    try:
        header = WebDriverWait(driver, WAIT_NORMAL).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#reportDailyHeader"))
        )
    except Exception:
        return False

    btns = header.find_elements(By.TAG_NAME, "button")
    for b in btns:
        if safe_click(driver, b):
            time.sleep(0.4)
            if driver.find_elements(By.CSS_SELECTOR, "[role='dialog']"):
                return True
    return False

def _get_dialog(driver):
    ds = driver.find_elements(By.CSS_SELECTOR, "[role='dialog']")
    return ds[-1] if ds else None

def _get_scroll_container(driver, dialog):
    try:
        return dialog.find_element(By.CSS_SELECTOR, ".Cal__MonthList__root")
    except Exception:
        pass
    return None

def _visible_date_range(dialog) -> Optional[Tuple[datetime.date, datetime.date]]:
    try:
        els = dialog.find_elements(By.CSS_SELECTOR, "[data-date]")
        ds = []
        for el in els:
            dd = el.get_attribute("data-date") or ""
            if re.match(r"^\d{4}-\d{2}-\d{2}$", dd):
                y, m, d = dd.split("-")
                ds.append(datetime.date(int(y), int(m), int(d)))
        if not ds:
            return None
        return min(ds), max(ds)
    except Exception:
        return None

def _click_confirm_if_any(driver, dialog) -> bool:
    for xp in [
        ".//button[contains(.,'確定する')]",
        ".//button[contains(.,'確定')]",
        ".//button[contains(.,'OK')]",
        ".//button[contains(.,'決定')]",
    ]:
        try:
            b = dialog.find_element(By.XPATH, xp)
            if safe_click(driver, b):
                time.sleep(0.2)
                return True
        except Exception:
            pass
    return False

def set_date_in_picker(driver, y: int, m: int, d: int) -> bool:
    target_date = datetime.date(y, m, d)
    target_iso = f"{y:04d}-{m:02d}-{d:02d}"
    sel_any = f"[data-date='{target_iso}']"

    def try_click_target() -> bool:
        dialog = _get_dialog(driver)
        if not dialog:
            return False
        try:
            el = dialog.find_element(By.CSS_SELECTOR, sel_any)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.05)
            driver.execute_script("arguments[0].click();", el)
            time.sleep(0.12)
            _click_confirm_if_any(driver, dialog)
            return True
        except Exception:
            return False

    if try_click_target():
        return True

    def month_index(dt: datetime.date) -> int:
        return dt.year * 12 + (dt.month - 1)

    for _ in range(50):
        dialog = _get_dialog(driver)
        if not dialog:
            return False
        sc = _get_scroll_container(driver, dialog)
        if not sc:
            return False

        rng = _visible_date_range(dialog)
        if not rng:
            driver.execute_script("arguments[0].scrollTop += 800;", sc)
            time.sleep(0.06)
            if try_click_target():
                return True
            continue

        vmin, vmax = rng
        if vmin <= target_date <= vmax:
            return try_click_target()

        if target_date > vmax:
            diff_m = month_index(target_date) - month_index(vmax)
            jump = min(max(diff_m * 260, 800), 24000)
            driver.execute_script("arguments[0].scrollTop += arguments[1];", sc, jump)
        else:
            diff_m = month_index(vmin) - month_index(target_date)
            jump = min(max(diff_m * 260, 800), 24000)
            driver.execute_script("arguments[0].scrollTop -= arguments[1];", sc, jump)
        time.sleep(0.06)

        if try_click_target():
            return True

    return False

def goto_report_date(driver, y: int, m: int, d: int):
    goto_report_daily(driver)

    cur = get_current_header_date(driver)
    if cur == (y, m, d):
        wait_table_stable_after_date_change(driver)
        return

    if not open_date_picker(driver):
        dump_debug(driver, "open_date_picker_fail")
        raise RuntimeError("[FATAL] 日付ピッカーが開けなかったある")

    if not set_date_in_picker(driver, y, m, d):
        dump_debug(driver, "set_date_in_picker_fail")
        raise RuntimeError("[FATAL] ピッカー内で指定日を選べなかったある（dbg参照）")

    for _ in range(100):
        cur = get_current_header_date(driver)
        if cur == (y, m, d):
            wait_table_stable_after_date_change(driver)
            return
        time.sleep(0.2)

    dump_debug(driver, "goto_report_date_fail")
    raise RuntimeError("[FATAL] 指定日へ移動できなかったある（dbg参照）")


# =========================
# テーブル行検索＆鉛筆クリック
# =========================
def normalize_name(name: str) -> str:
    return str(name).replace(" ", "").replace("　", "").replace("柳", "栁")

def _get_report_scroll_container(driver):
    """
    table を含む一番近いスクロール可能親を取る
    """
    script = r"""
    const table = document.querySelector("table");
    if (!table) return null;

    let el = table.parentElement;
    while (el) {
        const st = getComputedStyle(el);
        const oy = st.overflowY;
        if ((oy === 'auto' || oy === 'scroll') && el.scrollHeight > el.clientHeight) {
            return el;
        }
        el = el.parentElement;
    }

    return document.scrollingElement || document.documentElement || document.body;
    """
    return driver.execute_script(script)


from selenium.webdriver.common.action_chains import ActionChains

def find_row_by_name(driver, name: str):
    target = normalize_name(name)

    def scan_rows():
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        visible = []
        for r in rows:
            try:
                txt = normalize_name((r.text or "").replace("\n", " "))
                visible.append(txt)
                if target in txt:
                    return r, visible
            except Exception:
                continue
        return None, visible

    row, visible = scan_rows()

    # 先頭10件
    for i, txt in enumerate(visible[:10], 1):
        log(f"[DEBUG] visible row {i}: {txt}")

    # 末尾5件（10件以上あるときだけ）Q
    if len(visible) > 10:
        for i, txt in enumerate(visible[-5:], 1):
            log(f"[DEBUG] tail row {i}: {txt}")

    if row is not None:
        return row

    # 表の中央あたりを基点にホイールスクロール
    try:
        table = driver.find_element(By.CSS_SELECTOR, "table")
    except Exception:
        log("[DEBUG] table not found")
        return None

    last_sig = ""
    stuck = 0

    for step in range(40):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", table)
            time.sleep(0.15)
        except Exception:
            pass

        # まず現在の見えている行を取る
        row, visible = scan_rows()

        sig = " | ".join(visible)
        log(f"[DEBUG] pre-step={step+1} rows={len(visible)}")
        for i, txt in enumerate(visible[:10], 1):
            log(f"[DEBUG] pre visible row {i}: {txt}")

        if row is not None:
            return row

        # 同じ表示が続いたら小さく刻んで追加スクロール
        if sig == last_sig:
            stuck += 1
            scroll_amount = 120
        else:
            stuck = 0
            last_sig = sig
            scroll_amount = 220

        # Selenium 4 の wheel アクション
        moved = False
        try:
            ActionChains(driver).move_to_element(table).pause(0.08).scroll_by_amount(0, scroll_amount).perform()
            moved = True
        except Exception:
            pass

        # 保険：JSスクロール
        if not moved:
            try:
                driver.execute_script("window.scrollBy(0, arguments[0]);", scroll_amount)
            except Exception:
                pass

        time.sleep(0.45)

        # スクロール後に再取得
        row, visible = scan_rows()

        sig_after = " | ".join(visible)
        log(f"[DEBUG] post-step={step+1} rows={len(visible)} scroll={scroll_amount}")
        for i, txt in enumerate(visible[:10], 1):
            log(f"[DEBUG] post visible row {i}: {txt}")

        if row is not None:
            return row

        # スクロールしても画面が変わっていないなら、もう少しだけ細かく送る
        if sig_after == sig:
            try:
                ActionChains(driver).move_to_element(table).pause(0.05).scroll_by_amount(0, 80).perform()
            except Exception:
                try:
                    driver.execute_script("window.scrollBy(0, 80);")
                except Exception:
                    pass
            time.sleep(0.25)

            row, visible = scan_rows()
            if row is not None:
                return row

            sig_after2 = " | ".join(visible)
            if sig_after2 == sig_after:
                stuck += 1
            else:
                stuck = 0
                last_sig = sig_after2
        else:
            stuck = 0
            last_sig = sig_after

        if stuck >= 5:
            log("[DEBUG] scrolling appears stuck")
            break
        
        time.sleep(0.8)

        row, visible = scan_rows()

        sig = " | ".join(visible)
        log(f"[DEBUG] step={step+1} rows={len(visible)}")
        for i, txt in enumerate(visible[:10], 1):
            log(f"[DEBUG] visible row {i}: {txt}")

        if row is not None:
            return row

        if sig == last_sig:
            stuck += 1
        else:
            stuck = 0
            last_sig = sig

        if stuck >= 3:
            log("[DEBUG] scrolling appears stuck")
            break

    return None
def click_pencil_in_row(driver, row) -> bool:
    try:
        tds = row.find_elements(By.TAG_NAME, "td")
        if tds:
            last = tds[-1]
            for b in last.find_elements(By.TAG_NAME, "button"):
                if safe_click(driver, b):
                    time.sleep(0.5)  # ←ここ！！！！
                    return True
            for svg in last.find_elements(By.TAG_NAME, "svg"):
                if safe_click(driver, svg):
                    time.sleep(0.5)  # ←ここ！！！！
                    return True
    except Exception:
        pass
    return False

# =========================
# 利用者ごと → 支援記録ページ取得
# =========================
def goto_record_user_page(driver):
    """
    左メニューの
      記録 → 利用者ごと
    へ移動する
    """
    log("[STEP] goto_record_user_page start")

    # まず「記録」を押す
    record_clicked = False
    record_xpaths = [
        "//p[normalize-space(.)='記録']",
        "//span[normalize-space(.)='記録']",
        "//div[normalize-space(.)='記録']",
        "//*[contains(normalize-space(.), '記録')]",
    ]

    for xp in record_xpaths:
        try:
            elems = driver.find_elements(By.XPATH, xp)
            for el in elems:
                try:
                    if not el.is_displayed():
                        continue
                except Exception:
                    pass

                if safe_click(driver, el):
                    record_clicked = True
                    time.sleep(1.0)
                    break
            if record_clicked:
                break
        except Exception:
            pass

    if not record_clicked:
        dump_debug(driver, "goto_record_user_page_record_click_fail")
        raise RuntimeError("[FATAL] 左メニューの『記録』を押せなかったある")

    # 次に「利用者ごと」を押す
    user_clicked = False
    user_xpaths = [
        "//p[normalize-space(.)='利用者ごと']",
        "//span[normalize-space(.)='利用者ごと']",
        "//div[normalize-space(.)='利用者ごと']",
        "//*[contains(normalize-space(.), '利用者ごと')]",
    ]

    for xp in user_xpaths:
        try:
            elems = driver.find_elements(By.XPATH, xp)
            for el in elems:
                try:
                    if not el.is_displayed():
                        continue
                except Exception:
                    pass

                if safe_click(driver, el):
                    user_clicked = True
                    time.sleep(1.5)
                    break
            if user_clicked:
                break
        except Exception:
            pass

    if not user_clicked:
        dump_debug(driver, "goto_record_user_page_user_click_fail")
        raise RuntimeError("[FATAL] 『利用者ごと』を押せなかったある")

    log("[STEP] goto_record_user_page done")


def normalize_name_loose(name: str) -> str:
    """
    氏名比較用
    - 半角/全角スペース除去
    - 氏名以降の余計な文字は無視しやすくするため、
      比較時は「対象名が行テキストに含まれるか」で見る
    """
    s = "" if name is None else str(name)
    return s.replace(" ", "").replace("　", "").strip()


def find_user_row_in_record_page(driver, name: str):
    """
    利用者ごとページで、対象利用者の表示ブロックを探す
    - 名前のspanを直接探す
    - スペース無視
    - 氏名以降の文字は無視
    """
    target = normalize_name_loose(name)
    log(f"[STEP] find_user_row_in_record_page target={target}")

    try:
        spans = driver.find_elements(By.XPATH, "//span[contains(@class, 'jss509')]")
    except Exception:
        spans = []

    visible_texts = []

    for idx, sp in enumerate(spans, 1):
        try:
            txt = (sp.text or "").replace("\n", " ")
            txt_norm = normalize_name_loose(txt)
            visible_texts.append(txt_norm)

            log(f"[DEBUG] visible user span {idx}: {txt_norm}")

            if target and target in txt_norm:
                # 親方向へ登って、その人1件分のブロックを取る
                cur = sp
                for _ in range(8):
                    try:
                        cur = cur.find_element(By.XPATH, "..")
                    except Exception:
                        break

                    try:
                        block_text = normalize_name_loose((cur.text or "").replace("\n", " "))
                        if "支援記録" in block_text:
                            log(f"[DEBUG] matched user block idx={idx}: {block_text}")
                            return cur
                    except Exception:
                        pass

                # 最後の保険：spanそのものを返す
                return sp

        except Exception:
            continue

    return None


def click_support_record_button_in_row(driver, row) -> bool:
    """
    利用者ごとページの右側『支援記録』ボタンを押す
    """
    search_roots = [row]

    # spanだけ返ってきたときのため、親にも広げる
    cur = row
    for _ in range(6):
        try:
            cur = cur.find_element(By.XPATH, "..")
            search_roots.append(cur)
        except Exception:
            break

    for root in search_roots:
        xps = [
            ".//button[normalize-space(.)='支援記録']",
            ".//span[normalize-space(.)='支援記録']",
            ".//*[contains(normalize-space(.), '支援記録')]",
        ]

        for xp in xps:
            try:
                elems = root.find_elements(By.XPATH, xp)
            except Exception:
                elems = []

            for el in elems:
                try:
                    if not el.is_displayed():
                        continue
                except Exception:
                    pass

                try:
                    btn = el.find_element(By.XPATH, "./ancestor::button[1]")
                    if safe_click(driver, btn):
                        time.sleep(1.5)
                        return True
                except Exception:
                    pass

                if safe_click(driver, el):
                    time.sleep(1.5)
                    return True

    return False


def open_support_record_page_for_user(driver, resident_name: str):
    """
    利用者ごとページで対象利用者を見つけて『支援記録』を開く
    """
    log("[STEP] open_support_record_page_for_user start")

    row = find_user_row_in_record_page(driver, resident_name)
    if row is None:
        dump_debug(driver, "support_record_user_not_found")
        raise RuntimeError(f"[FATAL] 利用者ごとページで対象者が見つからないある: {resident_name}")

    ok = click_support_record_button_in_row(driver, row)
    if not ok:
        dump_debug(driver, "support_record_button_not_found")
        raise RuntimeError(f"[FATAL] 『支援記録』ボタンを押せなかったある: {resident_name}")

    # ページ遷移待ち
    time.sleep(2.0)
    log("[STEP] open_support_record_page_for_user done")


# =========================
# 支援記録ページ 年月移動
# =========================
def parse_support_record_ym(text: str):
    """
    '2026年3月' → (2026, 3)
    """
    s = "" if text is None else str(text).strip()
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def get_support_record_current_ym(driver):
    """
    支援記録ページ中央の年月表示を読む
    例: <span class="jss247">2026<span>年</span>3<span>月</span></span>
    """
    xpaths = [
        "//span[contains(normalize-space(.), '年') and contains(normalize-space(.), '月')]",
        "//*[contains(normalize-space(.), '年') and contains(normalize-space(.), '月')]",
    ]

    for xp in xpaths:
        try:
            elems = driver.find_elements(By.XPATH, xp)
        except Exception:
            elems = []

        for el in elems:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                pass

            try:
                txt = (el.text or "").strip()
                ym = parse_support_record_ym(txt)
                if ym:
                    log(f"[DEBUG] current support ym text={txt}")
                    return ym
            except Exception:
                continue

    return None


def _month_index(y: int, m: int) -> int:
    return y * 12 + (m - 1)


def get_support_record_month_nav_buttons(driver):
    """
    支援記録ページの左右ボタンを取る
    - left / right class を優先
    """
    left_btn = None
    right_btn = None

    try:
        btns = driver.find_elements(By.XPATH, "//button")
    except Exception:
        btns = []

    for b in btns:
        try:
            cls = (b.get_attribute("class") or "").strip().lower()
        except Exception:
            cls = ""

        if "left" in cls and left_btn is None:
            left_btn = b
        if "right" in cls and right_btn is None:
            right_btn = b

    return left_btn, right_btn


def wait_support_record_month_changed(driver, old_ym, timeout=10):
    t0 = time.time()
    while time.time() - t0 < timeout:
        cur = get_support_record_current_ym(driver)
        if cur and cur != old_ym:
            time.sleep(1.0)
            return True
        time.sleep(0.2)
    return False


# =========================
# 支援記録本文の取得
# =========================
def get_support_record_page_text(driver) -> str:
    """
    支援記録ページ本文を取得する
    """
    log("[STEP] get_support_record_page_text start")

    time.sleep(1.0)

    # まず「利用実績なし」を確認
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text or ""
    except Exception:
        body_text = ""

    if "利用実績がありません" in body_text and "利用実績を入力後、ご利用ください" in body_text:
        log("[DEBUG] 支援記録なし（利用実績なし）ある")
        return ""

    script = r"""
    function isVisible(el) {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        if (style.display === 'none' || style.visibility === 'hidden') return false;
        const rect = el.getBoundingClientRect();
        return rect.width > 0 && rect.height > 0;
    }

    function norm(s) {
        return (s || "").replace(/\r/g, "").trim();
    }

    const texts = [];
    const seen = new Set();

    const areas = Array.from(document.querySelectorAll("textarea[readonly]"))
        .filter(el => isVisible(el))
        .filter(el => (el.getAttribute("aria-hidden") || "").toLowerCase() !== "true");

    for (const ta of areas) {
        const v = norm(ta.value || ta.innerText || ta.textContent || "");
        if (!v) continue;
        if (!seen.has(v)) {
            seen.add(v);
            texts.push(v);
        }
    }

    if (texts.length === 0) {
        const all = Array.from(document.querySelectorAll("body *"))
            .filter(el => isVisible(el));

        for (const el of all) {
            const tag = (el.tagName || "").toLowerCase();
            if (["button", "svg", "path"].includes(tag)) continue;

            const rect = el.getBoundingClientRect();
            if (rect.left < 150) continue;

            const t = norm(el.innerText || el.textContent || "");
            if (!t) continue;
            if (t.length < 20) continue;

            if (!seen.has(t)) {
                seen.add(t);
                texts.push(t);
            }
        }
    }

    return texts.join("\n\n====================\n\n");
    """

    text = ""
    try:
        text = driver.execute_script(script)
    except Exception:
        text = ""

    text = "" if text is None else str(text).strip()

    if not text:
        log("[DEBUG] 支援記録本文なしある")
        return ""

    log(f"[DEBUG] support record text length={len(text)}")
    return text

""""
def fetch_support_record_text_for_month(driver, resident_name: str, year: int, month: int) -> str:
    log("[STEP] fetch_support_record_text_for_month start")

    ok = goto_users_summary(driver)
    if not ok:
        dump_debug(driver, "goto_users_summary_fail")
        raise RuntimeError("[FATAL] 利用者ごと一覧へ戻れません")

    time.sleep(1.0)

    ok = open_support_record_for_resident(driver, resident_name)
    if not ok:
        dump_debug(driver, "open_support_record_for_resident_fail")
        raise RuntimeError(f"[FATAL] 利用者一覧で対象利用者が見つかりません: {resident_name}")

    cur = driver.current_url or ""
    if "support_plan" in cur or "assessment" in cur or (not _is_real_support_record_url(cur)):
        dump_debug(driver, "wrong_page_after_open_support_record")
        raise RuntimeError(f"[FATAL] 支援記録ではないページに遷移しました: {cur}")

    ok = goto_support_record_month(driver, int(year), int(month))
    if not ok:
        dump_debug(driver, "goto_support_record_month_fail")
        raise RuntimeError(f"[FATAL] 支援記録の対象月へ移動できません: {year}/{month}")

    cur = driver.current_url or ""
    if "support_plan" in cur or "assessment" in cur or (not _is_real_support_record_url(cur)):
        dump_debug(driver, "wrong_page_after_month_jump")
        raise RuntimeError(f"[FATAL] 月移動後に支援記録ページではありません: {cur}")

    text = get_support_record_page_text(driver)

    log("[STEP] fetch_support_record_text_for_month done")
    return text
"""

# =========================
# 支援記録 → 利用形態一覧化（期間一括）
# =========================

DAY_HEADER_RE = re.compile(r"^(\d{1,2})日[（(]([^)）]+)[)）]$")
SUPPORT_ITEM_LABELS = ["就労先企業", "作業", "利用者状態", "職員考察", "面談", "その他"]


def _iter_year_months(start_year: int, start_month: int, end_year: int, end_month: int):
    y = int(start_year)
    m = int(start_month)
    ey = int(end_year)
    em = int(end_month)

    while (y, m) <= (ey, em):
        yield y, m
        m += 1
        if m >= 13:
            y += 1
            m = 1


def _clean_support_lines(text: str):
    text = "" if text is None else str(text)
    raw_lines = text.replace("\r", "\n").split("\n")

    lines = []
    for x in raw_lines:
        s = str(x).strip()
        if not s:
            continue
        lines.append(s)

    return lines


def _extract_support_body_text_for_parse(driver) -> str:
    """
    支援記録ページ全体の可視テキストを取る
    月一覧の1日ごとの表示をテキスト解析しやすい形にするため、body.text を使う
    """
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text or ""
    except Exception:
        body_text = ""

    body_text = body_text.strip()
    if not body_text:
        raise RuntimeError("[FATAL] 支援記録ページのbody text取得に失敗したある")

    return body_text


def _split_support_day_blocks_from_text(body_text: str):
    """
    支援記録ページの全文から、1日ごとのブロックへ分割する
    """
    lines = _clean_support_lines(body_text)

    start_idx = None
    for i, line in enumerate(lines):
        if DAY_HEADER_RE.match(line):
            start_idx = i
            break

    if start_idx is None:
        return []

    lines = lines[start_idx:]

    blocks = []
    current = None

    for line in lines:
        m = DAY_HEADER_RE.match(line)
        if m:
            if current:
                blocks.append(current)

            current = {
                "day": int(m.group(1)),
                "weekday": str(m.group(2)).strip(),
                "lines": [line],
            }
        else:
            if current is not None:
                current["lines"].append(line)

    if current:
        blocks.append(current)

    return blocks


def _detect_registered_kind_from_block_lines(lines):
    """
    画面の登録区分判定
    ルール（今回版）
    - 施設外就労 = 施設外
    - 通所 + 食事/あり がある = 通所
    - 通所 + それ以外 = 在宅
    """
    joined = "\n".join([str(x).strip() for x in lines if str(x).strip()])

    # 施設外
    if "施設外就労" in joined:
        return "施設外"

    # 通所
    if "通所" in joined:
        # 食事あり パターン
        for i, line in enumerate(lines):
            s = str(line).strip()

            # 「食事」→ 次行が「あり」
            if s == "食事":
                if i + 1 < len(lines):
                    nxt = str(lines[i + 1]).strip()
                    if nxt == "あり":
                        return "通所"

            # 1行にまとまってるパターン保険
            if "食事" in s and "あり" in s:
                return "通所"

            if "食事提供" in s and "あり" in s:
                return "通所"

        # 通所だけど食事ありが見えない → 在宅
        return "在宅"

    return "不明"


def _parse_support_sections_from_block_lines(lines):
    """
    1日ブロック内の項目を辞書にする
    例:
      利用者状態 -> 本文
      職員考察 -> 本文
    """
    sections = {}
    i = 0
    n = len(lines)

    while i < n:
        label = str(lines[i]).strip()

        if label in SUPPORT_ITEM_LABELS:
            i += 1
            vals = []

            while i < n:
                cur = str(lines[i]).strip()
                if cur in SUPPORT_ITEM_LABELS:
                    break
                vals.append(cur)
                i += 1

            sections[label] = "\n".join(vals).strip()
            continue

        i += 1

    return sections

def _contains_any(text: str, keywords):
    s = "" if text is None else str(text)
    for kw in keywords:
        if kw in s:
            return True
    return False


def _rule_based_diary_kind_safe(diary_text: str) -> str:
    """
    安全優先の事前判定
    - 自信があるときだけ返す
    - 少しでも怪しければ 判定できず
    """
    s = "" if diary_text is None else str(diary_text).strip()
    if not s:
        return "判定できず"

    # -------------------------
    # 施設外（強い根拠）
    # -------------------------
    outside_words = [
        "施設外就労",
        "就労先企業",
        "清掃",
        "拭き掃除",
        "ほうき",
        "モップ",
        "雑巾",
        "手すり",
        "ゴミ拾い",
        "マンション",
        "配電盤",
        "消火器",
        "エバーグリーン",
        "施設外",
    ]
    if _contains_any(s, outside_words):
        return "施設外"

    # -------------------------
    # 在宅（伝聞・電話/LINE連絡系）
    # -------------------------
    home_words = [
        "とのこと",
        "と言っていた",
        "と話していた",
        "とお話があり",
        "と連絡があり",
        "連絡があり",
        "電話",
        "LINE",
        "作業開始の連絡",
        "作業開始前の連絡",
        "作業終了の連絡",
        "開始の連絡",
        "終了の報告",
        "終了報告",
        "開始報告",
        "ご本人から",
        "在宅",
        "自宅",
    ]

    # -------------------------
    # 通所（来所・複数人関与・帰所系）
    # -------------------------
    day_words = [
        "来所",
        "通所",
        "談笑",
        "コミュニケーション",
        "役割分担",
        "他の利用者",
        "みんなと",
        "一緒に",
        "職員と",
        "周囲と",
        "帰る",
        "帰宅",
        "帰所",
        "退所",
        "出勤",
        "登所",
    ]

    has_home = _contains_any(s, home_words)
    has_day = _contains_any(s, day_words)

    # 在宅/通所が両方立つなら危険なので保留
    if has_home and has_day:
        return "判定できず"

    if has_home:
        return "在宅"

    if has_day:
        return "通所"

    return "判定できず"

def _build_diary_text_for_gemini(sections: dict) -> str:
    """
    Geminiに渡す日誌本文
    今回はなるべく判断材料を残すため
      作業 + 利用者状態 + 職員考察
    をまとめて渡す
    """
    parts = []

    work_text = str(sections.get("作業", "")).strip()
    user_state = str(sections.get("利用者状態", "")).strip()
    staff_note = str(sections.get("職員考察", "")).strip()

    if work_text:
        parts.append(f"【作業】\n{work_text}")

    if user_state:
        parts.append(f"【利用者状態】\n{user_state}")

    if staff_note:
        parts.append(f"【職員考察】\n{staff_note}")

    return "\n\n".join(parts).strip()


def _normalize_gemini_kind(text: str) -> str:
    s = "" if text is None else str(text).strip()

    if "施設外" in s:
        return "施設外"
    if "在宅" in s:
        return "在宅"
    if "通所" in s:
        return "通所"
    if "判定できず" in s:
        return "判定できず"
    if "不明" in s:
        return "判定できず"

    # 保険
    if "来所" in s:
        return "通所"

    return "判定できず"


def classify_support_diary_kind_with_gemini(client, diary_text: str) -> str:
    """
    diary_text の内容だけを見て
    在宅 / 通所 / 施設外 / 判定できず
    を返す

    方針:
    1) まずコード側で安全判定
    2) 判定できないときだけ Gemini
    3) Geminiでも怪しければ 判定できず
    """
    diary_text = "" if diary_text is None else str(diary_text).strip()
    if not diary_text:
        return "判定できず"

    # まず安全なルール判定
    safe_kind = _rule_based_diary_kind_safe(diary_text)
    if safe_kind != "判定できず":
        return safe_kind

    prompt = f"""以下は就労継続支援B型の支援記録の日誌本文です。
本文だけを見て、次の4択から最も安全に判定してください。

選択肢:
- 在宅
- 通所
- 施設外
- 判定できず

重要ルール:
- 間違った判定をするくらいなら、必ず「判定できず」を選ぶこと
- 自信が弱い場合は必ず「判定できず」
- 無理に推測しないこと
- 出力は1語のみ
- 理由は書かないこと

判定基準:
【在宅】
- 「〜とのこと」「〜と言っていた」など伝聞調
- 電話やLINEなどで連絡を取ったことが分かる
- 「連絡」「作業開始の連絡」「終了報告」など開始/終了を遠隔でやり取りしている
- 自宅での作業と読み取れる

【通所】
- 「来所」がある
- 「コミュニケーション」「談笑」「役割分担」など複数人との関わりが分かる
- 「帰る」「帰宅」「退所」など事業所から去ることが分かる
- 事業所内で一緒に作業している様子がある

【施設外】
- 「清掃」「拭き掃除」「ほうき」「モップ」「雑巾」など清掃系作業
- 就労先企業や施設外就労の文脈
- 内職ではなく外部作業先での作業と分かる

【判定できず】
- 上のどれとも断定しにくい
- 在宅と通所の要素が混ざる
- 決め手が弱い
- 少しでも怪しい

本文:
{diary_text}
"""

    try:
        result_text = _gemini_generate_text(client, prompt)
        normalized = _normalize_gemini_kind(result_text)

        # Geminiが曖昧なら即保留
        if normalized not in ["在宅", "通所", "施設外", "判定できず"]:
            return "判定できず"

        return normalized

    except Exception:
        return "判定できず"

def collect_support_record_daily_kind_rows_for_month(driver, resident_name: str, year: int, month: int, gemini_client=None):
    """
    いま開いている支援記録ページ（対象月）から、
    1日ごとの
      日付 / 曜日 / 登録判定 / 日誌判定 / 日誌本文
    を抽出する
    """
    body_text = _extract_support_body_text_for_parse(driver)
    day_blocks = _split_support_day_blocks_from_text(body_text)

    rows = []

    for block in day_blocks:
        block_lines = block.get("lines", [])
        if not block_lines:
            continue

        # 先頭の日付行を除いた内容
        content_lines = block_lines[1:]

        registered_kind = _detect_registered_kind_from_block_lines(content_lines)
        sections = _parse_support_sections_from_block_lines(content_lines)
        diary_text = _build_diary_text_for_gemini(sections)

        diary_kind = "判定できず"
        if gemini_client is not None:
            diary_kind = classify_support_diary_kind_with_gemini(gemini_client, diary_text)

        rows.append({
            "year": int(year),
            "month": int(month),
            "day": int(block.get("day", 0)),
            "weekday": str(block.get("weekday", "")).strip(),
            "registered_kind": registered_kind,
            "resident_name": str(resident_name).strip(),
            "diary_kind": diary_kind,
            "diary_text": diary_text,
            "user_state": str(sections.get("利用者状態", "")).strip(),
            "staff_note": str(sections.get("職員考察", "")).strip(),
            "work_text": str(sections.get("作業", "")).strip(),
            "company_text": str(sections.get("就労先企業", "")).strip(),
            "raw_block_text": "\n".join(block_lines).strip(),
        })

    return rows


def export_support_record_kind_rows_to_excel(rows, output_path: str):
    """
    簡易一覧Excel
    A: 利用者
    B: 日付
    C: 登録
    D: 日誌
    E: 判定できず
    F: 一致判定（一致=1 / 不一致=0）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 見出し
    ws["A1"] = "利用者"
    ws["B1"] = "日付"
    ws["C1"] = "登録"
    ws["D1"] = "日誌"
    ws["E1"] = "判定できず"
    ws["F1"] = "登録と一致する場合は 1\n登録と一致しない場合は 0"

    row_no = 2

    for item in rows:
        resident_name = str(item.get("resident_name", "")).strip()
        y = int(item.get("year", 0))
        m = int(item.get("month", 0))
        d = int(item.get("day", 0))

        registered_kind = str(item.get("registered_kind", "")).strip()
        diary_kind = str(item.get("diary_kind", "")).strip()

        # E列
        judge_cannot = "1" if diary_kind == "判定できず" else ""

        # F列
        match_flag = ""
        if diary_kind != "判定できず":
            match_flag = "1" if registered_kind == diary_kind else "0"

        ws.cell(row=row_no, column=1, value=resident_name)
        ws.cell(row=row_no, column=2, value=f"{y}/{m}/{d}")
        ws.cell(row=row_no, column=3, value=registered_kind)
        ws.cell(row=row_no, column=4, value=diary_kind)
        ws.cell(row=row_no, column=5, value=judge_cannot)
        ws.cell(row=row_no, column=6, value=match_flag)

        row_no += 1

    # 列幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 32

    # 折り返し
    from openpyxl.styles import Alignment
    ws["F1"].alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    return output_path


def fetch_support_record_kind_rows_for_range(
    driver,
    resident_name: str,
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
    gemini_client=None,
):
    """
    既存driverを使って、
    対象利用者の支援記録ページを1回開き、
    指定期間を月ごとに回して一覧データを返す
    """
    log("[STEP] fetch_support_record_kind_rows_for_range start")

    # 新しい安定ルートを使う
    goto_users_summary(driver)
    apply_users_summary_filter_show_expired(driver)

    ok = open_support_record_for_resident(driver, resident_name)
    if not ok:
        raise RuntimeError(f"[FATAL] 対象利用者の支援記録ページを開けなかったある: {resident_name}")

    all_rows = []

    for y, m in _iter_year_months(start_year, start_month, end_year, end_month):
        log(f"[STEP] support kind collect target={y}-{m:02d}")

        moved = goto_support_record_month(driver, y, m)
        if not moved:
            raise RuntimeError(f"[FATAL] 支援記録ページを {y}-{m:02d} に移動できなかったある: {resident_name}")

        time.sleep(1.0)

        month_rows = collect_support_record_daily_kind_rows_for_month(
            driver=driver,
            resident_name=resident_name,
            year=y,
            month=m,
            gemini_client=gemini_client,
        )

        all_rows.extend(month_rows)

    log("[STEP] fetch_support_record_kind_rows_for_range done")
    return all_rows


def run_support_record_kind_export(
    resident_name: str,
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
    output_path: str,
    login_username: str = "",
    login_password: str = "",
    gemini_client=None,
):
    """
    単発実行用
    - Knowbeログイン
    - 対象利用者の指定期間を全部読む
    - 登録判定 / 日誌判定一覧をExcel出力
    """
    resident_name = "" if resident_name is None else str(resident_name).strip()
    output_path = "" if output_path is None else str(output_path).strip()

    if not resident_name:
        raise RuntimeError("[FATAL] resident_name が空ある")
    if not output_path:
        raise RuntimeError("[FATAL] output_path が空ある")

    if not login_username or not login_password:
        login_username, login_password = get_knowbe_login_credentials()

    if not login_username or not login_password:
        raise RuntimeError("[FATAL] Knowbeログイン情報が空ある")

    driver = build_chrome_driver()

    try:
        log("[STEP] goto report daily")
        goto_report_daily(driver)

        log("[STEP] login")
        manual_login_wait(driver, login_username, login_password)

        rows = fetch_support_record_kind_rows_for_range(
            driver=driver,
            resident_name=resident_name,
            start_year=int(start_year),
            start_month=int(start_month),
            end_year=int(end_year),
            end_month=int(end_month),
            gemini_client=gemini_client,
        )

        export_support_record_kind_rows_to_excel(rows, output_path)
        return rows

    finally:
        try:
            driver.quit()
        except Exception:
            pass

# =========================
# 入力：強制クリア＆JS注入
# =========================
def _js_set_value_and_fire(driver, el, value: str):
    driver.execute_script("""
        const el = arguments[0];
        const v  = arguments[1];

        el.focus();
        const proto = Object.getPrototypeOf(el);
        const desc = Object.getOwnPropertyDescriptor(proto, 'value');
        if (desc && desc.set) desc.set.call(el, v);
        else el.value = v;

        el.dispatchEvent(new Event('input',  {bubbles:true}));
        el.dispatchEvent(new Event('change', {bubbles:true}));
        el.dispatchEvent(new Event('blur',   {bubbles:true}));
    """, el, value)

def _clear_input_strong(driver, el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    except Exception:
        pass
    safe_click(driver, el)
    time.sleep(0.05)

    try:
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
    except Exception:
        pass
    try:
        for _ in range(10):
            el.send_keys(Keys.BACKSPACE)
    except Exception:
        pass
    try:
        _js_set_value_and_fire(driver, el, "")
    except Exception:
        pass

def set_input_value(driver, el, value: str):
    value = "" if value is None else str(value)
    _clear_input_strong(driver, el)
    time.sleep(0.05)
    try:
        el.send_keys(value)
        el.send_keys(Keys.TAB)
    except Exception:
        pass
    try:
        _js_set_value_and_fire(driver, el, value)
    except Exception:
        pass

def enter_edit_mode(driver):
    import time
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    btns = driver.find_elements(By.TAG_NAME, "button")

    for b in btns:
        txt = (b.text or "").strip()
        if txt == "編集":
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b)
            time.sleep(0.3)
            if not safe_click(driver, b):
                driver.execute_script("arguments[0].click();", b)
            break
    else:
        raise RuntimeError("編集ボタンが見つからない")

    WebDriverWait(driver, 10).until(
        lambda d: any("保存" in ((x.text or "").strip()) for x in d.find_elements(By.TAG_NAME, "button"))
    )

    time.sleep(0.5)
    print("[FIX] 編集モードON", flush=True)


def save_all(driver):
    import time
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    btns = driver.find_elements(By.TAG_NAME, "button")

    for b in btns:
        txt = (b.text or "").strip()
        if "保存" in txt:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b)
            time.sleep(0.2)
            if not safe_click(driver, b):
                driver.execute_script("arguments[0].click();", b)
            break
    else:
        raise RuntimeError("保存ボタンが見つからない")

    time.sleep(1.0)
    print("[FIX] 保存完了", flush=True)

# =========================
# ドロップダウン選択（強化版）
# =========================
def _text_norm(t: str) -> str:
    return re.sub(r"\s+", " ", (t or "")).strip()

def _find_field_container_by_label(root, label_text: str):
    """
    ラベルの for 属性から、対応するフィールド本体を安定取得する。
    jss*** のような毎回変わる class 名には依存しない。
    """
    label_text = (label_text or "").strip()
    if not label_text:
        return None

    # まずラベル本文で候補を探す
    xps = [
        f".//label[normalize-space(.)='{label_text}']",
        f".//label[contains(normalize-space(.), '{label_text}')]",
        f".//*[self::label or @for][contains(normalize-space(.), '{label_text}')]",
    ]

    labels = []
    for xp in xps:
        try:
            labels.extend(root.find_elements(By.XPATH, xp))
        except Exception:
            pass

    for lab in labels:
        try:
            if not lab.is_displayed():
                continue
        except Exception:
            pass

        # もっとも安定：for="initial.status" → id="select-initial.status"
        try:
            target_for = (lab.get_attribute("for") or "").strip()
            if target_for:
                try:
                    el = root.find_element(By.ID, f"select-{target_for}")
                    if el.is_displayed():
                        return el
                except Exception:
                    pass
                try:
                    el = root.find_element(By.ID, target_for)
                    if el.is_displayed():
                        return el
                except Exception:
                    pass
        except Exception:
            pass

        # 保険：近い親の中の role=button を取る
        cur = lab
        for _ in range(6):
            try:
                cur = cur.find_element(By.XPATH, "..")
            except Exception:
                break

            try:
                btns = cur.find_elements(By.XPATH, ".//*[@role='button' and @aria-haspopup='true']")
            except Exception:
                btns = []

            for b in btns:
                try:
                    if b.is_displayed():
                        return b
                except Exception:
                    pass

    return None

def _open_dropdown(driver, container) -> bool:
    """
    select本体（role=button）を確実に開く。
    """
    candidates = []

    try:
        if container.get_attribute("role") == "button":
            candidates.append(container)
    except Exception:
        pass

    xps = [
        ".//*[@role='button' and @aria-haspopup='true']",
        ".//*[@role='button']",
        ".//*[@aria-haspopup='true']",
        ".//*[@aria-haspopup='listbox']",
        ".//div",
    ]

    for xp in xps:
        try:
            candidates.extend(container.find_elements(By.XPATH, xp))
        except Exception:
            pass

    seen = set()
    uniq = []
    for el in candidates:
        try:
            key = el.id
        except Exception:
            key = None
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        uniq.append(el)

    for el in uniq[:30]:
        try:
            if not el.is_displayed():
                continue
        except Exception:
            pass

        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.05)
        except Exception:
            pass

        if safe_click(driver, el):
            time.sleep(0.3)
            # メニューが出たか確認
            try:
                if (
                    driver.find_elements(By.XPATH, "//ul[@role='listbox']") or
                    driver.find_elements(By.XPATH, "//div[@role='listbox']") or
                    driver.find_elements(By.XPATH, "//li[@role='option']") or
                    driver.find_elements(By.XPATH, "//body//li")
                ):
                    return True
            except Exception:
                pass

    return False

def _get_open_menu(driver):
    """
    開いている Material-UI のメニュー本体を拾う
    """
    xps = [
        "//ul[@role='listbox']",
        "//div[@role='listbox']",
        "//ul[contains(@class,'MuiMenu-list')]",
        "//div[contains(@class,'MuiPopover-root')]",
        "//div[contains(@class,'MuiDialog-root')]//ul",
        "//body",
    ]

    found = []
    for xp in xps:
        try:
            found.extend(driver.find_elements(By.XPATH, xp))
        except Exception:
            pass

    for el in reversed(found):
        try:
            if el.is_displayed():
                return el
        except Exception:
            continue

    return None

def _choose_option_from_open_menu(driver, value_text: str) -> bool:
    """
    開いているメニューから表示文字で選ぶ
    """
    value_text = (value_text or "").strip()
    if not value_text:
        return False

    menu = _get_open_menu(driver)
    if not menu:
        return False

    option_xps = [
        f".//li[normalize-space(.)='{value_text}']",
        f".//li[contains(normalize-space(.), '{value_text}')]",
        f".//*[@role='option'][normalize-space(.)='{value_text}']",
        f".//*[@role='option'][contains(normalize-space(.), '{value_text}')]",
        f".//*[self::span or self::div][normalize-space(.)='{value_text}']",
        f".//*[self::span or self::div][contains(normalize-space(.), '{value_text}')]",
    ]

    for xp in option_xps:
        try:
            elems = menu.find_elements(By.XPATH, xp)
        except Exception:
            elems = []

        for el in elems[:50]:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                pass

            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.05)
            except Exception:
                pass

            if safe_click(driver, el):
                time.sleep(0.2)
                return True

            try:
                parent_li = el.find_element(By.XPATH, "./ancestor::li[1]")
                if safe_click(driver, parent_li):
                    time.sleep(0.2)
                    return True
            except Exception:
                pass

    return False

def _norm_name_for_match(s: str) -> str:
    s = "" if s is None else str(s)
    return (
        s.replace(" ", "")
         .replace("　", "")
         .replace("\n", "")
         .replace("\t", "")
         .strip()
    )

def find_user_card_by_name(driver, resident_name: str):
    """
    利用者ごと一覧で、対象利用者の表示ブロックを探す
    方針:
    1) まず画面上の短い名前候補を広く拾う
    2) 空白除去して完全一致
    3) 一致した要素から親をたどって「1人分のカード」を返す
    """
    target = _norm_name_for_match(resident_name)
    log(f"[STEP] find_user_card_by_name target={target}")

    # まず表示要素を広めに拾う
    xpaths = [
        "//span[normalize-space(.) != '']",
        "//p[normalize-space(.) != '']",
        "//div[normalize-space(.) != '']",
        "//a[normalize-space(.) != '']",
    ]

    candidates = []

    for xp in xpaths:
        try:
            elems = driver.find_elements(By.XPATH, xp)
        except Exception:
            elems = []

        for el in elems:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                pass

            try:
                txt = (el.text or "").strip()
            except Exception:
                txt = ""

            if not txt:
                continue

            txt_norm = _norm_name_for_match(txt)

            # 長すぎるブロックは除外
            if len(txt_norm) == 0 or len(txt_norm) > 20:
                continue

            candidates.append((el, txt, txt_norm))

    log(f"[DEBUG] candidate count = {len(candidates)}")

    for i, (_, raw, normed) in enumerate(candidates[:50], start=1):
        log(f"[DEBUG] candidate {i}: {normed}")

    # 完全一致を最優先
    for i, (el, raw, txt_norm) in enumerate(candidates, start=1):
        if txt_norm != target:
            continue

        log(f"[STEP] matched candidate {i}: {txt_norm}")

        # 親をたどって「その人のカード」を探す
        cur = el
        for _ in range(8):
            try:
                cur = cur.find_element(By.XPATH, "./..")
            except Exception:
                break

            try:
                block_text = _norm_name_for_match(cur.text or "")
            except Exception:
                block_text = ""

            # 本人名を含み、かつ支援記録などの操作要素を持つ親をカード扱い
            if target in block_text:
                try:
                    btns = cur.find_elements(By.XPATH, ".//button")
                except Exception:
                    btns = []

                if btns:
                    return cur

        # 最後の保険
        return el

    log(f"[DEBUG] user card not found by exact short-text match: {resident_name}")
    return None


def select_dropdown_skip_if_same(driver, root, label_text: str, value_text: str) -> bool:
    """
    jss番号ではなく、label for -> select-initial.xxx の流れで安定選択する。
    """
    value_text = _text_norm(value_text)
    if not value_text:
        return False

    cont = _find_field_container_by_label(root, label_text)
    if not cont:
        return False

    # 現在値確認
    try:
        now = _text_norm(cont.text or "")
        if value_text == now:
            return True
    except Exception:
        pass

    if not _open_dropdown(driver, cont):
        return False

    if _choose_option_from_open_menu(driver, value_text):
        return True

    # 保険：ESCして再試行
    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(0.2)
    except Exception:
        pass

    if not _open_dropdown(driver, cont):
        return False

    return _choose_option_from_open_menu(driver, value_text)


# =========================
# フィールド探索（時間・備考）
# =========================
def _find_time_input(root, kind: str):
    """
    利用実績モーダル上段の
      開始時間 / 終了時間
    を最優先で安定取得する。

    下段の
      作業開始時間 / 作業終了時間
    は絶対に取らない。
    """
    # ===== 1) まず name 属性で直接狙う =====
    direct_candidates = []
    if kind == "start":
        direct_candidates = [
            "input[name='initial.startTime']",
            "input[name='startTime']",
            "input[name='service.startTime']",
        ]
    else:
        direct_candidates = [
            "input[name='initial.endTime']",
            "input[name='endTime']",
            "input[name='service.endTime']",
        ]

    for css in direct_candidates:
        try:
            el = root.find_element(By.CSS_SELECTOR, css)
            if el and el.is_displayed():
                return el
        except Exception:
            pass

    # ===== 2) ラベル近傍で探す =====
    if kind == "start":
        target_labels = ["開始時間", "利用開始時間", "サービス開始時間"]
        ng_words = ["作業開始時間"]
    else:
        target_labels = ["終了時間", "利用終了時間", "サービス終了時間"]
        ng_words = ["作業終了時間"]

    for label_text in target_labels:
        try:
            labels = root.find_elements(
                By.XPATH,
                f".//*[normalize-space(text())='{label_text}' or contains(normalize-space(.), '{label_text}')]"
            )
        except Exception:
            labels = []

        for lab in labels:
            try:
                txt = (lab.text or "").strip()
                if not txt:
                    continue
                if any(ng in txt for ng in ng_words):
                    continue
                if "作業" in txt:
                    continue
            except Exception:
                continue

            cur = lab
            for _ in range(6):
                try:
                    cur = cur.find_element(By.XPATH, "..")
                except Exception:
                    break

                try:
                    area_text = (cur.text or "").strip()
                except Exception:
                    area_text = ""

                if "作業開始時間" in area_text or "作業終了時間" in area_text:
                    continue

                try:
                    inputs = cur.find_elements(By.XPATH, ".//input")
                except Exception:
                    inputs = []

                visible_inputs = []
                for inp in inputs:
                    try:
                        if inp.is_displayed():
                            visible_inputs.append(inp)
                    except Exception:
                        pass

                if len(visible_inputs) == 1:
                    return visible_inputs[0]

                if len(visible_inputs) >= 2:
                    try:
                        visible_inputs = sorted(visible_inputs, key=lambda e: e.location["x"])
                    except Exception:
                        pass
                    return visible_inputs[0] if kind == "start" else visible_inputs[-1]

            try:
                cand = lab.find_element(
                    By.XPATH,
                    "./following::input[not(ancestor::*[contains(., '作業開始時間') or contains(., '作業終了時間')])][1]"
                )
                if cand and cand.is_displayed():
                    return cand
            except Exception:
                pass

    # ===== 3) 最後の保険：上側だけ拾う =====
    try:
        all_inputs = root.find_elements(By.XPATH, ".//input")
    except Exception:
        all_inputs = []

    top_candidates = []
    for inp in all_inputs:
        try:
            if not inp.is_displayed():
                continue
        except Exception:
            continue

        try:
            parent = inp.find_element(By.XPATH, "./ancestor::*[self::div or self::td][1]")
            ptxt = (parent.text or "").strip()
            if "作業開始時間" in ptxt or "作業終了時間" in ptxt:
                continue
        except Exception:
            pass

        try:
            x = inp.location["x"]
            y = inp.location["y"]
        except Exception:
            x, y = 0, 0

        top_candidates.append((y, x, inp))

    if top_candidates:
        top_candidates.sort(key=lambda t: (t[0], t[1]))
        first_two = [t[2] for t in top_candidates[:2]]
        if len(first_two) == 1:
            return first_two[0]
        if len(first_two) >= 2:
            try:
                first_two = sorted(first_two, key=lambda e: e.location["x"])
            except Exception:
                pass
            return first_two[0] if kind == "start" else first_two[-1]

    return None

def _find_remark_area(root):
    for ta in root.find_elements(By.TAG_NAME, "textarea"):
        try:
            if ta.is_displayed():
                return ta
        except Exception:
            continue
    # inputしか無い場合の保険
    try:
        node = root.find_element(By.XPATH, ".//*[contains(normalize-space(.),'備考')]")
        cand = node.find_element(By.XPATH, ".//following::textarea[1] | .//following::input[1]")
        if cand and cand.is_displayed():
            return cand
    except Exception:
        pass
    return None


def process_report_edit(driver, it: PersonItem) -> bool:
    s = (it.service or "").strip()
    if s not in ("通所", "施設外就労"):
        log(f"⏭️ 対象外なのでスキップ（編集しない）: {it.name} / B列={s!r}")
        return True

    close_dialog_if_open(driver)

    row = find_row_by_name(driver, it.name)
    if row is None:
        log(f"⚠️ {it.name} 行が見つかりません")
        return False

    if not click_pencil_in_row(driver, row):
        log(f"⚠️ {it.name} 編集(鉛筆)ボタンが押せません")
        return False

    try:
        WebDriverWait(driver, 10).until(lambda d: get_top_dialog(d) is not None)
    except Exception:
        log(f"⚠️ {it.name} モーダルが開きません")
        return False

    dlg = get_top_dialog(driver)
    if not dlg:
        return False

    try:
        # B: サービス提供の状況
        if it.service:
            ok = select_dropdown_skip_if_same(driver, dlg, "サービス提供の状況", it.service)
            if not ok:
                ok = select_dropdown_skip_if_same(driver, dlg, "サービス提供の状況 *", it.service)
            if not ok:
                ok = select_dropdown_skip_if_same(driver, dlg, "サービス提供", it.service)

            if not ok:
                try:
                    dlg_text = (dlg.text or "").replace("\n", " ")
                except Exception:
                    dlg_text = ""

                if it.service in dlg_text:
                    log(f"ℹ️ {it.name} サービス提供は画面上で確認できたため続行")
                else:
                    dump_debug(driver, f"service_dropdown_fail_{it.name}")
                    log(f"⚠️ {it.name} サービス提供の選択に失敗しました")
                    close_dialog_if_open(driver)
                    return False

        meal = (it.meal or "").strip()
        if not meal:
            log(f"⚠️ {it.name} 食事提供(E列)が空欄です")
            close_dialog_if_open(driver)
            return False

        # ===== 上段の開始/終了時間 =====
        has_start = bool((it.start or "").strip())
        has_end = bool((it.end or "").strip())

        if has_start or has_end:
            print(f"[DEBUG] {it.name} it.start={it.start!r} it.end={it.end!r}", flush=True)

            inp_start = _find_time_input(dlg, "start")
            inp_end = _find_time_input(dlg, "end")

            print(
                f"[DEBUG] {it.name} inp_start found={inp_start is not None} inp_end found={inp_end is not None}",
                flush=True
            )

            if not inp_start or not inp_end:
                dump_debug(driver, f"time_input_not_found_{it.name}")
                log(f"⚠️ {it.name} 上側の開始/終了inputが取得できません")
                close_dialog_if_open(driver)
                return False

            if has_start:
                set_input_value(driver, inp_start, it.start)
                try:
                    print(
                        f"[DEBUG] {it.name} after set start value={inp_start.get_attribute('value')!r}",
                        flush=True
                    )
                except Exception as e:
                    print(f"[DEBUG] {it.name} after set start read error={e}", flush=True)

            if has_end:
                set_input_value(driver, inp_end, it.end)
                try:
                    print(
                        f"[DEBUG] {it.name} after set end value={inp_end.get_attribute('value')!r}",
                        flush=True
                    )
                except Exception as e:
                    print(f"[DEBUG] {it.name} after set end read error={e}", flush=True)

        # ===== 食事提供 =====
        ok = select_dropdown_skip_if_same(driver, dlg, "食事提供", meal)
        if not ok:
            dump_debug(driver, f"meal_dropdown_fail_{it.name}")
            log(f"⚠️ {it.name} 食事提供の選択に失敗しました")
            close_dialog_if_open(driver)
            return False

        # ===== 備考 =====
        note_src = (it.note or "").strip()
        final_note = "施設外就労(実施報告書等添付)" if s == "施設外就労" else note_src

        area = _find_remark_area(dlg)
        if area:
            set_input_value(driver, area, final_note)
        else:
            dump_debug(driver, f"remark_not_found_{it.name}")
            log(f"⚠️ {it.name} 備考欄が見つかりません")
            close_dialog_if_open(driver)
            return False

        # ===== 作業時間チェック制御 =====
        has_work_time = bool((it.work_start or "").strip()) and bool((it.work_end or "").strip())
        print(
            f"[DEBUG] process_report_edit has_work_time={has_work_time} "
            f"work_start={it.work_start!r} work_end={it.work_end!r}",
            flush=True
        )

        try:
            worked_chk = dlg.find_element(By.CSS_SELECTOR, "input[name='workRecord.worked']")
            is_checked = worked_chk.is_selected()
            print(f"[DEBUG] {it.name} worked checked before={is_checked}", flush=True)

            if has_work_time:
                # 作業時間あり → ONにする
                if not is_checked:
                    driver.execute_script("arguments[0].click();", worked_chk)
                    time.sleep(0.3)
                    print(f"[DEBUG] {it.name} 作業時間チェックON", flush=True)

                print("[DEBUG] before fill_work_record_section in process_report_edit", flush=True)
                try:
                    fill_work_record_section(driver, dlg, it)
                    print("[DEBUG] after fill_work_record_section in process_report_edit", flush=True)
                except Exception as e:
                    log(f"ℹ️ {it.name} 作業時間欄入力失敗のためスキップして続行します: {e}")

            else:
                # 作業時間なし → OFFにする
                if is_checked:
                    driver.execute_script("arguments[0].click();", worked_chk)
                    time.sleep(0.3)
                    print(f"[DEBUG] {it.name} 作業時間チェックOFF", flush=True)

                log(f"ℹ️ {it.name} 作業時間未入力のためチェックOFFで保存します")

        except Exception as e:
            print(f"[DEBUG] {it.name} 作業時間チェック制御失敗 {e}", flush=True)
            if has_work_time:
                try:
                    fill_work_record_section(driver, dlg, it)
                except Exception as ee:
                    log(f"ℹ️ {it.name} 作業時間欄入力失敗のためスキップして続行します: {ee}")
            else:
                log(f"ℹ️ {it.name} 作業時間未入力のためスキップ")

        # ===== 保存 =====
        save_btn = None
        for xp in [
            ".//button[contains(.,'保存する')]",
            ".//button[contains(.,'保存')]",
            ".//button[contains(.,'登録')]",
            ".//button[contains(.,'更新')]",
        ]:
            try:
                save_btn = dlg.find_element(By.XPATH, xp)
                break
            except Exception:
                continue

        if not save_btn:
            dump_debug(driver, f"save_btn_not_found_{it.name}")
            log(f"⚠️ {it.name} 保存ボタンが見つかりません")
            close_dialog_if_open(driver)
            return False

        for _ in range(25):
            disabled = save_btn.get_attribute("disabled")
            aria_disabled = (save_btn.get_attribute("aria-disabled") or "").lower()
            if disabled is None and aria_disabled != "true":
                break
            time.sleep(0.12)

        if not safe_click(driver, save_btn):
            try:
                driver.execute_script("arguments[0].click();", save_btn)
            except Exception:
                dump_debug(driver, f"save_click_fail_{it.name}")
                log(f"⚠️ {it.name} 保存ボタンが押せません")
                close_dialog_if_open(driver)
                return False

        # ===== 保存完了待ち =====
        save_done = False
        t0 = time.time()

        while time.time() - t0 < 15:
            try:
                if not dlg.is_displayed():
                    save_done = True
                    break
            except Exception:
                save_done = True
                break

            time.sleep(0.3)

        if not save_done:
            dump_debug(driver, f"save_wait_timeout_{it.name}")
            log(f"⚠️ {it.name} 保存完了待ちでタイムアウトしました")
            close_dialog_if_open(driver)
            return False

        # 保険
        time.sleep(1.0)

        return True

    except Exception as e:
        dump_debug(driver, f"exception_{it.name}")
        log(f"⚠️ {it.name} 例外: {e}")
        close_dialog_if_open(driver)
        return False


def update_report_note_only(driver, target_date, resident_name, note_text):
    """
    利用実績（日ごと）の対象利用者1件について、
    備考欄(note_text)だけを更新して保存する。
    既存の備考欄入力ロジック(_find_remark_area)をそのまま使う版。
    """
    if not target_date:
        raise RuntimeError("[FATAL] target_date が空ある")
    if not resident_name:
        raise RuntimeError("[FATAL] resident_name が空ある")

    m = re.match(r"^\s*(\d{4})-(\d{2})-(\d{2})\s*$", str(target_date))
    if not m:
        raise RuntimeError(f"[FATAL] target_date形式が不正ある: {target_date}")

    y = int(m.group(1))
    mo = int(m.group(2))
    d = int(m.group(3))

    close_dialog_if_open(driver)

    goto_report_date(driver, y, mo, d)
    wait_table_stable_after_date_change(driver)

    row = find_row_by_name(driver, resident_name)
    if row is None:
        log(f"⚠️ {resident_name} 行が見つかりません")
        return False

    if not click_pencil_in_row(driver, row):
        log(f"⚠️ {resident_name} 編集(鉛筆)ボタンが押せません")
        return False

    try:
        WebDriverWait(driver, 10).until(lambda drv: get_top_dialog(drv) is not None)
    except Exception:
        log(f"⚠️ {resident_name} モーダルが開きません")
        return False

    dlg = get_top_dialog(driver)
    if not dlg:
        log(f"⚠️ {resident_name} モーダル取得に失敗しました")
        return False

    try:
        area = _find_remark_area(dlg)
        if area is None:
            dump_debug(driver, f"remark_not_found_note_only_{resident_name}")
            log(f"⚠️ {resident_name} 備考欄が見つかりません")
            close_dialog_if_open(driver)
            return False

        set_input_value(driver, area, str(note_text or "").strip())
        time.sleep(0.2)

        save_all(driver)
        time.sleep(0.8)

        return True

    except Exception as e:
        dump_debug(driver, f"update_report_note_only_error_{resident_name}")
        log(f"⚠️ {resident_name} note only update error: {e}")
        close_dialog_if_open(driver)
        return False

# =========================
# 日々の記録 + Gemini
# =========================
def _daily_record_category(it: PersonItem) -> str:
    """
    スタッフ例文シート用のカテゴリ
      在宅 / 通所 / 施設外
    """
    s = (it.service or "").strip()
    meal = (it.meal or "").strip()

    if s == "施設外就労":
        return "施設外"
    if s == "通所" and meal == "提供なし":
        return "在宅"
    return "通所"


def _daily_record_work_label(it: PersonItem) -> str:
    """
    日々の記録ページの「作業」欄
      通所（提供あり/なし）→ 内職
      施設外就労           → 清掃
    """
    s = (it.service or "").strip()
    if s == "施設外就労":
        return "清掃"
    return "内職"


def _read_treaty_and_staff_examples(excel_path: str):
    """
    条約シート全文と、スタッフ例文シートを読む
    戻り値:
      treaty_text: str
      examples: {
        "スタッフ名": {
          "在宅":   {"利用者状態": "...", "職員考察": "..."},
          "通所":   {"利用者状態": "...", "職員考察": "..."},
          "施設外": {"利用者状態": "...", "職員考察": "..."},
        }
      }
    """
    staged = stage_excel_local(excel_path)
    wb = openpyxl.load_workbook(staged, data_only=True)

    # 条約
    treaty_text = ""
    if "条約" in wb.sheetnames:
        ws = wb["条約"]
        chunks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            a = norm(row[0] if len(row) > 0 else "")
            b = norm(row[1] if len(row) > 1 else "")
            c = norm(row[2] if len(row) > 2 else "")
            if a or b or c:
                chunks.append(f"{a} | {b} | {c}")
        treaty_text = "\n".join(chunks)

    # スタッフ例文
    examples = {}
    if "スタッフ例文" in wb.sheetnames:
        ws = wb["スタッフ例文"]
        row = 3
        while True:
            staff_name = norm(ws[f"A{row}"].value)
            if not staff_name:
                break

            examples[staff_name] = {
                "在宅": {
                    "利用者状態": norm(ws[f"B{row}"].value),
                    "職員考察": norm(ws[f"C{row}"].value),
                },
                "通所": {
                    "利用者状態": norm(ws[f"D{row}"].value),
                    "職員考察": norm(ws[f"E{row}"].value),
                },
                "施設外": {
                    "利用者状態": norm(ws[f"F{row}"].value),
                    "職員考察": norm(ws[f"G{row}"].value),
                },
            }
            row += 1

    wb.close()
    return treaty_text, examples


def _choose_daily_recorder(excel_path: str) -> str:
    """
    基本情報シートの「スタッフ名 / 日誌を書く人を[〇]」表から、
    〇 の付いたスタッフを1人選ぶ（今回は1日分固定）
    """
    staged = stage_excel_local(excel_path)
    wb = openpyxl.load_workbook(staged, data_only=True)

    if SHEET_MAIN not in wb.sheetnames:
        wb.close()
        return ""

    ws = wb[SHEET_MAIN]

    # K列=スタッフ名, L列=〇/空欄 の前提
    candidates = []
    row = 3
    while True:
        staff_name = norm(ws[f"K{row}"].value)
        flag = norm(ws[f"L{row}"].value)

        # KもLも空なら終端
        if not staff_name and not flag:
            break

        if staff_name and flag == "〇":
            candidates.append(staff_name)

        row += 1

    wb.close()

    if not candidates:
        return ""

    # 今回は1日分固定なので最初の1人を採用
    return candidates[0]

def _get_style_examples_for_staff(examples: dict, staff_name: str, category: str):
    """
    category: 在宅 / 通所 / 施設外
    """
    x = examples.get(staff_name, {})
    c = x.get(category, {})
    return c.get("利用者状態", ""), c.get("職員考察", "")


def _build_gemini_prompt(
    field_kind: str,
    base_memo: str,
    category: str,
    staff_name: str,
    style_example: str,
    treaty_text: str
) -> str:
    prompt = f"""あなたは就労継続支援B型の支援記録作成アシスタントです。
以下の「条約」を絶対遵守して、{field_kind}欄に入れる日本語文を1段落で作成してください。

【条約】
{treaty_text}

【今回の欄】
{field_kind}

【利用形態】
{category}

【参照するスタッフ文体】
スタッフ名: {staff_name}

【文体参考例】
{style_example}

【元メモ】
{base_memo}

【厳守事項】
- 出力は本文のみ
- 見出し、JSON、箇条書きは禁止
- 余計な説明は禁止
- {field_kind}欄にそのまま貼れる自然な文章にする
"""
    return prompt

def _gemini_generate_text(client, prompt: str) -> str:
    def _cleanup_text(text: str) -> str:
        text = (text or "").strip()
        text = text.replace("```json", "").replace("```", "").strip()
        return text

    def _looks_like_reasoning(text: str) -> bool:
        if not text:
            return False

        ng_patterns = [
            "考えられる最善の出力を生成するために",
            "思考プロセス",
            "依頼内容の確認",
            "開始メモの分析",
            "最終的な決定",
            "複数の候補を比較",
            "出力します",
            "1.  **",
            "2.  **",
            "3.  **",
            "4.  **",
            "5.  **",
            "6.  **",
            "7.  **",
        ]

        for p in ng_patterns:
            if p in text:
                return True

        # 箇条書きや番号だらけの応答も弾く
        lines = [x.strip() for x in text.splitlines() if x.strip()]
        numbered = sum(1 for x in lines if re.match(r"^\d+\.", x))
        bullets = sum(1 for x in lines if x.startswith("*") or x.startswith("-"))

        if numbered >= 2 or bullets >= 4:
            return True

        return False

    # 1回目
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
    )

    result_text = _cleanup_text(getattr(response, "text", "") or "")
    if not result_text:
        raise RuntimeError("Geminiの応答が空ある")

    # 思考プロセスっぽいなら再試行
    if _looks_like_reasoning(result_text):
        retry_prompt = (
            prompt
            + "\n\n【最重要・再指示】\n"
              "思考過程、分析、手順、箇条書き、番号付き説明は絶対に出力しないこと。\n"
              "出力は完成した支援記録の本文のみを1段落または2段落で返すこと。\n"
              "『依頼内容の確認』『分析』『最終的な決定』などの文言は禁止。"
        )

        response2 = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=retry_prompt,
        )

        result_text2 = _cleanup_text(getattr(response2, "text", "") or "")
        if result_text2 and not _looks_like_reasoning(result_text2):
            return result_text2

    return result_text

def _replace_placeholder_name(text: str, full_name: str) -> str:
    """
    スタッフ例文中の「〇〇さん」を利用者の苗字に置き換える
    例:
      荒木 和也 → 荒木さん
    """
    text = norm(text)
    full_name = norm(full_name)

    if not text or not full_name:
        return text

    surname = re.split(r"[ 　]+", full_name)[0].strip()
    if not surname:
        return text

    return text.replace("〇〇さん", f"{surname}さん")

def click_daily_edit_button(driver) -> bool:
    """
    右上の「編集」ボタンを押す
    """
    try:
        btn = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//button[.//span[contains(normalize-space(.),'編集')] or contains(normalize-space(.),'編集')]"
            ))
        )
    except Exception:
        dump_debug(driver, "click_daily_edit_button_not_found")
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.1)
    except Exception:
        pass

    if not safe_click(driver, btn):
        dump_debug(driver, "click_daily_edit_button_click_fail")
        return False

    # 編集後に「保存する」が見えたら成功
    t0 = time.time()
    while time.time() - t0 < 10:
        try:
            saves = driver.find_elements(
                By.XPATH,
                "//button[.//span[contains(normalize-space(.),'保存する')] or contains(normalize-space(.),'保存する')]"
            )
            for s in saves:
                try:
                    if s.is_displayed():
                        time.sleep(0.5)
                        return True
                except Exception:
                    pass
        except Exception:
            pass
        time.sleep(0.3)

    dump_debug(driver, "click_daily_edit_button_after_click_fail")
    return False
    
def _wait_daily_save_complete(driver, timeout=15) -> bool:
    """
    日々の記録の保存完了待ち
    どれか1つでも満たせば成功とみなす
      - 「編集」ボタンが再表示される
      - 「保存する」ボタンが消える
      - DOM上で編集状態っぽい要素が減る
    """
    t0 = time.time()
    while time.time() - t0 < timeout:
        # 編集ボタンが戻ってきたら保存完了扱い
        try:
            edits = driver.find_elements(
                By.XPATH,
                "//button[.//span[contains(normalize-space(.),'編集')] or contains(normalize-space(.),'編集')]"
            )
            for b in edits:
                try:
                    if b.is_displayed():
                        return True
                except Exception:
                    pass
        except Exception:
            pass

        # 保存するボタンが見えなくなっても完了扱い
        try:
            saves = driver.find_elements(
                By.XPATH,
                "//button[.//span[contains(normalize-space(.),'保存する')] or contains(normalize-space(.),'保存する')]"
            )
            visible = False
            for s in saves:
                try:
                    if s.is_displayed():
                        visible = True
                        break
                except Exception:
                    pass
            if not visible:
                return True
        except Exception:
            pass

        time.sleep(0.4)

    return False

def _find_daily_record_row_by_name(driver, name: str):
    """
    日々の記録テーブルの行を、利用者名で探す
    スペース差や氏名+付加情報の揺れに強くする
    """
    target = _norm_name_for_match(name)

    try:
        rows = driver.find_elements(By.XPATH, "//tbody/tr")
    except Exception:
        rows = []

    for r in rows:
        try:
            txt = _norm_name_for_match((r.text or "").replace("\n", " "))
            if target and target in txt:
                return r
        except Exception:
            pass

    return None


def _choose_option_from_open_listbox_text(driver, value_text: str) -> bool:
    """
    開いているリストボックス/メニューから、文字で選ぶ
    スタッフ名の全角/半角スペース揺れにも対応
    """
    value_text = norm(value_text)
    if not value_text:
        return False

    target_norm = _norm_name_for_match(value_text)

    # まず見えている候補を総当たり
    xps = [
        "//li[@role='option']",
        "//li",
        "//*[@role='option']",
    ]

    for xp in xps:
        try:
            els = driver.find_elements(By.XPATH, xp)
        except Exception:
            els = []

        for el in els:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                continue

            try:
                txt = norm(el.text)
            except Exception:
                txt = ""

            if not txt:
                continue

            txt_norm = _norm_name_for_match(txt)

            # 完全一致 or 含有一致
            if txt_norm == target_norm or target_norm in txt_norm:
                if safe_click(driver, el):
                    time.sleep(0.3)
                    return True

    return False

def _set_daily_work_for_row(driver, row, work_label: str) -> bool:
    """
    その行の作業欄を、内職/清掃にする
    """
    if not work_label:
        return True

    # まず現在値確認
    try:
        row_text = row.text or ""
        if work_label in row_text:
            return True
    except Exception:
        pass

    # 「内容」列の最初の role=button を狙う
    btn = None
    xps = [
        ".//*[@role='button' and contains(@id, 'work_history')]",
        ".//*[@role='button']",
    ]
    for xp in xps:
        try:
            els = row.find_elements(By.XPATH, xp)
        except Exception:
            els = []
        for el in els:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                pass
            btn = el
            break
        if btn:
            break

    if not btn:
        dump_debug(driver, "daily_work_button_not_found")
        return False

    if not safe_click(driver, btn):
        return False
    time.sleep(0.4)

    if not _choose_option_from_open_listbox_text(driver, work_label):
        dump_debug(driver, f"daily_work_option_fail_{work_label}")
        return False

    # メニューを閉じる
    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        time.sleep(0.2)
    except Exception:
        pass
    return True


def _set_daily_textareas_for_row(
    driver,
    row,
    user_status_text: str,
    staff_comment_text: str,
    send_user_status: bool = True,
    send_staff_comment: bool = True,
) -> bool:
    try:
        user_area = row.find_element(By.XPATH, ".//textarea[contains(@name,'user_status')]")
        staff_area = row.find_element(By.XPATH, ".//textarea[contains(@name,'staff_comment')]")
    except Exception:
        dump_debug(driver, "daily_textarea_not_found")
        return False

    try:
        if not user_area.is_displayed() or not staff_area.is_displayed():
            dump_debug(driver, "daily_textarea_hidden")
            return False
    except Exception:
        pass

    try:
        current_user_value = user_area.get_attribute("value") or user_area.get_attribute("textContent") or ""
    except Exception:
        current_user_value = ""

    try:
        current_staff_value = staff_area.get_attribute("value") or staff_area.get_attribute("textContent") or ""
    except Exception:
        current_staff_value = ""

    final_user_text = user_status_text if send_user_status else current_user_value
    final_staff_text = staff_comment_text if send_staff_comment else current_staff_value

    set_input_value(driver, user_area, final_user_text)
    time.sleep(0.2)
    set_input_value(driver, staff_area, final_staff_text)
    time.sleep(0.2)

    # 入力確認
    try:
        after_user = user_area.get_attribute("value") or ""
    except Exception:
        after_user = ""

    try:
        after_staff = staff_area.get_attribute("value") or ""
    except Exception:
        after_staff = ""

    print("[DEBUG] textarea after_user =", after_user, flush=True)
    print("[DEBUG] textarea after_staff =", after_staff, flush=True)

    if send_user_status and str(final_user_text).strip() and str(after_user).strip() != str(final_user_text).strip():
        dump_debug(driver, "daily_user_text_not_applied")
        return False

    if send_staff_comment and str(final_staff_text).strip() and str(after_staff).strip() != str(final_staff_text).strip():
        dump_debug(driver, "daily_staff_text_not_applied")
        return False

    return True

def _set_daily_recorder_for_row(driver, row, recorder_name: str) -> bool:
    if not recorder_name:
        return False

    btn = None

    xps = [
        ".//*[@role='button' and contains(@id, 'staff_id')]",
        ".//*[@role='button' and contains(normalize-space(.), '選択してください')]",
    ]

    for xp in xps:
        try:
            els = row.find_elements(By.XPATH, xp)
        except Exception:
            els = []

        for el in els:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                continue

            btn = el
            break

        if btn:
            break

    if not btn:
        dump_debug(driver, "daily_recorder_button_not_found")
        return False

    if not safe_click(driver, btn):
        dump_debug(driver, "daily_recorder_button_click_fail")
        return False

    time.sleep(0.5)

    if not _choose_option_from_open_listbox_text(driver, recorder_name):
        dump_debug(driver, f"daily_recorder_option_fail_{recorder_name}")
        return False

    time.sleep(0.3)
    return True


def click_daily_save_button(driver) -> bool:
    """
    右上の「保存する」ボタンを押して、保存完了まで待つ
    """
    try:
        span = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//span[contains(normalize-space(.),'保存する')]"
            ))
        )
    except Exception:
        dump_debug(driver, "click_daily_save_button_span_not_found")
        return False

    try:
        btn = span.find_element(By.XPATH, "./ancestor::button[1]")
    except Exception:
        dump_debug(driver, "click_daily_save_button_button_not_found")
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.1)
    except Exception:
        pass

    # disabled解除待ち
    t0 = time.time()
    while time.time() - t0 < 10:
        try:
            disabled = btn.get_attribute("disabled")
            aria_disabled = (btn.get_attribute("aria-disabled") or "").lower()
            if disabled is None and aria_disabled != "true":
                break
        except Exception:
            pass
        time.sleep(0.2)

    if not safe_click(driver, btn):
        try:
            driver.execute_script("arguments[0].click();", btn)
        except Exception:
            dump_debug(driver, "click_daily_save_button_click_fail")
            return False

    if not _wait_daily_save_complete(driver, timeout=15):
        dump_debug(driver, "click_daily_save_button_wait_fail")
        return False

    time.sleep(0.5)
    return True


def process_one_daily_record(
    driver,
    it: PersonItem,
    recorder_name: str,
    user_text: str,
    staff_text: str
) -> bool:

    print("🔥 実行開始したある！！！", flush=True)

    if not click_daily_edit_button(driver):
        return False

    row = find_row_by_name(driver, it.name)
    if row is None:
        log(f"⚠️ 日々の記録 行発見失敗ある: {it.name}")
        return False

    work_label = _daily_record_work_label(it)

    if not _set_daily_work_for_row(driver, row, work_label):
        log(f"⚠️ 作業欄の選択失敗ある: {it.name}")
        return False

    time.sleep(0.5)

    # ★ 作業欄変更後に行を取り直す
    row = find_row_by_name(driver, it.name)
    if row is None:
        log(f"⚠️ 作業欄変更後の行再取得失敗ある: {it.name}")
        return False

    if not _set_daily_textareas_for_row(driver, row, user_text, staff_text):
        log(f"⚠️ 日々の記録 textarea 入力失敗ある: {it.name}")
        return False

    time.sleep(0.3)

    # ★ recorder前にももう一度取り直す
    row = find_row_by_name(driver, it.name)
    if row is None:
        log(f"⚠️ 記録者入力前の行再取得失敗ある: {it.name}")
        return False

    if not _set_daily_recorder_for_row(driver, row, recorder_name):
        log(f"⚠️ 記録者選択失敗ある: {it.name}")
        return False

    # 作業時間（任意）
    has_work_time = bool(str(getattr(it, "work_start", "")).strip()) and bool(str(getattr(it, "work_end", "")).strip())

    if has_work_time:
        if not open_work_record_dialog_from_row(driver, row):
            log(f"ℹ️ 作業時間ダイアログなし/開けない → スキップ: {it.name}")
        else:
            dialog = get_top_dialog(driver)

            if dialog is None:
                log(f"ℹ️ 作業時間ダイアログ取得不可 → スキップ: {it.name}")
            else:
                try:
                    fill_work_record_section(driver, dialog, it)
                except Exception as e:
                    log(f"ℹ️ 作業時間入力失敗 → スキップ: {it.name} / {e}")

                work_save_btn = None
                for xp in [
                    ".//button[contains(.,'保存する')]",
                    ".//button[contains(.,'保存')]",
                    ".//button[contains(.,'登録')]",
                    ".//button[contains(.,'更新')]",
                ]:
                    try:
                        work_save_btn = dialog.find_element(By.XPATH, xp)
                        break
                    except Exception:
                        continue

                if work_save_btn:
                    if not safe_click(driver, work_save_btn):
                        try:
                            driver.execute_script("arguments[0].click();", work_save_btn)
                        except Exception:
                            log(f"ℹ️ 作業時間保存押下不可 → スキップ: {it.name}")
                else:
                    log(f"ℹ️ 作業時間保存ボタンなし → スキップ: {it.name}")

                try:
                    WebDriverWait(driver, 10).until(EC.invisibility_of_element(dialog))
                except Exception:
                    pass
    else:
        log(f"ℹ️ 作業時間未入力 → スキップ: {it.name}")

    if not click_daily_save_button(driver):
        log(f"⚠️ 日々の記録 保存失敗ある: {it.name}")
        return False

    log(f"✅ 日々の記録 保存成功ある: {it.name}")
    return True

def open_daily_record_page(driver):
    driver.get("https://mgr.knowbe.jp/v2/#/record/daily")
    time.sleep(2)

    print("[DEBUG] open_daily_record_page current_url =", driver.current_url, flush=True)
    print("[DEBUG] open_daily_record_page __file__ =", __file__, flush=True)

    body_text = driver.find_element(By.TAG_NAME, "body").text

    keywords = ["日々の記録", "業務日誌", "支援記録"]
    for k in keywords:
        if k in body_text:
            print(f"[DEBUG] open_daily_record_page hit keyword={k}", flush=True)
            return True

    dump_debug(driver, "open_daily_record_page_fail")
    return False

# =========================
# 利用者ごと → 支援記録 → 本文取得
# =========================
def goto_users_summary(driver):
    """
    Knowbe の「記録 → 利用者ごと」ページへ直接移動する
    """
    target_url = "https://mgr.knowbe.jp/v2/?_page=record/users_summary/#/record/users_summary"
    driver.get(target_url)
    time.sleep(1.5)

    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located(
            (By.XPATH, "//*[contains(normalize-space(.), '利用者ごとの記録')]")
        )
    )
    return True

def apply_users_summary_filter_show_expired(driver):
    """
    利用者ごと一覧で
    - 『退所者の非表示』のチェックを外す
    - 『この条件で絞り込む』を押す
    まで行う
    """
    log("[STEP] apply_users_summary_filter_show_expired start")

    try:
        expired_checkbox = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "expiredVisibility"))
        )

        if expired_checkbox.is_selected():
            driver.execute_script("arguments[0].click();", expired_checkbox)
            log("[DEBUG] 退所者の非表示 → OFFにしたある")
            time.sleep(0.7)
        else:
            log("[DEBUG] 退所者の非表示はすでにOFFある")

        filter_btn = None
        xpaths = [
            "//button[.//span[normalize-space(.)='この条件で絞り込む']]",
            "//button[normalize-space(.)='この条件で絞り込む']",
            "//*[self::button or self::div or self::span][contains(normalize-space(.), 'この条件で絞り込む')]",
        ]

        for xp in xpaths:
            try:
                elems = driver.find_elements(By.XPATH, xp)
            except Exception:
                elems = []

            for el in elems:
                try:
                    if not el.is_displayed():
                        continue
                except Exception:
                    pass

                try:
                    if el.tag_name.lower() == "button":
                        filter_btn = el
                    else:
                        filter_btn = el.find_element(By.XPATH, "./ancestor::button[1]")
                    break
                except Exception:
                    continue

            if filter_btn:
                break

        if not filter_btn:
            dump_debug(driver, "users_summary_filter_button_not_found")
            raise RuntimeError("[FATAL] 『この条件で絞り込む』ボタンが見つからないある")

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", filter_btn)
        time.sleep(0.2)

        if not safe_click(driver, filter_btn):
            driver.execute_script("arguments[0].click();", filter_btn)

        log("[DEBUG] 『この条件で絞り込む』を押したある")

        # 画面再描画のため少し待つだけ
        time.sleep(2.5)

        log("[STEP] apply_users_summary_filter_show_expired done")
        return True

    except Exception as e:
        dump_debug(driver, "apply_users_summary_filter_show_expired_fail")
        log(f"[DEBUG] 利用者一覧の絞り込み更新で失敗ある: {e}")
        return False

    
def normalize_resident_name_for_match(name: str) -> str:
    s = str(name or "")
    s = s.replace(" ", "").replace("　", "").strip()
    return s


def find_user_card_by_name(driver, resident_name: str):
    """
    利用者ごと一覧ページから、対象利用者の行（カード）を返す。
    完全一致ではなく、行本文に名前が含まれていればOKにする。
    """
    target = normalize_resident_name_for_match(resident_name)
    log(f"[STEP] find_user_card_by_name target={target}")

    # 一覧の行候補を広めに取る
    candidate_xpaths = [
        "//button[.//span[contains(normalize-space(.), '支援記録')]]/ancestor::*[self::div or self::li or self::section or self::tr][1]",
        "//span[contains(normalize-space(.), '支援記録')]/ancestor::*[self::div or self::li or self::section or self::tr][1]",
        "//*[contains(normalize-space(.), '支援記録') and (self::button or self::span or self::div)]/ancestor::*[self::div or self::li or self::section or self::tr][1]",
    ]

    candidates = []
    for xp in candidate_xpaths:
        try:
            found = driver.find_elements(By.XPATH, xp)
            for el in found:
                if el not in candidates:
                    candidates.append(el)
        except Exception:
            pass

    log(f"[DEBUG] candidate count = {len(candidates)}")

    for i, card in enumerate(candidates, start=1):
        try:
            txt = card.text or ""
        except Exception:
            txt = ""

        txt_norm = normalize_resident_name_for_match(txt)

        if i <= 15:
            log(f"[DEBUG] candidate {i}: {txt_norm[:200]}")

        # 名前が行本文に含まれていたらOK
        if target and target in txt_norm:
            log(f"[STEP] matched candidate {i}")
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
                time.sleep(0.3)
            except Exception:
                pass
            return card

    return None


def _is_real_support_record_url(url: str) -> bool:
    """
    本物の支援記録URLだけを許可する
    例:
      https://mgr.knowbe.jp/v2/#/record/365644/support
      https://mgr.knowbe.jp/v2/#/record/365644/support/2026/4
    """
    u = "" if url is None else str(url).strip()

    if "support_plan" in u:
        return False
    if "assessment" in u:
        return False

    return re.search(r"/record/\d+/support(?:$|[/?#])", u) is not None


def _button_text_loose(el) -> str:
    try:
        txt = (el.text or "").strip()
    except Exception:
        txt = ""

    if txt:
        return txt.replace(" ", "").replace("　", "").replace("\n", "").strip()

    # text が空のbutton対策
    try:
        txt = (el.get_attribute("innerText") or "").strip()
    except Exception:
        txt = ""

    return txt.replace(" ", "").replace("　", "").replace("\n", "").strip()


def _is_real_support_record_url(url: str) -> bool:
    u = "" if url is None else str(url).strip()
    if "support_plan" in u:
        return False
    if "assessment" in u:
        return False
    return re.search(r"/record/\d+/support(?:$|[/?#])", u) is not None


def open_support_record_for_resident(driver, resident_name: str) -> bool:
    """
    名前一致済みのカードから、そのカード内にある
    <span>支援記録</span> を起点に button を押す
    """
    log(f"[STEP] open_support_record_for_resident start resident={resident_name}")

    card = find_user_card_by_name(driver, resident_name)
    if card is None:
        log(f"[DEBUG] user card not found: {resident_name}")
        return False

    search_roots = [card]
    cur = card
    for _ in range(5):
        try:
            cur = cur.find_element(By.XPATH, "./..")
            search_roots.append(cur)
        except Exception:
            break

    xpaths = [
        ".//span[normalize-space(.)='支援記録']/ancestor::*[@role='button'][1]",
        ".//*[normalize-space(.)='支援記録']/ancestor::*[@role='button'][1]",
    ]

    for level, root in enumerate(search_roots, start=1):
        log(f"[DEBUG] search support button in container level {level}")

        for xp in xpaths:
            try:
                buttons = root.find_elements(By.XPATH, xp)
            except Exception:
                buttons = []

            log(f"[DEBUG] matched buttons level={level} xp={xp} count={len(buttons)}")

            for i, el in enumerate(buttons, start=1):
                try:
                    if not el.is_displayed():
                        continue
                except Exception:
                    pass

                href = ""
                try:
                    href = (el.get_attribute("href") or "").strip()
                except Exception:
                    href = ""

                log(f"[DEBUG] support href level={level} index={i} href={href}")

                if href:
                    driver.get(href)
                    time.sleep(1.5)

                    cur_url = driver.current_url or ""
                    log(f"[DEBUG] after href jump current_url = {cur_url}")

                    if "assessment" in cur_url:
                        continue

                    if "support_plan" in cur_url:
                        forced_url = re.sub(r"/support_plan(?:$|[/?#].*)", "/support", cur_url)
                        if forced_url != cur_url:
                            log(f"[DEBUG] force jump to support url = {forced_url}")
                            driver.get(forced_url)
                            time.sleep(1.5)
                            cur_url = driver.current_url or ""
                            log(f"[DEBUG] after force jump current_url = {cur_url}")

                    if _is_real_support_record_url(cur_url):
                        log(f"[STEP] open_support_record_for_resident done resident={resident_name}")
                        return True

                if safe_click(driver, el):
                    time.sleep(1.5)
                    cur_url = driver.current_url or ""

                    if "assessment" in cur_url:
                        continue

                    if "support_plan" in cur_url:
                        forced_url = re.sub(r"/support_plan(?:$|[/?#].*)", "/support", cur_url)
                        if forced_url != cur_url:
                            log(f"[DEBUG] force jump to support url = {forced_url}")
                            driver.get(forced_url)
                            time.sleep(1.5)
                            cur_url = driver.current_url or ""

                    if _is_real_support_record_url(cur_url):
                        log(f"[STEP] open_support_record_for_resident done resident={resident_name}")
                        return True

def goto_support_record_month(driver, target_year: int, target_month: int) -> bool:
    """
    本物の支援記録URLにだけ年月を付けて移動する
    例:
      /record/365644/support        -> /record/365644/support/2026/4
      /record/365644/support/2026/3 -> /record/365644/support/2026/4

    support_plan / assessment は即失敗
    """
    cur = driver.current_url or ""
    log(f"[DEBUG] goto_support_record_month start current_url = {cur}")

    # まずURLが安定するまで少し待つ
    for _ in range(30):
        cur = driver.current_url or ""
        log(f"[DEBUG] goto_support_record_month current_url = {cur}")

        if "support_plan" in cur or "assessment" in cur:
            log(f"[DEBUG] wrong page detected before month jump = {cur}")
            return False

        if _is_real_support_record_url(cur):
            break

        time.sleep(0.3)
    else:
        log("[DEBUG] real support URL not reached")
        return False

    cur = driver.current_url or ""

    if "support_plan" in cur or "assessment" in cur:
        log(f"[DEBUG] wrong page detected = {cur}")
        return False

    # /record/{id}/support を基点にする
    m = re.search(r"^(.*?/record/\d+/support)(?:/\d{4}/\d{1,2})?(?=$|[/?#])", cur)
    if not m:
        log(f"[DEBUG] support base regex not matched. current_url = {cur}")
        return False

    base_url = m.group(1)
    new_url = f"{base_url}/{int(target_year)}/{int(target_month)}"

    log(f"[DEBUG] goto_support_record_month new_url = {new_url}")

    driver.get(new_url)
    time.sleep(1.2)
    driver.refresh()
    time.sleep(2.0)

    after_url = driver.current_url or ""
    log(f"[DEBUG] goto_support_record_month after refresh current_url = {after_url}")

    if "support_plan" in after_url or "assessment" in after_url:
        log(f"[DEBUG] wrong page after refresh = {after_url}")
        return False

    if not _is_real_support_record_url(after_url):
        log(f"[DEBUG] still not real support url after refresh = {after_url}")
        return False

    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[contains(normalize-space(.), '支援記録')]")
            )
        )
    except Exception:
        log("[DEBUG] support page marker wait timeout")
        return False

    return True

def extract_support_record_text(driver) -> str:
    """
    第1版:
    まずは body 全体を取得し、その中から支援記録に関係する本文を返す。
    必要なら後で本文コンテナ限定に進化させる。
    """
    body_text = ""
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text or ""
    except Exception:
        body_text = ""

    body_text = body_text.strip()
    if not body_text:
        return ""

    # できるだけ支援記録本文っぽい位置から切り出す
    start_candidates = [
        "支援記録",
        "利用日数",
        "日付",
    ]

    start_idx = -1
    for key in start_candidates:
        idx = body_text.find(key)
        if idx != -1:
            start_idx = idx
            break

    if start_idx != -1:
        body_text = body_text[start_idx:].strip()

    return body_text

def uncheck_expired_visibility_if_needed(driver):
    """
    利用者ごと一覧の『退所者の非表示』チェックを外して、
    退所者も表示される状態にする
    """
    try:
        expired_checkbox = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.NAME, "expiredVisibility"))
        )

        if expired_checkbox.is_selected():
            driver.execute_script("arguments[0].click();", expired_checkbox)
            log("[DEBUG] 退所者の非表示 → OFFにしたある")
            time.sleep(1.5)  # 一覧の再描画待ち
        else:
            log("[DEBUG] 退所者の非表示はすでにOFFある")

    except Exception as e:
        log(f"[DEBUG] 退所者チェック操作スキップある: {e}")

def fetch_user_support_record_text_from_app(
    resident_name: str,
    target_year: int,
    target_month: int,
    login_username: str,
    login_password: str,
):
    """
    利用者ごと → 支援記録 → 対象月 の本文を取得する
    ※ 支援記録が存在しない月（入院・未利用など）は空文字を返す
    """
    if not resident_name:
        raise RuntimeError("[FATAL] resident_name が空です")
    if not target_year:
        raise RuntimeError("[FATAL] target_year が空です")
    if not target_month:
        raise RuntimeError("[FATAL] target_month が空です")
    if not login_username or not login_password:
        raise RuntimeError("[FATAL] Knowbeログイン情報が空です")

    driver = build_chrome_driver()

    try:
        log("[STEP] open base url")
        driver.get("https://mgr.knowbe.jp/v2/")
        time.sleep(1.0)

        log("[STEP] login")
        manual_login_wait(driver, login_username, login_password)

        # ログイン後に一覧へ行く
        log("[STEP] goto users summary")
        ok = goto_users_summary(driver)
        if not ok:
            dump_debug(driver, "goto_users_summary_fail")
            raise RuntimeError("[FATAL] 利用者ごと一覧へ戻れません")

        # 退所者も表示する条件で再絞り込み
        ok = apply_users_summary_filter_show_expired(driver)
        if not ok:
            raise RuntimeError("[FATAL] 利用者ごと一覧の絞り込み更新に失敗したある")

        log(f"[STEP] find resident: {resident_name}")
        ok = open_support_record_for_resident(driver, resident_name)
        if not ok:
            dump_debug(driver, "resident_not_found_in_users_summary")
            raise RuntimeError(f"[FATAL] 対象利用者カードは見つかったが支援記録ボタンを押せませんでした: {resident_name}")

        cur = driver.current_url or ""
        if "support_plan" in cur or "assessment" in cur or (not _is_real_support_record_url(cur)):
            dump_debug(driver, "wrong_page_after_open_support_record")
            raise RuntimeError(f"[FATAL] 支援記録ではないページに遷移しました: {cur}")

        log("[STEP] goto support target month")
        ok = goto_support_record_month(driver, int(target_year), int(target_month))
        if not ok:
            dump_debug(driver, "goto_support_record_month_fail")
            raise RuntimeError(f"[FATAL] 支援記録の対象月へ移動できません: {target_year}/{target_month}")

        cur = driver.current_url or ""
        if "support_plan" in cur or "assessment" in cur or (not _is_real_support_record_url(cur)):
            dump_debug(driver, "wrong_page_after_month_jump")
            raise RuntimeError(f"[FATAL] 月移動後に支援記録ページではありません: {cur}")

        # ★ 記録ゼロ月では「利用者状態」「職員考察」が出ないので、ここで厳しい待機はしない
        time.sleep(1.5)

        log("[STEP] extract support record text")
        support_text = get_support_record_page_text(driver)

        # ★ 入院・未利用などで支援記録が存在しない月は空文字をそのまま返す
        if not support_text or not str(support_text).strip():
            log("[DEBUG] support record text is empty (利用実績なし / 支援記録なし) ある")
            return ""

        return str(support_text).strip()

    finally:
        try:
            driver.quit()
        except Exception:
            pass

def run_daily_records(driver, excel_path: str, items: List[PersonItem], targets: List[PersonItem], y: int, m: int, d: int):
    """
    利用実績入力のあとに呼ぶ本体
    Gemini参照はこのファイルでは持たず、必要な前処理だけして各行を処理する
    """
    treaty_text, examples = _read_treaty_and_staff_examples(excel_path)
    recorder_name = _choose_daily_recorder(excel_path)

    if not recorder_name:
        raise RuntimeError("[FATAL] 日々の記録の記録者を決められないある")

    log(f"📝 日々の記録の記録者: {recorder_name}")

    if not open_daily_record_page(driver):
        raise RuntimeError("[FATAL] 日々の記録ページへ行けないある")

    for it in targets:
        s = norm(it.service)
        if s not in ("通所", "施設外就労"):
            continue

        log(f"🧾 日々の記録入力: {it.name}")

        category = _daily_record_category(it)

        style_user, style_staff = _get_style_examples_for_staff(
            examples,
            recorder_name,
            category
        )

        base_user_memo = norm(it.user_state)
        base_staff_memo = norm(it.staff_note)

        if not base_user_memo:
            base_user_memo = _replace_placeholder_name(style_user, it.name)

        if not base_staff_memo:
            base_staff_memo = _replace_placeholder_name(style_staff, it.name)

        user_text = base_user_memo
        staff_text = base_staff_memo

        ok = process_one_daily_record(
            driver=driver,
            it=it,
            recorder_name=recorder_name,
            user_text=user_text,
            staff_text=staff_text,
        )

        if not ok:
            dump_debug(driver, f"daily_record_fail_{it.name}")
            log(f"[WARN] 日々の記録失敗→次へ: {it.name}")

def _normalize_service_for_app(service_type: str, knowbe_target: str) -> str:
    s = norm(service_type)

    if s == "施設外就労":
        return "施設外就労"

    # 在宅も通所も、Knowbeの実績送信用には通所扱いに寄せる
    return "通所"


def _normalize_meal_for_app(meal_flag: str) -> str:
    m = norm(meal_flag)
    if m in ("あり", "提供あり", "有"):
        return "提供あり"
    return "提供なし"


def _build_single_item_from_app(
    resident_name: str,
    service_type: str,
    start_time: str,
    end_time: str,
    meal_flag: str,
    note_text: str,
    generated_status: str,
    generated_support: str,
    staff_name: str,
    knowbe_target: str,
    work_start_time="",
    work_end_time="",
    work_break_time="0",
    work_memo="",
) -> PersonItem:
    return PersonItem(
        name=norm(resident_name),
        service=_normalize_service_for_app(service_type, knowbe_target),
        start=normalize_hhmm(start_time),
        end=normalize_hhmm(end_time),
        meal=_normalize_meal_for_app(meal_flag),
        note=norm(note_text),
        user_state=norm(generated_status),
        staff_note=norm(generated_support),
        staff_name=norm(staff_name),
        staff_mark="〇",
        work_start=str(work_start_time).strip(),
        work_end=str(work_end_time).strip(),
        work_break=str(work_break_time).strip() or "0",
        work_memo=str(work_memo).strip(),
    )

def _find_work_time_inputs(dialog):
    start_el = None
    end_el = None

    try:
        start_el = dialog.find_element(By.CSS_SELECTOR, "input[name='workRecord.startTime']")
    except Exception:
        start_el = None

    try:
        end_el = dialog.find_element(By.CSS_SELECTOR, "input[name='workRecord.endTime']")
    except Exception:
        end_el = None

    if start_el is not None and end_el is not None:
        return start_el, end_el

    def _find_input_near_label(label_text: str):
        try:
            labels = dialog.find_elements(
                By.XPATH,
                f".//*[normalize-space(text())='{label_text}' or contains(normalize-space(.), '{label_text}')]"
            )
        except Exception:
            labels = []

        for lab in labels:
            try:
                if not lab.is_displayed():
                    continue
            except Exception:
                pass

            cur = lab
            for _ in range(6):
                try:
                    cur = cur.find_element(By.XPATH, "..")
                except Exception:
                    break

                try:
                    inputs = cur.find_elements(By.XPATH, ".//input")
                except Exception:
                    inputs = []

                visible_inputs = []
                for inp in inputs:
                    try:
                        if inp.is_displayed():
                            visible_inputs.append(inp)
                    except Exception:
                        pass

                if visible_inputs:
                    try:
                        visible_inputs = sorted(visible_inputs, key=lambda e: e.location["x"])
                    except Exception:
                        pass
                    return visible_inputs[0]

            try:
                cand = lab.find_element(By.XPATH, "./following::input[1]")
                if cand and cand.is_displayed():
                    return cand
            except Exception:
                pass

        return None

    if start_el is None:
        start_el = _find_input_near_label("作業開始時間")

    if end_el is None:
        end_el = _find_input_near_label("作業終了時間")

    return start_el, end_el

def open_work_record_dialog_from_row(driver, row) -> bool:
    """
    日々の記録の対象行から、作業時間モーダルを開く
    かなり強引に複数パターンを総当たりする完全版
    """
    def _dialog_has_work_inputs(dlg):
        if dlg is None:
            return False

        try:
            if dlg.find_elements(By.CSS_SELECTOR, "input[name='workRecord.startTime']"):
                return True
        except Exception:
            pass

        try:
            txt = (dlg.text or "").strip()
        except Exception:
            txt = ""

        return ("作業開始時間" in txt) or ("作業終了時間" in txt) or ("休憩時間" in txt)

    def _wait_work_dialog(timeout=6):
        t0 = time.time()
        while time.time() - t0 < timeout:
            dlg = get_top_dialog(driver)
            if _dialog_has_work_inputs(dlg):
                return True
            time.sleep(0.2)
        return False

    # まず行を中央へ
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row)
        time.sleep(0.2)
    except Exception:
        pass

    # 1. かなり具体的な候補
    xps = [
        ".//button[contains(., '作業時間')]",
        ".//*[contains(normalize-space(.), '作業時間')]/ancestor::button[1]",
        ".//*[@role='button' and contains(normalize-space(.), '作業時間')]",
        ".//*[@role='button' and contains(normalize-space(.), '作業')]",
        ".//*[contains(normalize-space(.), '作業')]/ancestor::*[@role='button'][1]",
        ".//*[@role='button' and contains(@id, 'workRecord')]",
        ".//*[@role='button' and contains(@name, 'workRecord')]",
        ".//button[.//*[contains(normalize-space(.), '作業')]]",
        ".//button[.//*[contains(normalize-space(.), '時間')]]",
        ".//td[last()]//button",
        ".//button",
        ".//*[@role='button']",
    ]

    tried = []

    for xp in xps:
        try:
            els = row.find_elements(By.XPATH, xp)
        except Exception:
            els = []

        for el in els:
            try:
                if not el.is_displayed():
                    continue
            except Exception:
                pass

            try:
                sig = (
                    (el.text or "").strip(),
                    el.get_attribute("id") or "",
                    el.get_attribute("name") or "",
                    el.get_attribute("class") or "",
                    el.get_attribute("aria-label") or "",
                )
                if sig in tried:
                    continue
                tried.append(sig)
            except Exception:
                pass

            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.1)
            except Exception:
                pass

            if safe_click(driver, el):
                time.sleep(0.4)
                if _wait_work_dialog(timeout=2.5):
                    print("[DEBUG] work record dialog opened by candidate button", flush=True)
                    return True

                # ダイアログじゃないものを開いた場合は ESC で閉じる
                try:
                    ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                    time.sleep(0.2)
                except Exception:
                    pass

    # 2. 行内の input / textarea / select の近くにある clickable 要素を総当たり
    try:
        clickables = row.find_elements(
            By.XPATH,
            ".//*[(self::button or @role='button' or self::svg or self::span or self::div)]"
        )
    except Exception:
        clickables = []

    for el in clickables[:80]:
        try:
            if not el.is_displayed():
                continue
        except Exception:
            continue

        try:
            txt = ((el.text or "") + " " + (el.get_attribute("aria-label") or "")).strip()
        except Exception:
            txt = ""

        # なるべく作業/時間に寄せる
        if txt and ("作業" not in txt and "時間" not in txt and "work" not in txt.lower()):
            continue

        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.05)
        except Exception:
            pass

        if safe_click(driver, el):
            time.sleep(0.4)
            if _wait_work_dialog(timeout=2.5):
                print("[DEBUG] work record dialog opened by fallback clickable", flush=True)
                return True

            try:
                ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                time.sleep(0.2)
            except Exception:
                pass

    print("[DEBUG] work record dialog open failed", flush=True)
    return False


def fill_work_record_section(driver, root, it):
    """
    作業時間ダイアログ内へ
      - 作業実施チェック
      - 作業開始時間
      - 作業終了時間
      - 休憩時間
      - メモ
    を入力する
    ※ 欄がない事業所でも落とさずスキップする
    """

    work_start = str(getattr(it, "work_start", "") or "").strip()
    work_end = str(getattr(it, "work_end", "") or "").strip()
    work_break = str(getattr(it, "work_break", "") or "").strip()
    work_memo = str(getattr(it, "work_memo", "") or "").strip()

    print("FILL_WORK_RECORD_SECTION ENTER", flush=True)
    print(f"[DEBUG] work_start={work_start!r}", flush=True)
    print(f"[DEBUG] work_end={work_end!r}", flush=True)
    print(f"[DEBUG] work_break={work_break!r}", flush=True)
    print(f"[DEBUG] work_memo={work_memo!r}", flush=True)

    # 最優先防御：作業開始/終了が両方そろっていないなら絶対スキップ
    if not work_start or not work_end:
        print("[DEBUG] work time empty -> skip", flush=True)
        return True

    def _find_break_input(dialog):
        try:
            el = dialog.find_element(By.CSS_SELECTOR, "input[name='workRecord.breakTime']")
            if el.is_displayed():
                return el
        except Exception:
            pass

        labels = []
        try:
            labels = dialog.find_elements(
                By.XPATH,
                ".//*[normalize-space(text())='休憩時間' or contains(normalize-space(.), '休憩時間')]"
            )
        except Exception:
            labels = []

        for lab in labels:
            cur = lab
            for _ in range(6):
                try:
                    cur = cur.find_element(By.XPATH, "..")
                except Exception:
                    break
                try:
                    inputs = cur.find_elements(By.XPATH, ".//input")
                except Exception:
                    inputs = []
                for inp in inputs:
                    try:
                        if inp.is_displayed():
                            return inp
                    except Exception:
                        pass

        try:
            inputs = dialog.find_elements(By.XPATH, ".//input")
        except Exception:
            inputs = []

        visible = []
        for inp in inputs:
            try:
                if inp.is_displayed():
                    visible.append(inp)
            except Exception:
                pass

        if len(visible) >= 3:
            try:
                visible = sorted(visible, key=lambda e: (e.location["y"], e.location["x"]))
            except Exception:
                pass
            return visible[2]

        return None

    def _find_memo_area(dialog):
        try:
            el = dialog.find_element(By.CSS_SELECTOR, "textarea[name='workRecord.memo']")
            if el.is_displayed():
                return el
        except Exception:
            pass

        try:
            areas = dialog.find_elements(By.TAG_NAME, "textarea")
        except Exception:
            areas = []

        for ta in areas:
            try:
                if ta.is_displayed():
                    return ta
            except Exception:
                pass

        return None

    try:
        worked_chk = root.find_element(By.CSS_SELECTOR, "input[name='workRecord.worked']")
        print(f"[DEBUG] worked checked before={worked_chk.is_selected()}", flush=True)
        if not worked_chk.is_selected():
            driver.execute_script("arguments[0].click();", worked_chk)
            time.sleep(0.3)
    except Exception as e:
        print(f"[DEBUG] worked checkbox not found or click fail={e}", flush=True)
        return True

    time.sleep(0.3)

    try:
        start_el, end_el = _find_work_time_inputs(root)
    except Exception as e:
        print(f"[DEBUG] _find_work_time_inputs error={e}", flush=True)
        return True

    print(f"[DEBUG] start_el found={start_el is not None}", flush=True)
    print(f"[DEBUG] end_el found={end_el is not None}", flush=True)

    if start_el is None or end_el is None:
        print("[DEBUG] time inputs not found -> skip", flush=True)
        return True

    try:
        set_input_value(driver, start_el, work_start)
        time.sleep(0.15)
    except Exception as e:
        print(f"[DEBUG] start input error={e}", flush=True)

    try:
        set_input_value(driver, end_el, work_end)
        time.sleep(0.15)
    except Exception as e:
        print(f"[DEBUG] end input error={e}", flush=True)

    try:
        break_el = _find_break_input(root)
        print(f"[DEBUG] break_el found={break_el is not None}", flush=True)
        if break_el is not None:
            set_input_value(driver, break_el, work_break or "0")
            time.sleep(0.15)
    except Exception as e:
        print(f"[DEBUG] break input error={e}", flush=True)

    try:
        memo_el = _find_memo_area(root)
        print(f"[DEBUG] memo_el found={memo_el is not None}", flush=True)
        if memo_el is not None:
            set_input_value(driver, memo_el, work_memo)
            time.sleep(0.1)
    except Exception as e:
        print(f"[DEBUG] memo input error={e}", flush=True)

    return True

def send_one_record_from_app(
    target_date,
    resident_name,
    service_type,
    start_time,
    end_time,
    meal_flag,
    note_text,
    generated_status,
    generated_support,
    staff_name,
    knowbe_target,
    login_username,
    login_password,
    work_start_time,
    work_end_time,
    work_break_time,
    work_memo="",
    send_user_status=True,
    send_staff_comment=True,
):
    print("RUN_ASSISTANCE_VERSION = 2026-03-29-row-root-worktime-final2", flush=True)
    """
    appから1件だけ渡されたデータを Knowbe に送る完全版
    作業時間は 利用実績モーダル(process_report_edit) 側で入力する。
    日々の記録ページでは、作業時間ダイアログは開かない。
    """
    if not target_date:
        raise RuntimeError("[FATAL] target_date が空ある")
    if not resident_name:
        raise RuntimeError("[FATAL] resident_name が空ある")
    if not start_time:
        raise RuntimeError("[FATAL] start_time が空ある")
    if not end_time:
        raise RuntimeError("[FATAL] end_time が空ある")
    if not staff_name:
        raise RuntimeError("[FATAL] staff_name が空ある")

    m = re.match(r"^\s*(\d{4})-(\d{2})-(\d{2})\s*$", str(target_date))
    if not m:
        raise RuntimeError(f"[FATAL] target_date形式が不正ある: {target_date}")

    y = int(m.group(1))
    mo = int(m.group(2))
    d = int(m.group(3))

    it = _build_single_item_from_app(
        resident_name=resident_name,
        service_type=service_type,
        start_time=start_time,
        end_time=end_time,
        meal_flag=meal_flag,
        note_text=note_text,
        generated_status=generated_status,
        generated_support=generated_support,
        staff_name=staff_name,
        knowbe_target=knowbe_target,
        work_start_time=work_start_time,
        work_end_time=work_end_time,
        work_break_time=work_break_time,
        work_memo=work_memo,
    )

    print("[STEP] build_chrome_driver start", flush=True)
    driver = build_chrome_driver()
    print("[STEP] driver created", flush=True)

    try:
        print("[STEP] goto_report_daily start", flush=True)
        goto_report_daily(driver)
        print("[STEP] goto_report_daily done", flush=True)

        print("[STEP] manual_login_wait start", flush=True)
        manual_login_wait(driver, login_username, login_password)
        print("[STEP] manual_login_wait done", flush=True)

        print("[STEP] goto_report_date start", flush=True)
        goto_report_date(driver, y, mo, d)
        print("[STEP] goto_report_date done", flush=True)

        # ① 利用実績
        # ここで開始時間・終了時間・休憩時間・作業メモまで入力する
        log(f"🏃 app単発 実績処理: {it.name}")
        ok = process_report_edit(driver, it)
        if not ok:
            raise RuntimeError(f"[FATAL] 利用実績の入力失敗ある: {it.name}")

        # ② 日々の記録ページ
        print("[STEP] open_daily_record_page start", flush=True)
        if not open_daily_record_page(driver):
            raise RuntimeError("[FATAL] 日々の記録ページへ行けないある")
        print("[STEP] open_daily_record_page done", flush=True)

        if not click_daily_edit_button(driver):
            raise RuntimeError("[FATAL] 日々の記録の編集ボタンが押せないある")

        row = _find_daily_record_row_by_name(driver, it.name)
        if row is None:
            dump_debug(driver, f"daily_row_not_found_{it.name}")
            raise RuntimeError(f"[FATAL] 日々の記録 行発見失敗ある: {it.name}")

        work_label = _daily_record_work_label(it)

        # ③ 作業欄
        if not _set_daily_work_for_row(driver, row, work_label):
            dump_debug(driver, f"daily_work_set_fail_{it.name}")
            raise RuntimeError(f"[FATAL] 作業欄の選択失敗ある: {it.name}")

        # ③.5 日誌生成ルール適用
        result = generate_journal_from_memo(
            memo=generated_status,
            work_label=work_label,
            start_time=start_time,
            end_time=end_time,
        )
        
        generated_status = result["user_state"]
        generated_support = result["staff_note"]

        # ④ 利用者状態 / 職員考察
        if not _set_daily_textareas_for_row(
            driver,
            row,
            generated_status,
            generated_support,
            send_user_status=bool(send_user_status),
            send_staff_comment=bool(send_staff_comment),
        ):
            dump_debug(driver, f"daily_textarea_fail_{it.name}")
            raise RuntimeError(f"[FATAL] 日々の記録 textarea 入力失敗ある: {it.name}")

        # ⑤ 記録者
        if not _set_daily_recorder_for_row(driver, row, staff_name):
            dump_debug(driver, f"daily_recorder_fail_{it.name}")
            raise RuntimeError(f"[FATAL] 記録者選択失敗ある: {it.name}")

        # ⑥ 日々の記録ページ全体の保存
        # ※ 作業時間ダイアログはここでは開かない
        if not click_daily_save_button(driver):
            dump_debug(driver, f"daily_save_fail_{it.name}")
            raise RuntimeError(f"[FATAL] 日々の記録 保存失敗ある: {it.name}")

        log("🎊 app単発送信 完了ある！")
        return True

    finally:
        try:
            driver.quit()
        except Exception:
            pass

def fetch_support_record_page_text(driver):
    return driver.find_element(By.TAG_NAME, "body").text


def enter_edit_mode(driver):
    import time
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    btns = driver.find_elements(By.TAG_NAME, "button")

    for b in btns:
        txt = (b.text or "").strip()
        if txt == "編集":
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b)
            time.sleep(0.3)
            if not safe_click(driver, b):
                driver.execute_script("arguments[0].click();", b)
            break
    else:
        raise RuntimeError("編集ボタンが見つからない")

    WebDriverWait(driver, 10).until(
        lambda d: any("保存" in ((x.text or "").strip()) for x in d.find_elements(By.TAG_NAME, "button"))
    )

    time.sleep(0.5)
    print("[FIX] 編集モードON", flush=True)
    return True

def update_day_fields(driver, user_state: str, staff_note: str):
    dlg = driver.find_elements(By.CSS_SELECTOR, "[role='dialog']")[-1]

    textareas = dlg.find_elements(By.TAG_NAME, "textarea")

    if len(textareas) < 2:
        raise RuntimeError("textarea不足（モーダル取得ミス）")

    # 利用者状態
    set_input_value(driver, textareas[0], user_state)

    # 職員考察
    set_input_value(driver, textareas[1], staff_note)

    print("[FIX] fields updated", flush=True)


def save_day(driver):
    dlg = driver.find_elements(By.CSS_SELECTOR, "[role='dialog']")[-1]

    # 👉 保存ボタンを探す（編集→保存に変わるやつ）
    buttons = dlg.find_elements(By.TAG_NAME, "button")

    saved = False

    for b in buttons:
        try:
            txt = (b.text or "").strip()

            if "保存" in txt or "更新" in txt:
                if safe_click(driver, b):
                    saved = True
                    break
        except Exception:
            continue

    if not saved:
        raise RuntimeError("保存ボタンが見つからない")

    # 👉 モーダルが閉じるまで待つ
    WebDriverWait(driver, 10).until(
        lambda d: len(d.find_elements(By.CSS_SELECTOR, "[role='dialog']")) == 0
    )

    print("[FIX] saved & modal closed", flush=True)
    time.sleep(0.5)

def main():
    # =========================================
    # app単発モード
    # =========================================
    if os.environ.get("KB_SINGLE_MODE", "") == "1":
        login_username, login_password = get_knowbe_login_credentials(
            company_id=os.environ.get("KB_COMPANY_ID", "")
        )

        if not login_username or not login_password:
            raise RuntimeError("[FATAL] KB_LOGIN_USERNAME / KB_LOGIN_PASSWORD が空ある")

        send_one_record_from_app(
            target_date=os.environ.get("KB_TARGET_DATE", ""),
            resident_name=os.environ.get("KB_RESIDENT_NAME", ""),
            service_type=os.environ.get("KB_SERVICE_TYPE", ""),
            start_time=os.environ.get("KB_START_TIME", ""),
            end_time=os.environ.get("KB_END_TIME", ""),
            meal_flag=os.environ.get("KB_MEAL_FLAG", ""),
            note_text=os.environ.get("KB_NOTE_TEXT", ""),
            generated_status=os.environ.get("KB_GENERATED_STATUS", ""),
            generated_support=os.environ.get("KB_GENERATED_SUPPORT", ""),
            staff_name=os.environ.get("KB_STAFF_NAME", ""),
            knowbe_target=os.environ.get("KB_KNOWBE_TARGET", ""),
            login_username=login_username,
            login_password=login_password,
            work_start_time=os.environ.get("KB_WORK_START_TIME", ""),
            work_end_time=os.environ.get("KB_WORK_END_TIME", ""),
            work_break_time=os.environ.get("KB_WORK_BREAK_TIME", "0"),
            work_memo=os.environ.get("KB_WORK_MEMO", ""),
        )
        return

    # =========================================
    # 旧：Excel一括モード
    # =========================================
    excel_path = os.environ.get("EXCEL_PATH", EXCEL_PATH_DEFAULT)
    if not os.path.exists(excel_path):
        raise RuntimeError(f"[FATAL] Excelが見つからないある: {excel_path}")

    log("エクセルを熟読中ある...")
    y, m, d, items = read_sheet_data(excel_path)

    targets = [
        it for it in items
        if (it.service or "").strip() in ("通所", "施設外就労")
        and (it.start or "").strip()
        and (it.end or "").strip()
    ]
    log(f"🚀 {len(targets)}名の処理を開始するある！ ターゲット: {y}/{m}/{d}")

    read_treaty_check(excel_path)

    driver = build_chrome_driver()

    try:
        login_username, login_password = get_knowbe_login_credentials()

        if not login_username or not login_password:
            raise RuntimeError("[FATAL] KB_LOGIN_USERNAME / KB_LOGIN_PASSWORD が空ある")

        goto_report_daily(driver)
        manual_login_wait(driver, login_username, login_password)
        goto_report_date(driver, y, m, d)

        any_ok = False
        for it in targets:
            read_treaty_check(excel_path)
            log(f"🏃 {it.name} 実績処理")
            ok = process_report_edit(driver, it)
            if not ok:
                log(f"[WARN] 実績入力失敗→次へ: {it.name}")
            else:
                any_ok = True

        if not any_ok:
            log("⚠️ 実績が1件も成功してないある（dbgを確認してくれある）")
            dump_debug(driver, "no_success_report")

        if any_ok:
            run_daily_records(driver, excel_path, items, targets, y, m, d)

        log("🎊 全行程完了ある！")

    finally:
        try:
            driver.quit()
        except Exception:
            pass
